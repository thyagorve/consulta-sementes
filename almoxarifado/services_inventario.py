# almoxarifado/services_inventario.py

from __future__ import annotations

import csv
import io
import re
import unicodedata
import xml.etree.ElementTree as ET
from datetime import datetime
from decimal import Decimal, InvalidOperation

from django.db import transaction
from openpyxl import load_workbook

from .models import (
    DadosValidadeItem,
    ImportacaoInventario,
    Item,
    LinhaComparacaoInventario,
)


# ============================================================
# COLUNAS ACEITAS NA PLANILHA
# ============================================================

ALIASES = {
    'codigo': {
        'codigo',
        'cod',
        'codproduto',
        'codigoitem',
        'coditem',
        'cprod',
    },
    'nome': {
        'nome',
        'produto',
        'item',
        'descricao',
        'xprod',
    },
    'quantidade': {
        'quantidade',
        'qtd',
        'saldo',
        'estoque',
        'qcom',
    },
    'unidade': {
        'unidade',
        'un',
        'um',
        'ucom',
    },
    'lote': {
        'lote',
        'nlote',
    },
    'fabricacao': {
        'fabricacao',
        'datafabricacao',
        'dtfabricacao',
        'dfab',
    },
    'vencimento': {
        'vencimento',
        'validade',
        'datavencimento',
        'dtvencimento',
        'dval',
    },
    'localizacao': {
        'localizacao',
        'local',
        'endereco',
    },
    'departamento': {
        'departamento',
        'dept',
        'setor',
    },
    'estoque_minimo': {
        'estoqueminimo',
        'minimo',
        'estoque_min',
        'min',
    },
    'fornecedor': {
        'fornecedor',
        'fabricante',
    },
}


# ============================================================
# FUNÇÕES BÁSICAS
# ============================================================

def normalizar_texto(valor):
    texto = str(
        valor or ''
    ).strip()

    texto = ''.join(
        caractere
        for caractere in unicodedata.normalize(
            'NFKD',
            texto,
        )
        if not unicodedata.combining(
            caractere
        )
    )

    return texto.lower()


def normalizar_coluna(valor):
    texto = normalizar_texto(
        valor
    )

    return re.sub(
        r'[^a-z0-9]+',
        '',
        texto,
    )


def mapear_cabecalho(cabecalho):
    mapa = {}

    normalizados = [
        normalizar_coluna(
            item
        )
        for item in cabecalho
    ]

    for campo, aliases in (
        ALIASES.items()
    ):

        aliases_normalizados = {
            normalizar_coluna(
                alias
            )
            for alias in aliases
        }

        for indice, nome in enumerate(
            normalizados
        ):

            if nome in aliases_normalizados:
                mapa[campo] = indice
                break

    return mapa


def decimal_seguro(
    valor,
    default=Decimal('0'),
):
    if valor is None:
        return default

    if valor == '':
        return default

    if isinstance(
        valor,
        Decimal,
    ):
        return valor

    texto = str(
        valor
    ).strip().replace(
        ' ',
        '',
    )

    if ',' in texto and '.' in texto:
        texto = (
            texto
            .replace(
                '.',
                '',
            )
            .replace(
                ',',
                '.',
            )
        )

    elif ',' in texto:
        texto = texto.replace(
            ',',
            '.',
        )

    try:
        return Decimal(
            texto
        )

    except (
        InvalidOperation,
        ValueError,
        TypeError,
    ):
        return default


def data_segura(valor):
    if not valor:
        return None

    if (
        hasattr(
            valor,
            'date',
        )
        and not isinstance(
            valor,
            str,
        )
    ):
        try:
            return valor.date()
        except Exception:
            pass

    if (
        hasattr(
            valor,
            'year',
        )
        and hasattr(
            valor,
            'month',
        )
        and hasattr(
            valor,
            'day',
        )
    ):
        return valor

    texto = str(
        valor
    ).strip()

    formatos = (
        '%d/%m/%Y',
        '%Y-%m-%d',
        '%d-%m-%Y',
        '%d/%m/%y',
        '%Y%m%d',
    )

    for formato in formatos:
        try:
            return datetime.strptime(
                texto[:10],
                formato,
            ).date()

        except ValueError:
            continue

    return None


def _valor_linha(
    linha,
    mapa,
    campo,
    default='',
):
    indice = mapa.get(
        campo
    )

    if indice is None:
        return default

    if indice >= len(
        linha
    ):
        return default

    return linha[
        indice
    ]


# ============================================================
# LEITURA DA PLANILHA
# ============================================================

def ler_planilha(upload):
    nome = (
        getattr(
            upload,
            'name',
            '',
        )
        or ''
    ).lower()

    conteudo = upload.read()

    if nome.endswith(
        '.csv'
    ):

        texto = conteudo.decode(
            'utf-8-sig',
            errors='replace',
        )

        amostra = texto[
            :4096
        ]

        delimitador = (
            ';'
            if amostra.count(';')
            >= amostra.count(',')
            else ','
        )

        linhas = list(
            csv.reader(
                io.StringIO(
                    texto
                ),
                delimiter=delimitador,
            )
        )

    elif (
        nome.endswith(
            '.xlsx'
        )
        or nome.endswith(
            '.xlsm'
        )
    ):

        workbook = load_workbook(
            io.BytesIO(
                conteudo
            ),
            data_only=True,
            read_only=True,
        )

        planilha = (
            workbook.active
        )

        linhas = [
            list(
                linha
            )
            for linha in planilha.iter_rows(
                values_only=True
            )
        ]

    else:
        raise ValueError(
            'Envie um arquivo '
            '.xlsx, .xlsm ou .csv.'
        )

    if not linhas:
        return []

    mapa = mapear_cabecalho(
        linhas[0]
    )

    if 'codigo' not in mapa:
        raise ValueError(
            'A planilha precisa '
            'possuir uma coluna '
            'de código.'
        )

    if (
        'quantidade'
        not in mapa
    ):
        raise ValueError(
            'A planilha precisa '
            'possuir uma coluna '
            'de quantidade.'
        )

    resultado = []

    for linha in linhas[
        1:
    ]:

        codigo = str(
            _valor_linha(
                linha,
                mapa,
                'codigo',
                '',
            )
            or ''
        ).strip()

        if not codigo:
            continue

        fabricacao = (
            data_segura(
                _valor_linha(
                    linha,
                    mapa,
                    'fabricacao',
                    None,
                )
            )
        )

        vencimento = (
            data_segura(
                _valor_linha(
                    linha,
                    mapa,
                    'vencimento',
                    None,
                )
            )
        )

        resultado.append({
            'codigo':
                codigo,

            'nome':
                str(
                    _valor_linha(
                        linha,
                        mapa,
                        'nome',
                        '',
                    )
                    or ''
                ).strip(),

            'quantidade':
                str(
                    decimal_seguro(
                        _valor_linha(
                            linha,
                            mapa,
                            'quantidade',
                            0,
                        )
                    )
                ),

            'unidade':
                str(
                    _valor_linha(
                        linha,
                        mapa,
                        'unidade',
                        '',
                    )
                    or ''
                ).strip(),

            'lote':
                str(
                    _valor_linha(
                        linha,
                        mapa,
                        'lote',
                        '',
                    )
                    or ''
                ).strip(),

            'data_fabricacao':
                (
                    fabricacao
                    .isoformat()
                    if fabricacao
                    else ''
                ),

            'data_vencimento':
                (
                    vencimento
                    .isoformat()
                    if vencimento
                    else ''
                ),

            'localizacao':
                str(
                    _valor_linha(
                        linha,
                        mapa,
                        'localizacao',
                        '',
                    )
                    or ''
                ).strip(),

            'departamento':
                str(
                    _valor_linha(
                        linha,
                        mapa,
                        'departamento',
                        '',
                    )
                    or ''
                ).strip(),

            'estoque_minimo':
                str(
                    decimal_seguro(
                        _valor_linha(
                            linha,
                            mapa,
                            'estoque_minimo',
                            0,
                        )
                    )
                ),

            'fornecedor':
                str(
                    _valor_linha(
                        linha,
                        mapa,
                        'fornecedor',
                        '',
                    )
                    or ''
                ).strip(),
        })

    return resultado


# ============================================================
# LEITURA DE NF-e XML
# ============================================================

def _tag_local(
    elemento,
    nome,
):
    for child in elemento.iter():

        if (
            child.tag
            .split('}')[-1]
            == nome
        ):
            return child.text

    return None


def ler_nfe_xml(upload):
    conteudo = upload.read()

    try:
        root = ET.fromstring(
            conteudo
        )

    except ET.ParseError as exc:
        raise ValueError(
            f'XML inválido: {exc}'
        )

    itens = []

    for det in root.iter():

        if (
            det.tag
            .split('}')[-1]
            != 'det'
        ):
            continue

        produto = None

        for child in det:

            if (
                child.tag
                .split('}')[-1]
                == 'prod'
            ):
                produto = child
                break

        if produto is None:
            continue

        def get(nome):
            return (
                _tag_local(
                    produto,
                    nome,
                )
                or ''
            )

        dados = {
            'codigo':
                str(
                    get(
                        'cProd'
                    )
                ).strip(),

            'nome':
                str(
                    get(
                        'xProd'
                    )
                ).strip(),

            'quantidade':
                str(
                    decimal_seguro(
                        get(
                            'qCom'
                        )
                    )
                ),

            'unidade':
                str(
                    get(
                        'uCom'
                    )
                ).strip(),

            'lote':
                '',

            'data_fabricacao':
                '',

            'data_vencimento':
                '',

            'localizacao':
                '',

            'departamento':
                '',

            'estoque_minimo':
                '0',

            'fornecedor':
                '',

            'ean':
                str(
                    get(
                        'cEAN'
                    )
                ).strip(),

            'ncm':
                str(
                    get(
                        'NCM'
                    )
                ).strip(),

            'cfop':
                str(
                    get(
                        'CFOP'
                    )
                ).strip(),

            'valor_unitario':
                str(
                    decimal_seguro(
                        get(
                            'vUnCom'
                        )
                    )
                ),
        }

        for rastro in produto.iter():

            if (
                rastro.tag
                .split('}')[-1]
                != 'rastro'
            ):
                continue

            dados['lote'] = str(
                _tag_local(
                    rastro,
                    'nLote',
                )
                or ''
            ).strip()

            fabricacao = (
                data_segura(
                    _tag_local(
                        rastro,
                        'dFab',
                    )
                )
            )

            vencimento = (
                data_segura(
                    _tag_local(
                        rastro,
                        'dVal',
                    )
                )
            )

            dados[
                'data_fabricacao'
            ] = (
                fabricacao.isoformat()
                if fabricacao
                else ''
            )

            dados[
                'data_vencimento'
            ] = (
                vencimento.isoformat()
                if vencimento
                else ''
            )

            break

        if dados[
            'codigo'
        ]:
            itens.append(
                dados
            )

    if not itens:
        raise ValueError(
            'Nenhum produto '
            'foi encontrado no XML.'
        )

    return itens


# ============================================================
# NORMALIZAÇÃO DO MODEL ITEM
# ============================================================

def _campos_item():
    return {
        campo.name
        for campo
        in Item._meta.get_fields()
        if getattr(
            campo,
            'concrete',
            False,
        )
    }


def _campo_item(
    nome
):
    try:
        return (
            Item._meta
            .get_field(
                nome
            )
        )

    except Exception:
        return None


def _limite_charfield(
    nome,
    valor,
):
    if valor is None:
        return valor

    campo = _campo_item(
        nome
    )

    if not campo:
        return valor

    max_length = getattr(
        campo,
        'max_length',
        None,
    )

    if not max_length:
        return valor

    texto = str(
        valor
    ).strip()

    if (
        len(
            texto
        )
        <= max_length
    ):
        return texto

    raise ValueError(
        f'O valor "{texto}" '
        f'para o campo "{nome}" '
        f'possui {len(texto)} '
        f'caracteres, mas o sistema '
        f'aceita no máximo '
        f'{max_length}.'
    )


def _normalizar_choice(
    campo_nome,
    valor,
):
    """
    Aceita tanto o código quanto
    o texto exibido no choices.

    Ex.:
    "Pacote" -> "PCT"
    "Produção" -> código real do
    departamento, se existir no model.
    """
    if valor is None:
        return ''

    texto_original = str(
        valor
    ).strip()

    if not texto_original:
        return ''

    campo = _campo_item(
        campo_nome
    )

    if not campo:
        return texto_original

    choices = list(
        getattr(
            campo,
            'choices',
            None,
        )
        or []
    )

    if not choices:
        return texto_original

    alvo = normalizar_texto(
        texto_original
    )

    # Primeiro tenta pelo código.
    for codigo, label in choices:

        if (
            normalizar_texto(
                codigo
            )
            == alvo
        ):
            return str(
                codigo
            )

    # Depois tenta pelo label.
    for codigo, label in choices:

        if (
            normalizar_texto(
                label
            )
            == alvo
        ):
            return str(
                codigo
            )

    # Compatibilidade com abreviações comuns.
    aliases = {
        'unidade': {
            'und': 'UN',
            'unidades': 'UN',
            'caixa': 'CX',
            'caixas': 'CX',
            'pacote': 'PCT',
            'pacotes': 'PCT',
            'quilograma': 'KG',
            'quilogramas': 'KG',
            'quilo': 'KG',
            'quilos': 'KG',
            'litro': 'L',
            'litros': 'L',
            'mililitro': 'ML',
            'mililitros': 'ML',
            'metro': 'M',
            'metros': 'M',
            'centimetro': 'CM',
            'centimetros': 'CM',
            'pares': 'PAR',
            'duzia': 'DZ',
            'duzias': 'DZ',
            'rolo': 'RL',
            'rolos': 'RL',
            'folha': 'FL',
            'folhas': 'FL',
        }
    }

    alias = (
        aliases
        .get(
            campo_nome,
            {},
        )
        .get(
            alvo
        )
    )

    if alias:
        # Só aceita se o alias
        # também existir nos choices.
        for codigo, label in choices:

            if str(
                codigo
            ) == alias:
                return alias

    raise ValueError(
        f'O valor "{texto_original}" '
        f'não é uma opção válida '
        f'para "{campo_nome}".'
    )


def normalizar_unidade(
    valor
):
    return _normalizar_choice(
        'unidade',
        valor,
    )


def normalizar_departamento(
    valor
):
    return _normalizar_choice(
        'departamento',
        valor,
    )


# ============================================================
# CHAVE DO INVENTÁRIO
# ============================================================

def chave_item(
    codigo,
    lote,
    modo,
):
    codigo = str(
        codigo or ''
    ).strip().upper()

    lote = str(
        lote or ''
    ).strip().upper()

    if modo == 'COM_LOTE':
        return (
            f'{codigo}::{lote}'
        )

    return codigo


# ============================================================
# CRIAÇÃO/ATUALIZAÇÃO DO ITEM
# ============================================================

def _kwargs_item(
    dados
):
    campos = _campos_item()

    unidade = (
        normalizar_unidade(
            dados.get(
                'unidade'
            )
        )
        if dados.get(
            'unidade'
        )
        else 'UN'
    )

    departamento = (
        normalizar_departamento(
            dados.get(
                'departamento'
            )
        )
        if dados.get(
            'departamento'
        )
        else ''
    )

    candidato = {
        'codigo':
            dados.get(
                'codigo',
                '',
            ),

        'nome':
            (
                dados.get(
                    'nome'
                )
                or dados.get(
                    'codigo'
                )
                or 'ITEM IMPORTADO'
            ),

        'descricao':
            (
                dados.get(
                    'nome'
                )
                or ''
            ),

        'quantidade':
            decimal_seguro(
                dados.get(
                    'quantidade'
                )
            ),

        'unidade':
            unidade,

        'lote':
            (
                dados.get(
                    'lote'
                )
                or ''
            ),

        'localizacao':
            (
                dados.get(
                    'localizacao'
                )
                or ''
            ),

        'departamento':
            departamento,

        'estoque_minimo':
            decimal_seguro(
                dados.get(
                    'estoque_minimo'
                )
            ),

        'fornecedor':
            (
                dados.get(
                    'fornecedor'
                )
                or ''
            ),

        'ativo':
            True,
    }

    resultado = {}

    for campo, valor in (
        candidato.items()
    ):

        if campo not in campos:
            continue

        # Não envia departamento vazio
        # se o campo tiver default no model.
        if (
            campo
            == 'departamento'
            and valor == ''
        ):
            continue

        if isinstance(
            valor,
            str,
        ):
            valor = (
                _limite_charfield(
                    campo,
                    valor,
                )
            )

        resultado[
            campo
        ] = valor

    return resultado


def atualizar_validade(
    item,
    dados,
):
    fabricacao = data_segura(
        dados.get(
            'data_fabricacao'
        )
    )

    vencimento = data_segura(
        dados.get(
            'data_vencimento'
        )
    )

    if (
        fabricacao
        and vencimento
        and fabricacao > vencimento
    ):
        raise ValueError(
            'A data de fabricação '
            'não pode ser posterior '
            'ao vencimento.'
        )

    objeto, _ = (
        DadosValidadeItem
        .objects
        .get_or_create(
            item=item
        )
    )

    objeto.data_fabricacao = (
        fabricacao
    )

    objeto.data_vencimento = (
        vencimento
    )

    objeto.save(
        update_fields=[
            'data_fabricacao',
            'data_vencimento',
            'atualizado_em',
        ]
    )


# ============================================================
# CRIAÇÃO DA COMPARAÇÃO
# ============================================================

def criar_comparacao(
    upload,
    usuario,
    modo='COM_LOTE',
    tipo='PLANILHA',
):
    modo = (
        modo
        if modo in {
            'COM_LOTE',
            'SEM_LOTE',
        }
        else 'COM_LOTE'
    )

    tipo = (
        tipo
        if tipo in {
            'PLANILHA',
            'NFE_XML',
        }
        else 'PLANILHA'
    )

    if tipo == 'NFE_XML':
        dados_arquivo = (
            ler_nfe_xml(
                upload
            )
        )
    else:
        dados_arquivo = (
            ler_planilha(
                upload
            )
        )

    importacao = (
        ImportacaoInventario
        .objects
        .create(
            tipo=tipo,
            modo_comparacao=modo,
            nome_arquivo=getattr(
                upload,
                'name',
                '',
            ),
            criado_por=usuario,
            status='PROCESSANDO',
        )
    )

    arquivo_por_chave = {}

    for dados in dados_arquivo:

        chave = chave_item(
            dados.get(
                'codigo'
            ),
            dados.get(
                'lote'
            ),
            modo,
        )

        if chave in arquivo_por_chave:

            anterior = (
                arquivo_por_chave[
                    chave
                ]
            )

            anterior[
                'quantidade'
            ] = str(
                decimal_seguro(
                    anterior.get(
                        'quantidade'
                    )
                )
                +
                decimal_seguro(
                    dados.get(
                        'quantidade'
                    )
                )
            )

        else:
            arquivo_por_chave[
                chave
            ] = dados

    sistema_por_chave = {}

    for item in (
        Item.objects
        .filter(
            ativo=True
        )
        .iterator()
    ):

        chave = chave_item(
            getattr(
                item,
                'codigo',
                '',
            ),
            getattr(
                item,
                'lote',
                '',
            ),
            modo,
        )

        sistema_por_chave.setdefault(
            chave,
            [],
        ).append(
            item
        )

    chaves = (
        set(
            arquivo_por_chave
        )
        |
        set(
            sistema_por_chave
        )
    )

    objetos = []

    resumo = {
        'iguais': 0,
        'saldo_divergente': 0,
        'unidade_divergente': 0,
        'dados_divergentes': 0,
        'so_sistema': 0,
        'so_arquivo': 0,
        'ambiguos': 0,
    }

    for chave in sorted(
        chaves
    ):

        arquivo = (
            arquivo_por_chave
            .get(
                chave
            )
        )

        encontrados = (
            sistema_por_chave
            .get(
                chave,
                [],
            )
        )

        # ----------------------------------------------------
        # SOMENTE NO SISTEMA
        # ----------------------------------------------------
        if arquivo is None:

            for item in encontrados:

                objetos.append(
                    LinhaComparacaoInventario(
                        importacao=importacao,
                        item=item,
                        chave=chave,
                        codigo=str(
                            getattr(
                                item,
                                'codigo',
                                '',
                            )
                            or ''
                        ),
                        lote=str(
                            getattr(
                                item,
                                'lote',
                                '',
                            )
                            or ''
                        ),
                        quantidade_sistema=(
                            decimal_seguro(
                                getattr(
                                    item,
                                    'quantidade',
                                    0,
                                )
                            )
                        ),
                        unidade_sistema=str(
                            getattr(
                                item,
                                'unidade',
                                '',
                            )
                            or ''
                        ),
                        status='SO_SISTEMA',
                    )
                )

                resumo[
                    'so_sistema'
                ] += 1

            continue

        # ----------------------------------------------------
        # SOMENTE NO ARQUIVO
        # ----------------------------------------------------
        if not encontrados:

            objetos.append(
                LinhaComparacaoInventario(
                    importacao=importacao,
                    chave=chave,
                    codigo=arquivo.get(
                        'codigo',
                        '',
                    ),
                    lote=arquivo.get(
                        'lote',
                        '',
                    ),
                    nome_arquivo=arquivo.get(
                        'nome',
                        '',
                    ),
                    quantidade_arquivo=(
                        decimal_seguro(
                            arquivo.get(
                                'quantidade'
                            )
                        )
                    ),
                    unidade_arquivo=(
                        arquivo.get(
                            'unidade',
                            '',
                        )
                    ),
                    dados_arquivo=arquivo,
                    status='SO_ARQUIVO',
                )
            )

            resumo[
                'so_arquivo'
            ] += 1

            continue

        # ----------------------------------------------------
        # AMBÍGUO
        # ----------------------------------------------------
        if len(
            encontrados
        ) > 1:

            objetos.append(
                LinhaComparacaoInventario(
                    importacao=importacao,
                    chave=chave,
                    codigo=arquivo.get(
                        'codigo',
                        '',
                    ),
                    lote=arquivo.get(
                        'lote',
                        '',
                    ),
                    nome_arquivo=arquivo.get(
                        'nome',
                        '',
                    ),
                    quantidade_arquivo=(
                        decimal_seguro(
                            arquivo.get(
                                'quantidade'
                            )
                        )
                    ),
                    unidade_arquivo=(
                        arquivo.get(
                            'unidade',
                            '',
                        )
                    ),
                    dados_arquivo=arquivo,
                    status='AMBIGUO',
                    mensagem=(
                        f'{len(encontrados)} '
                        f'itens do sistema '
                        f'possuem a mesma chave.'
                    ),
                )
            )

            resumo[
                'ambiguos'
            ] += 1

            continue

        # ----------------------------------------------------
        # ITEM ÚNICO
        # ----------------------------------------------------
        item = encontrados[
            0
        ]

        qtd_sistema = (
            decimal_seguro(
                getattr(
                    item,
                    'quantidade',
                    0,
                )
            )
        )

        qtd_arquivo = (
            decimal_seguro(
                arquivo.get(
                    'quantidade'
                )
            )
        )

        unidade_sistema = str(
            getattr(
                item,
                'unidade',
                '',
            )
            or ''
        ).strip().upper()

        try:
            unidade_arquivo = (
                normalizar_unidade(
                    arquivo.get(
                        'unidade'
                    )
                )
                if arquivo.get(
                    'unidade'
                )
                else ''
            )
        except ValueError:
            unidade_arquivo = str(
                arquivo.get(
                    'unidade',
                    '',
                )
                or ''
            ).strip().upper()

        if (
            qtd_sistema
            == qtd_arquivo
            and (
                not unidade_arquivo
                or unidade_sistema
                == unidade_arquivo
            )
        ):
            status = 'IGUAL'

            resumo[
                'iguais'
            ] += 1

        elif (
            qtd_sistema
            != qtd_arquivo
            and unidade_arquivo
            and unidade_sistema
            != unidade_arquivo
        ):
            status = (
                'DADOS_DIVERGENTES'
            )

            resumo[
                'dados_divergentes'
            ] += 1

        elif (
            qtd_sistema
            != qtd_arquivo
        ):
            status = (
                'SALDO_DIVERGENTE'
            )

            resumo[
                'saldo_divergente'
            ] += 1

        else:
            status = (
                'UNIDADE_DIVERGENTE'
            )

            resumo[
                'unidade_divergente'
            ] += 1

        objetos.append(
            LinhaComparacaoInventario(
                importacao=importacao,
                item=item,
                chave=chave,
                codigo=arquivo.get(
                    'codigo',
                    '',
                ),
                lote=arquivo.get(
                    'lote',
                    '',
                ),
                nome_arquivo=arquivo.get(
                    'nome',
                    '',
                ),
                quantidade_sistema=(
                    qtd_sistema
                ),
                quantidade_arquivo=(
                    qtd_arquivo
                ),
                unidade_sistema=(
                    unidade_sistema
                ),
                unidade_arquivo=(
                    unidade_arquivo
                ),
                dados_arquivo=arquivo,
                status=status,
            )
        )

    LinhaComparacaoInventario.objects.bulk_create(
        objetos,
        batch_size=500,
    )

    importacao.total_linhas = len(
        objetos
    )

    importacao.resumo = resumo

    importacao.status = (
        'PRONTA'
    )

    importacao.save(
        update_fields=[
            'total_linhas',
            'resumo',
            'status',
        ]
    )

    return importacao


# ============================================================
# APLICAÇÃO DA DECISÃO
# ============================================================

@transaction.atomic
def aplicar_linha(
    linha,
    acao,
):
    if linha.aplicado:
        return linha

    dados = (
        linha.dados_arquivo
        or {}
    )

    campos = (
        _campos_item()
    )

    # --------------------------------------------------------
    # IGNORAR / MANTER SISTEMA
    # --------------------------------------------------------
    if acao in {
        'IGNORAR',
        'MANTER_SISTEMA',
    }:

        linha.acao = acao
        linha.aplicado = True

        linha.save(
            update_fields=[
                'acao',
                'aplicado',
            ]
        )

        return linha

    # --------------------------------------------------------
    # CRIAR ITEM
    # --------------------------------------------------------
    if acao == 'CRIAR_ITEM':

        if linha.item_id:
            raise ValueError(
                'Esta linha já está '
                'vinculada a um item '
                'do sistema.'
            )

        if linha.status != 'SO_ARQUIVO':
            raise ValueError(
                'Criar item só é '
                'permitido para linhas '
                'que existem apenas '
                'no arquivo.'
            )

        item = Item.objects.create(
            **_kwargs_item(
                dados
            )
        )

        atualizar_validade(
            item,
            dados,
        )

        linha.item = item
        linha.acao = acao
        linha.aplicado = True

        linha.save(
            update_fields=[
                'item',
                'acao',
                'aplicado',
            ]
        )

        return linha

    # --------------------------------------------------------
    # USAR ARQUIVO
    # --------------------------------------------------------
    if acao == 'USAR_ARQUIVO':

        # Se ainda estiver ambígua e
        # nenhuma escolha foi vinculada,
        # bloqueia.
        if (
            linha.status
            == 'AMBIGUO'
            and not linha.item_id
        ):
            raise ValueError(
                'Esta linha é ambígua. '
                'Escolha primeiro qual '
                'item do sistema deve '
                'receber os dados '
                'do arquivo.'
            )

        # ----------------------------------------
        # ITEM EXISTENTE
        # ----------------------------------------
        if linha.item_id:

            item = linha.item

            unidade = (
                normalizar_unidade(
                    dados.get(
                        'unidade'
                    )
                )
                if dados.get(
                    'unidade'
                )
                else getattr(
                    item,
                    'unidade',
                    '',
                )
            )

            departamento = (
                normalizar_departamento(
                    dados.get(
                        'departamento'
                    )
                )
                if dados.get(
                    'departamento'
                )
                else ''
            )

            mapeamento = {
                'codigo':
                    dados.get(
                        'codigo'
                    ),

                'nome':
                    dados.get(
                        'nome'
                    ),

                'quantidade':
                    decimal_seguro(
                        dados.get(
                            'quantidade'
                        )
                    ),

                'unidade':
                    unidade,

                'lote':
                    dados.get(
                        'lote'
                    ),

                'localizacao':
                    dados.get(
                        'localizacao'
                    ),

                'estoque_minimo':
                    decimal_seguro(
                        dados.get(
                            'estoque_minimo'
                        )
                    ),

                'fornecedor':
                    dados.get(
                        'fornecedor'
                    ),
            }

            if departamento:
                mapeamento[
                    'departamento'
                ] = departamento

            alterados = []

            for campo, valor in (
                mapeamento.items()
            ):

                if campo not in campos:
                    continue

                # Campo vazio vindo da
                # planilha não apaga
                # dados importantes,
                # exceto lote/localização.
                if (
                    valor in (
                        None,
                        '',
                    )
                    and campo
                    not in {
                        'lote',
                        'localizacao',
                    }
                ):
                    continue

                if isinstance(
                    valor,
                    str,
                ):
                    valor = (
                        _limite_charfield(
                            campo,
                            valor,
                        )
                    )

                setattr(
                    item,
                    campo,
                    valor,
                )

                alterados.append(
                    campo
                )

            if alterados:
                item.save(
                    update_fields=list(
                        dict.fromkeys(
                            alterados
                        )
                    )
                )

            atualizar_validade(
                item,
                dados,
            )

        # ----------------------------------------
        # ITEM NÃO EXISTENTE
        # ----------------------------------------
        else:

            if (
                linha.status
                != 'SO_ARQUIVO'
            ):
                raise ValueError(
                    'Não foi possível '
                    'identificar um item '
                    'único para esta linha.'
                )

            item = Item.objects.create(
                **_kwargs_item(
                    dados
                )
            )

            atualizar_validade(
                item,
                dados,
            )

            linha.item = item

        linha.acao = acao
        linha.aplicado = True

        linha.save(
            update_fields=[
                'item',
                'acao',
                'aplicado',
            ]
        )

        return linha

    raise ValueError(
        f'Ação inválida: {acao}'
    )
