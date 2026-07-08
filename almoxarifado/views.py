import json
import csv
import io
import xmltodict
import requests
import logging
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt, ensure_csrf_cookie
from django.db.models import Q, Sum, F
from django.contrib import messages
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .models import (
    Item, Saida, CarrinhoSolicitacao, Departamento, UnidadeMedida,
    ConfiguracaoWhatsApp, HistoricoNotificacaoAlmoxarifado, AgendamentoNotificacao,
    EntradaNotaFiscal, ItemEntrada
)
from django.contrib.auth.decorators import login_required, permission_required
from .models import (
    Item, Saida, CarrinhoSolicitacao, Departamento, UnidadeMedida,
    ConfiguracaoWhatsApp, HistoricoNotificacaoAlmoxarifado, AgendamentoNotificacao,
    EntradaNotaFiscal, ItemEntrada, InstanciaWhatsApp  # <-- ADICIONE ESTE
)


logger = logging.getLogger(__name__)


def parse_decimal(value, default=None):
    if value is None or str(value).strip() == '':
        return default
    try:
        val_str = str(value).strip().replace(',', '.')
        d = Decimal(val_str)
        d = d.normalize()
        return d
    except (InvalidOperation, ValueError):
        return default


@login_required
@permission_required('almoxarifado.pode_ver_almoxarifado', raise_exception=True)
def lista_itens(request):
    mostrar_todos = request.GET.get('todos', '0') == '1'
    filtro_status = request.GET.get('status', '')
    
    itens = Item.objects.filter(ativo=True)
    
    # Filtro por status (Zerados ou Estoque Baixo)
    if filtro_status == 'zerados':
        itens = itens.filter(quantidade__lte=0)
    elif filtro_status == 'baixo':
        itens = itens.filter(quantidade__gt=0, quantidade__lte=F('estoque_minimo'))
    
    if not mostrar_todos:
        itens = itens.filter(quantidade__gt=0)
    
    busca = request.GET.get('busca', '')
    if busca:
        itens = itens.filter(
            Q(nome__icontains=busca) | 
            Q(codigo__icontains=busca) |
            Q(localizacao__icontains=busca) |
            Q(fornecedor__icontains=busca) |
            Q(lote__icontains=busca) |
            Q(ca__icontains=busca) |
            Q(descricao__icontains=busca) |
            Q(marca__icontains=busca)
        )
    
    departamento = request.GET.get('departamento', '')
    if departamento:
        itens = itens.filter(departamento=departamento)
    
    ordenar = request.GET.get('ordenar', 'nome')
    ordenacao_map = {
        'nome': 'nome',
        '-quantidade': '-quantidade',
        'quantidade': 'quantidade',
        'recente': '-updated_at',
    }
    itens = itens.order_by(ordenacao_map.get(ordenar, 'nome'))
    
    usuario = request.session.get('usuario_carrinho', request.user.username if request.user.is_authenticated else 'anonimo')
    carrinho_count = CarrinhoSolicitacao.objects.filter(usuario=usuario).count()
    
    total_itens = itens.count()
    total_quantidade = itens.aggregate(Sum('quantidade'))['quantidade__sum'] or 0
    
    # Contagem de zerados e baixo para o filtro
    zerados_count = Item.objects.filter(ativo=True, quantidade__lte=0).count()
    baixo_count = Item.objects.filter(ativo=True, quantidade__gt=0, quantidade__lte=F('estoque_minimo')).count()
    
    context = {
        'itens': itens,
        'busca': busca,
        'departamento': departamento,
        'mostrar_todos': mostrar_todos,
        'ordenar': ordenar,
        'departamentos': Departamento.choices,
        'unidades': UnidadeMedida.choices,
        'total_itens': total_itens,
        'total_quantidade': total_quantidade,
        'carrinho_count': carrinho_count,
        'zerados_count': zerados_count,
        'baixo_count': baixo_count,
        'filtro_status': filtro_status,
    }
    return render(request, 'almoxarifado/lista_itens.html', context)


def saidas_list(request):
    saidas = Saida.objects.select_related('item').all().order_by('-data', '-hora')
    
    busca = request.GET.get('busca', '')
    if busca:
        saidas = saidas.filter(
            Q(solicitante__icontains=busca) | 
            Q(item_nome__icontains=busca) |
            Q(item_codigo__icontains=busca)
        )
    
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')
    if data_inicio:
        saidas = saidas.filter(data__gte=data_inicio)
    if data_fim:
        saidas = saidas.filter(data__lte=data_fim)
    
    dept = request.GET.get('dept', '')
    if dept:
        saidas = saidas.filter(departamento=dept)
    
    total_saidas = saidas.count()
    total_retirado = saidas.aggregate(Sum('quantidade'))['quantidade__sum'] or 0
    
    context = {
        'saidas': saidas,
        'busca': busca,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'dept': dept,
        'departamentos': Departamento.choices,
        'total_saidas': total_saidas,
        'total_retirado': total_retirado,
    }
    return render(request, 'almoxarifado/saidas_list.html', context)


def buscar_itens_ajax(request):
    busca = request.GET.get('busca', '')
    departamento = request.GET.get('departamento', '')
    mostrar_todos = request.GET.get('todos', '0') == '1'
    ordenar = request.GET.get('ordenar', 'nome')
    filtro_status = request.GET.get('status', '')
    
    itens = Item.objects.filter(ativo=True)
    
    if filtro_status == 'zerados':
        itens = itens.filter(quantidade__lte=0)
    elif filtro_status == 'baixo':
        itens = itens.filter(quantidade__gt=0, quantidade__lte=F('estoque_minimo'))
    
    if not mostrar_todos:
        itens = itens.filter(quantidade__gt=0)
    
    if busca:
        itens = itens.filter(
            Q(nome__icontains=busca) | 
            Q(codigo__icontains=busca) |
            Q(localizacao__icontains=busca) |
            Q(fornecedor__icontains=busca) |
            Q(lote__icontains=busca) |
            Q(ca__icontains=busca)
        )
    
    if departamento:
        itens = itens.filter(departamento=departamento)
    
    ordenacao_map = {'nome': 'nome', '-quantidade': '-quantidade', 'quantidade': 'quantidade', 'recente': '-updated_at'}
    itens = itens.order_by(ordenacao_map.get(ordenar, 'nome'))
    
    data = {
        'itens': [{
            'id': i.id, 'codigo': i.codigo, 'nome': i.nome,
            'quantidade': float(i.quantidade), 'unidade': i.get_unidade_display(),
            'localizacao': i.localizacao or '-', 'departamento': i.get_departamento_display(),
            'status_estoque': i.status_estoque, 'lote': i.lote or '-', 'ca': i.ca or '-',
            'tamanho': i.tamanho or '-',
        } for i in itens],
    }
    return JsonResponse(data)


@require_http_methods(["GET"])
def buscar_por_codigo(request):
    codigo = request.GET.get('codigo', '').strip()
    if not codigo:
        return JsonResponse({'encontrado': False})
    
    try:
        item = Item.objects.filter(codigo=codigo, ativo=True).first()
        
        if item:
            return JsonResponse({
                'encontrado': True,
                'item': {
                    'id': item.id, 'codigo': item.codigo, 'nome': item.nome,
                    'descricao': item.descricao, 'departamento': item.departamento,
                    'unidade': item.unidade, 'localizacao': item.localizacao,
                    'estoque_minimo': float(item.estoque_minimo),
                    'fornecedor': item.fornecedor, 'quantidade': float(item.quantidade),
                    'lote': item.lote, 'ca': item.ca, 'categoria': item.categoria,
                    'marca': item.marca, 'tamanho': item.tamanho,
                }
            })
        else:
            return JsonResponse({'encontrado': False})
            
    except Exception as e:
        return JsonResponse({'encontrado': False, 'error': str(e)})


@login_required
@permission_required('almoxarifado.pode_gerenciar_almoxarifado', raise_exception=True)
def adicionar_item(request):
    try:
        nome = request.POST.get('nome', '').strip()
        if not nome:
            return JsonResponse({'success': False, 'error': 'Nome obrigatório'}, status=400)
        
        codigo = request.POST.get('codigo', '').strip() or None
        quantidade = parse_decimal(request.POST.get('quantidade', '0'), Decimal('0'))
        lote = request.POST.get('lote', '').strip() or None
        ca = request.POST.get('ca', '').strip() or None
        localizacao = request.POST.get('localizacao', '').strip() or None
        tamanho = request.POST.get('tamanho', '').strip() or None
        
        if codigo:
            item_existente = Item.objects.filter(
                codigo=codigo,
                lote=lote,
                ca=ca,
                localizacao=localizacao
            ).first()
            
            if item_existente:
                if request.POST.get('somar', 'false') == 'true':
                    item_existente.quantidade += quantidade
                    item_existente.save()
                    return JsonResponse({
                        'success': True,
                        'message': f'Quantidade somada! Total: {float(item_existente.quantidade)} {item_existente.get_unidade_display()}'
                    })
                else:
                    return JsonResponse({
                        'success': False,
                        'codigo_existente': True,
                        'error': f'Item {codigo} já existe. Deseja somar?'
                    }, status=409)
        
        item = Item.objects.create(
            nome=nome,
            codigo=codigo,
            quantidade=quantidade,
            departamento=request.POST.get('departamento', 'OUT'),
            unidade=request.POST.get('unidade', 'UN'),
            localizacao=localizacao,
            descricao=request.POST.get('descricao', '').strip() or None,
            estoque_minimo=parse_decimal(request.POST.get('estoque_minimo', '5'), Decimal('5')),
            fornecedor=request.POST.get('fornecedor', '').strip() or None,
            lote=lote,
            ca=ca,
            categoria=request.POST.get('categoria', '').strip() or None,
            marca=request.POST.get('marca', '').strip() or None,
            tamanho=tamanho,
        )
        
        if 'foto' in request.FILES:
            item.foto = request.FILES['foto']
            item.save()
        
        return JsonResponse({'success': True, 'message': f'Item {item.codigo} criado com sucesso!'})
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@permission_required('almoxarifado.pode_gerenciar_almoxarifado', raise_exception=True)
def editar_item(request, pk):
    try:
        item = get_object_or_404(Item, pk=pk)
        
        if request.POST.get('nome'): item.nome = str(request.POST['nome']).strip()
        if 'descricao' in request.POST: item.descricao = str(request.POST.get('descricao', '')).strip() or None
        if request.POST.get('departamento'): item.departamento = str(request.POST['departamento'])[:4]
        if request.POST.get('unidade'): item.unidade = str(request.POST['unidade'])[:3]
        if 'localizacao' in request.POST: item.localizacao = str(request.POST.get('localizacao', '')).strip() or None
        if request.POST.get('estoque_minimo'): item.estoque_minimo = parse_decimal(request.POST['estoque_minimo'], item.estoque_minimo)
        if 'fornecedor' in request.POST: item.fornecedor = str(request.POST.get('fornecedor', '')).strip() or None
        if 'lote' in request.POST: item.lote = str(request.POST.get('lote', '')).strip() or None
        if 'ca' in request.POST: item.ca = str(request.POST.get('ca', '')).strip() or None
        if 'categoria' in request.POST: item.categoria = str(request.POST.get('categoria', '')).strip() or None
        if 'marca' in request.POST: item.marca = str(request.POST.get('marca', '')).strip() or None
        if request.POST.get('quantidade'): item.quantidade = parse_decimal(request.POST['quantidade'], item.quantidade)
        if 'tamanho' in request.POST: item.tamanho = str(request.POST.get('tamanho', '')).strip() or None
        
        if 'foto' in request.FILES:
            item.foto = request.FILES['foto']
        
        item.save()
        
        return JsonResponse({'success': True, 'message': f'Item {item.codigo} atualizado!'})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def detalhe_item(request, pk):
    item = get_object_or_404(Item, pk=pk)
    ultimas_saidas = item.saidas.all().order_by('-data', '-hora')[:5]
    
    return JsonResponse({
        'id': item.id, 'codigo': item.codigo, 'nome': item.nome,
        'quantidade': float(item.quantidade), 'unidade': item.get_unidade_display(),
        'departamento': item.get_departamento_display(), 'localizacao': item.localizacao or 'Não definida',
        'descricao': item.descricao or 'Sem descrição', 'estoque_minimo': float(item.estoque_minimo),
        'fornecedor': item.fornecedor or 'Não informado', 'marca': item.marca or '-',
        'lote': item.lote or '-', 'ca': item.ca or '-', 'categoria': item.categoria or '-',
        'status_estoque': item.status_estoque,
        'foto_url': item.foto.url if item.foto else None,
        'tamanho': item.tamanho or '-',
        'ultimas_saidas': [{
            'data': s.data.strftime('%d/%m/%Y'), 'hora': s.hora.strftime('%H:%M'),
            'solicitante': s.solicitante, 'quantidade': float(s.quantidade),
            'observacao': s.observacao[:50] if s.observacao else '',
        } for s in ultimas_saidas]
    })


@login_required
@permission_required('almoxarifado.pode_gerenciar_almoxarifado', raise_exception=True)
def dar_baixa(request, pk):
    try:
        item = get_object_or_404(Item, pk=pk)
        data = json.loads(request.body) if request.content_type == 'application/json' else request.POST
        
        solicitante = data.get('solicitante', '').strip()
        quantidade = parse_decimal(data.get('quantidade', '0'))
        
        if not solicitante:
            return JsonResponse({'success': False, 'error': 'Solicitante obrigatório'}, status=400)
        if quantidade <= 0:
            return JsonResponse({'success': False, 'error': 'Quantidade deve ser maior que zero'}, status=400)
        if quantidade > item.quantidade:
            return JsonResponse({'success': False, 'error': f'Estoque insuficiente! Disponível: {float(item.quantidade)}'}, status=400)
        
        Saida.objects.create(
            item=item, item_nome=item.nome, item_codigo=item.codigo,
            solicitante=solicitante,
            departamento=data.get('departamento') or None,
            quantidade=quantidade,
            data=data.get('data', date.today().isoformat()),
            hora=data.get('hora', datetime.now().strftime('%H:%M')),
            observacao=data.get('observacao', '').strip()
        )
        
        item.quantidade -= quantidade
        item.save()
        
        return JsonResponse({'success': True, 'message': f'Baixa de {float(quantidade)} {item.get_unidade_display()} realizada!'})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@permission_required('almoxarifado.pode_gerenciar_almoxarifado', raise_exception=True)
def excluir_item(request, pk):
    """Desativa um item (soft delete)"""
    try:
        item = get_object_or_404(Item, pk=pk)
        item.ativo = False
        item.save()
        return JsonResponse({'success': True, 'message': f'Item {item.codigo} desativado com sucesso!'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ===== CARRINHO =====

def ver_carrinho(request):
    usuario = request.session.get('usuario_carrinho', 'anonimo')
    itens = CarrinhoSolicitacao.objects.filter(usuario=usuario).select_related('item')
    return JsonResponse({
        'itens': [{
            'id': i.id, 'item_id': i.item.id, 'codigo': i.item.codigo,
            'nome': i.item.nome, 'quantidade': float(i.quantidade),
            'unidade': i.item.get_unidade_display(), 'estoque': float(i.item.quantidade),
        } for i in itens]
    })


@require_http_methods(["POST"])
@login_required
@permission_required('almoxarifado.pode_gerenciar_almoxarifado', raise_exception=True)
def adicionar_ao_carrinho(request):
    try:
        usuario = request.session.get('usuario_carrinho', 'anonimo')
        item_id = request.POST.get('item_id')
        quantidade = parse_decimal(request.POST.get('quantidade', '1'), Decimal('1'))
        
        item = get_object_or_404(Item, pk=item_id, ativo=True)
        
        if quantidade > item.quantidade:
            return JsonResponse({'success': False, 'error': f'Estoque insuficiente! Disponível: {float(item.quantidade)}'}, status=400)
        
        carrinho_item, created = CarrinhoSolicitacao.objects.get_or_create(
            usuario=usuario, item=item,
            defaults={'quantidade': quantidade}
        )
        if not created:
            carrinho_item.quantidade += quantidade
            carrinho_item.save()
        
        return JsonResponse({'success': True, 'message': f'{item.nome} adicionado ao carrinho!', 'carrinho_count': CarrinhoSolicitacao.objects.filter(usuario=usuario).count()})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def remover_do_carrinho(request, pk):
    try:
        usuario = request.session.get('usuario_carrinho', 'anonimo')
        CarrinhoSolicitacao.objects.filter(usuario=usuario, pk=pk).delete()
        count = CarrinhoSolicitacao.objects.filter(usuario=usuario).count()
        return JsonResponse({'success': True, 'message': 'Item removido!', 'carrinho_count': count})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
@login_required
@permission_required('almoxarifado.pode_gerenciar_almoxarifado', raise_exception=True)
def finalizar_carrinho(request):
    try:
        usuario = request.session.get('usuario_carrinho', 'anonimo')
        itens_carrinho = CarrinhoSolicitacao.objects.filter(usuario=usuario).select_related('item')
        
        if not itens_carrinho.exists():
            return JsonResponse({'success': False, 'error': 'Carrinho vazio!'}, status=400)
        
        data = json.loads(request.body) if request.content_type == 'application/json' else request.POST
        solicitante = data.get('solicitante', '').strip()
        
        if not solicitante:
            return JsonResponse({'success': False, 'error': 'Solicitante obrigatório!'}, status=400)
        
        hoje = date.today()
        agora = datetime.now().strftime('%H:%M')
        
        for ci in itens_carrinho:
            if ci.quantidade > ci.item.quantidade:
                return JsonResponse({'success': False, 'error': f'Estoque insuficiente para {ci.item.nome}!'}, status=400)
        
        for ci in itens_carrinho:
            Saida.objects.create(
                item=ci.item, item_nome=ci.item.nome, item_codigo=ci.item.codigo,
                solicitante=solicitante,
                departamento=data.get('departamento') or None,
                quantidade=ci.quantidade, data=hoje, hora=agora,
                observacao=data.get('observacao', '')
            )
            ci.item.quantidade -= ci.quantidade
            ci.item.save()
            ci.delete()
        
        return JsonResponse({'success': True, 'message': 'Baixa concluída com sucesso!'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def limpar_carrinho(request):
    usuario = request.session.get('usuario_carrinho', 'anonimo')
    CarrinhoSolicitacao.objects.filter(usuario=usuario).delete()
    return JsonResponse({'success': True, 'message': 'Carrinho limpo!'})


# ===== EXPORTAÇÃO =====

def exportar_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Estoque"
    
    cabecalhos = [
        'Código', 'Nome', 'Descrição', 'Departamento', 'Categoria',
        'Quantidade', 'Unidade', 'Localização', 'Estoque Mínimo',
        'Fornecedor', 'Marca', 'Lote', 'CA', 'Tamanho'
    ]
    
    header_font = Font(bold=True, color='FFFFFF', size=11, name='Arial')
    header_fill = PatternFill(start_color='059669', end_color='059669', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col, header in enumerate(cabecalhos, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    itens = Item.objects.filter(ativo=True).order_by('nome')
    
    for row, item in enumerate(itens, 2):
        dados = [
            item.codigo or '',
            item.nome,
            item.descricao or '',
            item.get_departamento_display(),
            item.categoria or '',
            float(item.quantidade),
            item.get_unidade_display(),
            item.localizacao or '',
            float(item.estoque_minimo),
            item.fornecedor or '',
            item.marca or '',
            item.lote or '',
            item.ca or '',
            item.tamanho or '',
        ]
        
        for col, valor in enumerate(dados, 1):
            cell = ws.cell(row=row, column=col, value=valor)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=estoque_{date.today().strftime("%Y%m%d")}.xlsx'
    return response


def baixar_modelo_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Modelo Importação"
    
    cabecalhos = [
        'Código', 'Nome', 'Descrição', 'Departamento', 'Categoria',
        'Quantidade', 'Unidade', 'Localização', 'Estoque Mínimo',
        'Fornecedor', 'Marca', 'Lote', 'CA', 'Tamanho'
    ]
    
    header_font = Font(bold=True, color='FFFFFF', size=11, name='Arial')
    header_fill = PatternFill(start_color='2f8f4e', end_color='2f8f4e', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col, header in enumerate(cabecalhos, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    exemplos = [
        ['001', 'Parafuso 10mm', 'Parafuso sextavado', 'Administrativo', 'Ferragens',
         100, 'Unidade', 'Prateleira A2', 10, 'Fornecedor ABC', 'Tramontina',
         'LOTE-001', 'CA-12345', 'Pequeno'],
    ]
    
    for row, exemplo in enumerate(exemplos, 2):
        for col, valor in enumerate(exemplo, 1):
            cell = ws.cell(row=row, column=col, value=valor)
            cell.border = thin_border
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=modelo_importacao_almoxarifado.xlsx'
    return response


def exportar_saidas_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Saídas"
    
    cabecalhos = ['Data', 'Hora', 'Solicitante', 'Departamento', 'Código Item', 'Item', 'Quantidade', 'Observação']
    
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='dc3545', end_color='dc3545', fill_type='solid')
    
    for col, header in enumerate(cabecalhos, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    saidas = Saida.objects.all().order_by('-data', '-hora')
    for row, saida in enumerate(saidas, 2):
        dados = [
            saida.data.strftime('%d/%m/%Y'), saida.hora.strftime('%H:%M'),
            saida.solicitante, saida.get_departamento_display() or '',
            saida.item_codigo or '', saida.item_nome, float(saida.quantidade),
            saida.observacao or ''
        ]
        for col, valor in enumerate(dados, 1):
            ws.cell(row=row, column=col, value=valor)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=saidas_{date.today().strftime("%Y%m%d")}.xlsx'
    return response


# ===== WHATSAPP API VIEWS =====

@login_required
@require_http_methods(["GET", "POST"])
def api_config_whatsapp(request):
    if request.method == "GET":
        config = ConfiguracaoWhatsApp.get_config()
        return JsonResponse({
            'success': True,
            'config': {
                'ativo': config.ativo,
                'api_url': config.api_url,
                'api_key': config.api_key,
                'instance_name': config.instance_name,
                'numeros_padrao': config.numeros_padrao or '',
                'numeros_por_departamento': config.numeros_por_departamento or {},
                'notificar_estoque_baixo': config.notificar_estoque_baixo,
                'notificar_estoque_zerado': config.notificar_estoque_zerado,
                'notificar_reposicao': config.notificar_reposicao,
                'template_estoque_baixo': config.template_estoque_baixo,
                'template_estoque_zerado': config.template_estoque_zerado,
                'template_reposicao': config.template_reposicao,
                'tipo_envio': config.tipo_envio,
                'notificar_baixo': config.notificar_baixo,
                'notificar_zerado': config.notificar_zerado,
                'notificar_reposicao': config.notificar_reposicao,
                'repetir_notificacoes': config.repetir_notificacoes,
                'intervalo_repeticao': config.intervalo_repeticao,
                'departamentos_ativos': config.departamentos_ativos or [],
                'template_resumo': config.template_resumo,
            }
        })
    
    elif request.method == "POST":
        try:
            data = json.loads(request.body)
            config = ConfiguracaoWhatsApp.get_config()
            
            config.ativo = data.get('ativo', config.ativo)
            config.api_url = data.get('api_url', config.api_url)
            config.api_key = data.get('api_key', config.api_key)
            config.instance_name = data.get('instance_name', config.instance_name)
            config.numeros_padrao = data.get('numeros_padrao', config.numeros_padrao)
            config.numeros_por_departamento = data.get('numeros_por_departamento', {})
            config.notificar_estoque_baixo = data.get('notificar_estoque_baixo', True)
            config.notificar_estoque_zerado = data.get('notificar_estoque_zerado', True)
            config.notificar_reposicao = data.get('notificar_reposicao', True)
            config.template_estoque_baixo = data.get('template_estoque_baixo', config.template_estoque_baixo)
            config.template_estoque_zerado = data.get('template_estoque_zerado', config.template_estoque_zerado)
            config.template_reposicao = data.get('template_reposicao', config.template_reposicao)
            config.tipo_envio = data.get('tipo_envio', 'tempo-real')
            config.notificar_baixo = data.get('notificar_baixo', True)
            config.notificar_zerado = data.get('notificar_zerado', True)
            config.notificar_reposicao = data.get('notificar_reposicao', True)
            config.repetir_notificacoes = data.get('repetir_notificacoes', False)
            config.intervalo_repeticao = data.get('intervalo_repeticao', 24)
            config.departamentos_ativos = data.get('departamentos_ativos', [])
            config.template_resumo = data.get('template_resumo', config.template_resumo)
            
            config.save()
            
            return JsonResponse({'success': True, 'message': 'Configurações salvas'})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_http_methods(["POST"])
def api_listar_instancias(request):
    """Lista apenas as instâncias criadas pelo sistema com status atualizado"""
    try:
        data = json.loads(request.body)
        api_url = data.get('api_url')
        api_key = data.get('api_key')
        
        # Buscar instâncias salvas no banco de dados
        instancias_db = InstanciaWhatsApp.objects.all().order_by('nome')
        
        instancias = []
        
        for inst in instancias_db:
            # Tentar obter status atual da Evolution API
            status_atual = inst.status
            
            if api_url:
                try:
                    if not api_url.startswith(('http://', 'https://')):
                        api_url_tmp = 'https://' + api_url
                    else:
                        api_url_tmp = api_url
                    
                    headers = {'apikey': api_key} if api_key else {}
                    url = f"{api_url_tmp.rstrip('/')}/instance/fetchInstances?instanceName={inst.nome}"
                    
                    response = requests.get(url, headers=headers, timeout=5, verify=False)
                    
                    if response.status_code == 200:
                        inst_data = response.json()
                        if isinstance(inst_data, list) and len(inst_data) > 0:
                            inst_data = inst_data[0]
                        api_status = inst_data.get('connectionStatus', 'close')
                        status_atual = 'connected' if api_status == 'open' else 'disconnected'
                        
                        # Atualizar status no banco se mudou
                        if inst.status != status_atual:
                            inst.status = status_atual
                            inst.save()
                except Exception as e:
                    print(f"Erro ao verificar status da instância {inst.nome}: {e}")
            
            instancias.append({
                'name': inst.nome,
                'instanceName': inst.nome,
                'connectionStatus': 'open' if status_atual == 'connected' else 'close',
                'created_by_system': True,
                'status_text': '✅ Conectado' if status_atual == 'connected' else '❌ Desconectado'
            })
        
        return JsonResponse({'success': True, 'instancias': instancias})
        
    except Exception as e:
        print(f"Erro em api_listar_instancias: {e}")
        return JsonResponse({'success': True, 'instancias': []})

@login_required
@require_http_methods(["POST"])
def api_criar_instancia(request):
    try:
        data = json.loads(request.body)
        api_url = data.get('api_url')
        api_key = data.get('api_key')
        instance_name = data.get('instance_name')
        
        if not api_url or not instance_name:
            return JsonResponse({'success': False, 'error': 'URL e nome da instância são obrigatórios'})
        
        if not api_url.startswith(('http://', 'https://')):
            api_url = 'https://' + api_url
        
        # Verificar se a instância já existe no banco
        if InstanciaWhatsApp.objects.filter(nome=instance_name).exists():
            return JsonResponse({'success': False, 'error': f'Instância "{instance_name}" já existe!'})
        
        # Tentar criar na Evolution API
        headers = {'Content-Type': 'application/json'}
        if api_key:
            headers['apikey'] = api_key
            
        url = f"{api_url.rstrip('/')}/instance/create"
        payload = {
            "instanceName": instance_name,
            "qrcode": True,
            "integration": "WHATSAPP-BAILEYS"
        }
        
        response = requests.post(url, json=payload, headers=headers, timeout=30, verify=False)
        
        if response.status_code in [200, 201]:
            # Salvar no banco de dados
            InstanciaWhatsApp.objects.create(
                nome=instance_name,
                status='disconnected',
                api_url=api_url
            )
            return JsonResponse({'success': True, 'message': f'Instância "{instance_name}" criada'})
        else:
            return JsonResponse({'success': False, 'error': f'HTTP {response.status_code}: {response.text[:100]}'})
            
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_http_methods(["POST"])
def api_qrcode_instancia(request):
    try:
        data = json.loads(request.body)
        api_url = data.get('api_url')
        api_key = data.get('api_key')
        instance_name = data.get('instance_name')
        
        if not api_url or not instance_name:
            return JsonResponse({'success': False, 'error': 'URL e nome da instância são obrigatórios'})
        
        if not api_url.startswith(('http://', 'https://')):
            api_url = 'https://' + api_url
        
        headers = {'apikey': api_key} if api_key else {}
        url = f"{api_url.rstrip('/')}/instance/connect/{instance_name}"
        
        response = requests.get(url, headers=headers, timeout=30, verify=False)
        
        if response.status_code == 200:
            qrcode_data = response.json()
            
            pairing_code = qrcode_data.get('pairingCode')
            qrcode = qrcode_data.get('base64') or qrcode_data.get('qrcode')
            
            if pairing_code and pairing_code != 'null':
                # Atualizar status para connecting
                InstanciaWhatsApp.objects.filter(nome=instance_name).update(status='connecting')
                return JsonResponse({'success': True, 'pairingCode': pairing_code})
            elif qrcode:
                InstanciaWhatsApp.objects.filter(nome=instance_name).update(status='connecting')
                return JsonResponse({'success': True, 'qrcode': qrcode})
            else:
                return JsonResponse({'success': False, 'error': 'Nenhum código disponível'})
        else:
            return JsonResponse({'success': False, 'error': f'HTTP {response.status_code}'})
            
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_http_methods(["POST"])
def api_status_instancia(request):
    try:
        data = json.loads(request.body)
        api_url = data.get('api_url')
        api_key = data.get('api_key')
        instance_name = data.get('instance_name')
        
        if not api_url or not instance_name:
            return JsonResponse({'success': False, 'error': 'URL e nome da instância são obrigatórios'})
        
        if not api_url.startswith(('http://', 'https://')):
            api_url = 'https://' + api_url
        
        headers = {'apikey': api_key} if api_key else {}
        url = f"{api_url.rstrip('/')}/instance/fetchInstances?instanceName={instance_name}"
        
        response = requests.get(url, headers=headers, timeout=10, verify=False)
        
        status = 'disconnected'
        if response.status_code == 200:
            inst_data = response.json()
            if isinstance(inst_data, list) and len(inst_data) > 0:
                inst_data = inst_data[0]
            api_status = inst_data.get('connectionStatus', 'close')
            status = 'connected' if api_status == 'open' else 'disconnected'
            
            # Atualizar status no banco
            InstanciaWhatsApp.objects.filter(nome=instance_name).update(status=status)
        
        return JsonResponse({'success': True, 'status': status})
            
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_http_methods(["POST"])
def api_deletar_instancia(request):
    try:
        data = json.loads(request.body)
        api_url = data.get('api_url')
        api_key = data.get('api_key')
        instance_name = data.get('instance_name')
        
        if not api_url or not instance_name:
            return JsonResponse({'success': False, 'error': 'URL e nome da instância são obrigatórios'})
        
        if not api_url.startswith(('http://', 'https://')):
            api_url = 'https://' + api_url
        
        headers = {'apikey': api_key} if api_key else {}
        
        endpoints = [
            f"{api_url}/instance/delete/{instance_name}",
            f"{api_url}/instance/logout/{instance_name}",
        ]
        
        api_deleted = False
        for url in endpoints:
            try:
                response = requests.delete(url, headers=headers, timeout=10, verify=False)
                if response.status_code in [200, 204]:
                    api_deleted = True
                    break
            except:
                continue
        
        # Remover do banco de dados
        InstanciaWhatsApp.objects.filter(nome=instance_name).delete()
        
        return JsonResponse({'success': True, 'message': f'Instância "{instance_name}" deletada'})
            
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
# ===== AGENDAMENTOS =====

@login_required
@require_http_methods(["GET"])
def listar_agendamentos(request):
    try:
        config = ConfiguracaoWhatsApp.get_config()
        agendamentos = config.agendamentos.filter(ativo=True).order_by('horario')
        
        agendamentos_data = []
        for ag in agendamentos:
            agendamentos_data.append({
                'id': ag.id,
                'horario': ag.horario.strftime('%H:%M'),
                'dias_semana': ag.dias_semana or [],
                'ativo': ag.ativo,
                'descricao': ag.descricao or '',
            })
        
        return JsonResponse({'success': True, 'agendamentos': agendamentos_data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_http_methods(["POST"])
def criar_agendamento(request):
    try:
        data = json.loads(request.body)
        config = ConfiguracaoWhatsApp.get_config()
        
        horario = data.get('horario')
        if not horario:
            return JsonResponse({'success': False, 'error': 'Horário obrigatório'})
        
        dias_semana = data.get('dias_semana', [])
        if dias_semana:
            dias_semana = [int(d) for d in dias_semana if 0 <= int(d) <= 6]
        
        # Verificar se já existe
        for ag in config.agendamentos.filter(horario=horario):
            return JsonResponse({'success': False, 'error': f'Horário {horario} já está agendado!'})
        
        agendamento = AgendamentoNotificacao.objects.create(
            config=config,
            horario=horario,
            dias_semana=dias_semana if dias_semana else None,
            descricao=data.get('descricao', ''),
            ativo=True
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Agendamento criado para {horario}',
            'agendamento': {
                'id': agendamento.id,
                'horario': agendamento.horario.strftime('%H:%M'),
                'dias_semana': agendamento.dias_semana or [],
                'descricao': agendamento.descricao,
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_http_methods(["POST"])
def deletar_agendamento(request, agendamento_id):
    try:
        agendamento = get_object_or_404(AgendamentoNotificacao, id=agendamento_id)
        horario_str = agendamento.horario.strftime('%H:%M')
        agendamento.delete()
        
        return JsonResponse({'success': True, 'message': f'Agendamento {horario_str} removido'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_http_methods(["POST"])
def api_enviar_notificacao_agora(request):
    try:
        from django.core.management import call_command
        from io import StringIO
        
        out = StringIO()
        call_command('enviar_notificacoes_almoxarifado', '--now', stdout=out)
        
        return JsonResponse({'success': True, 'message': 'Notificações enviadas com sucesso!'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})