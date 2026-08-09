# Django imports
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from django.db.models import Q, Sum, Count
from django.contrib import messages
from django.http import JsonResponse
from django.utils import timezone
from django.core.paginator import Paginator
from django.contrib.auth.models import User
from django.contrib.auth import update_session_auth_hash
from django.core.serializers.json import DjangoJSONEncoder 
from .models import HistoricoMovimentacao
from .models import OrigemDestino
from .models import Estoque, Cultivar, Peneira, Categoria, StatusSistemico
from django.db.models import Q, Sum, Prefetch, F
# Adicione no topo com os outros imports
import datetime
from django import forms  #
# Python imports
from decimal import Decimal, InvalidOperation
from datetime import timedelta
import random
from django.contrib.admin.views.decorators import staff_member_required
from django.http import JsonResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from .models import ArmazemLayout, ElementoMapa, Estoque
import json
from django.utils import timezone

from django.db.models import (
    Case,
    F,
    IntegerField,
    Q,
    Sum,
    Value,
    When,
)
from django.core.cache import cache

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
# App imports
from .models import (
    Estoque, HistoricoMovimentacao, Configuracao, Cultivar, 
    Peneira, Categoria, Tratamento, PerfilUsuario, Especie, OrigemDestino,Armazem, Endereco, Solicitacao, ColunaKanban, 
    RegraWorkflow,
    HistoricoCard,
    ConfiguracaoAtualizacao,  


)
from collections import defaultdict
from .forms import (
    NovaEntradaForm, ConfiguracaoForm, CultivarForm, PeneiraForm, 
    CategoriaForm, TratamentoForm, NovoConferenteUserForm, MudarSenhaForm  
)


# sapp/views.py - No início do arquivo, adicione:

from django.shortcuts import render, redirect, reverse  # Adicione 'reverse' aqui
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User, Group, Permission
from django.contrib.auth.hashers import make_password
from django.http import JsonResponse
from django.db import transaction
from django.urls import reverse  # Também pode importar assim
from .models import (
    Produto, Cultivar, Peneira, Especie, Categoria, Tratamento, 
    Armazem, Endereco, OrigemDestino, Configuracao
)
from .forms import ConfiguracaoForm, NovoConferenteUserForm

# Pandas e outros imports
import pandas as pd
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import io
import json
from django.db import transaction
import tempfile
import os

from django.db import transaction
from .models import FotoMovimentacao # e os outros models   
    

# No início de views.py, com os outros imports de models
from .models import (
    Estoque, HistoricoMovimentacao, Configuracao, Cultivar, 
    Peneira, Categoria, Tratamento, PerfilUsuario,
    # Adicione estes:
    Empenho, ItemEmpenho, EmpenhoStatus
)


from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db.models import Q, Sum
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from sapp.models import Estoque, Cultivar, Peneira, Categoria, Tratamento, Especie
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth.models import User
from .models import (
    Configuracao, Cultivar, Peneira, Categoria, 
    Tratamento, Especie, Produto
)
from .forms import (
    ConfiguracaoForm, NovoConferenteUserForm,
    CultivarForm, PeneiraForm, CategoriaForm, TratamentoForm
)

# views.py - ARQUIVO COMPLETO
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth.models import User
from django.db.models import Q
from .models import (
    Configuracao, Cultivar, Peneira, Categoria, 
    Tratamento, Especie, Produto, Estoque,
    HistoricoMovimentacao, Empenho, ItemEmpenho,
    ArmazemLayout, ElementoMapa
)
from .forms import (
    ConfiguracaoForm, NovoConferenteUserForm,
    CultivarForm, PeneiraForm, CategoriaForm, 
    TratamentoForm, ProdutoForm, NovaEntradaForm
)

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth.models import User
from django.db.models import Q
from .models import (
    Configuracao, Cultivar, Peneira, Categoria, 
    Tratamento, Especie, Produto, Estoque
)
from .forms import (
    ConfiguracaoForm, NovoConferenteUserForm,
    CultivarForm, PeneiraForm, CategoriaForm, 
    TratamentoForm, ProdutoForm
)

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction, models
from django.db.models import Q

from .models import (
    Estoque,
    Empenho,
    ItemEmpenho,
    EmpenhoStatus,
    HistoricoMovimentacao
)
from django.contrib.admin.views.decorators import staff_member_required 
from django.contrib.admin.views.decorators import staff_member_required
from django.http import JsonResponse
from django.shortcuts import render, get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.contrib.auth.decorators import login_required, permission_required
from .models import ArmazemLayout, ElementoMapa, Estoque
import json




# ================================================================
# FUNÇÕES AUXILIARES (ADICIONAR NO TOPO DO ARQUIVO views.py)
# ================================================================
def processar_inteiro(valor, default=0):
    """Converte valor para inteiro com segurança"""
    if valor is None or valor == '':
        return default
    
    try:
        if isinstance(valor, str):
            # Remove caracteres não numéricos, mantendo ponto decimal para conversão
            valor_limpo = ''
            for char in valor:
                if char.isdigit() or char in '.,':
                    valor_limpo += char
            valor = valor_limpo.replace(',', '.')
            
            if '.' in valor:
                # Se tiver decimal, arredonda para baixo
                return int(float(valor))
            else:
                return int(valor) if valor else default
        else:
            return int(valor)
    except (ValueError, TypeError, AttributeError):
        return default

def processar_decimal(valor, default=Decimal('0.00')):
    """Converte valor para Decimal com segurança"""
    if valor is None:
        return default
    
    try:
        if isinstance(valor, str):
            valor = valor.replace(',', '.')
            # Remove caracteres não numéricos, mantendo ponto decimal
            valor = ''.join(c for c in valor if c.isdigit() or c == '.' or c == '-')
            if not valor:
                return default
        
        # Converte para Decimal, limitando casas decimais
        return Decimal(str(valor)).quantize(Decimal('0.01'))
    except (InvalidOperation, ValueError, TypeError):
        return default


from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, Q
from django.db.models.functions import TruncDate
from django.utils import timezone
from django.http import JsonResponse
from datetime import datetime, timedelta
from .models import Estoque, HistoricoMovimentacao, Especie, Cultivar, Peneira




@login_required
@permission_required('sapp.pode_ver_estoque', raise_exception=True)
def dashboard_data(request):
    """Endpoint AJAX para dados do dashboard"""
    try:
        # Receber filtros (listas)
        tipos_semente = request.GET.getlist('tipo_semente[]')
        cultivares = request.GET.getlist('cultivar[]')
        peneiras = request.GET.getlist('peneira[]')
        unidades = request.GET.getlist('unidade[]')
        armazens = request.GET.getlist('armazem[]')
        data_inicio = request.GET.get('data_inicio', '').strip()
        data_fim = request.GET.get('data_fim', '').strip()
        tipo_mov = request.GET.get('tipo_mov', '').strip()
        search = request.GET.get('search', '').strip()
        
        # Base queries
        est_qs = Estoque.objects.filter(saldo__gt=0)
        mov_qs = HistoricoMovimentacao.objects.select_related('estoque', 'usuario')
        
        # Aplicar filtros de estoque
        if tipos_semente:
            est_qs = est_qs.filter(especie__nome__in=tipos_semente)
        
        if cultivares:
            est_qs = est_qs.filter(cultivar_id__in=cultivares)
        
        if peneiras:
            est_qs = est_qs.filter(peneira_id__in=peneiras)
        
        if unidades:
            est_qs = est_qs.filter(embalagem__in=unidades)
        
        if armazens:
            est_qs = est_qs.filter(az__in=armazens)
        
        if search:
            est_qs = est_qs.filter(
                Q(lote__icontains=search) | 
                Q(cultivar__nome__icontains=search) |
                Q(especie__nome__icontains=search)
            )
        
        # Aplicar filtros de movimentação
        mov_qs = mov_qs.filter(estoque__in=est_qs)
        
        # Filtros de data
        if data_inicio:
            try:
                data_inicio_obj = datetime.strptime(data_inicio, '%Y-%m-%d').date()
                mov_qs = mov_qs.filter(data_hora__date__gte=data_inicio_obj)
            except:
                pass
        
        if data_fim:
            try:
                data_fim_obj = datetime.strptime(data_fim, '%Y-%m-%d').date()
                mov_qs = mov_qs.filter(data_hora__date__lte=data_fim_obj)
            except:
                pass
        
        if tipo_mov:
            mov_qs = mov_qs.filter(tipo__iexact=tipo_mov)
        
        # Calcular KPIs
        bags = est_qs.filter(embalagem='BAG').aggregate(s=Sum('saldo'))['s'] or 0
        scs = est_qs.filter(embalagem='SC').aggregate(s=Sum('saldo'))['s'] or 0
        total_sc = (bags * 25) + scs
        
        kpis = {
            'total_sc': int(total_sc),
            'bags': int(bags),
            'scs': int(scs),
            'peso': float(est_qs.aggregate(s=Sum('peso_total'))['s'] or 0),
            'ativos': est_qs.count(),
            'parados': est_qs.filter(data_ultima_movimentacao__lt=timezone.now() - timedelta(days=30)).count()
        }
        
        # Dados dos gráficos
        cores_padrao = ['#10b981', '#3b82f6', '#f59e0b', '#ef4444', '#8b5cf6', '#ec4899', '#06b6d4', '#84cc16', '#f97316', '#6366f1']
        
        cultivares_data = list(est_qs.filter(cultivar__isnull=False)
                               .values('cultivar__nome')
                               .annotate(volume=Sum('saldo'))
                               .filter(volume__gt=0)
                               .order_by('-volume')[:10])
        
        peneiras_data = list(est_qs.filter(peneira__isnull=False)
                             .values('peneira__nome')
                             .annotate(volume=Sum('saldo'))
                             .filter(volume__gt=0)
                             .order_by('-volume'))
        
        armazens_data = list(est_qs.exclude(az__isnull=True).exclude(az='')
                             .values('az')
                             .annotate(volume=Sum('saldo'))
                             .filter(volume__gt=0)
                             .order_by('az'))
        
        # Tendência
        data_limite = timezone.now() - timedelta(days=15)
        tendencia_data = list(mov_qs.filter(data_hora__date__gte=data_limite.date())
                              .annotate(dia=TruncDate('data_hora'))
                              .values('dia')
                              .annotate(
                                  entradas=Count('id', filter=Q(tipo__iexact='Entrada')),
                                  saidas=Count('id', filter=Q(tipo__iexact='Saída'))
                              )
                              .order_by('dia'))
        
        graficos = {
            'cultivar': {
                'labels': [d['cultivar__nome'] for d in cultivares_data],
                'values': [int(d['volume']) for d in cultivares_data],
                'colors': cores_padrao[:len(cultivares_data)]
            },
            'peneira': {
                'labels': [d['peneira__nome'] for d in peneiras_data],
                'values': [int(d['volume']) for d in peneiras_data],
                'colors': cores_padrao[:len(peneiras_data)]
            },
            'armazem': {
                'labels': [d['az'] for d in armazens_data],
                'values': [int(d['volume']) for d in armazens_data],
                'colors': cores_padrao[:len(armazens_data)]
            },
            'tendencia': {
                'labels': [d['dia'].strftime('%d/%m') for d in tendencia_data],
                'entradas': [d['entradas'] for d in tendencia_data],
                'saidas': [d['saidas'] for d in tendencia_data]
            }
        }
        
        # Opções de filtros (encadeamento)
        opcoes_filtros = {
            'tipos_semente': list(est_qs.values_list('especie__nome', flat=True).distinct().order_by('especie__nome')),
            'cultivares': list(est_qs.filter(cultivar__isnull=False).values('cultivar_id', 'cultivar__nome').distinct().order_by('cultivar__nome')),
            'peneiras': list(est_qs.filter(peneira__isnull=False).values('peneira_id', 'peneira__nome').distinct().order_by('peneira__nome')),
            'armazens': list(est_qs.exclude(az__isnull=True).exclude(az='').values_list('az', flat=True).distinct().order_by('az'))
        }
        
        # Movimentações recentes
        movimentacoes = []
        for mov in mov_qs.order_by('-data_hora')[:10]:
            movimentacoes.append({
                'dt': mov.data_hora.strftime('%d/%m/%Y %H:%M') if mov.data_hora else '--',
                'tp': mov.tipo or '--',
                'lt': mov.lote_ref or (mov.estoque.lote if mov.estoque else '--'),
                'unidade': mov.estoque.embalagem if mov.estoque else '--',
                'qtd': getattr(mov, 'quantidade', 0),
                'us': mov.usuario.username if mov.usuario else 'Sistema'
            })
        
        return JsonResponse({
            'success': True,
            'kpis': kpis,
            'graficos': graficos,
            'recentes': movimentacoes,
            'opcoes_filtros': opcoes_filtros
        })
        
    except Exception as e:
        print(f"[ERROR] {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
# ================================================================
# LISTA DE ESTOQUE (TABELA PRINCIPAL)
# ================================================================

@login_required
@permission_required('sapp.pode_ver_estoque', raise_exception=True)
def lista_estoque(request, template_name='sapp/tabela_estoque.html'):
    """
    View para a página principal de estoque - MOSTRA TODOS OS LOTES
    """
    
    # QuerySet Base - TODOS os lotes (inclusive zerados) PARA EXIBIÇÃO
    qs = Estoque.objects.all().select_related(
        'cultivar', 'peneira', 'categoria', 'tratamento', 'especie', 'conferente'
    ).order_by('-data_ultima_movimentacao', '-id')
    
    # QuerySet Base para MÉTRICAS - TODOS os lotes (para os cards)
    qs_metrics = Estoque.objects.all()
    
    
    # FILTRO POR STATUS
    status = request.GET.get('status', 'todos')
    if status == 'disponivel':
        qs = qs.filter(saldo__gt=0)
    elif status == 'esgotado':
        qs = qs.filter(saldo=0)

    # Busca Global
    busca = request.GET.get('busca', '').strip()
    if busca:
        for termo in busca.split():
            qs = qs.filter(
                Q(lote__icontains=termo) | 
                Q(produto__icontains=termo) |
                Q(cultivar__nome__icontains=termo) | 
                Q(especie__nome__icontains=termo) |
                Q(endereco__icontains=termo) | 
                Q(cliente__icontains=termo) |
                Q(empresa__icontains=termo)
            )

    # Aplicar filtros sequenciais - COM SUPORTE A VALORES VAZIOS (__null__)
    filter_map = {
        'az': 'az__in',
        'lote': 'lote__in',
        'produto': 'produto__in',
        'cultivar': 'cultivar__nome__in',
        'peneira': 'peneira__nome__in',
        'categoria': 'categoria__nome__in',
        'endereco': 'endereco__in',
        'especie': 'especie__nome__in',
        'tratamento': 'tratamento__nome__in',
        'embalagem': 'embalagem__in',
        'cliente': 'cliente__in',
        'empresa': 'empresa__in',
        'conferente': 'conferente__username__in'
    }

    for param, lookup in filter_map.items():
        values = request.GET.getlist(param)
        # REMOVER VALORES VAZIOS
        values = [v for v in values if v and v.strip()]
        
        # VERIFICAR SE TEM O VALOR ESPECIAL __null__ (VAZIO)
        tem_null = '__null__' in values
        if tem_null:
            values.remove('__null__')
        
        if values and tem_null:
            # CASO: TEM VALORES ESPECÍFICOS E TAMBÉM QUER VAZIOS
            q = Q(**{lookup: values}) | Q(**{f"{param}__isnull": True}) | Q(**{f"{param}": ''})
            qs = qs.filter(q)
        elif values:
            # CASO: SÓ VALORES ESPECÍFICOS
            qs = qs.filter(**{lookup: values})
        elif tem_null:
            # CASO: SÓ VAZIOS
            qs = qs.filter(Q(**{f"{param}__isnull": True}) | Q(**{f"{param}": ''}))

    # Filtros numéricos
    for field in ['saldo', 'peso_unitario', 'peso_total']:
        min_val = request.GET.get(f'min_{field}')
        max_val = request.GET.get(f'max_{field}')
        if min_val:
            qs = qs.filter(**{f'{field}__gte': min_val})
        if max_val:
            qs = qs.filter(**{f'{field}__lte': max_val})

    # MÉTRICAS PARA OS CARDS - Usando o queryset NÃO FILTRADO (qs_metrics)
    # CARD 1: Lotes Ativos (APENAS saldo > 0)
    total_itens_ativos = qs_metrics.filter(saldo__gt=0).count()
    
    # CARD 2: SC Equivalente (somente saldo > 0)
    saldo_bags = qs_metrics.filter(embalagem='BAG', saldo__gt=0).aggregate(s=Sum('saldo'))['s'] or 0
    saldo_sc = qs_metrics.filter(embalagem='SC', saldo__gt=0).aggregate(s=Sum('saldo'))['s'] or 0
    saldo_total_sc = (saldo_bags * 25) + saldo_sc
    origens = OrigemDestino.objects.all().order_by('nome')
    # CARD 3: Unidades BAG (somente saldo > 0)
    saldo_bags_total = qs_metrics.filter(embalagem='BAG', saldo__gt=0).aggregate(s=Sum('saldo'))['s'] or 0
    
    # CARD 4: PME Total (KG)
    pme_total = qs_metrics.filter(saldo__gt=0).aggregate(s=Sum('peso_total'))['s'] or Decimal('0.00')
    
    # CARD 5: Clientes Únicos (somente saldo > 0)
    clientes_unicos = qs_metrics.filter(
        saldo__gt=0
    ).exclude(
        cliente__isnull=True
    ).exclude(
        cliente=''
    ).values('cliente').distinct().count()

    # Opções de Filtro (baseadas no queryset filtrado qs, NÃO no qs_metrics)
    def get_options_list(field_lookup, param_name):
        vals = qs.values_list(field_lookup, flat=True).distinct().order_by(field_lookup)
        options = [str(v) for v in vals if v is not None and str(v).strip() != '']
        # Ordenar e retornar
        return sorted(options)

    filter_options = {
        'az': get_options_list('az', 'az'),
        'lote': get_options_list('lote', 'lote'),
        'produto': get_options_list('produto', 'produto'),
        'cultivar': get_options_list('cultivar__nome', 'cultivar'),
        'peneira': get_options_list('peneira__nome', 'peneira'),
        'categoria': get_options_list('categoria__nome', 'categoria'),
        'endereco': get_options_list('endereco', 'endereco'),
        'especie': get_options_list('especie__nome', 'especie'),
        'tratamento': get_options_list('tratamento__nome', 'tratamento'),
        'embalagem': get_options_list('embalagem', 'embalagem'),
        'cliente': get_options_list('cliente', 'cliente'),
        'empresa': get_options_list('empresa', 'empresa'),
        'conferente': get_options_list('conferente__username', 'conferente')
    }

    # Paginação
    page_size = request.GET.get('page_size', 25)
    try:
        page_size = int(page_size)
    except (ValueError, TypeError):
        page_size = 25
    
    paginator = Paginator(qs, page_size)
    page_number = request.GET.get('page', 1)
    
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']
    
    context = {
        'estoque': page_obj,
        'itens': page_obj,
        'status': status,
        'busca': busca,
        'total_itens': total_itens_ativos,  # CARD 1: APENAS saldo > 0
        'total_sc': saldo_total_sc,          # CARD 2: APENAS saldo > 0
        'total_bags': saldo_bags_total,      # CARD 3: APENAS saldo > 0
        'total_pme': pme_total,              # CARD 4: NOVO CARD
        'clientes_unicos': clientes_unicos,  # CARD 5: APENAS saldo > 0
        'filter_options': filter_options,
        'url_params': query_params.urlencode(),
        'page_sizes': [10, 25, 50, 100, 200],
        'page_size': page_size,
        'all_cultivares': Cultivar.objects.all(),
        'all_peneiras': Peneira.objects.all(),
        'all_categorias': Categoria.objects.all(),
        'all_tratamentos': Tratamento.objects.all(),
        'all_especies': Especie.objects.all(),
        'origens': origens,
    }
    
    return render(request, template_name, context)


@login_required
@permission_required('sapp.pode_movimentar_estoque', raise_exception=True)
def gestao_estoque(request, template_name='sapp/gestao_estoque.html'):
    """
    View para gestão de estoque - MOSTRA APENAS LOTES COM SALDO > 0
    """
    
    # QuerySet Base - APENAS LOTES COM SALDO > 0 - NUNCA mostrar saldo zero
    qs = Estoque.objects.filter(saldo__gt=0).select_related(
        'cultivar', 'peneira', 'categoria', 'tratamento', 'especie', 'conferente'
    ).order_by('-data_ultima_movimentacao', '-id')
    
    # NÃO existe qs_metrics separado - tudo deve usar o mesmo filtro
    
    # FILTRO POR STATUS SISTÊMICO
    status_filter = request.GET.getlist('status_sistemico')

    if status_filter:
        status_ids = []

        for status_value in status_filter:
            try:
                if str(status_value).isdigit():
                    status_ids.append(int(status_value))
                else:
                    status_obj = StatusSistemico.objects.get(nome=status_value)
                    status_ids.append(status_obj.id)
            except (StatusSistemico.DoesNotExist, ValueError):
                pass

        if status_ids:
            qs = qs.filter(status_sistemico__in=status_ids)
    
    # Busca Global
    busca = request.GET.get('busca', '').strip()
    if busca:
        for termo in busca.split():
            qs = qs.filter(
                Q(lote__icontains=termo) | 
                Q(produto__icontains=termo) |
                Q(cultivar__nome__icontains=termo) | 
                Q(especie__nome__icontains=termo) |
                Q(endereco__icontains=termo) | 
                Q(cliente__icontains=termo)
            )

    # Aplicar filtros sequenciais
    filter_map = {
        'az': 'az__in',
        'lote': 'lote__in',
        'produto': 'produto__in',
        'cultivar': 'cultivar__nome__in',
        'peneira': 'peneira__nome__in',
        'categoria': 'categoria__nome__in',
        'endereco': 'endereco__in',
        'especie': 'especie__nome__in',
        'tratamento': 'tratamento__nome__in',
        'embalagem': 'embalagem__in',
        'cliente': 'cliente__in',
        'empresa': 'empresa__in',
        'conferente': 'conferente__username__in'
    }

    for param, lookup in filter_map.items():
        values = request.GET.getlist(param)
        values = [v for v in values if v and v.strip()]
        
        if values:
            if '__null__' in values:
                specific_values = [v for v in values if v != '__null__']
                if specific_values:
                    qs = qs.filter(
                        Q(**{lookup: specific_values}) | 
                        Q(**{lookup.replace('__in', '__isnull'): True})
                    )
                else:
                    qs = qs.filter(**{lookup.replace('__in', '__isnull'): True})
            else:
                qs = qs.filter(**{lookup: values})

    # Filtros numéricos
    for field in ['saldo', 'peso_unitario', 'peso_total']:
        min_val = request.GET.get(f'min_{field}')
        max_val = request.GET.get(f'max_{field}')
        if min_val:
            try:
                qs = qs.filter(**{f'{field}__gte': float(min_val)})
            except ValueError:
                pass
        if max_val:
            try:
                qs = qs.filter(**{f'{field}__lte': float(max_val)})
            except ValueError:
                pass

    # MÉTRICAS - usando o mesmo queryset filtrado
    saldo_bags = qs.filter(embalagem='BAG').aggregate(s=Sum('saldo'))['s'] or 0
    saldo_sc = qs.filter(embalagem='SC').aggregate(s=Sum('saldo'))['s'] or 0
    saldo_total_sc = (saldo_bags * 25) + saldo_sc
    
    total_pme = qs.aggregate(s=Sum('peso_total'))['s'] or 0
    
    
    # Opções de Filtro - baseadas no queryset COMPLETO (com saldo > 0)
    base_options_qs = Estoque.objects.filter(saldo__gt=0)
    
    def get_options_list(field_lookup):
        vals = base_options_qs.values_list(field_lookup, flat=True).distinct().order_by(field_lookup)
        options = []
        for v in vals:
            if v is not None and str(v).strip() != '':
                options.append(str(v))
        return options
    status_ids_em_uso = qs.exclude(
        status_sistemico__isnull=True
    ).values_list(
        'status_sistemico_id',
        flat=True
    ).distinct()

    status_em_uso = StatusSistemico.objects.filter(
        ativo=True,
        id__in=status_ids_em_uso
    ).order_by('ordem', 'nome')

    status_options = [
        {
            'value': str(s.id),
            'label': f"{s.icone or ''} {s.nome}".strip()
        }
        for s in status_em_uso
    ]
    
    filter_options = {
        'status_sistemico': status_options,
        'az': get_options_list('az'),
        'lote': get_options_list('lote'),
        'produto': get_options_list('produto'),
        'cultivar': get_options_list('cultivar__nome'),
        'peneira': get_options_list('peneira__nome'),
        'categoria': get_options_list('categoria__nome'),
        'endereco': get_options_list('endereco'),
        'especie': get_options_list('especie__nome'),
        'tratamento': get_options_list('tratamento__nome'),
        'embalagem': get_options_list('embalagem'),
        'cliente': get_options_list('cliente'),
        'empresa': get_options_list('empresa'),
        'conferente': get_options_list('conferente__username')
    }

    # Paginação
    page_size = request.GET.get('page_size', 25)
    try:
        page_size = int(page_size)
    except (ValueError, TypeError):
        page_size = 25
    
    paginator = Paginator(qs, page_size)
    page_number = request.GET.get('page', 1)
    
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']
    
    total_itens = qs.count()
    clientes_unicos = qs.exclude(cliente__isnull=True).exclude(cliente='').values('cliente').distinct().count()

    total_empenhado = qs.aggregate(e=Sum('empenhado'))['e'] or 0
    total_disponivel = (qs.aggregate(s=Sum('saldo'))['s'] or 0) - total_empenhado

    context = {
        'estoque': page_obj,
        'itens': page_obj,
        'busca': busca,
        'total_itens': total_itens,
        'total_sc': saldo_total_sc,
        'total_bags': saldo_bags,
        'total_sc_fisico': saldo_sc,
        'total_pme': total_pme,
        'clientes_unicos': clientes_unicos,
        'filter_options': filter_options,
        'url_params': query_params.urlencode(),
        'page_sizes': [10, 25, 50, 100, 200],
        'page_size': page_size,
        'total_empenhado': total_empenhado,
        'total_disponivel': total_disponivel,
    }
    
    return render(request, template_name, context)



@login_required
def opcoes_filtro_api(request):
    coluna = request.GET.get('coluna')

    if not coluna:
        return JsonResponse({
            'success': False,
            'error': 'Coluna não especificada'
        })

    qs = Estoque.objects.filter(saldo__gt=0)

    busca = request.GET.get('busca', '').strip()
    if busca:
        for termo in busca.split():
            qs = qs.filter(
                Q(lote__icontains=termo) |
                Q(produto__icontains=termo) |
                Q(cultivar__nome__icontains=termo) |
                Q(especie__nome__icontains=termo) |
                Q(endereco__icontains=termo) |
                Q(cliente__icontains=termo)
            )

    field_map = {
        'az': 'az',
        'lote': 'lote',
        'produto': 'produto',
        'cultivar': 'cultivar__nome',
        'peneira': 'peneira__nome',
        'categoria': 'categoria__nome',
        'endereco': 'endereco',
        'saldo': 'saldo',
        'peso_unitario': 'peso_unitario',
        'peso_total': 'peso_total',
        'especie': 'especie__nome',
        'tratamento': 'tratamento__nome',
        'embalagem': 'embalagem',
        'cliente': 'cliente',
        'empresa': 'empresa',
        'conferente': 'conferente__username',
    }

    filter_map = {
        'az': 'az__in',
        'lote': 'lote__in',
        'produto': 'produto__in',
        'cultivar': 'cultivar__nome__in',
        'peneira': 'peneira__nome__in',
        'categoria': 'categoria__nome__in',
        'endereco': 'endereco__in',
        'especie': 'especie__nome__in',
        'tratamento': 'tratamento__nome__in',
        'embalagem': 'embalagem__in',
        'cliente': 'cliente__in',
        'empresa': 'empresa__in',
        'conferente': 'conferente__username__in',
    }

    for param, lookup in filter_map.items():
        if param == coluna:
            continue

        values = request.GET.getlist(param)
        values = [v for v in values if v and v.strip()]

        if values:
            if '__null__' in values:
                specific_values = [v for v in values if v != '__null__']
                null_lookup = lookup.replace('__in', '__isnull')

                if specific_values:
                    qs = qs.filter(
                        Q(**{lookup: specific_values}) |
                        Q(**{null_lookup: True})
                    )
                else:
                    qs = qs.filter(**{null_lookup: True})
            else:
                qs = qs.filter(**{lookup: values})

    for field in ['saldo', 'peso_unitario', 'peso_total']:
        if field == coluna:
            continue

        min_val = request.GET.get(f'min_{field}')
        max_val = request.GET.get(f'max_{field}')

        if min_val:
            try:
                qs = qs.filter(**{f'{field}__gte': float(min_val)})
            except ValueError:
                pass

        if max_val:
            try:
                qs = qs.filter(**{f'{field}__lte': float(max_val)})
            except ValueError:
                pass

    status_filter = request.GET.getlist('status_sistemico')
    if status_filter and coluna != 'status_sistemico':
        status_ids = []

        for status_value in status_filter:
            try:
                if str(status_value).isdigit():
                    status_ids.append(int(status_value))
                else:
                    status_obj = StatusSistemico.objects.get(nome=status_value)
                    status_ids.append(status_obj.id)
            except StatusSistemico.DoesNotExist:
                pass

        if status_ids:
            qs = qs.filter(status_sistemico__in=status_ids)

    if coluna == 'status_sistemico':
        status_ids_em_uso = qs.exclude(
            status_sistemico__isnull=True
        ).values_list(
            'status_sistemico_id',
            flat=True
        ).distinct()

        status_em_uso = StatusSistemico.objects.filter(
            ativo=True,
            id__in=status_ids_em_uso
        ).order_by('ordem', 'nome')

        opcoes = [
            {
                'value': str(s.id),
                'label': f"{s.icone or ''} {s.nome}".strip()
            }
            for s in status_em_uso
        ]

        return JsonResponse({
            'success': True,
            'opcoes': opcoes,
            'tem_null': qs.filter(status_sistemico__isnull=True).exists()
        })

    if coluna not in field_map:
        return JsonResponse({
            'success': False,
            'error': f'Coluna inválida: {coluna}'
        })

    field = field_map[coluna]

    try:
        values = qs.filter(
            **{f'{field}__isnull': False}
        ).exclude(
            **{f'{field}': ''}
        ).values_list(
            field, flat=True
        ).distinct().order_by(field)

        values = [v for v in values if v is not None and str(v).strip() != '']

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

    opcoes = [
        {
            'value': str(v),
            'label': str(v)
        }
        for v in values
    ]

    tem_null = qs.filter(**{f'{field}__isnull': True}).exists()

    return JsonResponse({
        'success': True,
        'opcoes': opcoes,
        'tem_null': tem_null
    })

@login_required
@permission_required('sapp.pode_movimentar_estoque', raise_exception=True)
def registrar_saida(request, id):
    print("🔍 [REGISTRAR SAÍDA] Iniciando processamento da expedição")
    
    item = get_object_or_404(Estoque, id=id)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # 1. Captura de Dados
                qtd = int(request.POST.get('quantidade_saida', 0))
                carga = request.POST.get('numero_carga', '')
                motorista = request.POST.get('motorista', '')
                placa = request.POST.get('placa', '')
                cliente = request.POST.get('cliente', '')
                obs = request.POST.get('observacao', '')
                fotos = request.FILES.getlist('fotos')

                print(f"📦 Dados recebidos:")
                print(f"   Quantidade: {qtd}")
                print(f"   Carga: {carga}")
                print(f"   Motorista: {motorista}")
                print(f"   Placa: {placa}")
                print(f"   Cliente: {cliente}")
                print(f"   Obs: {obs}")
                print(f"   Fotos recebidas: {len(fotos)}")

                # 2. Validação Rigorosa
                erros = []
                if qtd <= 0: 
                    erros.append("❌ Quantidade inválida.")
                if qtd > item.saldo: 
                    erros.append(f"❌ Saldo insuficiente. Disponível: {item.saldo}.")
                if not motorista.strip(): 
                    erros.append("❌ Motorista é obrigatório.")
                if not placa.strip(): 
                    erros.append("❌ Placa é obrigatória.")
                if not carga.strip(): 
                    erros.append("❌ Número da Carga é obrigatório.")
                
                # Fotos são obrigatórias para expedição
                if len(fotos) == 0:
                    erros.append("❌ Pelo menos uma foto é obrigatória na expedição.")
                
                if erros:
                    for e in erros: 
                        print(f"⚠️ {e}")
                        messages.error(request, e)
                    return redirect('sapp:lista_estoque')

                # 3. Salvar estado anterior para histórico
                saldo_anterior = item.saldo
                print(f"💰 Saldo anterior: {saldo_anterior}")

                # 4. Processamento da Saída
                item.saida += qtd
                item.saldo = item.entrada - item.saida
                item.conferente = request.user
                item.data_ultima_saida = timezone.now()
                
                # Atualizar peso total
                if item.peso_unitario and item.peso_unitario > 0:
                    item.peso_total = Decimal(str(item.saldo)) * Decimal(str(item.peso_unitario))
                    item.peso_total = item.peso_total.quantize(Decimal('0.01'))
                
                # Atualizar observação
                obs_historico = f"[EXPEDIÇÃO {timezone.now().strftime('%d/%m/%Y %H:%M')}] Carga: {carga}, Motorista: {motorista}"
                if obs:
                    obs_historico += f" | Obs: {obs}"
                
                if item.observacao:
                    item.observacao += f"\n\n{obs_historico}"
                else:
                    item.observacao = obs_historico
                
                item.save()
                print(f"✅ Item atualizado: {item.lote} | Saldo anterior: {saldo_anterior} → Novo saldo: {item.saldo}")

                # 5. Descrição Rica em HTML para o Histórico
                desc_html = f"""
                    <div class="d-flex flex-column gap-1">
                        <div class="d-flex justify-content-between border-bottom pb-1">
                            <span><strong>Qtd Expedida:</strong> <span class="text-danger">-{qtd}</span></span>
                            <span><strong>Saldo Restante:</strong> {item.saldo}</span>
                        </div>
                        <div class="small text-muted mt-1">
                            <i class="fas fa-truck"></i> <strong>Carga:</strong> {carga} | <strong>Placa:</strong> {placa}<br>
                            <i class="fas fa-id-card"></i> <strong>Motorista:</strong> {motorista}<br>
                            <i class="fas fa-building"></i> <strong>Cliente:</strong> {cliente or 'N/A'}<br>
                            <i class="fas fa-user"></i> <strong>Responsável:</strong> {request.user.get_full_name() or request.user.username}<br>
                            <i class="fas fa-clock"></i> <strong>Data/Hora:</strong> {timezone.now().strftime('%d/%m/%Y %H:%M')}
                        </div>
                        {f'<div class="mt-1 p-1 bg-light rounded small"><strong>Obs:</strong> {obs}</div>' if obs else ''}
                    </div>
                """

                print(f"📝 Criando histórico de movimentação...")

                # 6. Criar histórico de movimentação
                historico = HistoricoMovimentacao.objects.create(
                    estoque=item,
                    usuario=request.user,
                    tipo='Expedição',
                    descricao=desc_html,
                    quantidade=qtd,
                    numero_carga=carga,
                    motorista=motorista,
                    placa=placa,
                    cliente=cliente
                )

                print(f"✅ Histórico criado: ID {historico.id}")

                # 7. **CORREÇÃO CRÍTICA: Salvar Fotos**
                fotos_salvas = 0
                for foto in fotos:
                    try:
                        # CORREÇÃO AQUI: Use o objeto 'historico' diretamente
                        FotoMovimentacao.objects.create(
                            historico=historico,  # Usando o objeto já salvo
                            arquivo=foto,
                            legenda=f"Expedição {carga} - {item.lote} - {timezone.now().strftime('%d/%m/%Y')}"
                        )
                        fotos_salvas += 1
                        print(f"📸 Foto salva: {foto.name} (ID: {historico.id})")
                    except Exception as foto_error:
                        print(f"⚠️ Erro ao salvar foto {foto.name}: {foto_error}")
                        # Não falha a operação por causa de uma foto

                print(f"✅ Fotos salvas: {fotos_salvas}/{len(fotos)}")

                # 8. Mensagem de sucesso
                mensagem_sucesso = f"✅ Expedição da Carga {carga} registrada com sucesso!"
                if fotos_salvas < len(fotos):
                    mensagem_sucesso += f" ({fotos_salvas}/{len(fotos)} fotos salvas)"
                
                messages.success(request, mensagem_sucesso)
                print(f"🎉 Expedição concluída com sucesso!")
                
                # 9. DEBUG: Verificar se fotos foram realmente salvas
                fotos_salvas_query = FotoMovimentacao.objects.filter(historico=historico).count()
                print(f"🔍 DEBUG - Fotos no banco para histórico {historico.id}: {fotos_salvas_query}")

        except Exception as e:
            import traceback
            print(f"💥 ERRO CRÍTICO NA EXPEDIÇÃO:")
            print(f"   Mensagem: {str(e)}")
            print(f"   Traceback: {traceback.format_exc()}")
            messages.error(request, f"❌ Erro crítico ao registrar expedição: {str(e)}")
            
    return redirect('sapp:lista_estoque')

from django.views.decorators.csrf import csrf_protect

@csrf_protect
@login_required
@permission_required('sapp.pode_movimentar_estoque', raise_exception=True)
def transferir(request, id):
    origem = get_object_or_404(Estoque, id=id)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                qtd = int(request.POST.get('quantidade', 0))
                tipo_transferencia = request.POST.get('tipo_transferencia', 'normal')
                novo_end = request.POST.get('novo_endereco', '').strip().upper()
                
                # === VALIDAÇÕES BÁSICAS (COMUNS A AMBOS OS TIPOS) ===
                if qtd <= 0:
                    messages.error(request, "❌ Quantidade deve ser maior que zero!")
                    return redirect('sapp:lista_estoque')
                
                if qtd > origem.saldo:
                    messages.error(request, f"❌ Saldo insuficiente. Disponível: {origem.saldo}")
                    return redirect('sapp:lista_estoque')
                
                # Validação de endereço apenas para transferência normal
                if tipo_transferencia == 'normal' and not novo_end:
                    messages.error(request, "❌ Novo endereço é obrigatório para transferência normal!")
                    return redirect('sapp:lista_estoque')
                
                # === 1. SEMPRE DAR BAIXA NA ORIGEM ===
                origem.saida += qtd
                origem.save()  # Saldo é recalculado automaticamente no save()
                
                # === 2. PROCESSAMENTO POR TIPO DE TRANSFERÊNCIA ===
                if tipo_transferencia == 'beneficiamento':
                    # ============================================
                    # CASO 1: ENVIO PARA BENEFICIAMENTO
                    # ============================================
                    
                    # Criar histórico de beneficiamento (NÃO cria destino)
                    descricao_beneficiamento = f"Enviado para beneficiamento – Quantidade: {qtd} {origem.embalagem}"
                    if novo_end:
                        descricao_beneficiamento += f" | Destino referência: {novo_end}"
                    
                    historico_beneficiamento = HistoricoMovimentacao.objects.create(
                        estoque=origem,
                        usuario=request.user,
                        tipo='Beneficiamento',
                        descricao=descricao_beneficiamento
                    )
                    
                    # Salvar fotos no histórico de beneficiamento
                    for f in request.FILES.getlist('fotos'):
                        FotoMovimentacao.objects.create(historico=historico_beneficiamento, arquivo=f)
                    
                    messages.success(
                        request, 
                        f"✅ Lote enviado para beneficiamento! Quantidade baixada: {qtd} {origem.embalagem}"
                    )
                    
                else:  # tipo_transferencia == 'normal'
                    # ============================================
                    # CASO 2: TRANSFERÊNCIA NORMAL (FLUXO ORIGINAL)
                    # ============================================
                    
                    # BUSCAR OBJETOS RELACIONADOS
                    # Espécie
                    novo_especie_id = request.POST.get('especie')
                    if novo_especie_id and novo_especie_id.strip() != '':
                        obj_especie = get_object_or_404(Especie, id=novo_especie_id)
                    else:
                        obj_especie = origem.especie
                    
                    # Cultivar
                    cultivar_id = request.POST.get('cultivar')
                    if cultivar_id and cultivar_id.strip() != '':
                        obj_cultivar = get_object_or_404(Cultivar, id=cultivar_id)
                    else:
                        obj_cultivar = origem.cultivar
                    
                    # Peneira
                    peneira_id = request.POST.get('peneira')
                    if peneira_id and peneira_id.strip() != '':
                        obj_peneira = get_object_or_404(Peneira, id=peneira_id)
                    else:
                        obj_peneira = origem.peneira
                    
                    # Categoria
                    categoria_id = request.POST.get('categoria')
                    if categoria_id and categoria_id.strip() != '':
                        obj_categoria = get_object_or_404(Categoria, id=categoria_id)
                    else:
                        obj_categoria = origem.categoria
                    
                    # Tratamento
                    tratamento_id = request.POST.get('tratamento')
                    if tratamento_id and tratamento_id.strip() != '':
                        obj_tratamento = get_object_or_404(Tratamento, id=tratamento_id)
                    else:
                        obj_tratamento = origem.tratamento
                    
                    # Processar peso unitário
                    peso_raw = request.POST.get('peso_unitario', origem.peso_unitario or '0')
                    try:
                        peso_raw = str(peso_raw).replace(',', '.')
                        if peso_raw.count('.') > 1:
                            partes = peso_raw.split('.')
                            peso_raw = f"{partes[0]}.{''.join(partes[1:])}"
                        novo_peso = Decimal(peso_raw).quantize(Decimal('0.01'))
                    except:
                        novo_peso = origem.peso_unitario or Decimal('0.00')
                    
                    # 🔥 CORREÇÃO: Separar os filtros corretamente
                    # Primeiro, montar dicionário com todos os campos EXCETO saldo__gt
                    campos_base = {
                        'lote': origem.lote,
                        'cultivar': obj_cultivar,
                        'especie': obj_especie,
                        'peneira': obj_peneira,
                        'categoria': obj_categoria,
                        'tratamento': obj_tratamento,
                        'embalagem': request.POST.get('embalagem', origem.embalagem),
                        'empresa': request.POST.get('empresa', origem.empresa or ''),
                        'cliente': request.POST.get('cliente', origem.cliente or ''),
                        'endereco': novo_end,
                    }
                    
                    # Buscar registro existente com MESMO PESO
                    destino_existente = Estoque.objects.filter(
                        **campos_base,
                        peso_unitario=novo_peso,
                        saldo__gt=0  # 🔥 AGORA CORRETO: um único argumento saldo__gt
                    ).first()
                    
                    # 🔥 CORREÇÃO: Buscar registro com PESO DIFERENTE
                    destino_peso_diferente = None
                    if not destino_existente:
                        destino_peso_diferente = Estoque.objects.filter(
                            **campos_base,  # Mesmos campos base
                            saldo__gt=0  # 🔥 AGORA CORRETO
                        ).exclude(
                            peso_unitario=novo_peso  # Exclui quem tem o mesmo peso
                        ).first()
                    
                    if destino_existente:
                        # 🔥 CASO 1: MESMO PESO - PODE SOMAR
                        saldo_anterior = destino_existente.saldo
                        destino_existente.entrada += qtd
                        destino_existente.saldo += qtd
                        
                        # Atualizar campos que podem ter mudado
                        destino_existente.peso_unitario = novo_peso  # Mantém o mesmo peso
                        destino_existente.empresa = request.POST.get('empresa', destino_existente.empresa or '')
                        destino_existente.cliente = request.POST.get('cliente', destino_existente.cliente or '')
                        destino_existente.az = request.POST.get('az', destino_existente.az or '')
                        destino_existente.conferente = request.user
                        
                        # Atualizar observação
                        obs_atual = destino_existente.observacao or ''
                        nova_obs = request.POST.get('observacao', '')
                        if nova_obs:
                            if obs_atual:
                                destino_existente.observacao = f"{obs_atual}\n[TRANSFERÊNCIA {timezone.now().strftime('%d/%m %H:%M')}]: {nova_obs}"
                            else:
                                destino_existente.observacao = f"[TRANSFERÊNCIA {timezone.now().strftime('%d/%m %H:%M')}]: {nova_obs}"
                        
                        destino_existente.save()
                        
                        destino = destino_existente
                        mensagem_tipo = f"somado ao registro existente (Saldo anterior: {saldo_anterior}, Peso: {novo_peso} kg)"
                        
                        print(f"✅ Somando ao lote existente com mesmo peso: {origem.lote} | Peso: {novo_peso} kg")
                        
                    elif destino_peso_diferente:
                        # 🔥 CASO 2: PESO DIFERENTE - NÃO SOMA, CRIA NOVO REGISTRO
                        print(f"⚠️ Lote {origem.lote} já existe em {novo_end} com peso DIFERENTE ({destino_peso_diferente.peso_unitario} kg vs {novo_peso} kg)")
                        
                        # Avisar ao usuário
                        messages.warning(
                            request,
                            f"⚠️ Já existe um lote {origem.lote} em {novo_end} com peso {destino_peso_diferente.peso_unitario} kg. "
                            f"Como o peso é diferente ({novo_peso} kg), foi criado um NOVO registro."
                        )
                        
                        # Criar NOVO registro (não somar)
                        destino = Estoque.objects.create(
                            lote=origem.lote,
                            endereco=novo_end,
                            entrada=qtd,
                            saldo=qtd,
                            conferente=request.user,
                            origem_destino=f"Transferência de {origem.endereco}",
                            
                            # Campos de texto com fallback
                            produto=request.POST.get('produto', origem.produto or ''),
                            cliente=request.POST.get('cliente', origem.cliente or ''),
                            empresa=request.POST.get('empresa', origem.empresa or ''),
                            az=request.POST.get('az', origem.az or ''),
                            peso_unitario=novo_peso,  # Peso NOVO
                            embalagem=request.POST.get('embalagem', origem.embalagem),
                            observacao=request.POST.get('observacao', origem.observacao or '') + f" [Peso: {novo_peso} kg - DIFERENTE DO EXISTENTE]",
                            
                            # Foreign Keys (Objetos, não IDs)
                            especie=obj_especie,
                            cultivar=obj_cultivar,
                            peneira=obj_peneira,
                            categoria=obj_categoria,
                            tratamento=obj_tratamento,
                        )
                        mensagem_tipo = f"criado no novo endereço (peso diferente: {novo_peso} kg)"
                        
                    else:
                        # 🔥 CASO 3: NÃO EXISTE - CRIAR NOVO REGISTRO
                        destino = Estoque.objects.create(
                            lote=origem.lote,
                            endereco=novo_end,
                            entrada=qtd,
                            saldo=qtd,
                            conferente=request.user,
                            origem_destino=f"Transferência de {origem.endereco}",
                            
                            # Campos de texto com fallback
                            produto=request.POST.get('produto', origem.produto or ''),
                            cliente=request.POST.get('cliente', origem.cliente or ''),
                            empresa=request.POST.get('empresa', origem.empresa or ''),
                            az=request.POST.get('az', origem.az or ''),
                            peso_unitario=novo_peso,
                            embalagem=request.POST.get('embalagem', origem.embalagem),
                            observacao=request.POST.get('observacao', origem.observacao or ''),
                            
                            # Foreign Keys (Objetos, não IDs)
                            especie=obj_especie,
                            cultivar=obj_cultivar,
                            peneira=obj_peneira,
                            categoria=obj_categoria,
                            tratamento=obj_tratamento,
                        )
                        mensagem_tipo = "criado no novo endereço"
                    
                    # Históricos (Saída da origem)
                    hist_saida = HistoricoMovimentacao.objects.create(
                        estoque=origem,
                        usuario=request.user,
                        tipo='Transferência (Saída)',
                        descricao=f"Transferido para {novo_end} ({destino.lote}) - Quantidade: {qtd} {origem.embalagem} | {mensagem_tipo}"
                    )
                    
                    # Histórico (Entrada no destino)
                    hist_entrada = HistoricoMovimentacao.objects.create(
                        estoque=destino,
                        usuario=request.user,
                        tipo='Transferência (Entrada)',
                        descricao=f"Recebido de {origem.endereco} ({origem.lote}) - Quantidade: {qtd} {origem.embalagem} | Peso: {novo_peso} kg | Novo saldo: {destino.saldo}"
                    )
                    
                    # Salvar fotos na saída (origem)
                    for f in request.FILES.getlist('fotos'):
                        FotoMovimentacao.objects.create(historico=hist_saida, arquivo=f)
                    
                    messages.success(request, f"✅ Transferência concluída! {qtd} unidades {mensagem_tipo} em {novo_end}")
                
        except Exception as e:
            import traceback
            print(f"❌ ERRO NA TRANSFERÊNCIA: {e}")
            print(traceback.format_exc())
            messages.error(request, f"❌ Erro ao transferir: {str(e)}")
            
    return redirect('sapp:lista_estoque')



# No seu views.py
from django.http import JsonResponse

@login_required
@permission_required('sapp.pode_ver_estoque', raise_exception=True)
def detalhes_estoque_api(request, id):
    """API para retornar dados de um item do estoque"""
    try:
        item = Estoque.objects.get(id=id)
        data = {
            'id': item.id,
            'lote': item.lote,
            'endereco': item.endereco,
            'saldo': item.saldo,
            'entrada': item.entrada,
            'produto': item.produto,
            'cliente': item.cliente,
            'empresa': item.empresa,
            'az': item.az,
            'peso_unitario': str(item.peso_unitario) if item.peso_unitario else '',
            'embalagem': item.embalagem,
            'observacao': item.observacao or '',
            'especie_id': item.especie.id if item.especie else '',
            'cultivar_id': item.cultivar.id if item.cultivar else '',
            'peneira_id': item.peneira.id if item.peneira else '',
            'categoria_id': item.categoria.id if item.categoria else '',
            'tratamento_id': item.tratamento.id if item.tratamento else '',
        }
        return JsonResponse(data)
    except Estoque.DoesNotExist:
        return JsonResponse({'error': 'Item não encontrado'}, status=404)

@login_required
@permission_required('sapp.pode_movimentar_estoque', raise_exception=True)
def nova_entrada(request):
    if request.method == 'POST':
        try:
            with transaction.atomic():
                lote = request.POST.get('lote', '').strip()
                endereco = request.POST.get('endereco', '').strip().upper()
                produto = request.POST.get('produto', '').strip()
                qtd = int(request.POST.get('entrada', 0))
                
                # 🔥 NOVO: Capturar o checkbox
                ultimo_lote_linha = request.POST.get('ultimo_lote_linha') == 'on'
                
                # Processar peso unitário
                peso_raw = request.POST.get('peso_unitario', '0')
                try:
                    peso_raw = str(peso_raw).replace(',', '.')
                    if peso_raw.count('.') > 1:
                        partes = peso_raw.split('.')
                        peso_raw = f"{partes[0]}.{''.join(partes[1:])}"
                    novo_peso = Decimal(peso_raw).quantize(Decimal('0.01'))
                except:
                    novo_peso = Decimal('0.00')
                
                # Buscar objetos relacionados
                especie_id = request.POST.get('especie')
                if especie_id:
                    especie_obj = get_object_or_404(Especie, id=especie_id)
                else:
                    especie_obj, _ = Especie.objects.get_or_create(nome='SOJA')

                cultivar = get_object_or_404(Cultivar, id=request.POST.get('cultivar'))
                peneira = get_object_or_404(Peneira, id=request.POST.get('peneira'))
                categoria = get_object_or_404(Categoria, id=request.POST.get('categoria'))
                
                trat_id = request.POST.get('tratamento')
                tratamento = Tratamento.objects.filter(id=trat_id).first() if trat_id else None

                # Buscar item existente
                item = Estoque.objects.filter(
                    lote=lote, 
                    endereco=endereco,
                    produto=produto,
                    cultivar=cultivar,
                    peso_unitario=novo_peso
                ).first()
                
                if item:
                    # Soma ao existente
                    item.entrada += qtd
                    item.observacao += f"\n[+ENTRADA {qtd} em {timezone.now().strftime('%d/%m')}]"
                    item.especie = especie_obj
                    
                    # 🔥 IMPORTANTE: Se for marcar como último lote
                    if ultimo_lote_linha:
                        # Verificar se já existe outro último na mesma linha
                        dados_end = extrair_ln_p(endereco)
                        if dados_end:
                            outro_ultimo = Estoque.objects.filter(
                                endereco__startswith=f"{dados_end['rua']} {dados_end['ln']} P",
                                ultimo_lote_linha=True
                            ).exclude(id=item.id).first()
                            
                            if outro_ultimo:
                                outro_ultimo.ultimo_lote_linha = False
                                outro_ultimo.save()
                        
                        item.ultimo_lote_linha = True
                    
                    msg = "adicionados ao lote existente"
                    print(f"✅ Somando ao lote existente: {lote}")
                else:
                    # Criar novo lote
                    item = Estoque(
                        lote=lote, 
                        endereco=endereco, 
                        entrada=qtd, 
                        saldo=qtd,
                        cultivar=cultivar, 
                        peneira=peneira, 
                        categoria=categoria, 
                        tratamento=tratamento,
                        especie=especie_obj,
                        conferente=request.user,
                        produto=produto,
                        cliente=request.POST.get('cliente', ''),
                        empresa=request.POST.get('empresa', ''),
                        az=request.POST.get('az', ''),
                        origem_destino=request.POST.get('origem_destino', ''),
                        peso_unitario=novo_peso,
                        embalagem=request.POST.get('embalagem', 'BAG'),
                        observacao=request.POST.get('observacao', ''),
                        ultimo_lote_linha=ultimo_lote_linha  # 🔥 NOVO
                    )
                    
                    # Se for marcar como último, verificar conflitos
                    if ultimo_lote_linha:
                        dados_end = extrair_ln_p(endereco)
                        if dados_end:
                            outro_ultimo = Estoque.objects.filter(
                                endereco__startswith=f"{dados_end['rua']} {dados_end['ln']} P",
                                ultimo_lote_linha=True
                            ).first()
                            
                            if outro_ultimo:
                                outro_ultimo.ultimo_lote_linha = False
                                outro_ultimo.save()
                    
                    msg = "criado com sucesso"
                    print(f"🆕 Novo lote criado: {lote}")
                
                item.save()
                
                # Calcular peso total
                if item.peso_unitario and item.peso_unitario > 0:
                    item.peso_total = Decimal(str(item.saldo)) * item.peso_unitario
                    item.peso_total = item.peso_total.quantize(Decimal('0.01'))
                    item.save()
                
                # Histórico
                status_ultimo = " e marcado como ÚLTIMO DA LINHA" if ultimo_lote_linha else ""
                descricao_historico = f"Entrada de {qtd} unidades. ({msg}{status_ultimo}) | Produto: {produto} | Peso: {novo_peso} kg"
                hist = HistoricoMovimentacao.objects.create(
                    estoque=item, 
                    usuario=request.user, 
                    tipo='Entrada',
                    descricao=descricao_historico
                )
                
                for f in request.FILES.getlist('fotos'):
                    FotoMovimentacao.objects.create(historico=hist, arquivo=f)
                
                messages.success(request, f"✅ Lote {lote} {msg}!{status_ultimo}")
                
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            messages.error(request, f"Erro ao processar entrada: {str(e)}")
            
    return redirect('sapp:lista_estoque')

@login_required
@permission_required('sapp.pode_ver_estoque', raise_exception=True)
def nova_saida(request):
    print("veio aqui na função  nova_saida")
    """Registra uma nova saída geral (para qualquer lote)"""
    if request.method == 'POST':
        try:
            lote_id = request.POST.get('lote_id')
            quantidade = int(request.POST.get('quantidade', 0))
            numero_carga = request.POST.get('numero_carga', '')
            motorista = request.POST.get('motorista', '')
            cliente = request.POST.get('cliente', '')
            observacao = request.POST.get('observacao', '')
            
            if not lote_id or quantidade <= 0:
                messages.error(request, "❌ Dados inválidos.")
                return redirect('sapp:lista_estoque')
            
            item = Estoque.objects.get(id=lote_id)
            
            if quantidade > item.saldo:
                messages.error(request, f"❌ Quantidade excede o saldo disponível ({item.saldo}).")
                return redirect('sapp:lista_estoque')
            
            # Salvar estado anterior
            saldo_anterior = item.saldo
            
            # Atualizar saída e saldo
            item.saida += quantidade  # CORRETO
            item.saldo = item.entrada - item.saida  # CORRETO
            
            # Recalcular peso total
            if item.peso_unitario:
                item.peso_total = Decimal(item.saldo) * Decimal(item.peso_unitario)
            
            # Atualizar data da última saída
            item.data_ultima_saida = timezone.now()
            
            # Atualizar observação
            if observacao:
                if item.observacao:
                    item.observacao += f"\n\n[EXPEDIÇÃO GERAL {timezone.now().strftime('%d/%m/%Y %H:%M')}]: {observacao}"
                else:
                    item.observacao = f"[EXPEDIÇÃO GERAL {timezone.now().strftime('%d/%m/%Y %H:%M')}]: {observacao}"
            
            item.save()
            
            # Registrar histórico
            historico = HistoricoMovimentacao.objects.create(
                estoque=item,
                usuario=request.user,
                tipo='Expedição via Sistema',
                descricao=(
                    f"<b>📤 EXPEDIÇÃO REGISTRADA</b><br>"
                    f"<b>Método:</b> Formulário Geral<br>"
                    f"<b>Quantidade:</b> {quantidade} unidades<br>"
                    f"<b>Carga:</b> {numero_carga}<br>"
                    f"<b>Motorista:</b> {motorista}<br>"
                    f"<b>Cliente:</b> {cliente}<br>"
                    f"<b>Saldo anterior:</b> {saldo_anterior}<br>"
                    f"<b>Novo saldo:</b> {item.saldo}<br>"
                    f"<b>Observação:</b> {observacao or 'Nenhuma'}<br>"
                    f"<b>Responsável:</b> {request.user.get_full_name() or request.user.username}"
                ),
                numero_carga=numero_carga,
                motorista=motorista,
                cliente=cliente
            )
            
            # Salvar foto se existir
            if 'foto' in request.FILES:
                historico.foto = request.FILES['foto']
                historico.save()
            
            messages.success(request, f"✅ Expedição de {quantidade} unidades registrada para o lote {item.lote}!")
            
        except Estoque.DoesNotExist:
            messages.error(request, "❌ Lote não encontrado.")
        except Exception as e:
            messages.error(request, f"❌ Erro ao registrar expedição: {str(e)}")
            import traceback
            print(f"🔍 Erro detalhado: {traceback.format_exc()}")
    
    return redirect('sapp:lista_estoque')

@login_required
@permission_required('sapp.pode_ver_estoque', raise_exception=True)
def relatorio_saidas(request):
    """Relatório detalhado de todas as saídas"""
    if request.method == 'POST':
        # Filtros por período
        data_inicio = request.POST.get('data_inicio')
        data_fim = request.POST.get('data_fim')
        
        saidas = HistoricoMovimentacao.objects.filter(tipo__contains='Saída')
        
        if data_inicio:
            saidas = saidas.filter(data_hora__gte=data_inicio)
        if data_fim:
            saidas = saidas.filter(data_hora__lte=data_fim)
        
        # Agrupar por destino
        saidas_por_destino = saidas.values('descricao').annotate(
            total=Count('id'),
            ultima_data=Max('data_hora')
        )
        
        context = {
            'saidas': saidas,
            'saidas_por_destino': saidas_por_destino,
            'total_saidas': saidas.count(),
            'periodo': f"{data_inicio} a {data_fim}" if data_inicio and data_fim else "Todos os períodos"
        }
        
        return render(request, 'sapp/relatorio_saidas.html', context)
    
    return render(request, 'sapp/relatorio_saidas.html')


from django.http import JsonResponse
from django.db.models import Sum, Q
from django.contrib.auth.decorators import login_required

@login_required
@permission_required('sapp.pode_ver_estoque', raise_exception=True)
def api_estoque_estatisticas(request):
    """API para atualizar os cards de estatísticas com base nos filtros atuais"""
    
    # Query base - apenas saldo > 0
    qs = Estoque.objects.filter(saldo__gt=0)
    
    # Aplicar os mesmos filtros da view principal
    # Status sistêmico
    status_filter = request.GET.getlist('status_sistemico')
    if status_filter:
        if '__null__' in status_filter:
            qs = qs.filter(
                Q(status_sistemico__in=[s for s in status_filter if s != '__null__']) | 
                Q(status_sistemico__isnull=True)
            )
        else:
            qs = qs.filter(status_sistemico__in=status_filter)
    
    # Busca
    busca = request.GET.get('busca', '').strip()
    if busca:
        for termo in busca.split():
            qs = qs.filter(
                Q(lote__icontains=termo) | 
                Q(produto__icontains=termo) |
                Q(cultivar__nome__icontains=termo) | 
                Q(endereco__icontains=termo) | 
                Q(cliente__icontains=termo)
            )
    
    # Filtros de seleção
    filter_map = {
        'az': 'az__in',
        'lote': 'lote__in',
        'produto': 'produto__in',
        'cultivar': 'cultivar__nome__in',
        'peneira': 'peneira__nome__in',
        'categoria': 'categoria__nome__in',
        'endereco': 'endereco__in',
        'especie': 'especie__nome__in',
        'tratamento': 'tratamento__nome__in',
        'embalagem': 'embalagem__in',
        'cliente': 'cliente__in',
        'empresa': 'empresa__in',
        'conferente': 'conferente__username__in'
    }

    for param, lookup in filter_map.items():
        values = request.GET.getlist(param)
        values = [v for v in values if v and v.strip()]
        if values:
            if '__null__' in values:
                specific_values = [v for v in values if v != '__null__']
                if specific_values:
                    qs = qs.filter(
                        Q(**{lookup: specific_values}) | 
                        Q(**{lookup.replace('__in', '__isnull'): True})
                    )
                else:
                    qs = qs.filter(**{lookup.replace('__in', '__isnull'): True})
            else:
                qs = qs.filter(**{lookup: values})
    
    # Filtros numéricos
    for field in ['saldo', 'peso_unitario', 'peso_total']:
        min_val = request.GET.get(f'min_{field}')
        max_val = request.GET.get(f'max_{field}')
        if min_val:
            try:
                qs = qs.filter(**{f'{field}__gte': float(min_val)})
            except ValueError:
                pass
        if max_val:
            try:
                qs = qs.filter(**{f'{field}__lte': float(max_val)})
            except ValueError:
                pass
    
    # Calcular estatísticas
    total_itens = qs.count()
    
    saldo_bags = qs.filter(embalagem='BAG').aggregate(s=Sum('saldo'))['s'] or 0
    saldo_sc = qs.filter(embalagem='SC').aggregate(s=Sum('saldo'))['s'] or 0
    total_sc = (saldo_bags * 25) + saldo_sc
    
    total_pme = qs.aggregate(s=Sum('peso_total'))['s'] or 0
    
    clientes_unicos = qs.exclude(cliente__isnull=True).exclude(cliente='').values('cliente').distinct().count()
    
    return JsonResponse({
        'success': True,
        'total_itens': total_itens,
        'total_sc': total_sc,
        'total_bags': saldo_bags,
        'total_pme': total_pme,
        'clientes_unicos': clientes_unicos
    })




@login_required
@permission_required('sapp.pode_ver_estoque', raise_exception=True)
def api_opcoes_filtro(request):
    """Retorna opções de filtro baseadas nos filtros atuais (encadeamento)"""
    coluna = request.GET.get('coluna')
    if not coluna:
        return JsonResponse({'success': False, 'error': 'Coluna não especificada'})

    # Query base - APENAS saldo > 0
    qs = Estoque.objects.filter(saldo__gt=0)

    # Aplicar TODOS os filtros atuais, exceto a coluna que estamos abrindo
    # Status sistêmico
    status_filter = request.GET.getlist('status_sistemico')
    if status_filter and coluna != 'status_sistemico':
        if '__null__' in status_filter:
            qs = qs.filter(
                Q(status_sistemico__in=[s for s in status_filter if s != '__null__']) | 
                Q(status_sistemico__isnull=True)
            )
        else:
            qs = qs.filter(status_sistemico__in=status_filter)

    # Busca
    busca = request.GET.get('busca', '').strip()
    if busca:
        for termo in busca.split():
            qs = qs.filter(
                Q(lote__icontains=termo) | 
                Q(produto__icontains=termo) |
                Q(cultivar__nome__icontains=termo) | 
                Q(endereco__icontains=termo) | 
                Q(cliente__icontains=termo)
            )

    # Mapeamento de filtros
    filter_map = {
        'az': 'az__in',
        'lote': 'lote__in',
        'produto': 'produto__in',
        'cultivar': 'cultivar__nome__in',
        'peneira': 'peneira__nome__in',
        'categoria': 'categoria__nome__in',
        'endereco': 'endereco__in',
        'especie': 'especie__nome__in',
        'tratamento': 'tratamento__nome__in',
        'embalagem': 'embalagem__in',
        'cliente': 'cliente__in',
        'empresa': 'empresa__in',
        'conferente': 'conferente__username__in'
    }

    # Aplicar outros filtros (exceto a coluna atual)
    for param, lookup in filter_map.items():
        if param == coluna:
            continue

        values = request.GET.getlist(param)
        values = [v for v in values if v and v.strip()]
        if values:
            if '__null__' in values:
                specific_values = [v for v in values if v != '__null__']
                if specific_values:
                    qs = qs.filter(
                        Q(**{lookup: specific_values}) | 
                        Q(**{lookup.replace('__in', '__isnull'): True})
                    )
                else:
                    qs = qs.filter(**{lookup.replace('__in', '__isnull'): True})
            else:
                qs = qs.filter(**{lookup: values})

    # Filtros numéricos (exceto a coluna atual)
    for field in ['saldo', 'peso_unitario', 'peso_total']:
        if field == coluna:
            continue
        min_val = request.GET.get(f'min_{field}')
        max_val = request.GET.get(f'max_{field}')
        if min_val:
            try:
                qs = qs.filter(**{f'{field}__gte': float(min_val)})
            except ValueError:
                pass
        if max_val:
            try:
                qs = qs.filter(**{f'{field}__lte': float(max_val)})
            except ValueError:
                pass

    # Mapeamento para buscar os valores distintos
    field_lookup_map = {
        'az': 'az',
        'lote': 'lote',
        'produto': 'produto',
        'cultivar': 'cultivar__nome',
        'peneira': 'peneira__nome',
        'categoria': 'categoria__nome',
        'endereco': 'endereco',
        'especie': 'especie__nome',
        'tratamento': 'tratamento__nome',
        'embalagem': 'embalagem',
        'cliente': 'cliente',
        'empresa': 'empresa',
        'conferente': 'conferente__username'
    }

    if coluna in field_lookup_map:
        lookup = field_lookup_map[coluna]
        tem_null = qs.filter(**{lookup + '__isnull': True}).exists() or qs.filter(**{lookup: ''}).exists()
        valores = qs.exclude(**{lookup: None}).exclude(**{lookup: ''}).values_list(lookup, flat=True).distinct().order_by(lookup)
        opcoes = [str(v) for v in valores if v is not None and str(v).strip() != '']

        return JsonResponse({
            'success': True,
            'opcoes': opcoes,
            'tem_null': tem_null
        })

    return JsonResponse({'success': False, 'error': 'Coluna inválida'})


############################################################################
# NO VIEWS.PY - CORRIGIR A FUNÇÃO editar COMPLETAMENTE:
@login_required
@permission_required('sapp.pode_movimentar_estoque', raise_exception=True)
def editar(request, id):
    item = get_object_or_404(Estoque, id=id)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # 1. CAPTURA O ESTADO ANTIGO (Para histórico)
                antigo = {
                    'lote': item.lote,
                    'endereco': item.endereco,
                    'empresa': item.empresa or "",
                    'origem_destino': item.origem_destino or "",
                    'peso_unitario': item.peso_unitario,
                    'entrada': item.entrada,  # NOVO
                    'saida': item.saida,      # NOVO (para referência)
                    'saldo': item.saldo,      # NOVO (para referência)
                    'embalagem': item.embalagem,
                    'az': item.az or "",
                    'observacao': item.observacao or "",
                    'cliente': item.cliente or "", 
                    'cultivar': item.cultivar.nome if item.cultivar else "",
                    'peneira': item.peneira.nome if item.peneira else "",
                    'categoria': item.categoria.nome if item.categoria else "",
                    'especie': item.especie.nome if item.especie else "SOJA",
                    'tratamento': item.tratamento.nome if item.tratamento else "Sem Tratamento",
                    'produto': item.produto or "", 
                }

                # 2. CAPTURA OS NOVOS VALORES
                novo_lote = request.POST.get('lote', '').strip()
                novo_endereco = request.POST.get('endereco', '').strip().upper()
                novo_empresa = request.POST.get('empresa', '').strip()
                novo_origem_destino = request.POST.get('origem_destino', '').strip()
                novo_produto = request.POST.get('produto', '').strip()
                novo_cliente = request.POST.get('cliente', '').strip()
                
                # NOVO: Capturar quantidade
                nova_entrada_raw = request.POST.get('entrada', '0')
                try:
                    nova_entrada = int(float(nova_entrada_raw))
                    if nova_entrada < 0:
                        nova_entrada = 0
                except:
                    nova_entrada = item.entrada
                
                # Tratamento do peso
                peso_raw = request.POST.get('peso_unitario', '0')
                try:
                    peso_raw = str(peso_raw).replace(',', '.')
                    if peso_raw.count('.') > 1:
                        partes = peso_raw.split('.')
                        peso_raw = f"{partes[0]}.{''.join(partes[1:])}"
                    novo_peso = Decimal(peso_raw)
                except:
                    novo_peso = Decimal('0.00')
                
                novo_emb = request.POST.get('embalagem', 'BAG')
                novo_az = request.POST.get('az', '').strip()
                novo_obs = request.POST.get('observacao', '').strip()

                # 3. BUSCAR OBJETOS RELACIONADOS
                # Espécie
                novo_especie_id = request.POST.get('especie')
                if novo_especie_id:
                    obj_especie = get_object_or_404(Especie, id=novo_especie_id)
                else:
                    obj_especie = item.especie

                # Cultivar
                try:
                    obj_cultivar = get_object_or_404(Cultivar, id=request.POST.get('cultivar'))
                except:
                    obj_cultivar = item.cultivar
                    
                # Peneira
                try:
                    obj_peneira = get_object_or_404(Peneira, id=request.POST.get('peneira'))
                except:
                    obj_peneira = item.peneira
                    
                # Categoria
                try:
                    obj_categoria = get_object_or_404(Categoria, id=request.POST.get('categoria'))
                except:
                    obj_categoria = item.categoria
                
                # Tratamento
                tratamento_id = request.POST.get('tratamento')
                if tratamento_id:
                    try:
                        obj_tratamento = get_object_or_404(Tratamento, id=tratamento_id)
                    except:
                        obj_tratamento = item.tratamento
                else:
                    obj_tratamento = None

                # 4. COMPARAÇÃO DETALHADA PARA O HISTÓRICO
                mudancas = []
                
                # Campos básicos (incluindo entrada)
                campos_para_comparar = [
                    ('lote', 'Lote', antigo['lote'], novo_lote),
                    ('endereco', 'Endereço', antigo['endereco'], novo_endereco),
                    ('empresa', 'Empresa', antigo['empresa'], novo_empresa),
                    ('origem_destino', 'Origem/Destino', antigo['origem_destino'], novo_origem_destino),
                    ('produto', 'Produto', antigo['produto'], novo_produto),
                    ('cliente', 'Cliente', antigo['cliente'], novo_cliente),
                    ('peso_unitario', 'Peso Unitário', antigo['peso_unitario'], novo_peso),
                    ('entrada', 'Quantidade (Entrada)', antigo['entrada'], nova_entrada),  # NOVO
                    ('embalagem', 'Embalagem', antigo['embalagem'], novo_emb),
                    ('az', 'AZ', antigo['az'], novo_az),
                    ('observacao', 'Observação', antigo['observacao'], novo_obs),
                    ('cultivar', 'Cultivar', antigo['cultivar'], obj_cultivar.nome if obj_cultivar else '-'),
                    ('peneira', 'Peneira', antigo['peneira'], obj_peneira.nome if obj_peneira else '-'),
                    ('categoria', 'Categoria', antigo['categoria'], obj_categoria.nome if obj_categoria else '-'),
                    ('especie', 'Espécie', antigo['especie'], obj_especie.nome if obj_especie else '-'),
                    ('tratamento', 'Tratamento', antigo['tratamento'], obj_tratamento.nome if obj_tratamento else 'Sem Tratamento'),
                ]
                
                for campo_nome, label, valor_antigo, valor_novo in campos_para_comparar:
                    if str(valor_antigo or '') != str(valor_novo or ''):
                        mudancas.append(f"{label}: {valor_antigo} → <b>{valor_novo}</b>")

                # 5. ATUALIZAR O OBJETO
                item.lote = novo_lote
                item.endereco = novo_endereco
                item.empresa = novo_empresa
                item.origem_destino = novo_origem_destino
                item.produto = novo_produto
                item.cliente = novo_cliente
                item.peso_unitario = novo_peso
                item.entrada = nova_entrada  # NOVO: Atualiza a entrada
                # NÃO altera a saída - mantém o valor original
                item.embalagem = novo_emb
                item.az = novo_az
                item.observacao = novo_obs
                
                # Atualizando Foreign Keys
                item.cultivar = obj_cultivar
                item.peneira = obj_peneira
                item.categoria = obj_categoria
                item.tratamento = obj_tratamento
                item.especie = obj_especie
                
                item.conferente = request.user
                
                # 6. SALVAR (o método save() recalcula saldo e peso_total automaticamente)
                item.save()

                # 7. VERIFICAR SE HOUVE MUDANÇA NA QUANTIDADE E CRIAR HISTÓRICO ESPECÍFICO
                if antigo['entrada'] != nova_entrada:
                    diferenca = nova_entrada - antigo['entrada']
                    if diferenca > 0:
                        tipo_historico = 'Ajuste de Estoque (Adição)'
                        descricao_adicional = f"<br><span class='text-success'>📦 Quantidade aumentada em <b>{diferenca}</b> unidades (entrada: {antigo['entrada']} → {nova_entrada})</span>"
                    else:
                        tipo_historico = 'Ajuste de Estoque (Redução)'
                        descricao_adicional = f"<br><span class='text-danger'>📦 Quantidade reduzida em <b>{abs(diferenca)}</b> unidades (entrada: {antigo['entrada']} → {nova_entrada})</span>"
                    
                    # Adiciona ao histórico principal ou cria um separado
                    HistoricoMovimentacao.objects.create(
                        estoque=item,
                        usuario=request.user,
                        tipo=tipo_historico,
                        descricao=f"<b>AJUSTE MANUAL DE QUANTIDADE:</b><br>{descricao_adicional}"
                    )

                # 8. REGISTRAR HISTÓRICO PRINCIPAL
                if mudancas:
                    descricao_html = "<br>".join(mudancas)
                    HistoricoMovimentacao.objects.create(
                        estoque=item,
                        usuario=request.user,
                        tipo='Edição de Lote',
                        descricao=f"<b>EDIÇÃO REALIZADA:</b><br>{descricao_html}"
                    )
                elif antigo['entrada'] == nova_entrada:  # Só cria se não houve mudança na quantidade
                    HistoricoMovimentacao.objects.create(
                        estoque=item,
                        usuario=request.user,
                        tipo='Edição (Sem mudanças)',
                        descricao="Salvo sem alterações visíveis."
                    )

                messages.success(request, f"✅ Lote {item.lote} atualizado com sucesso! Saldo atual: {item.saldo} unidades")
                
        except Exception as e:
            import traceback
            print(f"❌ ERRO NA EDIÇÃO: {e}")
            print(traceback.format_exc())
            messages.error(request, f"Erro ao editar lote: {str(e)}")
            
    return redirect('sapp:lista_estoque')
      






@login_required
@permission_required('sapp.pode_ver_estoque', raise_exception=True)
def excluir_lote(request, id):
    item = get_object_or_404(Estoque, id=id)
    if request.method == 'POST':
        HistoricoMovimentacao.objects.create(
            estoque=None, 
            lote_ref=f"{item.lote} (Excluído)",
            usuario=request.user,
            tipo='EXCLUSÃO',
            descricao=f"Lote <b>{item.lote}</b> do endereço <b>{item.endereco}</b> foi excluído."
        )
        item.delete()
        messages.success(request, "Lote excluído.")
    return redirect('sapp:lista_estoque')

def logout_view(request):
    """
    Realiza o logout e redireciona para o login.
    Aceita POST (padrão recomendado) ou GET se necessário.
    """
    logout(request)
    return redirect('sapp:login')





# sapp/views.py - Função completa corrigida

# sapp/views.py - Substitua a função configuracoes por esta versão SIMPLIFICADA

@login_required
@permission_required('sapp.pode_configuracoes', raise_exception=True)
def configuracoes(request):
    """
    View completa de configurações do sistema
    Gerencia: Produtos, Armazéns, Endereços, Usuários, Permissões
    """
    
    config = Configuracao.get_solo()
    
    # =============================
    # QUERYSETS PRINCIPAIS
    # =============================
    
    # Usuários (todos exceto o próprio usuário logado)
    usuarios_conferentes = User.objects.filter(
        is_superuser=False
    ).exclude(id=request.user.id).order_by('username')
    
    # Produtos com relacionamentos
    produtos = Produto.objects.select_related(
        'cultivar', 'peneira', 'especie', 'categoria', 'tratamento'
    ).all().order_by('-data_cadastro')
    
    # Parâmetros
    cultivares = Cultivar.objects.all().order_by('nome')
    peneiras = Peneira.objects.all().order_by('nome')
    especies = Especie.objects.all().order_by('nome')
    categorias = Categoria.objects.all().order_by('nome')
    tratamentos = Tratamento.objects.all().order_by('nome')
    
    # Armazéns
    armazens_lista = Armazem.objects.all().order_by('nome')
    
    # Endereços com armazém
    enderecos_lista = Endereco.objects.select_related('armazem').all().order_by('codigo')
    
    # Origens/Destinos
    origens_lista = OrigemDestino.objects.all().order_by('nome')
    
    # =============================
    # PROCESSAMENTO POST
    # =============================
    
    if request.method == 'POST':
        
        acao = request.POST.get('acao')
        active_tab = request.POST.get('active_tab', 'produto')
        
        # ====================================
        # 1. PRODUTOS
        # ====================================
        
        if acao == 'add_produto':
            try:
                cultivar_id = request.POST.get('cultivar')
                codigo = request.POST.get('codigo', '').strip().upper()
                descricao = request.POST.get('descricao', '').strip()
                
                if not cultivar_id or not codigo:
                    messages.error(request, "❌ Cultivar e Código são obrigatórios!")
                elif Produto.objects.filter(codigo=codigo).exists():
                    messages.error(request, f"❌ Código '{codigo}' já existe!")
                else:
                    with transaction.atomic():
                        produto = Produto.objects.create(
                            cultivar_id=cultivar_id,
                            codigo=codigo,
                            descricao=descricao,
                            tipo=request.POST.get('tipo', '').strip(),
                            empresa=request.POST.get('empresa', '').strip(),
                            ativo=request.POST.get('ativo') == 'on'
                        )
                        produto.peneira_id = request.POST.get('peneira') or None
                        produto.especie_id = request.POST.get('especie') or None
                        produto.categoria_id = request.POST.get('categoria') or None
                        produto.tratamento_id = request.POST.get('tratamento') or None
                        produto.save()
                        messages.success(request, f"✅ Produto '{codigo}' cadastrado com sucesso!")
            except Exception as e:
                messages.error(request, f"❌ Erro ao cadastrar produto: {str(e)}")
        
        elif acao == 'delete_produto':
            try:
                item_id = request.POST.get('id_item')
                if not item_id:
                    messages.error(request, "❌ Produto não identificado!")
                else:
                    produto = Produto.objects.get(id=item_id)
                    codigo = produto.codigo
                    produto.delete()
                    messages.success(request, f"✅ Produto '{codigo}' excluído com sucesso!")
            except Produto.DoesNotExist:
                messages.error(request, "❌ Produto não encontrado!")
            except Exception as e:
                messages.error(request, f"❌ Erro ao excluir produto: {str(e)}")
        
        # ====================================
        # 2. USUÁRIOS E PERMISSÕES (APENAS INDIVIDUAIS - SEM GRUPOS)
        # ====================================
        
        elif acao == 'create_conferente_user':
            # Verifica permissão para criar usuários
            if not request.user.is_superuser and not request.user.has_perm('sapp.pode_gerenciar_usuarios'):
                messages.error(request, "❌ Você não tem permissão para criar usuários!")
            else:
                username = request.POST.get('username', '').strip().lower()
                first_name = request.POST.get('first_name', '').strip()
                password = request.POST.get('password', '').strip()
                
                # Validações
                if not username or not first_name:
                    messages.error(request, "❌ Nome de usuário e nome completo são obrigatórios!")
                elif User.objects.filter(username=username).exists():
                    messages.error(request, f"❌ Usuário '{username}' já existe!")
                else:
                    try:
                        with transaction.atomic():
                            # Define senha padrão se não for fornecida
                            if not password:
                                password = 'conceito123'
                            elif len(password) < 6:
                                messages.error(request, "❌ A senha deve ter no mínimo 6 caracteres!")
                                return redirect(f"{reverse('sapp:configuracoes')}#{active_tab}")
                            
                            # Cria o usuário (SEM grupos)
                            user = User.objects.create_user(
                                username=username,
                                first_name=first_name,
                                password=password
                            )
                            
                            # NÃO ADICIONA GRUPOS - apenas permissões individuais
                            
                            # 🔥 Adiciona permissões específicas selecionadas nos checkboxes
                            permissions_added = []
                            for key, value in request.POST.items():
                                if key.startswith('pode_') and value == 'on':
                                    try:
                                        # Buscar permissão no app sapp
                                        permission = Permission.objects.filter(
                                            codename=key,
                                            content_type__app_label='sapp'
                                        ).first()
                                        
                                        # Se não encontrar, buscar no almoxarifado
                                        if not permission:
                                            permission = Permission.objects.filter(
                                                codename=key,
                                                content_type__app_label='almoxarifado'
                                            ).first()
                                        
                                        if permission:
                                            user.user_permissions.add(permission)
                                            permissions_added.append(key)
                                    except Exception as e:
                                        print(f"Erro ao adicionar permissão {key}: {e}")
                            
                            messages.success(request, f"✅ Usuário '{first_name}' criado com sucesso! Senha: {password}")
                            if permissions_added:
                                messages.info(request, f"📋 Permissões adicionadas: {', '.join(permissions_added)}")
                    
                    except Exception as e:
                        messages.error(request, f"❌ Erro ao criar usuário: {str(e)}")
        
# sapp/views.py - Substitua a função update_user_permissions

        elif acao == 'update_user_permissions':
            # Verifica permissão para editar permissões
            if not request.user.is_superuser and not request.user.has_perm('sapp.pode_gerenciar_usuarios'):
                messages.error(request, "❌ Você não tem permissão para editar permissões!")
            else:
                user_id = request.POST.get('user_id')
                
                print(f"\n🔍 [DEBUG] Recebida requisição update_user_permissions")
                print(f"   user_id: {user_id}")
                print(f"   POST keys: {list(request.POST.keys())}")
                
                try:
                    user = User.objects.get(id=user_id)
                    
                    if user == request.user and not request.user.is_superuser:
                        messages.error(request, "❌ Você não pode editar suas próprias permissões!")
                    else:
                        with transaction.atomic():
                            # 🔥 LIMPA TODAS as permissões atuais
                            user.user_permissions.clear()
                            print(f"   ✅ Permissões antigas removidas")
                            
                            # 🔥 Lista para guardar as permissões adicionadas
                            permissions_added = []
                            
                            # 🔥 Percorre todos os campos do POST
                            for key, value in request.POST.items():
                                # Ignora campos que não são permissões
                                if key in ['csrfmiddlewaretoken', 'acao', 'user_id', 'active_tab', 'group_name']:
                                    continue
                                
                                print(f"   Campo: {key} = {value}")
                                
                                # sapp/views.py - Substitua a parte de busca de permissão

                                if value == 'on':  # Checkbox marcado
                                    permission = None
                                    
                                    # 🔥 CORREÇÃO: Buscar em ORDEM CORRETA
                                    # Primeiro no app almoxarifado (para permissões de almoxarifado)
                                    if key in ['pode_ver_almoxarifado', 'pode_gerenciar_almoxarifado']:
                                        permission = Permission.objects.filter(
                                            codename=key,
                                            content_type__app_label='almoxarifado'
                                        ).first()
                                    
                                    # Depois no app sapp
                                    if not permission:
                                        permission = Permission.objects.filter(
                                            codename=key,
                                            content_type__app_label='sapp'
                                        ).first()
                                    
                                    if permission:
                                        user.user_permissions.add(permission)
                                        permissions_added.append(key)
                                        print(f"   ✅ Adicionada permissão: {key} (app: {permission.content_type.app_label})")
                                    else:
                                        print(f"   ❌ Permissão não encontrada: {key}")
                            
                            # 🔥 Salvar (garantir que foi salvo)
                            user.save()
                            
                            # 🔥 Verificar se salvou
                            saved_perms = list(user.user_permissions.values_list('codename', flat=True))
                            print(f"   📋 Permissões salvas no banco: {saved_perms}")
                            
                            if permissions_added:
                                messages.success(request, f"✅ Permissões de '{user.first_name}' atualizadas! ({len(permissions_added)} permissões)")
                            else:
                                messages.success(request, f"✅ Todas as permissões de '{user.first_name}' foram removidas!")
                            
                except User.DoesNotExist:
                    messages.error(request, "❌ Usuário não encontrado!")
                except Exception as e:
                    print(f"❌ Erro: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    messages.error(request, f"❌ Erro ao atualizar permissões: {str(e)}")
        elif acao == 'reset_password':
            # Verifica permissão para resetar senha
            if not request.user.is_superuser and not request.user.has_perm('sapp.pode_gerenciar_usuarios'):
                messages.error(request, "❌ Você não tem permissão para resetar senhas!")
            else:
                user_id = request.POST.get('user_id')
                try:
                    user = User.objects.get(id=user_id)
                    
                    # Não permite resetar próprio usuário (exceto se for superusuário)
                    if user == request.user and not request.user.is_superuser:
                        messages.error(request, "❌ Você não pode resetar sua própria senha!")
                    else:
                        new_password = 'conceito123'
                        user.set_password(new_password)
                        user.save()
                        messages.success(request, f"✅ Senha de '{user.first_name}' resetada para: {new_password}")
                
                except User.DoesNotExist:
                    messages.error(request, "❌ Usuário não encontrado!")
                except Exception as e:
                    messages.error(request, f"❌ Erro ao resetar senha: {str(e)}")
        
        elif acao == 'delete_user':
            # Verifica permissão para excluir usuário
            if not request.user.is_superuser and not request.user.has_perm('sapp.pode_gerenciar_usuarios'):
                messages.error(request, "❌ Você não tem permissão para excluir usuários!")
            else:
                user_id = request.POST.get('user_id')
                try:
                    user = User.objects.get(id=user_id)
                    
                    # Não permite excluir próprio usuário
                    if user == request.user:
                        messages.error(request, "❌ Você não pode excluir sua própria conta!")
                    else:
                        username = user.username
                        user.delete()
                        messages.success(request, f"✅ Usuário '{username}' excluído com sucesso!")
                
                except User.DoesNotExist:
                    messages.error(request, "❌ Usuário não encontrado!")
                except Exception as e:
                    messages.error(request, f"❌ Erro ao excluir usuário: {str(e)}")
        
        # ====================================
        # 3. ARMAZÉNS
        # ====================================
        
        elif acao == 'add_armazem':
            nome = request.POST.get('nome', '').strip().upper()
            if nome:
                obj, created = Armazem.objects.get_or_create(nome=nome)
                if created:
                    messages.success(request, f"✅ Armazém '{nome}' criado com sucesso!")
                else:
                    messages.warning(request, f"⚠️ Armazém '{nome}' já existe!")
            else:
                messages.error(request, "❌ Nome do armazém não informado!")
        
        # ====================================
        # 4. ENDEREÇOS
        # ====================================
        
        elif acao == 'add_endereco':
            endereco_codigo = request.POST.get('endereco_codigo', '').strip().upper()
            armazem_id = request.POST.get('armazem_id')
            
            if not endereco_codigo:
                messages.error(request, "❌ Endereço não informado!")
            elif not armazem_id:
                messages.error(request, "❌ Selecione um armazém!")
            else:
                try:
                    armazem = Armazem.objects.get(id=armazem_id)
                    
                    if Endereco.objects.filter(codigo=endereco_codigo).exists():
                        messages.warning(request, f"⚠️ Endereço '{endereco_codigo}' já cadastrado!")
                    else:
                        Endereco.objects.create(
                            codigo=endereco_codigo,
                            armazem=armazem
                        )
                        messages.success(request, f"✅ Endereço '{endereco_codigo}' cadastrado no armazém '{armazem.nome}'!")
                        
                except Armazem.DoesNotExist:
                    messages.error(request, "❌ Armazém não encontrado!")
                except Exception as e:
                    messages.error(request, f"❌ Erro ao cadastrar endereço: {str(e)}")
        
        # ====================================
        # 5. CONFIGURAÇÃO GERAL
        # ====================================
        
        elif acao == 'config_geral':
            form = ConfiguracaoForm(request.POST, instance=config)
            if form.is_valid():
                form.save()
                messages.success(request, "✅ Configurações gerais salvas com sucesso!")
            else:
                messages.error(request, "❌ Erro ao salvar configurações. Verifique os dados.")
        
        # ====================================
        # 6. CADASTROS SIMPLES (CRUD)
        # ====================================
        
        elif acao in ['add_cultivar', 'add_peneira', 'add_especie', 'add_categoria', 'add_tratamento', 'add_origem']:
            model_map = {
                'add_cultivar': Cultivar,
                'add_peneira': Peneira,
                'add_especie': Especie,
                'add_categoria': Categoria,
                'add_tratamento': Tratamento,
                'add_origem': OrigemDestino
            }
            model = model_map.get(acao)
            nome_display = {
                'add_cultivar': 'Cultivar',
                'add_peneira': 'Peneira',
                'add_especie': 'Espécie',
                'add_categoria': 'Categoria',
                'add_tratamento': 'Tratamento',
                'add_origem': 'Origem/Destino'
            }
            
            if model:
                nome = request.POST.get('nome', '').strip()
                if nome:
                    obj, created = model.objects.get_or_create(nome=nome)
                    if created:
                        messages.success(request, f"✅ {nome_display[acao]} '{nome}' adicionado com sucesso!")
                    else:
                        messages.warning(request, f"⚠️ {nome_display[acao]} '{nome}' já existe!")
                else:
                    messages.error(request, f"❌ Nome do {nome_display[acao]} não informado!")
        
        # ====================================
        # 7. EXCLUSÃO GENÉRICA
        # ====================================
        
        elif acao == 'delete_item':
            tipo = request.POST.get('tipo_item')
            item_id = request.POST.get('id_item')
            
            if not item_id:
                messages.error(request, "❌ Item não identificado!")
            else:
                model_map = {
                    'cultivar': (Cultivar, 'Cultivar'),
                    'especie': (Especie, 'Espécie'),
                    'peneira': (Peneira, 'Peneira'),
                    'categoria': (Categoria, 'Categoria'),
                    'tratamento': (Tratamento, 'Tratamento'),
                    'armazem': (Armazem, 'Armazém'),
                    'endereco': (Endereco, 'Endereço'),
                    'origem': (OrigemDestino, 'Origem/Destino'),
                    'produto': (Produto, 'Produto'),
                }
                
                if tipo in model_map:
                    model, nome_tipo = model_map[tipo]
                    try:
                        item = model.objects.get(id=item_id)
                        nome_excluido = str(item)
                        
                        # Validações de integridade referencial
                        if tipo == 'endereco' and hasattr(item, 'estoque_set') and item.estoque_set.exists():
                            messages.error(request, f"❌ Endereço '{nome_excluido}' está sendo usado em lotes de estoque!")
                        elif tipo == 'armazem' and hasattr(item, 'enderecos') and item.enderecos.exists():
                            messages.error(request, f"❌ Armazém '{nome_excluido}' possui endereços vinculados!")
                        elif tipo == 'cultivar' and Produto.objects.filter(cultivar=item).exists():
                            messages.error(request, f"❌ Cultivar '{nome_excluido}' está sendo usado em produtos!")
                        elif tipo == 'especie' and Produto.objects.filter(especie=item).exists():
                            messages.error(request, f"❌ Espécie '{nome_excluido}' está sendo usada em produtos!")
                        elif tipo == 'peneira' and Produto.objects.filter(peneira=item).exists():
                            messages.error(request, f"❌ Peneira '{nome_excluido}' está sendo usada em produtos!")
                        elif tipo == 'categoria' and Produto.objects.filter(categoria=item).exists():
                            messages.error(request, f"❌ Categoria '{nome_excluido}' está sendo usada em produtos!")
                        elif tipo == 'tratamento' and Produto.objects.filter(tratamento=item).exists():
                            messages.error(request, f"❌ Tratamento '{nome_excluido}' está sendo usado em produtos!")
                        else:
                            item.delete()
                            messages.success(request, f"✅ {nome_tipo} '{nome_excluido}' removido com sucesso!")
                            
                    except model.DoesNotExist:
                        messages.error(request, f"❌ {nome_tipo} não encontrado!")
                    except Exception as e:
                        messages.error(request, f"❌ Erro ao remover {nome_tipo.lower()}: {str(e)}")
                else:
                    messages.error(request, "❌ Tipo de item inválido!")
        
        # Redireciona para a mesma aba
        return redirect(f"{reverse('sapp:configuracoes')}#{active_tab}")
    
    # =============================
    # CONTEXT PARA RENDERIZAÇÃO
    # =============================
    
    context = {
        'form_config': ConfiguracaoForm(instance=config),
        
        'cultivares': cultivares,
        'especies': especies,
        'peneiras': peneiras,
        'categorias': categorias,
        'tratamentos': tratamentos,
        
        'usuarios_conferentes': usuarios_conferentes,
        
        'form_conf_user': NovoConferenteUserForm(),
        
        'produtos': produtos,
        
        'armazens': armazens_lista,
        'enderecos': enderecos_lista,
        'origens': origens_lista,
    }
    
    return render(request, 'sapp/configuracoes.html', context)



@login_required
@permission_required('sapp.pode_gerenciar_usuarios', raise_exception=True)
def api_user_permissions(request, user_id):
    """
    API para buscar as permissões atuais de um usuário
    """
    if not request.user.is_superuser and not request.user.has_perm('sapp.pode_gerenciar_usuarios'):
        return JsonResponse({'success': False, 'error': 'Permissão negada'}, status=403)
    
    try:
        user = User.objects.get(id=user_id)
        
        # 🔥 Retornar o nome completo da permissão (com app)
        permissions = []
        for perm in user.user_permissions.all():
            permissions.append(perm.codename)
        
        return JsonResponse({
            'success': True,
            'permissions': permissions,
            'username': user.username,
            'first_name': user.first_name,
            'is_superuser': user.is_superuser,
            'groups': []
        })
        
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Usuário não encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q, Count, Min
from django.contrib.auth.decorators import login_required, permission_required
from sapp.models import HistoricoItemEmpenho

from collections import defaultdict

from django.contrib.auth.decorators import login_required, permission_required
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db.models import Count, Max, Q, Sum
from django.shortcuts import render
from django.utils import timezone
from django.utils.dateparse import parse_date

from sapp.models import HistoricoItemEmpenho


def campo_existe(model, caminho):
    """
    Verifica se um caminho de campo existe no model.

    Exemplos:
        cliente__nome
        empenho__cliente__nome
        processado_por__username
    """
    model_atual = model

    try:
        partes = caminho.split('__')

        for indice, parte in enumerate(partes):
            campo = model_atual._meta.get_field(parte)

            if indice < len(partes) - 1:
                if not campo.is_relation or not campo.related_model:
                    return False

                model_atual = campo.related_model

        return True

    except Exception:
        return False


def encontrar_campo_cliente():
    """
    Procura automaticamente onde está armazenado o nome do cliente.

    Adicione outros caminhos nesta lista caso seu model use outro nome.
    """
    campos_possiveis = [
        'cliente_nome',
        'nome_cliente',
        'cliente__nome',
        'cliente__razao_social',
        'cliente__nome_fantasia',

        'empenho__cliente_nome',
        'empenho__nome_cliente',
        'empenho__cliente__nome',
        'empenho__cliente__razao_social',
        'empenho__cliente__nome_fantasia',

        'estoque_origem__cliente__nome',
        'estoque_destino__cliente__nome',
    ]

    for caminho in campos_possiveis:
        if campo_existe(HistoricoItemEmpenho, caminho):
            return caminho

    return None


def obter_valor_atributo(objeto, caminho):
    """
    Obtém um valor usando caminhos como:
        empenho.cliente.nome
        cliente.razao_social
    """
    if objeto is None or not caminho:
        return ''

    valor = objeto

    for parte in caminho.split('__'):
        if valor is None:
            return ''

        valor = getattr(valor, parte, None)

        if callable(valor):
            try:
                valor = valor()
            except Exception:
                return ''

    if valor is None:
        return ''

    return str(valor).strip()


def obter_nome_cliente(movimentacao, campo_cliente=None):
    """
    Retorna o nome do cliente da movimentação.
    """
    if campo_cliente:
        nome = obter_valor_atributo(movimentacao, campo_cliente)

        if nome:
            return nome

    # Fallback para propriedades ou atributos que não sejam campos do banco.
    caminhos_fallback = [
        'cliente_nome',
        'nome_cliente',
        'cliente__nome',
        'cliente__razao_social',
        'cliente__nome_fantasia',

        'empenho__cliente_nome',
        'empenho__nome_cliente',
        'empenho__cliente__nome',
        'empenho__cliente__razao_social',
        'empenho__cliente__nome_fantasia',
    ]

    for caminho in caminhos_fallback:
        nome = obter_valor_atributo(movimentacao, caminho)

        if nome:
            return nome

    return ''


def obter_nome_usuario(usuario):
    """
    Retorna o nome completo do usuário ou o username.
    """
    if not usuario:
        return ''

    nome_completo = usuario.get_full_name().strip()

    if nome_completo:
        return nome_completo

    return usuario.username or ''


from collections import defaultdict
from types import SimpleNamespace

from django.contrib.auth.decorators import login_required, permission_required
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db.models import Q
from django.shortcuts import render
from django.utils import timezone
from django.utils.dateparse import parse_date

from .models import HistoricoMovimentacao, HistoricoItemEmpenho


def nome_usuario_historico(usuario):
    """Nome completo do usuário ou username; 'Sistema' se nulo."""
    if not usuario:
        return 'Sistema'
    nome = usuario.get_full_name().strip()
    return nome or usuario.username or 'Sistema'


def normalizar_tipo_historico(tipo):
    """
    Retorna o tipo normalizado (sem acentos, sem espaços extras, minúsculo).
    Mapeia variações comuns para uma chave única.
    """
    tipo_original = str(tipo or '').strip()
    tipo_lower = tipo_original.lower()

    mapa = {
        'entrada': 'entrada',
        'nova entrada': 'entrada',
        'saida': 'saida',
        'saída': 'saida',
        'baixa': 'saida',
        'transferencia': 'transferencia',
        'transferência': 'transferencia',
        'expedicao': 'expedicao',
        'expedição': 'expedicao',
        'edicao': 'edicao',
        'edição': 'edicao',
        'exclusao': 'exclusao',
        'exclusão': 'exclusao',
    }
    return mapa.get(tipo_lower, tipo_lower.replace(' ', '_'))


def nome_tipo_historico(tipo):
    """Nome amigável para exibição."""
    nomes = {
        'entrada': 'Entrada',
        'saida': 'Saída',
        'transferencia': 'Transferência',
        'expedicao': 'Expedição',
        'edicao': 'Edição',
        'exclusao': 'Exclusão',
    }
    return nomes.get(tipo, str(tipo or 'Não informado').replace('_', ' ').title())


def obter_enderecos_historico_antigo(movimentacao, tipo):
    """
    No histórico antigo não há endereços separados.
    Utiliza o endereço do estoque quando disponível.
    """
    endereco = movimentacao.estoque.endereco if movimentacao.estoque else ''

    if tipo == 'entrada':
        return '', endereco
    if tipo in ('saida', 'expedicao'):
        return endereco, ''
    if tipo == 'transferencia':
        return endereco, ''  # origem; destino normalmente na descrição
    return endereco, ''


@login_required
@permission_required('sapp.pode_ver_estoque', raise_exception=True)
def historico_geral(request):
    # ----------------------------------------------------------
    # 1. Captura dos parâmetros de filtro
    # ----------------------------------------------------------
    busca = request.GET.get('busca', '').strip()
    filtro_lote = request.GET.get('lote', '').strip()
    filtro_produto = request.GET.get('produto', '').strip()
    filtro_cliente = request.GET.get('cliente', '').strip()
    filtro_tipo = request.GET.get('tipo', '').strip()
    filtro_usuario = request.GET.get('usuario', '').strip()
    data_inicial_txt = request.GET.get('data_inicial', '').strip()
    data_final_txt = request.GET.get('data_final', '').strip()

    data_inicial = parse_date(data_inicial_txt)
    data_final = parse_date(data_final_txt)

    # ----------------------------------------------------------
    # 2. Querysets base
    # ----------------------------------------------------------
    qs_antigo = HistoricoMovimentacao.objects.select_related(
        'estoque', 'usuario'
    ).all()

    qs_novo = HistoricoItemEmpenho.objects.select_related(
        'estoque_origem', 'estoque_destino', 'processado_por', 'empenho'
    ).all()

    # ----------------------------------------------------------
    # 3. Busca textual (múltiplos termos → AND)
    # ----------------------------------------------------------
    if busca:
        termos = [t for t in busca.split() if t]
        for termo in termos:
            qs_antigo = qs_antigo.filter(
                Q(lote_ref__icontains=termo) |
                Q(estoque__lote__icontains=termo) |
                Q(estoque__produto__icontains=termo) |
                Q(estoque__cultivar__nome__icontains=termo) |
                Q(estoque__cliente__icontains=termo) |
                Q(estoque__empresa__icontains=termo) |
                Q(cliente__icontains=termo) |
                Q(tipo__icontains=termo) |
                Q(descricao__icontains=termo) |
                Q(numero_carga__icontains=termo) |
                Q(motorista__icontains=termo) |
                Q(placa__icontains=termo) |
                Q(ordem_entrega__icontains=termo) |
                Q(usuario__username__icontains=termo) |
                Q(usuario__first_name__icontains=termo) |
                Q(usuario__last_name__icontains=termo)
            )
            qs_novo = qs_novo.filter(
                Q(lote__icontains=termo) |
                Q(produto__icontains=termo) |
                Q(cultivar__icontains=termo) |
                Q(cliente__icontains=termo) |
                Q(empresa__icontains=termo) |
                Q(tipo__icontains=termo) |
                Q(observacao__icontains=termo) |
                Q(numero_carga__icontains=termo) |
                Q(placa__icontains=termo) |
                Q(endereco_origem__icontains=termo) |
                Q(endereco_destino__icontains=termo) |
                Q(processado_por__username__icontains=termo) |
                Q(processado_por__first_name__icontains=termo) |
                Q(processado_por__last_name__icontains=termo)
            )

    # ----------------------------------------------------------
    # 4. Filtros exatos
    # ----------------------------------------------------------
    if filtro_lote:
        qs_antigo = qs_antigo.filter(
            Q(lote_ref=filtro_lote) | Q(estoque__lote=filtro_lote)
        )
        qs_novo = qs_novo.filter(lote=filtro_lote)

    if filtro_produto:
        qs_antigo = qs_antigo.filter(estoque__produto=filtro_produto)
        qs_novo = qs_novo.filter(produto=filtro_produto)

    if filtro_cliente:
        qs_antigo = qs_antigo.filter(
            Q(cliente=filtro_cliente) | Q(estoque__cliente=filtro_cliente)
        )
        qs_novo = qs_novo.filter(cliente=filtro_cliente)

    if filtro_usuario:
        qs_antigo = qs_antigo.filter(usuario__username=filtro_usuario)
        qs_novo = qs_novo.filter(processado_por__username=filtro_usuario)

    # Filtro por tipo – agora 100% flexível: usamos icontains com a string normalizada,
    # o que casa com qualquer variação de maiúsculas/minúsculas e acentos.
    if filtro_tipo:
        tipo_normalizado = normalizar_tipo_historico(filtro_tipo)
        # Mapeamos as grafias mais comuns que podem estar no banco para o termo normalizado
        mapa_busca = {
            'entrada': ['entrada', 'nova entrada'],
            'saida': ['saida', 'saída', 'baixa'],
            'transferencia': ['transferencia', 'transferência'],
            'expedicao': ['expedicao', 'expedição'],
            'edicao': ['edicao', 'edição'],
            'exclusao': ['exclusao', 'exclusão'],
        }
        variantes = mapa_busca.get(tipo_normalizado, [filtro_tipo])
        tipo_q = Q()
        for v in variantes:
            tipo_q |= Q(tipo__icontains=v)  # icontains ignora case e acentos parcialmente
        qs_antigo = qs_antigo.filter(tipo_q)
        qs_novo = qs_novo.filter(tipo_q)

    # Datas
    if data_inicial:
        qs_antigo = qs_antigo.filter(data_hora__date__gte=data_inicial)
        qs_novo = qs_novo.filter(processado_em__date__gte=data_inicial)
    if data_final:
        qs_antigo = qs_antigo.filter(data_hora__date__lte=data_final)
        qs_novo = qs_novo.filter(processado_em__date__lte=data_final)

    # ----------------------------------------------------------
    # 5. Normalização dos registros (histórico antigo + novo)
    # ----------------------------------------------------------
    movimentacoes = []

    for m in qs_antigo.iterator(chunk_size=1000):
        estoque = m.estoque
        lote = m.lote_ref or (estoque.lote if estoque else '') or 'Sem lote'
        produto = estoque.produto if estoque else ''
        cliente = m.cliente or (estoque.cliente if estoque else '') or ''
        empresa = estoque.empresa if estoque else ''
        tipo = normalizar_tipo_historico(m.tipo)
        end_orig, end_dest = obter_enderecos_historico_antigo(m, tipo)

        movimentacoes.append(SimpleNamespace(
            chave=f'antigo-{m.pk}',
            origem_historico='Histórico geral',
            lote=lote,
            produto=produto,
            cliente_exibicao=cliente,
            empresa=empresa,
            cultivar=estoque.cultivar.nome if estoque and estoque.cultivar else '',
            peneira=estoque.peneira.nome if estoque and estoque.peneira else '',
            categoria=estoque.categoria.nome if estoque and estoque.categoria else '',
            tratamento=estoque.tratamento.nome if estoque and estoque.tratamento else '',
            especie=estoque.especie.nome if estoque and estoque.especie else '',
            embalagem=estoque.embalagem if estoque else '',
            quantidade=m.quantidade or 0,
            tipo=tipo,
            tipo_exibicao=nome_tipo_historico(tipo),
            endereco_origem=end_orig,
            endereco_destino=end_dest,
            usuario_exibicao=nome_usuario_historico(m.usuario),
            processado_em=m.data_hora,
            observacao=m.descricao or '',
            numero_carga=m.numero_carga or '',
            motorista=m.motorista or '',
            placa=m.placa or '',
            ordem_entrega=m.ordem_entrega or '',
        ))

    for m in qs_novo.iterator(chunk_size=1000):
        tipo = normalizar_tipo_historico(m.tipo)
        movimentacoes.append(SimpleNamespace(
            chave=f'novo-{m.pk}',
            origem_historico='Cards e empenhos',
            lote=m.lote or 'Sem lote',
            produto=m.produto or '',
            cliente_exibicao=m.cliente or '',
            empresa=m.empresa or '',
            cultivar=m.cultivar or '',
            peneira=m.peneira or '',
            categoria=m.categoria or '',
            tratamento=m.tratamento or '',
            especie=m.especie or '',
            embalagem=m.embalagem or '',
            quantidade=m.quantidade or 0,
            tipo=tipo,
            tipo_exibicao=nome_tipo_historico(tipo),
            endereco_origem=m.endereco_origem or '',
            endereco_destino=m.endereco_destino or '',
            usuario_exibicao=nome_usuario_historico(m.processado_por),
            processado_em=m.processado_em,
            observacao=m.observacao or '',
            numero_carga=m.numero_carga or '',
            motorista=m.empenho.motorista if m.empenho else '',
            placa=m.placa or '',
            ordem_entrega='',
        ))

    # ----------------------------------------------------------
    # 6. Ordenação e deduplicação leve
    # ----------------------------------------------------------
    data_minima = timezone.make_aware(timezone.datetime.min)
    movimentacoes.sort(key=lambda x: x.processado_em or data_minima, reverse=True)

    # Remove eventos duplicados (mesmo lote, tipo, qtd, data/minuto e carga)
    vistos = set()
    unicos = []
    for mov in movimentacoes:
        data_chave = mov.processado_em.strftime('%Y-%m-%d %H:%M') if mov.processado_em else ''
        chave = (
            str(mov.lote).strip().upper(),
            mov.tipo,
            int(mov.quantidade or 0),
            data_chave,
            str(mov.numero_carga or '').strip().upper(),
        )
        if chave not in vistos:
            vistos.add(chave)
            unicos.append(mov)
    movimentacoes = unicos

    # ----------------------------------------------------------
    # 7. Agrupamento por lote
    # ----------------------------------------------------------
    por_lote = defaultdict(list)
    for mov in movimentacoes:
        por_lote[str(mov.lote or 'Sem lote').strip()].append(mov)

    lotes_agrupados = []
    for idx, (lote_ref, movs) in enumerate(por_lote.items(), start=1):
        movs.sort(key=lambda x: x.processado_em or data_minima, reverse=True)
        ultima = movs[0]
        quantidade_total = sum(int(m.quantidade or 0) for m in movs)

        # Clientes e produtos únicos (resumo)
        clientes_unicos = list(dict.fromkeys(m.cliente_exibicao for m in movs if m.cliente_exibicao))
        cliente_resumo = ', '.join(clientes_unicos[:3])
        if len(clientes_unicos) > 3:
            cliente_resumo += f' +{len(clientes_unicos)-3}'

        produtos_unicos = list(dict.fromkeys(m.produto for m in movs if m.produto))
        produto_resumo = ', '.join(produtos_unicos[:2])
        if len(produtos_unicos) > 2:
            produto_resumo += f' +{len(produtos_unicos)-2}'

        lotes_agrupados.append({
            'grupo_id': f'grupo-{idx}',
            'lote_ref': lote_ref,
            'produto': produto_resumo,
            'cliente': cliente_resumo,
            'total_mov': len(movs),
            'quantidade_total': quantidade_total,
            'ultima_data': ultima.processado_em,
            'ultimo_end_origem': ultima.endereco_origem,
            'ultimo_end_destino': ultima.endereco_destino,
            'ultimo_usuario': ultima.usuario_exibicao,
            'ultimo_tipo': ultima.tipo,
            'ultimo_tipo_exibicao': ultima.tipo_exibicao,
            'movimentacoes': movs,
        })

    lotes_agrupados.sort(key=lambda g: g['ultima_data'] or data_minima, reverse=True)

    # ----------------------------------------------------------
    # 8. Cards informativos
    # ----------------------------------------------------------
    hoje = timezone.localdate()
    total_mov = len(movimentacoes)
    total_lotes = len(lotes_agrupados)
    total_exp = sum(1 for m in movimentacoes if m.tipo == 'expedicao')
    mov_hoje = sum(
        1 for m in movimentacoes
        if m.processado_em and timezone.localtime(m.processado_em).date() == hoje
    )

    # ----------------------------------------------------------
    # 9. Opções para os selects (sempre completas, sem filtro)
    # ----------------------------------------------------------
    def _lista_distinta(qs, campo, modelo_rel=None):
        """Extrai valores distintos de um campo, aceitando relacionamento."""
        if modelo_rel:
            return set(
                qs.exclude(**{f'{modelo_rel}__isnull': True})
                .values_list(f'{modelo_rel}__{campo}', flat=True)
            )
        return set(qs.exclude(**{campo: ''}).values_list(campo, flat=True))

    lotes_ant = _lista_distinta(HistoricoMovimentacao.objects, 'lote_ref')
    lotes_ant_est = _lista_distinta(HistoricoMovimentacao.objects, 'lote', modelo_rel='estoque')
    lotes_nov = _lista_distinta(HistoricoItemEmpenho.objects, 'lote')
    opcoes_lotes = sorted({x.strip() for x in (lotes_ant | lotes_ant_est | lotes_nov) if x and x.strip()}, key=str.lower)

    prod_ant = _lista_distinta(HistoricoMovimentacao.objects, 'produto', modelo_rel='estoque')
    prod_nov = _lista_distinta(HistoricoItemEmpenho.objects, 'produto')
    opcoes_produtos = sorted({x.strip() for x in (prod_ant | prod_nov) if x and x.strip()}, key=str.lower)

    cli_ant = _lista_distinta(HistoricoMovimentacao.objects, 'cliente')
    cli_ant_est = _lista_distinta(HistoricoMovimentacao.objects, 'cliente', modelo_rel='estoque')
    cli_nov = _lista_distinta(HistoricoItemEmpenho.objects, 'cliente')
    opcoes_clientes = sorted({x.strip() for x in (cli_ant | cli_ant_est | cli_nov) if x and x.strip()}, key=str.lower)

    # Usuários (dicionário username -> nome)
    usuarios_dict = {}
    for u in HistoricoMovimentacao.objects.select_related('usuario').exclude(usuario__isnull=True).values('usuario__username', 'usuario__first_name', 'usuario__last_name').distinct():
        username = u['usuario__username']
        nome = f"{u['usuario__first_name']} {u['usuario__last_name']}".strip() or username
        usuarios_dict[username] = nome
    for u in HistoricoItemEmpenho.objects.select_related('processado_por').exclude(processado_por__isnull=True).values('processado_por__username', 'processado_por__first_name', 'processado_por__last_name').distinct():
        username = u['processado_por__username']
        nome = f"{u['processado_por__first_name']} {u['processado_por__last_name']}".strip() or username
        usuarios_dict[username] = nome
    opcoes_usuarios = [{'valor': k, 'nome': v} for k, v in sorted(usuarios_dict.items(), key=lambda item: item[1].lower())]

    choices_tipo = [
        ('entrada', 'Entrada'),
        ('saida', 'Saída'),
        ('transferencia', 'Transferência'),
        ('expedicao', 'Expedição'),
        ('edicao', 'Edição'),
        ('exclusao', 'Exclusão'),
    ]

    # ----------------------------------------------------------
    # 10. Paginação dos lotes agrupados
    # ----------------------------------------------------------
    try:
        page_size = int(request.GET.get('page_size', 25))
    except (ValueError, TypeError):
        page_size = 25
    if page_size not in (10, 25, 50, 100, 200):
        page_size = 25

    paginator = Paginator(lotes_agrupados, page_size)
    page_number = request.GET.get('page', 1)
    try:
        pagina = paginator.page(page_number)
    except PageNotAnInteger:
        pagina = paginator.page(1)
    except EmptyPage:
        pagina = paginator.page(paginator.num_pages)

    # Ajusta IDs dos grupos por página
    for i, grupo in enumerate(pagina.object_list, start=1):
        grupo['grupo_id'] = f'grupo-{pagina.number}-{i}'

    query_params = request.GET.copy()
    query_params.pop('page', None)
    url_params = query_params.urlencode()

    context = {
        'lotes': pagina,
        'busca': busca,
        'filtro_lote': filtro_lote,
        'filtro_produto': filtro_produto,
        'filtro_cliente': filtro_cliente,
        'filtro_tipo': filtro_tipo,
        'filtro_usuario': filtro_usuario,
        'data_inicial': data_inicial_txt,
        'data_final': data_final_txt,
        'opcoes_lotes': opcoes_lotes,
        'opcoes_produtos': opcoes_produtos,
        'opcoes_clientes': opcoes_clientes,
        'opcoes_usuarios': opcoes_usuarios,
        'choices_tipo': choices_tipo,
        'total_movimentacoes': total_mov,
        'total_lotes': total_lotes,
        'total_expedicoes': total_exp,
        'movimentacoes_hoje': mov_hoje,
        'page_size': page_size,
        'page_sizes': [10, 25, 50, 100, 200],
        'url_params': url_params,
    }
    return render(request, 'sapp/historico_geral.html', context)


from django.contrib.auth import update_session_auth_hash
from django.contrib import messages
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .forms import MudarSenhaForm

@login_required
def mudar_senha(request):
    if request.method == 'POST':
        form = MudarSenhaForm(request.POST)
        if form.is_valid():
            nova_senha = form.cleaned_data['nova_senha']
            
            # Impede que o usuário use a senha padrão novamente
            if nova_senha == 'conceito123':
                messages.error(request, "❌ Não utilize a senha padrão. Escolha uma senha segura.")
                return render(request, 'sapp/mudar_senha.html', {'form': form})
            
            request.user.set_password(nova_senha)
            request.user.save()
            
            try:
                perfil = request.user.perfil
                perfil.primeiro_acesso = False
                perfil.save()
            except:
                pass
            
            update_session_auth_hash(request, request.user)
            messages.success(request, "✅ Senha atualizada com sucesso!")
            return redirect('sapp:redirecionar')
    else:
        form = MudarSenhaForm()
    
    return render(request, 'sapp/mudar_senha.html', {'form': form})



def exportar_excel(request):
    estoque = Estoque.objects.filter(saldo__gt=0).select_related(
        'cultivar', 'peneira', 'categoria', 'tratamento', 'conferente'
    )
    
    # Criar DataFrame
    data = []
    for item in estoque:
        data.append({
            'Lote': item.lote,
            'Produto': item.produto or '',  # 🔥 NOVO CAMPO
            'Cultivar': item.cultivar.nome,
            'Peneira': item.peneira.nome,
            'Categoria': item.categoria.nome,
            'Endereço': item.endereco,
            'Saldo': item.saldo,
            'Peso Unitário (kg)': float(item.peso_unitario),
            'Peso Total (kg)': float(item.peso_total),
            'Tratamento': item.tratamento.nome if item.tratamento else '',
            'Embalagem': item.get_embalagem_display(),
            'Conferente': item.conferente.first_name,
            'Data Entrada': item.data_entrada.strftime('%d/%m/%Y'),
            'AZ': item.az or '',
            'Origem/Destino': item.origem_destino,
            'Empresa': item.empresa,
            'Espécie': item.especie,
            'Observação': item.observacao or ''
        })
    
    df = pd.DataFrame(data)
    
    # Criar resposta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="estoque_sementes.xlsx"'
    
    # Exportar para Excel
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Estoque', index=False)
        
        # Formatar a planilha
        workbook = writer.book
        worksheet = writer.sheets['Estoque']
        
        # Ajustar largura das colunas
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    return response

def exportar_pdf(request):
    estoque = Estoque.objects.filter(saldo__gt=0).select_related(
        'cultivar', 'peneira', 'categoria', 'tratamento', 'conferente'
    )[:100]  # Limitar para não sobrecarregar o PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=30)
    elements = []
    styles = getSampleStyleSheet()
    title = Paragraph("RELATÓRIO DE ESTOQUE - SEMENTES", styles['Title'])
    elements.append(title)
    elements.append(Paragraph(f"Data: {timezone.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elements.append(Paragraph("<br/>", styles['Normal']))
    
    # Dados da tabela ATUALIZADOS
    data = [['Lote', 'Produto', 'Cultivar', 'Peneira', 'Endereço', 'Saldo', 'Peso Total']]  # 🔥 ADICIONADO PRODUTO
    
    for item in estoque:
        data.append([
            item.lote,
            item.produto or '',  # 🔥 NOVO CAMPO
            item.cultivar.nome,
            item.peneira.nome,
            item.endereco,
            str(item.saldo),
            f"{item.peso_total:.2f} kg"
        ])
    
    # Criar tabela
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2f8f4e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    
    # Rodapé
    elements.append(Paragraph(f"<br/>Total de itens: {estoque.count()}", styles['Normal']))
    
    # Gerar PDF
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="estoque_sementes.pdf"'
    
    return response

################ DEBUG #####################
@login_required
@permission_required('sapp.pode_configuracoes', raise_exception=True)
def debug_estoque_completo(request):
    """Debug COMPLETO do estoque atual"""
    estoque = Estoque.objects.all().select_related('peneira', 'cultivar', 'tratamento', 'categoria')
    
    print("🔍 [DEBUG COMPLETO DO ESTOQUE]")
    print("=" * 80)
    
    for item in estoque:
        print(f"Lote: {item.lote}")
        print(f"  Peneira: '{item.peneira.nome if item.peneira else 'None'}'")
        print(f"  Cultivar: '{item.cultivar.nome if item.cultivar else 'None'}'")
        print(f"  Tratamento: '{item.tratamento.nome if item.tratamento else 'None'}'")
        print(f"  Categoria: '{item.categoria.nome if item.categoria else 'None'}'")
        print(f"  Endereço: '{item.endereco}'")
        print(f"  Saldo: {item.saldo}")
        print("-" * 40)
    
    return JsonResponse({'success': True, 'message': 'Check console for debug info'})

@login_required
@permission_required('sapp.pode_configuracoes', raise_exception=True)
def debug_estoque_status(request):
    """Debug para ver status do estoque"""
    total_lotes = Estoque.objects.count()
    lotes_com_saldo = Estoque.objects.filter(saldo__gt=0).count()
    lotes_sem_saldo = Estoque.objects.filter(saldo=0).count()
    
    print("🔍 [DEBUG ESTOQUE STATUS]")
    print(f"📊 Total de lotes: {total_lotes}")
    print(f"✅ Com saldo > 0: {lotes_com_saldo}")
    print(f"❌ Com saldo = 0: {lotes_sem_saldo}")
    
    # Listar alguns lotes com saldo 0
    lotes_zerados = Estoque.objects.filter(saldo=0).values('lote', 'endereco', 'id')[:10]
    print("\n📝 Primeiros 10 lotes com saldo 0:")
    for lote in lotes_zerados:
        print(f"   Lote: {lote['lote']} | Endereço: {lote['endereco']} | ID: {lote['id']}")
    
    return JsonResponse({
        'success': True,
        'total_lotes': total_lotes,
        'com_saldo': lotes_com_saldo,
        'sem_saldo': lotes_sem_saldo
    })
################     API    ############################
@login_required
@permission_required('sapp.pode_ver_estoque', raise_exception=True)
def api_saldo_lote(request, id):
    """API para obter saldo de um lote específico"""
    try:
        item = get_object_or_404(Estoque, id=id)
        return JsonResponse({
            'success': True,
            'lote': item.lote,
            'saldo': item.saldo,
            'entrada': item.entrada,
            'saida': item.saida,
            'cultivar': item.cultivar.nome if item.cultivar else '',
            'endereco': item.endereco,
            'embalagem': item.embalagem,
            'peso_unitario': float(item.peso_unitario) if item.peso_unitario else 0,
            'peso_total': float(item.peso_total) if item.peso_total else 0
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
@permission_required('sapp.pode_ver_estoque', raise_exception=True)
def api_buscar_lotes(request):

    query = request.GET.get('q', '')

    if not query:
        return JsonResponse({'results': []})

    lotes = (
        Estoque.objects
        .filter(Q(lote__icontains=query))
        .select_related(
            'cultivar',
            'peneira',
            'categoria',
            'tratamento',
            'especie'
        )
        .order_by('-data_ultima_movimentacao')[:10]
    )

    results = []

    for item in lotes:
        results.append({
            "id": item.id,
            "lote": item.lote,
            "produto": item.produto,
            "cultivar": item.cultivar.nome if item.cultivar else "",
            "cultivar_id": item.cultivar.id if item.cultivar else None,
            "especie_id": item.especie.id if item.especie else None,
            "peneira_id": item.peneira.id if item.peneira else None,
            "categoria_id": item.categoria.id if item.categoria else None,
            "tratamento_id": item.tratamento.id if item.tratamento else None,
            "empresa": item.empresa,
            "cliente": item.cliente,
            "peso_unitario": float(item.peso_unitario) if item.peso_unitario else "",
            "embalagem": item.embalagem,
            "az": item.az,
            "endereco": item.endereco,
            "saldo": float(item.saldo)
        })

    return JsonResponse({"results": results})


@login_required
@permission_required('sapp.pode_ver_estoque', raise_exception=True)
def api_buscar_lote_completo(request):
    """API para buscar todos os dados de um lote existente"""
    lote = request.GET.get('lote', '')
    
    if not lote:
        return JsonResponse({'encontrado': False, 'error': 'Lote não especificado'})
    item = Estoque.objects.filter(lote=lote).order_by('-id').first()
    
    if item:
        data = {
            'encontrado': True,
            'lote': item.lote,
            'produto': item.produto or '',
            'cultivar_id': item.cultivar.id if item.cultivar else None,
            'peneira_id': item.peneira.id if item.peneira else None,
            'categoria_id': item.categoria.id if item.categoria else None,
            'tratamento_id': item.tratamento.id if item.tratamento else None,
            'empresa': item.empresa or '',
            'origem_destino': item.origem_destino or '',
            'especie_id': item.especie.id if item.especie else None,
            'peso_unitario': str(item.peso_unitario),
            'embalagem': item.embalagem or 'BAG',
            'az': item.az or '',
            'observacao': item.observacao or ''
        }
        return JsonResponse(data)
    
    return JsonResponse({'encontrado': False})

@login_required
@permission_required('sapp.pode_configuracoes', raise_exception=True)
def api_verificar_lote(request):
    """API para verificar se um lote existe"""
    lote = request.GET.get('lote', '')
    
    if not lote:
        return JsonResponse({'existe': False})
    
    existe = Estoque.objects.filter(lote=lote).exists()
    
    return JsonResponse({'existe': existe, 'lote': lote})

@login_required
@permission_required('sapp.pode_ver_estoque', raise_exception=True)
def api_estoque_resumo(request):
    """API para resumo do estoque (usado no dashboard)"""
    total_lotes = Estoque.objects.count()
    lotes_ativos = Estoque.objects.filter(saldo__gt=0).count()
    lotes_esgotados = Estoque.objects.filter(saldo=0).count()
    total_entrada = Estoque.objects.aggregate(total=Sum('entrada'))['total'] or 0
    total_saida = Estoque.objects.aggregate(total=Sum('saida'))['total'] or 0
    
    # Top 5 cultivares
    top_cultivares = Estoque.objects.filter(saldo__gt=0).values(
        'cultivar__nome'
    ).annotate(
        total_saldo=Sum('saldo'),
        total_lotes=Count('id')
    ).order_by('-total_saldo')[:5]
    
    return JsonResponse({
        'success': True,
        'total_lotes': total_lotes,
        'lotes_ativos': lotes_ativos,
        'lotes_esgotados': lotes_esgotados,
        'total_entrada': total_entrada,
        'total_saida': total_saida,
        'top_cultivares': list(top_cultivares)
    })

@login_required
@permission_required('sapp.pode_configuracoes', raise_exception=True)
def api_ultimas_movimentacoes(request):
    """API para últimas movimentações"""
    movimentacoes = HistoricoMovimentacao.objects.select_related(
        'estoque', 'usuario'
    ).order_by('-data_hora')[:10]
    
    data = []
    for mov in movimentacoes:
        data.append({
            'id': mov.id,
            'data_hora': mov.data_hora.strftime('%d/%m/%Y %H:%M'),
            'tipo': mov.tipo,
            'descricao': mov.descricao,
            'usuario': mov.usuario.username if mov.usuario else 'Sistema',
            'lote': mov.lote_ref
        })
    
    return JsonResponse({
        'success': True,
        'movimentacoes': data
    })
    
@login_required
@permission_required('sapp.pode_ver_empenhos', raise_exception=True)
def pagina_rascunho(request):
    """
    View para gestão de rascunhos, transferências e expedições.
    Implementa processamento em lote com preservação de histórico.
    """
    user = request.user
    MARCA_ORIGEM = "[REP]"

    # =====================================================
    # POST (AÇÕES)
    # =====================================================
    if request.method == 'POST':
        
        # -------------------------
        # EXCLUIR CARD
        # -------------------------
        if 'excluir_card' in request.POST:
            empenho_id = request.POST.get('empenho_id')
            empenho = get_object_or_404(Empenho, id=empenho_id)
            
            if empenho.historico_itens.exists():
                messages.warning(
                    request, 
                    "Este card possui itens já processados e não pode ser excluído. "
                    "Ele permanecerá disponível para consulta e impressão."
                )
            else:
                nome_card = empenho.observacao or f'Card #{empenho.id}'
                empenho.delete()
                messages.success(request, f"Card '{nome_card}' excluído com sucesso.")
            
            return redirect('sapp:pagina_rascunho')

        # -------------------------
        # EXCLUIR ITEM DO CARD
        # -------------------------
        if 'excluir_item' in request.POST:
            item_id = request.POST.get('item_id')
            item = get_object_or_404(
                ItemEmpenho.objects.select_related('empenho'), 
                id=item_id
            )
            
            empenho = item.empenho
            lote_info = item.lote
            item.delete()
            
            if not empenho.itens.exists() and not empenho.historico_itens.exists():
                empenho.delete()
                messages.success(request, f"Item '{lote_info}' removido. Card estava vazio e foi excluído.")
            else:
                messages.success(request, f"Item '{lote_info}' removido do card.")
            
            return redirect('sapp:pagina_rascunho')

        # -------------------------
        # TRANSFERIR / EXPEDIR EM MASSA
        # -------------------------
        if request.POST.get('origem_acao') == 'cards':
            acao = request.POST.get('acao_tipo')
            empenho_id = request.POST.get('empenho_id')
            obs_global = request.POST.get('obs_global', '').strip()
            
            if acao not in ['transferir', 'expedir']:
                messages.error(request, "Ação inválida. Use 'transferir' ou 'expedir'.")
                return redirect('sapp:pagina_rascunho')
            
            try:
                selected_items_json = request.POST.get('selected_items', '[]')
                selected_items = json.loads(selected_items_json)
                
                if not isinstance(selected_items, list):
                    raise ValueError("Formato inválido: esperado array")
                
                selected_ids = []
                for item in selected_items:
                    if isinstance(item, dict):
                        item_id = item.get('item_id')
                    else:
                        item_id = item
                    
                    try:
                        item_id = int(item_id)
                        if item_id > 0:
                            selected_ids.append(item_id)
                    except (ValueError, TypeError):
                        continue
                
                selected_ids = list(set(selected_ids))
                
                if not selected_ids:
                    raise ValueError("Nenhum ID válido encontrado")
                    
            except (json.JSONDecodeError, ValueError) as e:
                messages.error(request, f"Dados de seleção inválidos: {str(e)}")
                return redirect('sapp:pagina_rascunho')
            
            try:
                with transaction.atomic():
                    try:
                        empenho = Empenho.objects.select_for_update().get(id=empenho_id)
                    except Empenho.DoesNotExist:
                        raise ValueError("Card não encontrado.")
                    
                    itens = list(
                        ItemEmpenho.objects
                        .filter(id__in=selected_ids, empenho=empenho)
                        .select_related('estoque')
                        .select_for_update()
                    )
                    
                    if not itens:
                        raise ValueError("Nenhum item válido encontrado para processamento.")
                    
                    encontrados_ids = {item.id for item in itens}
                    nao_encontrados = set(selected_ids) - encontrados_ids
                    if nao_encontrados:
                        raise ValueError(
                            f"Alguns itens não pertencem a este card ou não existem: "
                            f"{sorted(nao_encontrados)}"
                        )
                    
                    for item in itens:
                        item.estoque.refresh_from_db()
                        
                        if item.quantidade <= 0:
                            raise ValueError(
                                f"Item ID {item.id} (lote {item.lote}) "
                                f"com quantidade inválida: {item.quantidade}"
                            )
                        
                        if item.quantidade > item.estoque.saldo:
                            raise ValueError(
                                f"Saldo insuficiente para lote {item.lote}. "
                                f"Disponível: {item.estoque.saldo}, "
                                f"Solicitado: {item.quantidade}."
                            )
                    
                    for item in itens:
                        if acao == 'transferir':
                            # Processar transferência
                            origem = item.estoque
                            qtd = item.quantidade
                            
                            novo_end = request.POST.get('novo_endereco', '').strip().upper()
                            novo_az = request.POST.get('az', '').strip().upper() or origem.az
                            obs_transferencia = request.POST.get('obs_transferencia', '').strip()
                            
                            if not novo_end:
                                raise ValueError("Novo endereço não informado.")
                            
                            if novo_end == origem.endereco:
                                raise ValueError(
                                    f"Endereço de destino igual ao de origem para lote {origem.lote}."
                                )
                            
                            destino = (
                                Estoque.objects
                                .select_for_update()
                                .filter(
                                    lote=origem.lote,
                                    produto=origem.produto,
                                    cultivar=origem.cultivar,
                                    peneira=origem.peneira,
                                    categoria=origem.categoria,
                                    tratamento=origem.tratamento,
                                    especie=origem.especie,
                                    endereco=novo_end,
                                    az=novo_az,
                                    empresa=origem.empresa,
                                    embalagem=origem.embalagem
                                )
                                .first()
                            )
                            
                            if destino:
                                destino.entrada += qtd
                                destino.save()
                            else:
                                destino = Estoque.objects.create(
                                    lote=origem.lote,
                                    produto=origem.produto,
                                    cultivar=origem.cultivar,
                                    peneira=origem.peneira,
                                    categoria=origem.categoria,
                                    tratamento=origem.tratamento,
                                    especie=origem.especie,
                                    endereco=novo_end,
                                    az=novo_az,
                                    entrada=qtd,
                                    peso_unitario=origem.peso_unitario,
                                    embalagem=origem.embalagem,
                                    conferente=user,
                                    empresa=origem.empresa,
                                    cliente=origem.cliente,
                                    observacao=f"{MARCA_ORIGEM} {obs_global} {obs_transferencia}".strip()
                                )
                            
                            origem.saida += qtd
                            origem.save()
                            
                            HistoricoMovimentacao.objects.create(
                                estoque=origem,
                                usuario=user,
                                quantidade=qtd,
                                tipo='Transferência (Saída)',
                                descricao=(
                                    f"{MARCA_ORIGEM} Transferido {qtd} un de "
                                    f"{origem.endereco} para {novo_end}. "
                                    f"{obs_transferencia}"
                                ).strip()
                            )
                            
                            HistoricoMovimentacao.objects.create(
                                estoque=destino,
                                usuario=user,
                                quantidade=qtd,
                                tipo='Transferência (Entrada)',
                                descricao=(
                                    f"{MARCA_ORIGEM} Recebido {qtd} un de "
                                    f"{origem.endereco} em {novo_end}. "
                                    f"{obs_transferencia}"
                                ).strip()
                            )
                            
                            HistoricoItemEmpenho.objects.create(
                                empenho=empenho,
                                item_empenho_id_original=item.id,
                                estoque_origem=origem,
                                estoque_destino=destino,
                                lote=origem.lote,
                                produto=origem.produto or '',
                                cultivar=origem.cultivar.nome if origem.cultivar else '',
                                peneira=origem.peneira.nome if origem.peneira else '',
                                categoria=origem.categoria.nome if origem.categoria else '',
                                tratamento=origem.tratamento.nome if origem.tratamento else '',
                                especie=origem.especie.nome if origem.especie else '',
                                embalagem=origem.embalagem or '',
                                empresa=origem.empresa or '',
                                cliente=origem.cliente or '',
                                endereco_origem=origem.endereco,
                                endereco_destino=novo_end,
                                quantidade=qtd,
                                tipo='transferencia',
                                observacao=obs_transferencia,
                                processado_por=user
                            )
                        else:
                            # Processar expedição
                            origem = item.estoque
                            qtd = item.quantidade
                            
                            obs_expedicao = request.POST.get('obs_expedicao', '').strip()
                            numero_carga = request.POST.get('numero_carga', '').strip()
                            cliente = request.POST.get('cliente', '').strip()
                            placa = request.POST.get('placa', '').strip()
                            
                            if not numero_carga:
                                raise ValueError("Número da carga/pedido não informado.")
                            
                            origem.saida += qtd
                            origem.save()
                            
                            HistoricoMovimentacao.objects.create(
                                estoque=origem,
                                usuario=user,
                                quantidade=qtd,
                                tipo='Expedição',
                                descricao=(
                                    f"{MARCA_ORIGEM} Expedido {qtd} un. "
                                    f"Carga: {numero_carga}. "
                                    f"{obs_global} {obs_expedicao}"
                                ).strip(),
                                numero_carga=numero_carga,
                                cliente=cliente or origem.cliente,
                                placa=placa
                            )
                            
                            HistoricoItemEmpenho.objects.create(
                                empenho=empenho,
                                item_empenho_id_original=item.id,
                                estoque_origem=origem,
                                lote=origem.lote,
                                produto=origem.produto or '',
                                cultivar=origem.cultivar.nome if origem.cultivar else '',
                                peneira=origem.peneira.nome if origem.peneira else '',
                                categoria=origem.categoria.nome if origem.categoria else '',
                                tratamento=origem.tratamento.nome if origem.tratamento else '',
                                especie=origem.especie.nome if origem.especie else '',
                                embalagem=origem.embalagem or '',
                                empresa=origem.empresa or '',
                                cliente=cliente or origem.cliente or '',
                                endereco_origem=origem.endereco,
                                quantidade=qtd,
                                tipo='expedicao',
                                observacao=obs_expedicao,
                                numero_carga=numero_carga,
                                placa=placa,
                                processado_por=user
                            )
                        
                        # SÓ AGORA excluir o item processado
                        item.delete()
                    
                    # Atualizar status do card
                    empenho.refresh_from_db()
                    if not empenho.itens.exists():
                        if empenho.historico_itens.exists():
                            status_concluido, _ = EmpenhoStatus.objects.get_or_create(
                                nome='Concluído',
                                defaults={'descricao': 'Card processado completamente'}
                            )
                            empenho.status = status_concluido
                            empenho.save()
                            messages.info(request, "Todos os itens foram processados. Card marcado como concluído.")
                        else:
                            empenho.delete()
                    
                    acao_nome = 'Transferência' if acao == 'transferir' else 'Expedição'
                    messages.success(
                        request, 
                        f"{acao_nome} realizada com sucesso! "
                        f"{len(itens)} item(ns) processado(s)."
                    )
                    
            except ValueError as e:
                messages.error(request, str(e))
            except Exception as e:
                messages.error(request, f"Erro inesperado ao processar: {str(e)}")
            
            return redirect('sapp:pagina_rascunho')

    # =====================================================
    # GET (DADOS)
    # =====================================================
    
    ids_rascunhos = ItemEmpenho.objects.all().values_list('estoque_id', flat=True)

    estoque_qs = (
        Estoque.objects
        .filter(Q(saldo__gt=0) | Q(id__in=ids_rascunhos))
        .select_related(
            'cultivar', 'peneira', 'categoria', 
            'tratamento', 'especie', 'conferente',
            'status_sistemico'
        )
        .order_by('lote', 'endereco')
    )

    todos_itens = list(
        ItemEmpenho.objects
        .select_related('empenho', 'estoque')
    )

    itens_por_estoque = defaultdict(list)
    for item in todos_itens:
        itens_por_estoque[item.estoque_id].append(item)

    lotes_contexto = []
    for lote in estoque_qs:
        itens = itens_por_estoque.get(lote.id, [])
        empenhado = lote.empenhado
        disponivel = lote.disponivel
        
        for item in itens:
            item.inconsistente = item.quantidade > lote.saldo
        
        tem_inconsistencia = any(
            item.quantidade > lote.saldo for item in itens
        )
        
        lotes_contexto.append({
            'lote': lote,
            'empenhado': empenhado,
            'disponivel': disponivel,
            'itens_empenho': itens,
            'tem_inconsistencia': tem_inconsistencia
        })

    cards_ativos = (
        Empenho.objects
        .filter(status__nome='Rascunho')
        .prefetch_related(
            'itens',
            'itens__estoque',
            'historico_itens'
        )
        .order_by('-id')
    )
    
    cards_concluidos = (
        Empenho.objects
        .filter(status__nome='Concluído')
        .prefetch_related('historico_itens')
        .order_by('-id')
    )
    
    # ================================================================
    # DADOS PARA O MODAL (JSON)
    # ================================================================
    cards_impressao = {}
    for card in list(cards_ativos) + list(cards_concluidos):
        itens_pendentes = []
        for item in card.itens.all():
            estoque = item.estoque
            if estoque is None:
                continue
            itens_pendentes.append({
                'item_id': item.id,
                'empenho_id': card.id,
                'estoque_id': estoque.id,
                'lote': item.lote or estoque.lote,
                'quantidade': item.quantidade,
                'endereco': item.endereco_origem or estoque.endereco,
                'produto': estoque.produto or '',
                'cultivar': item.cultivar or (estoque.cultivar.nome if estoque.cultivar else ''),
                'peneira': item.peneira or (estoque.peneira.nome if estoque.peneira else ''),
                'categoria': item.categoria or (estoque.categoria.nome if estoque.categoria else ''),
                'especie': estoque.especie.nome if estoque.especie else '',
                'tratamento': estoque.tratamento.nome if estoque.tratamento else '',
                'embalagem': estoque.embalagem or '',
                'empresa': estoque.empresa or '',
                'cliente': estoque.cliente or '',
                'saldo_atual': estoque.saldo,
                'peso_unitario': str(estoque.peso_unitario) if estoque.peso_unitario else '0',
                'peso_total': str(estoque.peso_total) if estoque.peso_total else '0',
                'az': estoque.az or '',
                'conferente': estoque.conferente.get_full_name() if estoque.conferente else '',
                'observacao': item.observacao or estoque.observacao or '',
                'status_sistemico': estoque.status_sistemico.nome if estoque.status_sistemico else '',
                'situacao': 'pendente',
                'processado_em': None
            })
        
        itens_processados = []
        for hist in card.historico_itens.all():
            itens_processados.append({
                'item_id': hist.id,
                'empenho_id': card.id,
                'estoque_id': hist.estoque_origem_id,
                'lote': hist.lote,
                'quantidade': hist.quantidade,
                'endereco': hist.endereco_origem,
                'produto': hist.produto,
                'cultivar': hist.cultivar,
                'peneira': hist.peneira,
                'categoria': hist.categoria,
                'especie': hist.especie,
                'tratamento': hist.tratamento,
                'embalagem': hist.embalagem,
                'empresa': hist.empresa,
                'cliente': hist.cliente,
                'saldo_atual': 0,
                'peso_unitario': '0',
                'peso_total': '0',
                'az': '',
                'conferente': '',
                'observacao': hist.observacao,
                'status_sistemico': '',
                'situacao': 'transferido' if hist.tipo == 'transferencia' else 'expedido',
                'processado_em': hist.processado_em.strftime('%d/%m/%Y %H:%M') if hist.processado_em else '',
                'tipo': hist.get_tipo_display(),
                'endereco_destino': hist.endereco_destino if hist.tipo == 'transferencia' else '',
            })
        
        cards_impressao[str(card.id)] = {
            'card_id': card.id,
            'card_nome': card.observacao or f'Card #{card.id}',
            'itens_pendentes': itens_pendentes,
            'itens_processados': itens_processados,
            'total_pendentes': len(itens_pendentes),
            'total_processados': len(itens_processados),
            'itens_ids': [item['item_id'] for item in itens_pendentes],
        }

    return render(request, 'sapp/pagina_rascunho.html', {
        'lotes': lotes_contexto,
        'cards_ativos': cards_ativos,
        'cards_concluidos': cards_concluidos,
        'cards_impressao': cards_impressao,
        'cards_impressao_json': json.dumps(cards_impressao, ensure_ascii=False),
    })

    
def processar_transferencia_item(request, item, user, MARCA_ORIGEM, obs_global, empenho):
    """Processa a transferência de um item específico."""
    origem = item.estoque
    qtd = item.quantidade  # Quantidade total do item, nunca parcial
    
    novo_end = request.POST.get('novo_endereco', '').strip().upper()
    novo_az = request.POST.get('az', '').strip().upper() or origem.az
    obs_transferencia = request.POST.get('obs_transferencia', '').strip()
    
    if not novo_end:
        raise ValueError("Novo endereço não informado.")
    
    if novo_end == origem.endereco:
        raise ValueError(
            f"Endereço de destino igual ao de origem para lote {origem.lote}."
        )
    
    # Buscar ou criar destino com lock para evitar duplicação
    destino = (
        Estoque.objects
        .select_for_update()
        .filter(
            lote=origem.lote,
            produto=origem.produto,
            cultivar=origem.cultivar,
            peneira=origem.peneira,
            categoria=origem.categoria,
            tratamento=origem.tratamento,
            especie=origem.especie,
            endereco=novo_end,
            az=novo_az,
            empresa=origem.empresa,
            embalagem=origem.embalagem
        )
        .first()
    )
    
    if destino:
        destino.entrada += qtd
        destino.save()
    else:
        destino = Estoque.objects.create(
            lote=origem.lote,
            produto=origem.produto,
            cultivar=origem.cultivar,
            peneira=origem.peneira,
            categoria=origem.categoria,
            tratamento=origem.tratamento,
            especie=origem.especie,
            endereco=novo_end,
            az=novo_az,
            entrada=qtd,
            peso_unitario=origem.peso_unitario,
            embalagem=origem.embalagem,
            conferente=user,
            empresa=origem.empresa,
            cliente=origem.cliente,
            observacao=f"{MARCA_ORIGEM} {obs_global} {obs_transferencia}".strip()
        )
    
    # Atualizar saída da origem
    origem.saida += qtd
    origem.save()
    
    # Criar histórico de movimentação (saída)
    HistoricoMovimentacao.objects.create(
        estoque=origem,
        usuario=user,
        quantidade=qtd,
        tipo='Transferência (Saída)',
        descricao=(
            f"{MARCA_ORIGEM} Transferido {qtd} un de "
            f"{origem.endereco} para {novo_end}. "
            f"{obs_transferencia}"
        ).strip()
    )
    
    # Criar histórico de movimentação (entrada)
    HistoricoMovimentacao.objects.create(
        estoque=destino,
        usuario=user,
        quantidade=qtd,
        tipo='Transferência (Entrada)',
        descricao=(
            f"{MARCA_ORIGEM} Recebido {qtd} un de "
            f"{origem.endereco} em {novo_end}. "
            f"{obs_transferencia}"
        ).strip()
    )
    
    # Criar histórico do item empenho (ANTES de excluir o item)
    HistoricoItemEmpenho.objects.create(
        empenho=empenho,
        item_empenho_id_original=item.id,
        estoque_origem=origem,
        estoque_destino=destino,
        lote=origem.lote,
        produto=origem.produto or '',
        cultivar=origem.cultivar.nome if origem.cultivar else '',
        peneira=origem.peneira.nome if origem.peneira else '',
        categoria=origem.categoria.nome if origem.categoria else '',
        tratamento=origem.tratamento.nome if origem.tratamento else '',
        especie=origem.especie.nome if origem.especie else '',
        embalagem=origem.embalagem or '',
        empresa=origem.empresa or '',
        cliente=origem.cliente or '',
        endereco_origem=origem.endereco,
        endereco_destino=novo_end,
        quantidade=qtd,
        tipo='transferencia',
        observacao=obs_transferencia,
        processado_por=user
    )
    
    # SÓ AGORA excluir o item original
    item.delete()


def processar_expedicao_item(request, item, user, MARCA_ORIGEM, obs_global, empenho):
    """Processa a expedição de um item específico."""
    origem = item.estoque
    qtd = item.quantidade  # Quantidade total do item, nunca parcial
    
    obs_expedicao = request.POST.get('obs_expedicao', '').strip()
    numero_carga = request.POST.get('numero_carga', '').strip()
    cliente = request.POST.get('cliente', '').strip()
    placa = request.POST.get('placa', '').strip()
    
    if not numero_carga:
        raise ValueError("Número da carga/pedido não informado.")
    
    # Atualizar saída da origem
    origem.saida += qtd
    origem.save()
    
    # Criar histórico de movimentação
    HistoricoMovimentacao.objects.create(
        estoque=origem,
        usuario=user,
        quantidade=qtd,
        tipo='Expedição',
        descricao=(
            f"{MARCA_ORIGEM} Expedido {qtd} un. "
            f"Carga: {numero_carga}. "
            f"{obs_global} {obs_expedicao}"
        ).strip(),
        numero_carga=numero_carga,
        cliente=cliente or origem.cliente,
        placa=placa
    )
    
    # Criar histórico do item empenho (ANTES de excluir o item)
    HistoricoItemEmpenho.objects.create(
        empenho=empenho,
        item_empenho_id_original=item.id,
        estoque_origem=origem,
        lote=origem.lote,
        produto=origem.produto or '',
        cultivar=origem.cultivar.nome if origem.cultivar else '',
        peneira=origem.peneira.nome if origem.peneira else '',
        categoria=origem.categoria.nome if origem.categoria else '',
        tratamento=origem.tratamento.nome if origem.tratamento else '',
        especie=origem.especie.nome if origem.especie else '',
        embalagem=origem.embalagem or '',
        empresa=origem.empresa or '',
        cliente=cliente or origem.cliente or '',
        endereco_origem=origem.endereco,
        quantidade=qtd,
        tipo='expedicao',
        observacao=obs_expedicao,
        numero_carga=numero_carga,
        placa=placa,
        processado_por=user
    )
    
    # SÓ AGORA excluir o item original
    item.delete()


@login_required
@permission_required('sapp.pode_ver_estoque', raise_exception=True)  # CORRIGIDO
def api_buscar_dados_lote(request):
    item_id = request.GET.get('item_id')
    
    try:
        item = Estoque.objects.select_related(
            'cultivar', 'peneira', 'categoria', 'tratamento', 'especie'
        ).get(id=item_id)
        data = {
            'encontrado': True,
            'id': item.id,
            'lote': item.lote,
            'endereco': item.endereco,
            'saldo': item.saldo,
            'entrada': item.entrada,
            'produto': item.produto or '',
            'cliente': item.cliente or '',
            'empresa': item.empresa or '',
            'az': item.az or '',
            'peso_unitario': str(item.peso_unitario).replace(',', '.') if item.peso_unitario else '0.00',
            'embalagem': item.embalagem,
            'observacao': item.observacao or '',
            'especie_id': item.especie.id if item.especie else '',
            'cultivar_id': item.cultivar.id if item.cultivar else '',
            'peneira_id': item.peneira.id if item.peneira else '',
            'categoria_id': item.categoria.id if item.categoria else '',
            'tratamento_id': item.tratamento.id if item.tratamento else '',
        }
        return JsonResponse(data)
    except Estoque.DoesNotExist:
        return JsonResponse({'encontrado': False, 'erro': 'Lote não encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'encontrado': False, 'erro': str(e)}, status=500)

#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@        ESTA COM DECORADOR         @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
@login_required
@permission_required('sapp.pode_ver_empenhos', raise_exception=True)
def api_itens_empenhos(request):
    """API para buscar itens dos empenhos selecionados"""
    empenhos_ids = request.GET.get('empenhos_ids', '')
    
    if not empenhos_ids:
        return JsonResponse({'itens': []})
    
    ids_list = [int(id) for id in empenhos_ids.split(',') if id.isdigit()]
    
    itens = ItemEmpenho.objects.filter(
        empenho_id__in=ids_list,
        empenho__usuario=request.user
    ).select_related('estoque', 'empenho')
    
    itens_data = []
    for item in itens:
        itens_data.append({
            'lote': item.estoque.lote,
            'quantidade': item.quantidade,
            'empenho': item.empenho.observacao,
            'endereco': item.estoque.endereco
        })
    
    return JsonResponse({
        'itens': itens_data,
        'total': len(itens_data)
    })



@login_required
@permission_required('sapp.pode_movimentar_estoque', raise_exception=True)
def api_autocomplete_nova_entrada(request):
    
    # No views.py, dentro de nova_entrada ou editar:

    endereco_raw = request.POST.get('endereco', '').strip().upper() # R-A LN01 P01
    # Regex para separar: (Rua) (Linha) (Posição)
    import re
    match = re.match(r'^(R-[A-Z]+)\s+(LN\d+)\s+(P\d+)$', endereco_raw)

    if not match:
        messages.error(request, "Formato de endereço inválido! Use: R-A LN01 P01")
        return redirect('sapp:lista_estoque')

    rua_nome, linha_nome, posicao_str = match.groups()

    # 1. Validar Posição (01 a 06)
    posicao_num = int(re.search(r'\d+', posicao_str).group())
    if posicao_num < 1 or posicao_num > 6:
        messages.error(request, f"Posição {posicao_str} inválida! Use de 01 a 06.")
        return redirect('sapp:lista_estoque')

    # 2. Verificar se Rua e Linha existem no cadastro
    rua_obj = Rua.objects.filter(nome=rua_nome).first()
    if not rua_obj:
        messages.error(request, f"Rua {rua_nome} não cadastrada!")
        return redirect('sapp:lista_estoque')

    if not Linha.objects.filter(nome=linha_nome).exists():
        messages.error(request, f"Linha {linha_nome} não cadastrada!")
        return redirect('sapp:lista_estoque')

    # 3. SETAR ARMAZÉM AUTOMÁTICO (Puxa da Rua)
    item.az = rua_obj.armazem.nome
    item.endereco = endereco_raw
    termo = request.GET.get('term', '').strip()
    
    if len(termo) < 2:
        return JsonResponse([], safe=False)
    
    # Busca lotes que contenham o texto digitado
    qs = Estoque.objects.filter(lote__icontains=termo).select_related(
        'especie', 'cultivar', 'peneira', 'categoria', 'tratamento'
    ).order_by('-id')
    
    resultados = []
    lotes_vistos = set()
    
    for item in qs:
        if item.lote not in lotes_vistos:
            dados_item = {
                'lote': item.lote,
                'produto': item.produto or '',
                'cultivar__id': item.cultivar.id if item.cultivar else None,
                'peneira__id': item.peneira.id if item.peneira else None,
                'categoria__id': item.categoria.id if item.categoria else None,
                'tratamento__id': item.tratamento.id if item.tratamento else None,
                'especie__id': item.especie.id if item.especie else None,
                
                'empresa': item.empresa or '',
                'origem_destino': item.origem_destino or '',
                'cliente': item.cliente or '',
                'peso_unitario': str(item.peso_unitario),
                'embalagem': item.embalagem,
                'az': item.az or '',
                'observacao': item.observacao or ''
            }

            resultados.append({
                'label': item.lote,
                'dados': dados_item
            })
            lotes_vistos.add(item.lote)
        
        if len(resultados) >= 10: 
            break
            
    return JsonResponse(resultados, safe=False)

    


@staff_member_required
def api_status_enderecos(request):
    enderecos = MapeamentoEndereco.objects.filter(ativo=True)
    resultado = {}
    
    for mapa in enderecos:
        tem_saldo = Estoque.objects.filter(
            endereco=mapa.endereco, 
            saldo__gt=0
        ).exists()
        
        resultado[mapa.endereco] = {
            'tem_saldo': tem_saldo,
            'cor_padrao': mapa.cor_padrao,
            'cor_positivo': mapa.cor_positivo
        }
    
    return JsonResponse(resultado)

# ============================================================================
# APIs PARA O CANVAS (ADMIN APENAS)
# ============================================================================


@login_required
@permission_required('sapp.pode_ver_estoque', raise_exception=True)
def verificar_estoque_endereco(request, endereco):
    """API para verificar se existe estoque em um endereço"""
    if request.method == 'GET':
        try:
            # Decodifica o endereço (pode ter espaços ou caracteres especiais)
            endereco_decodificado = endereco
            
            # Verifica se há estoque
            tem_estoque = Estoque.objects.filter(
                endereco__iexact=endereco_decodificado,
                saldo__gt=0
            ).exists()
            
            # Verifica se existe cadastro (mesmo com saldo zero)
            existe_cadastro = Estoque.objects.filter(
                endereco__iexact=endereco_decodificado
            ).exists()
            
            return JsonResponse({
                'success': True,
                'endereco': endereco_decodificado,
                'tem_estoque': tem_estoque,
                'existe_cadastro': existe_cadastro,
                'mensagem': f'Endereço {endereco_decodificado} tem estoque' if tem_estoque else f'Endereço {endereco_decodificado} está vazio'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False, 
                'error': str(e),
                'endereco': endereco
            })
    
    return JsonResponse({'success': False, 'error': 'Método não permitido'})



def exportar_mapa_json(request, armazem_numero):
   
    if not request.user.is_staff:
        return JsonResponse({'error': 'Acesso negado'}, status=403)
    
    armazem = get_object_or_404(ArmazemLayout, numero=armazem_numero)
    elementos = armazem.elementos.all().order_by('ordem_z')
    
    dados = {
        'armazem': {
            'id': armazem.id,
            'numero': armazem.numero,
            'nome': armazem.nome,
            'largura_canvas': armazem.largura_canvas,
            'altura_canvas': armazem.altura_canvas,
        },
        'elementos': [
            {
                'id': elem.id,
                'tipo': elem.tipo,
                'pos_x': elem.pos_x,
                'pos_y': elem.pos_y,
                'largura': elem.largura,
                'altura': elem.altura,
                'cor_preenchimento': elem.cor_preenchimento,
                'cor_borda': elem.cor_borda,
                'espessura_borda': elem.espessura_borda,
                'conteudo_texto': elem.conteudo_texto,
                'fonte_nome': elem.fonte_nome,
                'fonte_tamanho': elem.fonte_tamanho,
                'texto_negrito': elem.texto_negrito,
                'texto_italico': elem.texto_italico,
                'texto_direcao': elem.texto_direcao,
                'linha_tipo': elem.linha_tipo,
                'identificador': elem.identificador,
                'ordem_z': elem.ordem_z,
            }
            for elem in elementos
        ],
        'total_elementos': elementos.count(),
        'exportado_em': timezone.now().isoformat()
    }
    
    return JsonResponse(dados, json_dumps_params={'indent': 2})


@staff_member_required
@csrf_exempt
def importar_mapa_json(request, armazem_numero):
   
    if request.method == 'POST':
        try:
            armazem = get_object_or_404(ArmazemLayout, numero=armazem_numero)
            data = json.loads(request.body)
            
            # Limpa elementos existentes
            ElementoMapa.objects.filter(armazem=armazem).delete()
            
            # Cria novos elementos
            elementos_criados = []
            for idx, elem_data in enumerate(data.get('elementos', [])):
                elemento = ElementoMapa.objects.create(
                    armazem=armazem,
                    tipo=elem_data.get('tipo', 'RETANGULO'),
                    pos_x=elem_data.get('pos_x', 0),
                    pos_y=elem_data.get('pos_y', 0),
                    largura=elem_data.get('largura', 100),
                    altura=elem_data.get('altura', 60),
                    cor_preenchimento=elem_data.get('cor_preenchimento', '#CCCCCC'),
                    cor_borda=elem_data.get('cor_borda', '#000000'),
                    espessura_borda=elem_data.get('espessura_borda', 2),
                    conteudo_texto=elem_data.get('conteudo_texto', ''),
                    fonte_nome=elem_data.get('fonte_nome', 'Arial'),
                    fonte_tamanho=elem_data.get('fonte_tamanho', 14),
                    texto_negrito=elem_data.get('texto_negrito', False),
                    texto_italico=elem_data.get('texto_italico', False),
                    texto_direcao=elem_data.get('texto_direcao', 'horizontal'),
                    linha_tipo=elem_data.get('linha_tipo', 'solida'),
                    identificador=elem_data.get('identificador', ''),
                    ordem_z=elem_data.get('ordem_z', idx + 1),
                )
                elementos_criados.append(elemento.id)
            
            return JsonResponse({
                'success': True,
                'message': f'Mapa importado com sucesso! {len(elementos_criados)} elementos criados.',
                'armazem': armazem.numero,
                'total_elementos': len(elementos_criados)
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Método não permitido'})

# ============================================================================
# VIEW DE FALLBACK (para compatibilidade)
# ============================================================================

@login_required
@permission_required('sapp.pode_ver_mapa', raise_exception=True)  # 🔥 ADICIONAR
def lista_armazens(request):
  
    armazens = ArmazemLayout.objects.filter(ativo=True).order_by('numero')
    
    context = {
        'armazens': armazens,
        'is_admin': request.user.is_staff,
        'titulo_pagina': 'Mapas dos Armazéns'
    }
    return render(request, 'sapp/lista_armazens.html', context)

@staff_member_required
def criar_armazem(request):
    """Cria um novo AZ e redireciona para o editor dele"""
    if request.method == 'POST':
        numero = request.POST.get('numero')
        nome = request.POST.get('nome')
        largura = request.POST.get('largura_canvas', 1200)
        altura = request.POST.get('altura_canvas', 800)
        
        novo_az = ArmazemLayout.objects.create(
            numero=numero,
            nome=nome,
            largura_canvas=largura,
            altura_canvas=altura
        )
        messages.success(request, f"Armazém {novo_az.numero} criado com sucesso!")
        return redirect('sapp:editor_avancado', armazem_numero=novo_az.numero)
    return redirect('sapp:lista_armazens')

@staff_member_required
def editar_config_armazem(request, armazem_id):
  
    if request.method == 'POST':
        armazem = get_object_or_404(ArmazemLayout, id=armazem_id)
        armazem.numero = request.POST.get('numero')
        armazem.nome = request.POST.get('nome')
        armazem.largura_canvas = request.POST.get('largura_canvas')
        armazem.altura_canvas = request.POST.get('altura_canvas')
        armazem.save()
        
        messages.success(request, "Configurações do mapa atualizadas!")
        return redirect('sapp:editor_avancado', armazem_numero=armazem.numero)
    return redirect('sapp:lista_armazens')

# ============================================================================
# EDITOR DE MAPA (ADMIN)
# ============================================================================

@login_required
@permission_required('sapp.pode_ver_mapa', raise_exception=True)
def mapa_ocupacao_canvas(request, armazem_numero=1):
    # 1. Busca Armazém e Elementos
    armazem = get_object_or_404(ArmazemLayout, numero=armazem_numero, ativo=True)
    elementos_db = armazem.elementos.all().order_by('ordem_z')
    armazens_disponiveis = ArmazemLayout.objects.filter(ativo=True).order_by('numero')
    
    # 2. Busca Estoque com Saldo > 0
    itens_estoque = Estoque.objects.filter(saldo__gt=0)

    # 3. Mapeia Estoque (Normalizando Endereço: Tira espaços e põe Maiúsculo)
    dados_ocupacao = {}
    
    for item in itens_estoque:
        if item.endereco:
            # A MÁGICA: .strip().upper() garante que " a-01" seja igual a "A-01"
            chave = item.endereco.strip().upper()
            
            if chave not in dados_ocupacao:
                dados_ocupacao[chave] = []
            
            dados_ocupacao[chave].append({
                'lote': item.lote,
                'produto': str(item.produto or 'S/ Produto'),
                'qtd': float(item.saldo),
                'embalagem': str(item.embalagem),
                'cliente': str(item.cliente or '-')
            })

    # 4. Prepara Elementos para o Mapa (Já definindo a cor aqui)
    elementos_render = []
    
    for el in elementos_db:
        # Dados básicos
        item_dict = {
            'tipo': el.tipo,
            'x': el.pos_x, 'y': el.pos_y, 'w': el.largura, 'h': el.altura, 'rot': el.rotacao,
            'texto': el.conteudo_texto,
            'id': el.identificador
        }

        # SE FOR RETÂNGULO: Verifica se deve pintar
        if el.tipo == 'RETANGULO' and el.identificador:
            chave_mapa = el.identificador.strip().upper() # Normaliza também
            
            if chave_mapa in dados_ocupacao:
                # TEM ESTOQUE -> VERDE
                item_dict['cor'] = '#10b981' 
                item_dict['stroke'] = '#065f46'
                item_dict['ocupado'] = True
            else:
                # VAZIO -> CINZA (Ou a cor que você escolheu no editor)
                item_dict['cor'] = el.cor_preenchimento or '#f3f4f6'
                item_dict['stroke'] = el.cor_borda or '#9ca3af'
                item_dict['ocupado'] = False
        else:
            # TEXTOS e LINHAS -> Cor original
            item_dict['cor'] = el.cor_preenchimento
            item_dict['stroke'] = el.cor_borda
            item_dict['ocupado'] = False

        elementos_render.append(item_dict)

    # 5. Renderiza
    context = {
        'armazem': armazem,
        'armazens_disponiveis': armazens_disponiveis,
        'elementos_json': json.dumps(elementos_render, cls=DjangoJSONEncoder),
        'dados_ocupacao_json': json.dumps(dados_ocupacao, cls=DjangoJSONEncoder),
        'is_admin': request.user.is_staff,
    }
    
    return render(request, 'sapp/mapa_visualizacao.html', context)

@staff_member_required
@csrf_exempt
def salvar_todos_elementos(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            armazem_id = data.get('armazem_id')
            elementos_data = data.get('elementos', [])
            
            armazem = ArmazemLayout.objects.get(id=armazem_id)
            
            ElementoMapa.objects.filter(armazem=armazem).delete()
            
            novos_objetos = []
            for idx, item in enumerate(elementos_data):
                novo = ElementoMapa(
                    armazem=armazem,
                    tipo=item.get('tipo', 'RETANGULO'),
                    pos_x=item.get('pos_x'),
                    pos_y=item.get('pos_y'),
                    largura=item.get('largura'),
                    altura=item.get('altura'),
                    rotacao=item.get('rotacao', 0),
                    ordem_z=idx, # A ordem que vem do array é a ordem visual
                    
                    # Dados visuais
                    cor_preenchimento=item.get('cor_preenchimento'),
                    conteudo_texto=item.get('conteudo_texto', ''),
                    fonte_tamanho=item.get('fonte_tamanho', 14),
                    
                    # O MAIS IMPORTANTE: O ENDEREÇO
                    identificador=item.get('identificador', '').strip().upper() 
                )
                novos_objetos.append(novo)
            
            # Bulk create é muito mais rápido
            ElementoMapa.objects.bulk_create(novos_objetos)
            
            return JsonResponse({'success': True, 'total': len(novos_objetos)})
            
        except Exception as e:
            print(f"Erro ao salvar mapa: {e}")
            return JsonResponse({'success': False, 'error': str(e)})
            
    return JsonResponse({'success': False, 'error': 'Método inválido'})

# ============================================================================
# FUNÇÕES AUXILIARES
# ============================================================================
"""
def lista_armazens(request):
    armazens = ArmazemLayout.objects.filter(ativo=True).order_by('numero')
    
    context = {
        'armazens': armazens,
        'is_admin': request.user.is_staff,
        'titulo_pagina': 'Mapas dos Armazéns'
    }
    return render(request, 'sapp/lista_armazens.html', context)
"""
@staff_member_required
@csrf_exempt
def criar_armazens_automaticos(request):
    """API para criar armazéns automaticamente"""
    if request.method == 'POST':
        try:
            armazens_padrao = [
                {'numero': 1, 'nome': 'Armazém Principal', 'largura_canvas': 1200, 'altura_canvas': 800},
                {'numero': 2, 'nome': 'Armazém Secundário', 'largura_canvas': 1000, 'altura_canvas': 600},
                {'numero': 3, 'nome': 'Armazém de Reserva', 'largura_canvas': 800, 'altura_canvas': 500},
            ]
            
            criados = []
            for data in armazens_padrao:
                armazem, created = ArmazemLayout.objects.get_or_create(
                    numero=data['numero'],
                    defaults=data
                )
                if created:
                    criados.append(f"Armazém {armazem.numero} - {armazem.nome}")
            
            return JsonResponse({
                'success': True,
                'message': f'{len(criados)} armazéns criados com sucesso!',
                'armazens': criados
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Método não permitido'})


@staff_member_required
def editor_avancado(request, armazem_numero=1):
    armazem = get_object_or_404(ArmazemLayout, numero=armazem_numero, ativo=True)
    elementos = armazem.elementos.all().order_by('ordem_z')
    
    # ADICIONE ESTA LINHA ABAIXO se não tiver:
    armazens_disponiveis = ArmazemLayout.objects.filter(ativo=True).order_by('numero')
    
    context = {
        'armazem': armazem,
        'elementos': elementos,
        'armazens_disponiveis': armazens_disponiveis, # ENVIE PARA O CONTEXTO
        'titulo_pagina': f'Editor Gráfico - Armazém {armazem.numero}',
    }
    return render(request, 'sapp/editor_avancado.html', context)


@login_required
@permission_required('sapp.pode_ver_estoque', raise_exception=True)  # CORRIGIDO
def api_buscar_produto(request):

    try:
        # Log para debug
        print("=" * 50)
        print("API: Recebida requisição para buscar produto")
        print(f"API: Método: {request.method}")
        print(f"API: GET params: {dict(request.GET)}")
        
        # Apenas aceita GET
        if request.method != 'GET':
            return JsonResponse({
                'encontrado': False,
                'erro': 'Método não permitido. Use GET.'
            }, status=405)
        
        # Pegar código da query string
        codigo = request.GET.get('codigo', '').strip()
        
        if not codigo:
            print("API: Erro - Código não fornecido")
            return JsonResponse({
                'encontrado': False, 
                'erro': 'Código não fornecido'
            }, status=400)
        
        print(f"API: Buscando produto com código: '{codigo}'")
        
        # Importar dentro da função para evitar problemas de importação circular
        from .models import Produto
        
        # Buscar produto ativo pelo código
        produto = Produto.objects.filter(codigo=codigo, ativo=True).first()
        
        if not produto:
            print(f"API: Produto '{codigo}' não encontrado ou inativo")
            return JsonResponse({
                'encontrado': False, 
                'erro': f'Produto "{codigo}" não encontrado ou inativo'
            })
        
        print(f"API: Produto encontrado - ID: {produto.id}, Código: {produto.codigo}")
        
        # Preparar dados para resposta - USANDO OS CAMPOS REAIS DO SEU MODELO
        dados = {
            'codigo': produto.codigo,
            'cultivar_id': str(produto.cultivar.id) if produto.cultivar else None,
            'cultivar_nome': produto.cultivar.nome if produto.cultivar else '',
            'peneira_id': str(produto.peneira.id) if produto.peneira else None,
            'peneira_nome': produto.peneira.nome if produto.peneira else '',
            'especie_id': str(produto.especie.id) if produto.especie else None,
            'especie_nome': produto.especie.nome if produto.especie else '',
            'categoria_id': str(produto.categoria.id) if produto.categoria else None,
            'categoria_nome': produto.categoria.nome if produto.categoria else '',
            'tratamento_id': str(produto.tratamento.id) if produto.tratamento else None,
            'tratamento_nome': produto.tratamento.nome if produto.tratamento else '',
            'empresa': produto.empresa or '',
            'tipo': produto.tipo or '',
            'descricao': produto.descricao or ''
        }
        
        print(f"API: Dados preparados para retorno: {json.dumps(dados, indent=2, ensure_ascii=False)}")
        
        # Criar resposta
        response_data = {
            'encontrado': True, 
            'dados': dados
        }
        
        print("API: Retornando dados com sucesso")
        print("=" * 50)
        
        return JsonResponse(response_data)
        
    except Exception as e:
        print(f"API: ERRO INTERNO: {str(e)}")
        import traceback
        traceback.print_exc()
        
        return JsonResponse({
            'encontrado': False, 
            'erro': f'Erro interno do servidor: {str(e)}'
        }, status=500)
    



# sapp/views.py

@login_required
@permission_required('sapp.pode_movimentar_estoque', raise_exception=True)
def api_atualizar_status_sistemico(request):
    """API para atualizar o status sistêmico com cores personalizadas"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            lote_id = data.get('lote_id')
            status_id = data.get('status_id')
            observacao = data.get('observacao', '').strip()
            
            if not lote_id or not status_id:
                return JsonResponse({
                    'success': False, 
                    'error': 'ID do lote e status são obrigatórios'
                })
            
            with transaction.atomic():
                lote = get_object_or_404(Estoque, id=lote_id)
                status_anterior = lote.status_sistemico
                novo_status = get_object_or_404(StatusSistemico, id=status_id)
                
                # Salvar histórico
                HistoricoStatusSistemico.objects.create(
                    estoque=lote,
                    status_anterior=status_anterior,
                    status_novo=novo_status,
                    observacao=observacao or '',
                    alterado_por=request.user
                )
                
                # Atualizar o lote
                lote.status_sistemico = novo_status
                lote.status_sistemico_alterado_por = request.user
                lote.status_sistemico_alterado_em = timezone.now()
                lote.status_sistemico_observacao = observacao or ''
                lote.save()
                
                # Registrar no histórico geral
                HistoricoMovimentacao.objects.create(
                    estoque=lote,
                    usuario=request.user,
                    tipo='Status Sistêmico',
                    descricao=f'Status alterado para: {novo_status.nome} - {observacao or "Sem observação"}'
                )
                
                return JsonResponse({
                    'success': True,
                    'status': {
                        'id': novo_status.id,
                        'nome': novo_status.nome,
                        'cor': novo_status.cor,
                        'icone': novo_status.icone or '',
                        'legenda': novo_status.legenda or '',
                    },
                    'alterado_por': request.user.get_full_name() or request.user.username,
                    'alterado_em': timezone.now().strftime('%d/%m/%Y %H:%M'),
                    'observacao': observacao or ''
                })
                
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Método não permitido'})

# sapp/views.py - Adicione esta função

@login_required
@permission_required('sapp.pode_movimentar_estoque', raise_exception=True)
def api_excluir_status(request, status_id):
    """Exclui um status personalizado (apenas se não estiver em uso)"""
    if request.method != 'DELETE':
        return JsonResponse({'success': False, 'error': 'Método não permitido'})
    
    try:
        status = get_object_or_404(StatusSistemico, id=status_id)
        
        # Verificar se é status padrão
        if status.e_padrao:
            return JsonResponse({
                'success': False, 
                'error': 'Status padrão não pode ser excluído'
            })
        
        # Verificar se está em uso
        em_uso = Estoque.objects.filter(status_sistemico=status).exists()
        if em_uso:
            return JsonResponse({
                'success': False, 
                'error': 'Status está em uso por um ou mais lotes. Remova os lotes primeiro.'
            })
        
        # Excluir
        status.delete()
        return JsonResponse({'success': True})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
    
# sapp/views.py

from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required, permission_required
from django.views.decorators.csrf import csrf_exempt
from django.db import transaction
import json
from .models import Estoque, StatusSistemico, HistoricoStatusSistemico

@login_required
@permission_required('sapp.pode_movimentar_estoque', raise_exception=True)
def api_atualizar_status_sistemico(request):
    """API para atualizar o status sistêmico com cores personalizadas"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            lote_id = data.get('lote_id')
            status_id = data.get('status_id')
            observacao = data.get('observacao', '').strip()
            nova_legenda = data.get('nova_legenda', '').strip()
            
            if not lote_id or not status_id:
                return JsonResponse({
                    'success': False, 
                    'error': 'ID do lote e status são obrigatórios'
                })
            
            with transaction.atomic():
                lote = get_object_or_404(Estoque, id=lote_id)
                status_anterior = lote.status_sistemico
                novo_status = get_object_or_404(StatusSistemico, id=status_id)
                
                # Atualizar legenda se fornecida
                if nova_legenda:
                    novo_status.legenda = nova_legenda
                    novo_status.save(update_fields=['legenda'])
                
                # Salvar histórico antes de alterar
                HistoricoStatusSistemico.objects.create(
                    estoque=lote,
                    status_anterior=status_anterior,
                    status_novo=novo_status,
                    observacao=observacao or '',
                    alterado_por=request.user
                )
                
                # Atualizar o lote
                lote.status_sistemico = novo_status
                lote.status_sistemico_alterado_por = request.user
                lote.status_sistemico_alterado_em = timezone.now()
                lote.status_sistemico_observacao = observacao or ''
                lote.save()
                
                # Registrar no histórico geral
                HistoricoMovimentacao.objects.create(
                    estoque=lote,
                    usuario=request.user,
                    tipo='Status Sistêmico',
                    descricao=f'Status alterado para: {novo_status.nome} - {observacao or "Sem observação"}'
                )
                
                return JsonResponse({
                    'success': True,
                    'status': {
                        'id': novo_status.id,
                        'nome': novo_status.nome,
                        'cor': novo_status.cor,
                        'icone': novo_status.icone or '',
                        'legenda': novo_status.legenda or '',
                    },
                    'alterado_por': request.user.get_full_name() or request.user.username,
                    'alterado_em': timezone.now().strftime('%d/%m/%Y %H:%M'),
                    'observacao': observacao or ''
                })
                
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Método não permitido'})

@login_required
@permission_required('sapp.pode_movimentar_estoque', raise_exception=True)
def api_listar_status(request):
    """Lista todos os status disponíveis"""
    status_list = StatusSistemico.objects.filter(ativo=True).order_by('ordem', 'nome')
    return JsonResponse({
        'success': True,
        'status': [{
            'id': s.id,
            'nome': s.nome,
            'cor': s.cor,
            'icone': s.icone or '',
            'legenda': s.legenda or '',
            'e_padrao': s.e_padrao,
        } for s in status_list]
    })

@login_required
@permission_required('sapp.pode_movimentar_estoque', raise_exception=True)
def api_criar_status(request):
    """Cria um novo status personalizado"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            nome = data.get('nome', '').strip()
            cor = data.get('cor', '#6c757d')
            legenda = data.get('legenda', '').strip()
            icone = data.get('icone', '')
            
            if not nome:
                return JsonResponse({'success': False, 'error': 'Nome do status é obrigatório'})
            
            # Verificar se já existe
            if StatusSistemico.objects.filter(nome__iexact=nome).exists():
                return JsonResponse({
                    'success': False, 
                    'error': f'Já existe um status com o nome "{nome}"'
                })
            
            status = StatusSistemico.objects.create(
                nome=nome,
                cor=cor,
                legenda=legenda or nome,
                icone=icone or '🔹',
                e_padrao=False,
                ativo=True,
                criado_por=request.user
            )
            
            return JsonResponse({
                'success': True,
                'status': {
                    'id': status.id,
                    'nome': status.nome,
                    'cor': status.cor,
                    'icone': status.icone,
                    'legenda': status.legenda,
                }
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Método não permitido'})

@login_required
@permission_required('sapp.pode_movimentar_estoque', raise_exception=True)
def api_editar_status(request, status_id):
    """Edita um status existente"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            status = get_object_or_404(StatusSistemico, id=status_id)
            
            # Atualizar campos
            if 'nome' in data:
                status.nome = data['nome'].strip()
            if 'cor' in data:
                status.cor = data['cor']
            if 'legenda' in data:
                status.legenda = data['legenda'].strip()
            if 'icone' in data:
                status.icone = data['icone']
            if 'ativo' in data:
                status.ativo = data['ativo']
            
            status.save()
            
            return JsonResponse({
                'success': True,
                'status': {
                    'id': status.id,
                    'nome': status.nome,
                    'cor': status.cor,
                    'icone': status.icone,
                    'legenda': status.legenda,
                }
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Método não permitido'})



################# dashboard ###########################################
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse
from django.db.models import Sum, Count, Q, Avg
from django.utils import timezone
from datetime import timedelta
import json
from .models import (
    Estoque, Cultivar, Peneira, Categoria, Especie, Tratamento, Produto,
    DashboardConfig, HistoricoMovimentacao, ArmazemLayout
)
from django.contrib import messages

def is_admin(user):
    return user.is_superuser or user.groups.filter(name='Administradores').exists()

@login_required
def dashboard(request):
    """Dashboard principal com gráficos dinâmicos"""
    
    if not request.user.is_superuser and not request.user.has_perm('sapp.pode_ver_dashboard') and not request.user.has_perm('sapp.pode_ver_estoque'):
        # Redireciona para a primeira página que o usuário tem permissão
        return redirect('sapp:redirecionar')
    
    # ==================== CONFIGURAÇÃO DO DASHBOARD ====================
    try:
        config = DashboardConfig.objects.get(criado_por=request.user)
    except DashboardConfig.DoesNotExist:
        config = DashboardConfig.objects.create(criado_por=request.user)
    
    # ==================== QUERYSET BASE COM FILTROS ====================
    queryset = Estoque.objects.all()
    
    # APLICAR FILTRO PL (peneira é null ou nome='sp')
    tipo_filtro = request.GET.get('tipo', 'todos')
    if tipo_filtro == 'pl':
        # PL = Sem peneira (peneira_id is null) OU peneira.nome = 'sp'
        queryset = queryset.filter(
            Q(peneira__isnull=True) | 
            Q(peneira__nome__iexact='sp')
        )
    elif tipo_filtro == 'nao_pl':
        # Não PL = Tem peneira e não é 'sp'
        queryset = queryset.filter(
            peneira__isnull=False
        ).exclude(peneira__nome__iexact='sp')
    
    # Outros filtros
    if request.GET.get('cultivar'):
        queryset = queryset.filter(cultivar_id=request.GET.get('cultivar'))
    if request.GET.get('peneira') and request.GET.get('peneira') != 'sp':
        queryset = queryset.filter(peneira_id=request.GET.get('peneira'))
    if request.GET.get('armazem'):
        queryset = queryset.filter(az=request.GET.get('armazem'))
    if request.GET.get('especie'):
        queryset = queryset.filter(especie_id=request.GET.get('especie'))
    
    # ==================== DADOS PARA O TEMPLATE ====================
    
    # KPIs principais
    total_sc = queryset.aggregate(total=Sum('saldo'))['total'] or 0
    total_bag = queryset.filter(embalagem='BAG').aggregate(total=Sum('saldo'))['total'] or 0
    peso_total = queryset.aggregate(total=Sum('peso_total'))['total'] or 0
    
    # Totais PL e Não PL
    total_pl_geral = Estoque.objects.filter(
        Q(peneira__isnull=True) | Q(peneira__nome__iexact='sp')
    ).count()
    total_nao_pl_geral = Estoque.objects.filter(
        peneira__isnull=False
    ).exclude(peneira__nome__iexact='sp').count()
    
    # Lotes ativos e esgotados
    itens_ativos = queryset.filter(saldo__gt=0).count()
    itens_esgotados = queryset.filter(saldo=0).count()
    
    # Movimentação do mês
    inicio_mes = timezone.now().replace(day=1, hour=0, minute=0, second=0)
    movimentacao_mes = HistoricoMovimentacao.objects.filter(data_hora__gte=inicio_mes).count()
    
    # TOP CULTIVARES
    top_cultivares = list(queryset.filter(
        saldo__gt=0, cultivar__isnull=False
    ).values('cultivar__nome').annotate(
        total_saldo=Sum('saldo')
    ).order_by('-total_saldo')[:10])
    
    # Dados para gráfico de ESPÉCIE
    dados_especie = list(queryset.filter(
        especie__isnull=False, saldo__gt=0
    ).values('especie__nome').annotate(
        total=Sum('saldo')
    ).order_by('-total')[:10])
    
    # Dados para gráfico de PENEIRA
    categorias_distribuicao = list(queryset.filter(
        saldo__gt=0, peneira__isnull=False
    ).exclude(peneira__nome__iexact='sp').values('peneira__nome').annotate(
        total=Sum('saldo')
    ).order_by('-total'))
    
    # ==================== GRÁFICO DE ARMAZÉM COM FILTROS ====================
    # Aplicar filtros ao queryset de armazém
    armazem_queryset = queryset.filter(az__isnull=False).exclude(az='')
    
    # Aplicar filtro por espécie no armazém
    if request.GET.get('armazem_especie'):
        armazem_queryset = armazem_queryset.filter(especie_id=request.GET.get('armazem_especie'))
    
    # Aplicar filtro por peneira no armazém
    if request.GET.get('armazem_peneira'):
        if request.GET.get('armazem_peneira') == 'pl':
            armazem_queryset = armazem_queryset.filter(
                Q(peneira__isnull=True) | Q(peneira__nome__iexact='sp')
            )
        elif request.GET.get('armazem_peneira') == 'nao_pl':
            armazem_queryset = armazem_queryset.filter(
                peneira__isnull=False
            ).exclude(peneira__nome__iexact='sp')
        else:
            armazem_queryset = armazem_queryset.filter(peneira_id=request.GET.get('armazem_peneira'))
    
    # Dados para gráfico de ARMAZÉM
    capacidade_armazem = list(armazem_queryset.values('az').annotate(
        total_sc=Sum('saldo'),
        total_lotes=Count('id'),
        peso_total=Sum('peso_total')
    ).order_by('az'))
    
    # ==================== GRÁFICO DE TENDÊNCIA CORRIGIDO ====================
    # Período baseado na configuração ou parâmetro da URL
    dias_tendencia = int(request.GET.get('tendencia_dias', 7))
    data_limite = timezone.now() - timedelta(days=dias_tendencia)
    
    from django.db.models.functions import TruncDate
    
    # Base queryset para tendência
    tendencia_queryset = HistoricoMovimentacao.objects.filter(
        data_hora__gte=data_limite
    )
    
    # Aplicar filtros à tendência
    if request.GET.get('tendencia_tipo'):
        if request.GET.get('tendencia_tipo') == 'pl':
            tendencia_queryset = tendencia_queryset.filter(
                Q(estoque__peneira__isnull=True) | 
                Q(estoque__peneira__nome__iexact='sp')
            )
        elif request.GET.get('tendencia_tipo') == 'nao_pl':
            tendencia_queryset = tendencia_queryset.filter(
                estoque__peneira__isnull=False
            ).exclude(estoque__peneira__nome__iexact='sp')
    
    if request.GET.get('tendencia_especie'):
        tendencia_queryset = tendencia_queryset.filter(
            estoque__especie_id=request.GET.get('tendencia_especie')
        )
    
    # Entradas por dia
    entradas = tendencia_queryset.filter(
        tipo__icontains='Entrada'
    ).annotate(
        dia=TruncDate('data_hora')
    ).values('dia').annotate(
        total=Count('id')
    ).order_by('dia')
    
    # Saídas por dia
    saidas = tendencia_queryset.filter(
        Q(tipo__icontains='Saída') | Q(tipo__icontains='Expedição')
    ).annotate(
        dia=TruncDate('data_hora')
    ).values('dia').annotate(
        total=Count('id')
    ).order_by('dia')
    
    # Criar dicionários para fácil acesso
    entradas_dict = {item['dia']: item['total'] for item in entradas}
    saidas_dict = {item['dia']: item['total'] for item in saidas}
    
    # Gerar lista de dias do período
    dias = []
    for i in range(dias_tendencia):
        dia = (timezone.now() - timedelta(days=i)).date()
        dias.append(dia)
    dias.reverse()
    
    movimentacoes_diarias = []
    for dia in dias:
        movimentacoes_diarias.append({
            'dia': dia.strftime('%d/%m'),
            'entradas': entradas_dict.get(dia, 0),
            'saidas': saidas_dict.get(dia, 0)
        })
    
    # Clientes únicos
    clientes_unicos = queryset.exclude(
        cliente__isnull=True
    ).exclude(cliente='').values('cliente').distinct().count()
    
    # Taxa de ocupação
    total_armazens = ArmazemLayout.objects.filter(ativo=True).count()
    if capacidade_armazem and total_armazens > 0:
        total_ocupado = sum([item['total_sc'] for item in capacidade_armazem])
        # Considerando capacidade média de 1000 SC por armazém
        taxa_ocupacao = min(round((total_ocupado / (total_armazens * 1000)) * 100), 100)
    else:
        taxa_ocupacao = 0
    
    # Movimentações recentes
    movimentacao_recente = HistoricoMovimentacao.objects.select_related(
        'usuario', 'estoque'
    ).order_by('-data_hora')[:10]
    
    # ==================== CONTEXTO ====================
    context = {
        # Configuração
        'config': config,
        'is_admin': is_admin(request.user),
        
        # KPIs principais
        'total_sc': total_sc,
        'total_sc_convertido': total_sc,
        'total_bag': total_bag,
        'peso_total': peso_total,
        'itens_ativos': itens_ativos,
        'itens_esgotados': itens_esgotados,
        'movimentacao_mes': movimentacao_mes,
        'clientes_unicos': clientes_unicos,
        'taxa_ocupacao': taxa_ocupacao,
        
        # Totais PL/Não PL
        'total_pl': total_pl_geral,
        'total_nao_pl': total_nao_pl_geral,
        
        # Dados para gráficos
        'top_cultivares': top_cultivares,
        'dados_especie': dados_especie,
        'categorias_distribuicao': categorias_distribuicao,
        'capacidade_armazem': capacidade_armazem,
        'movimentacoes_diarias': movimentacoes_diarias,
        'movimentacao_recente': movimentacao_recente,
        
        # Dados para filtros
        'cultivares': Cultivar.objects.all().order_by('nome'),
        'peneiras': Peneira.objects.all().order_by('nome'),
        'armazens': ArmazemLayout.objects.filter(ativo=True).order_by('numero'),
        'especies': Especie.objects.all().order_by('nome'),
        
        # Filtro ativo
        'tipo_filtro_ativo': tipo_filtro,
        'tendencia_dias_atual': dias_tendencia,
        
        'page_title': 'Dashboard Analítico',
    }
    
    return render(request, 'sapp/dashboard.html', context)


@login_required
@user_passes_test(is_admin)
def salvar_config_dashboard(request):
    """Salva configurações do dashboard (apenas admin)"""
    if request.method == 'POST':
        config, created = DashboardConfig.objects.update_or_create(
            criado_por=request.user,
            defaults={
                'cultivar_tipo': request.POST.get('cultivar_tipo', 'doughnut'),
                'cultivar_qtd': int(request.POST.get('cultivar_qtd', 10)),
                'cultivar_ordem': request.POST.get('cultivar_ordem', 'valor_desc'),
                'cultivar_zerados': request.POST.get('cultivar_zerados') == 'on',
                'cultivar_agrupar_outros': request.POST.get('cultivar_agrupar_outros') == 'on',
                'peneira_tipo': request.POST.get('peneira_tipo', 'pie'),
                'peneira_qtd': int(request.POST.get('peneira_qtd', 8)),
                'peneira_ordem': request.POST.get('peneira_ordem', 'valor_desc'),
                'armazem_tipo': request.POST.get('armazem_tipo', 'bar'),
                'armazem_ordem': request.POST.get('armazem_ordem', 'nome_asc'),
                'armazem_metrica': request.POST.get('armazem_metrica', 'volume'),
                'tendencia_periodo': int(request.POST.get('tendencia_periodo', 7)),
                'tendencia_saidas': request.POST.get('tendencia_saidas') == 'on',
                'tendencia_transferencias': request.POST.get('tendencia_transferencias') == 'on',
                'tendencia_agrupamento': request.POST.get('tendencia_agrupamento', 'day'),
                'auto_refresh': int(request.POST.get('auto_refresh', 0)),
                'unidade_padrao': request.POST.get('unidade_padrao', 'sc'),
                'tema_cores': request.POST.get('tema_cores', 'default'),
                'mostrar_legendas': request.POST.get('mostrar_legendas') == 'on',
                'mostrar_percentuais': request.POST.get('mostrar_percentuais') == 'on',
                'filtro_cultivar': request.POST.get('filtro_cultivar') == 'on',
                'filtro_peneira': request.POST.get('filtro_peneira') == 'on',
                'filtro_armazem': request.POST.get('filtro_armazem') == 'on',
                'filtro_periodo': request.POST.get('filtro_periodo') == 'on',
            }
        )
        
        messages.success(request, 'Configurações salvas com sucesso!')
        return redirect('sapp:dashboard')
    
    return redirect('sapp:dashboard')
################################################## fim dashbord ##################
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q
from .models import Estoque, Produto, ConfiguracaoLogo
import re

@login_required
@permission_required('sapp.pode_ver_estoque', raise_exception=True)
def ficha_rastreabilidade(request):
    """
    View para exibir a ficha de rastreabilidade
    PRIORIDADE: 
    1. parâmetro 'item_id' (para pegar a linha específica)
    2. parâmetro 'lote' (fallback para compatibilidade)
    3. filtros normais (exatamente 1 resultado)
    """
    
    import re
    from django.db.models import Q
    from .models import Estoque, Produto, ConfiguracaoLogo
    
    # ========== CASO 1: TEM ITEM_ID ESPECÍFICO ==========
    item_id = request.GET.get('item_id', '').strip()
    
    if item_id and item_id.isdigit():
        try:
            item = Estoque.objects.filter(id=item_id).first()
            
            if not item:
                messages.error(request, f"Item ID '{item_id}' não encontrado.")
                return redirect('sapp:gestao_estoque')
            
            # Processar o item e renderizar a ficha
            return processar_item_ficha(request, item)
            
        except Exception as e:
            messages.error(request, f"Erro ao buscar item: {str(e)}")
            return redirect('sapp:gestao_estoque')
    
    # ========== CASO 2: TEM LOTE ESPECÍFICO (FALLBACK) ==========
    lote_especifico = request.GET.get('lote', '').strip()
    
    if lote_especifico:
        try:
            # Busca o primeiro item com este lote
            item = Estoque.objects.filter(lote=lote_especifico).first()
            
            if not item:
                messages.error(request, f"Lote '{lote_especifico}' não encontrado.")
                return redirect('sapp:gestao_estoque')
            
            # Processar o item e renderizar a ficha
            return processar_item_ficha(request, item)
            
        except Exception as e:
            messages.error(request, f"Erro ao buscar lote: {str(e)}")
            return redirect('sapp:gestao_estoque')
    
    # ========== CASO 3: SEM LOTE ESPECÍFICO - USAR FILTROS NORMAIS ==========
    filtros = Q()
    
    # Campos de texto (busca parcial)
    campos_texto = ['lote', 'az', 'produto', 'endereco', 'cliente', 'empresa']
    for campo in campos_texto:
        valor = request.GET.get(campo, '')
        if valor and valor.strip():
            filtros &= Q(**{f'{campo}__icontains': valor.strip()})
    
    # Busca global
    busca = request.GET.get('busca', '')
    if busca and busca.strip():
        for termo in busca.split():
            filtros &= (
                Q(lote__icontains=termo) | 
                Q(produto__icontains=termo) |
                Q(cultivar__nome__icontains=termo) | 
                Q(especie__nome__icontains=termo) |
                Q(endereco__icontains=termo) | 
                Q(cliente__icontains=termo) |
                Q(empresa__icontains=termo)
            )
    
    # Filtros de seleção
    campos_selecao = [
        ('cultivar', 'cultivar__id__in'),
        ('peneira', 'peneira__id__in'),
        ('categoria', 'categoria__id__in'),
        ('especie', 'especie__id__in'),
        ('tratamento', 'tratamento__id__in'),
        ('embalagem', 'embalagem__in'),
    ]
    
    for param, lookup in campos_selecao:
        values = request.GET.getlist(param)
        values = [v for v in values if v and str(v).strip()]
        if values:
            filtros &= Q(**{lookup: values})
    
    # Filtro por status
    status = request.GET.get('status', 'todos')
    if status == 'disponivel':
        filtros &= Q(saldo__gt=0)
    elif status == 'esgotado':
        filtros &= Q(saldo=0)
    
    # Filtros numéricos
    for field in ['saldo', 'peso_unitario', 'peso_total']:
        min_val = request.GET.get(f'min_{field}')
        max_val = request.GET.get(f'max_{field}')
        if min_val and min_val.strip():
            try:
                filtros &= Q(**{f'{field}__gte': float(min_val)})
            except:
                pass
        if max_val and max_val.strip():
            try:
                filtros &= Q(**{f'{field}__lte': float(max_val)})
            except:
                pass
    
    # ========== BUSCAR ITENS COM OS FILTROS ==========
    itens_filtrados = Estoque.objects.filter(filtros).distinct()
    
    if itens_filtrados.count() != 1:
        messages.error(
            request, 
            f"É necessário ter exatamente 1 lote filtrado. Encontrados: {itens_filtrados.count()}"
        )
        return redirect(request.META.get('HTTP_REFERER', 'sapp:gestao_estoque'))
    
    # PEGAR O ÚNICO ITEM
    item = itens_filtrados.first()
    
    # Processar o item
    return processar_item_ficha(request, item)


def processar_item_ficha(request, item):
    """
    Função auxiliar para processar os dados do item e renderizar a ficha
    """
    import re
    from .models import Produto, ConfiguracaoLogo
    
    # Extrair código do produto do campo produto
    codigo_produto = ''
    descricao_completa = ''
    produto_obj = None
    
    # Tenta encontrar o código do produto
    if item.produto:
        # Primeiro tenta encontrar padrão de 10 dígitos
        match = re.search(r'\b(\d{10})\b', item.produto)
        if match:
            codigo_produto = match.group(1)
            produto_obj = Produto.objects.filter(codigo=codigo_produto).first()
        
        # Se não achou com 10 dígitos, usa o próprio produto como código
        if not produto_obj:
            codigo_produto = item.produto
            produto_obj = Produto.objects.filter(codigo=item.produto).first()
    
    # Se encontrou o produto, usa a descrição EXATA que está no model
    if produto_obj and produto_obj.descricao:
        descricao_completa = produto_obj.descricao
    
    # QR Code: código_produto/lote
    qrcode_texto = item.lote
    if codigo_produto:
        qrcode_texto = f"{codigo_produto}/{item.lote}"
    
    # Safra padrão 2025/2026
    safra = "2025/2026"
    
    # Extrair AZ do endereço se necessário
    az = item.az
    if not az and item.endereco:
        # Pega as primeiras letras do endereço como AZ
        az = ''.join([c for c in item.endereco[:2] if c.isalpha()]).upper()
    
    # Extrair RUA, LN, PS do endereço (formato: AZ RUA LN PS)
    rua = ''
    ln = ''
    ps = ''
    
    if item.endereco:
        partes = item.endereco.split()
        if len(partes) >= 4:
            rua = partes[1] if len(partes) > 1 else ''
            ln = partes[2] if len(partes) > 2 else ''
            ps = partes[3] if len(partes) > 3 else ''
    
    # Dados completos
    item_data = {
        'id': item.id,
        'lote': item.lote,
        'safra': safra,
        'codigo_produto': codigo_produto,
        'descricao': descricao_completa,
        'produto': descricao_completa,
        'az': az or '',
        'rua': rua,
        'ln': ln,
        'ps': ps,
        'endereco': item.endereco or '',
        'empresa': item.empresa or 'GRUPO CONCEITO',
        'peneira': item.peneira.nome if item.peneira else '',
        'categoria': item.categoria.nome if item.categoria else '',
        'cultivar': item.cultivar.nome if item.cultivar else '',
        'peso_unitario': item.peso_unitario,
        'peso_total': item.peso_total,
        'embalagem': item.get_embalagem_display() if hasattr(item, 'get_embalagem_display') else item.embalagem,
        'cliente': item.cliente or '',
        'status': item.status,
        'status_sistemico': item.status_sistemico,
        'saldo': item.saldo,
        'qrcode_texto': qrcode_texto,
    }
    
    # Buscar configuração da logo
    config_logo = ConfiguracaoLogo.get_logo()
    
    context = {
        'item': item_data,
        'config_logo': config_logo,
        'erro': None,
        'item_id': item.id,
        'lote_buscado': item.lote,
        'filtros_aplicados': request.GET.urlencode(),
    }
    
    return render(request, 'sapp/ficha_rastreabilidade.html', context)


def processar_item_ficha(request, item):
    """
    Função auxiliar para processar os dados do item e renderizar a ficha
    """
    # Extrair código do produto do campo produto
    codigo_produto = ''
    descricao_completa = ''
    produto_obj = None
    
    print(f"🔍 Debug - item.produto: '{item.produto}'")  # Debug
    
    # Tenta encontrar o código do produto
    if item.produto:
        # Primeiro tenta encontrar padrão de 10 dígitos
        match = re.search(r'\b(\d{10})\b', item.produto)
        if match:
            codigo_produto = match.group(1)
            produto_obj = Produto.objects.filter(codigo=codigo_produto).first()
            print(f"🔍 Debug - Código extraído (10 dígitos): '{codigo_produto}'")
        
        # Se não achou com 10 dígitos, usa o próprio produto como código
        if not produto_obj:
            codigo_produto = item.produto
            produto_obj = Produto.objects.filter(codigo=item.produto).first()
            print(f"🔍 Debug - Usando produto como código: '{codigo_produto}'")
    
    # Se encontrou o produto, usa a descrição EXATA que está no model
    if produto_obj and produto_obj.descricao:
        descricao_completa = produto_obj.descricao
        print(f"✅ Descrição encontrada no Produto: '{descricao_completa}'")
    else:
        # Fallback: vazio
        descricao_completa = ''
        print(f"⚠️ Nenhuma descrição encontrada")
    
    # QR Code: código_produto/lote
    qrcode_texto = item.lote
    if codigo_produto:
        qrcode_texto = f"{codigo_produto}/{item.lote}"
    
    # Safra padrão 2025/2026
    safra = "2025/2026"
    
    # Dados completos
    item_data = {
        'id': item.id,
        'lote': item.lote,
        'safra': safra,
        'codigo_produto': codigo_produto,
        'descricao': descricao_completa,
        'produto': descricao_completa,
        'az': item.az or (item.endereco[:2] if item.endereco else ''),
        'endereco': item.endereco or '',
        'empresa': item.empresa or 'GRUPO CONCEITO',
        'peneira': item.peneira.nome if item.peneira else '',
        'categoria': item.categoria.nome if item.categoria else '',
        'cultivar': item.cultivar.nome if item.cultivar else '',
        'peso_unitario': item.peso_unitario,
        'peso_total': item.peso_total,
        'embalagem': item.get_embalagem_display() if hasattr(item, 'get_embalagem_display') else item.embalagem,
        'cliente': item.cliente or '',
        'status': item.status,
        'status_sistemico': item.status_sistemico,
        'saldo': item.saldo,
        'qrcode_texto': qrcode_texto,
    }
    
    # Buscar configuração da logo
    config_logo = ConfiguracaoLogo.get_logo()
    
    context = {
        'item': item_data,
        'config_logo': config_logo,
        'erro': None,
        'lote_buscado': item.lote,
        'filtros_aplicados': request.GET.urlencode(),
    }
    
    return render(request, 'sapp/ficha_rastreabilidade.html', context)




def extrair_safra(lote):
    """
    Extrai a safra do número do lote
    Exemplos: 2025/2026, 2025, 25/26, SAFRA25
    """
    if not lote:
        return '______________'
    
    lote_str = str(lote)
    
    # Padrão: 2025/2026
    padrao1 = r'(20\d{2}[/-]20\d{2})'
    match = re.search(padrao1, lote_str)
    if match:
        return match.group(1)
    
    # Padrão: 25/26
    padrao2 = r'(\d{2}[/-]\d{2})'
    match = re.search(padrao2, lote_str)
    if match:
        ano1 = match.group(1)[:2]
        ano2 = match.group(1)[-2:]
        return f"20{ano1}/20{ano2}"
    
    # Padrão: SAFRA25 ou SAFRA2025
    padrao3 = r'SAFRA[-\s]*(\d{2,4})'
    match = re.search(padrao3, lote_str, re.IGNORECASE)
    if match:
        ano = match.group(1)
        if len(ano) == 2:
            return f"20{ano}"
        return ano
    
    # Padrão: apenas ano 2025
    padrao4 = r'(20\d{2})'
    match = re.search(padrao4, lote_str)
    if match:
        return match.group(1)
    
    return '______________'


@login_required
@permission_required('sapp.pode_ver_estoque', raise_exception=True)
def ficha_rastreabilidade_por_id(request, estoque_id):
    """
    View para exibir ficha de rastreabilidade por ID do estoque
    URL: /ficha-rastreabilidade/<int:estoque_id>/
    """
    try:
        item = get_object_or_404(Estoque, id=estoque_id)
        
        item_data = {
            'lote': item.lote,
            'safra': extrair_safra(item.lote),
            'produto': str(item.cultivar) if item.cultivar else item.produto,
            'az': item.az or item.endereco[:2] if item.endereco else '',
            'empresa': item.empresa or 'GRUPO CONCEITO',
            'peneira': item.peneira,
            'categoria': item.categoria,
            'cultivar': item.cultivar,
            'endereco': item.endereco,
            'saldo': item.saldo,
            'peso_unitario': item.peso_unitario,
            'peso_total': item.peso_total,
            'embalagem': item.get_embalagem_display(),
            'cliente': item.cliente,
            'status': item.status,
            'status_sistemico': item.status_sistemico,
        }
        
        context = {
            'item': item_data,
            'erro': None,
            'lote_buscado': item.lote,
        }
    except Exception as e:
        context = {
            'item': {
                'lote': '______________',
                'safra': '______________',
                'produto': '______________',
                'az': '______________',
                'empresa': '______________',
                'peneira': None,
                'categoria': None,
                'cultivar': None,
                'endereco': '______________',
                'saldo': 0,
                'peso_unitario': 0,
                'peso_total': 0,
                'embalagem': '---',
                'cliente': '______________',
                'status': '---',
                'status_sistemico': 'critico',
            },
            'erro': f"Erro ao buscar item: {str(e)}",
            'lote_buscado': None,
        }
    
    return render(request, 'ficha_rastreabilidade.html', context)

# View para múltiplos lotes (caso queira uma ficha com vários itens)
@login_required
@permission_required('sapp.pode_ver_estoque', raise_exception=True)  # CORRIGIDO
def ficha_rastreabilidade_multipla(request):
    """
    View para exibir fichas de múltiplos lotes
    Uso: /ficha-rastreabilidade/multipla/?lotes=123,456,789
    """
    lotes_param = request.GET.get('lotes', '')
    itens = []
    
    if lotes_param:
        lista_lotes = [l.strip() for l in lotes_param.split(',') if l.strip()]
        for lote in lista_lotes:
            item = Estoque.objects.filter(lote=lote).first()
            if item:
                itens.append({
                    'lote': item.lote,
                    'safra': extrair_safra(item.lote),
                    'produto': str(item.cultivar) if item.cultivar else item.produto,
                    'az': item.az or item.endereco[:2] if item.endereco else '',
                    'empresa': item.empresa or 'GRUPO CONCEITO',
                    'peneira': item.peneira,
                    'categoria': item.categoria,
                    'endereco': item.endereco,
                    'saldo': item.saldo,
                })
    
    context = {
        'itens': itens,
        'total_itens': len(itens),
    }
    return render(request, 'ficha_rastreabilidade_multipla.html', context)



import re

def extrair_ln_p(endereco):
    """
    Extrai LN e P de um endereço no formato R-X LN## P##
    Retorna (rua, ln, posicao) ou None se não seguir o padrão
    """
    if not endereco:
        return None
    
    # Padrão: R-X LN## P## (ex: R-A LN10 P03)
    pattern = r'^(R-[A-Z])\s+(LN\d+)\s+(P\d+)$'
    match = re.match(pattern, endereco.strip().upper())
    
    if match:
        rua = match.group(1)  # R-A
        ln = match.group(2)   # LN10
        p = match.group(3)    # P03
        posicao = int(re.search(r'\d+', p).group())  # 3
        return {
            'rua': rua,
            'ln': ln,
            'posicao': posicao,
            'endereco_completo': endereco
        }
    return None

def get_posicoes_linha(rua, ln):
    """
    Retorna todas as posições existentes de uma rua+linha
    """
    enderecos = Estoque.objects.filter(
        endereco__startswith=f"{rua} {ln} P"
    ).values_list('endereco', flat=True).distinct()
    
    posicoes = []
    for end in enderecos:
        dados = extrair_ln_p(end)
        if dados:
            posicoes.append({
                'endereco': end,
                'posicao': dados['posicao']
            })
    
    # Ordenar por posição numérica
    return sorted(posicoes, key=lambda x: x['posicao'])

@login_required
@permission_required('sapp.pode_movimentar_estoque', raise_exception=True)
def marcar_ultimo_lote_linha(request, estoque_id):
    """
    Marca/desmarca um lote como último da linha
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método não permitido'})
    
    try:
        lote = Estoque.objects.get(id=estoque_id)
        
        # Verificar se o endereço segue o padrão
        dados_end = extrair_ln_p(lote.endereco)
        if not dados_end:
            return JsonResponse({
                'success': False,
                'error': 'Endereço não segue padrão LN + P'
            })
        
        # Se já está marcado, desmarcar
        if lote.ultimo_lote_linha:
            lote.ultimo_lote_linha = False
            lote.save()
            
            # Limpar marcações da linha
            posicoes = get_posicoes_linha(dados_end['rua'], dados_end['ln'])
            for pos in posicoes:
                if pos['posicao'] >= dados_end['posicao']:
                    # Aqui você pode limpar alguma flag visual se necessário
                    pass
            
            return JsonResponse({
                'success': True,
                'marcado': False,
                'mensagem': 'Marca removida'
            })
        
        # Verificar se já existe outro último na mesma linha
        outro_ultimo = Estoque.objects.filter(
            endereco__startswith=f"{dados_end['rua']} {dados_end['ln']} P",
            ultimo_lote_linha=True
        ).exclude(id=estoque_id).first()
        
        if outro_ultimo:
            # Desmarcar o outro
            outro_ultimo.ultimo_lote_linha = False
            outro_ultimo.save()
        
        # Marcar este como último
        lote.ultimo_lote_linha = True
        lote.save()
        
        return JsonResponse({
            'success': True,
            'marcado': True,
            'mensagem': 'Marcado como último lote da linha'
        })
        
    except Estoque.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Lote não encontrado'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@permission_required('sapp.pode_ver_estoque', raise_exception=True)
def get_marcacoes_linha(request, rua, ln):
    """
    Retorna as posições afetadas pela marcação
    """
    try:
        # Encontrar o último lote marcado nesta linha
        ultimo = Estoque.objects.filter(
            endereco__startswith=f"{rua} {ln} P",
            ultimo_lote_linha=True
        ).first()
        
        if not ultimo:
            return JsonResponse({
                'success': True,
                'tem_marcacao': False,
                'posicoes_afetadas': []
            })
        
        dados_ultimo = extrair_ln_p(ultimo.endereco)
        if not dados_ultimo:
            return JsonResponse({
                'success': True,
                'tem_marcacao': False,
                'posicoes_afetadas': []
            })
        
        # Todas as posições da linha
        posicoes = get_posicoes_linha(rua, ln)
        
        # Filtrar posições >= a posição marcada
        posicoes_afetadas = [
            p['endereco'] for p in posicoes 
            if p['posicao'] >= dados_ultimo['posicao']
        ]
        
        return JsonResponse({
            'success': True,
            'tem_marcacao': True,
            'lote_marcado': ultimo.lote,
            'posicao_marcada': dados_ultimo['posicao'],
            'posicoes_afetadas': posicoes_afetadas
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
    


@login_required
@permission_required('sapp.pode_ver_mapa', raise_exception=True)
def api_mapa_dados(request, armazem_numero):
    """API para retornar dados do mapa em formato JSON"""
    try:
        armazem = get_object_or_404(ArmazemLayout, numero=armazem_numero)
        elementos = armazem.elementos.all()
        
        # Buscar estoque
        itens_estoque = Estoque.objects.filter(saldo__gt=0)
        ocupacao = {}
        
        for el in elementos:
            if el.tipo == 'RETANGULO' and el.identificador:
                chave = el.identificador.strip().upper()
                tem_estoque = itens_estoque.filter(endereco__iexact=chave).exists()
                if tem_estoque:
                    ocupacao[el.id] = True
        
        # Converter elementos para dicionário
        elementos_list = []
        for el in elementos:
            el_dict = {
                'id': el.id,
                'tipo': el.tipo,
                'x': el.pos_x,
                'y': el.pos_y,
                'w': el.largura,
                'h': el.altura,
                'rot': el.rotacao,
                'cor': el.cor_preenchimento,
                'stroke': el.cor_borda,
                'texto': el.conteudo_texto,
                'identificador': el.identificador,
            }
            elementos_list.append(el_dict)
        
        return JsonResponse({
            'success': True,
            'armazem': {
                'id': armazem.id,
                'numero': armazem.numero,
                'nome': armazem.nome,
                'largura_canvas': armazem.largura_canvas,
                'altura_canvas': armazem.altura_canvas,
            },
            'elementos': elementos_list,
            'ocupacao': ocupacao
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@permission_required('sapp.pode_ver_mapa', raise_exception=True) 
def api_marcacoes_ultimo_lote(request):
    """
    Retorna todas as posições que devem receber marcação de X
    (posições posteriores à posição marcada como último lote)
    """
    try:
        from .models import ArmazemLayout, ElementoMapa
        
        # Buscar todos os lotes marcados como último
        lotes_marcados = Estoque.objects.filter(
            ultimo_lote_linha=True,
            saldo__gt=0
        )
        
        marcacoes = {}
        
        for lote in lotes_marcados:
            # Extrair informações do endereço usando regex
            dados_end = extrair_info_endereco(lote.endereco)
            if not dados_end:
                continue
            
            rua = dados_end.get('rua')        # R-A
            ln = dados_end.get('linha')       # LN01
            posicao_marcada = dados_end.get('posicao')  # 4
            
            # Se não tiver posição, pula
            if not posicao_marcada:
                continue
            
            # Buscar no MAPA todos os endereços desta linha
            padrao = f"{rua} {ln} P"
            elementos = ElementoMapa.objects.filter(
                tipo='RETANGULO',
                identificador__startswith=padrao
            ).values_list('identificador', flat=True).distinct()
            
            # Para cada endereço do mapa, verificar se é posterior
            for endereco in elementos:
                dados_pos = extrair_info_endereco(endereco)
                if dados_pos and dados_pos.get('posicao', 0) > posicao_marcada:
                    marcacoes[endereco.strip().upper()] = True
        
        return JsonResponse({
            'success': True,
            'marcacoes': marcacoes,
            'total': len(marcacoes)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@login_required
@permission_required('sapp.pode_ver_estoque', raise_exception=True)
def api_get_armazem_by_rua(request):
    """
    Obtém o armazém a partir de um endereço
    """
    endereco_codigo = request.GET.get('endereco', '').strip().upper()
    
    if not endereco_codigo:
        return JsonResponse({'sucesso': False, 'msg': 'Endereço não informado'}, status=400)
    
    try:
        # Busca o endereço pelo código exato ou contém
        endereco = Endereco.objects.select_related('armazem').filter(
            codigo__icontains=endereco_codigo
        ).first()
        
        if endereco:
            # Extrai informações adicionais do endereço (se possível)
            dados = extrair_info_endereco(endereco.codigo)
            
            return JsonResponse({
                'sucesso': True, 
                'az': endereco.armazem.nome,
                'endereco': endereco.codigo,
                'rua': dados.get('rua') if dados else None,
                'linha': dados.get('linha') if dados else None,
                'posicao': dados.get('posicao') if dados else None
            })
        
        return JsonResponse({'sucesso': False, 'msg': 'Endereço não cadastrado'}, status=404)
        
    except Exception as e:
        return JsonResponse({'sucesso': False, 'msg': str(e)}, status=500)


def extrair_info_endereco(endereco_str):
    """
    Função auxiliar para extrair rua, linha e posição de um endereço
    Usa regex para extrair informações mesmo sem campos no banco
    
    Exemplos:
    - "R-A LN10 P02" -> {'rua': 'R-A', 'linha': 'LN10', 'posicao': 2}
    - "R-A LN10" -> {'rua': 'R-A', 'linha': 'LN10', 'posicao': None}
    - "R-A GERAL" -> {'rua': 'R-A', 'linha': 'GERAL', 'posicao': None}
    - "R-A" -> {'rua': 'R-A', 'linha': None, 'posicao': None}
    """
    import re
    
    if not endereco_str:
        return None
    
    endereco_str = endereco_str.strip().upper()
    
    # Padrão completo: R-A LN10 P02
    match_completo = re.match(r'^(R-[A-Z])\s+LN(\d{2})\s+P(\d{2})$', endereco_str)
    if match_completo:
        return {
            'rua': match_completo.group(1),
            'linha': f"LN{match_completo.group(2)}",
            'posicao': int(match_completo.group(3))
        }
    
    # Padrão sem posição: R-A LN10
    match_linha = re.match(r'^(R-[A-Z])\s+LN(\d{2})$', endereco_str)
    if match_linha:
        return {
            'rua': match_linha.group(1),
            'linha': f"LN{match_linha.group(2)}",
            'posicao': None
        }
    
    # Padrão geral: R-A GERAL
    match_geral = re.match(r'^(R-[A-Z])\s+GERAL$', endereco_str)
    if match_geral:
        return {
            'rua': match_geral.group(1),
            'linha': 'GERAL',
            'posicao': None
        }
    
    # Apenas a rua: R-A
    match_rua = re.match(r'^(R-[A-Z])$', endereco_str)
    if match_rua:
        return {
            'rua': match_rua.group(1),
            'linha': None,
            'posicao': None
        }
    
    # Formato livre: tenta extrair o primeiro como rua
    partes = endereco_str.split()
    return {
        'rua': partes[0] if partes else endereco_str,
        'linha': None,
        'posicao': None
    }


@login_required
@permission_required('sapp.pode_ver_estoque', raise_exception=True)  # Mude
def validar_endereco(request):
    """
    Valida endereço - busca nos endereços cadastrados e sugere cadastro se não existir
    """
    
    endereco_raw = request.GET.get('endereco', '').strip().upper()
    
    if not endereco_raw:
        return JsonResponse({
            'valido': False,
            'erro': 'Endereço não informado'
        })
    
    try:
        # Busca exata primeiro
        endereco_obj = Endereco.objects.filter(codigo=endereco_raw).select_related('armazem').first()
        
        if endereco_obj:
            return JsonResponse({
                'valido': True,
                'mensagem': f'✅ Endereço válido! Localizado no armazém {endereco_obj.armazem.nome}',
                'endereco_formatado': endereco_obj.codigo,
                'dados': {
                    'codigo': endereco_obj.codigo,
                    'id': endereco_obj.id,
                    'armazem': endereco_obj.armazem.nome,
                    'armazem_id': endereco_obj.armazem.id
                }
            })
        
        # Busca parcial (se digitar só parte do endereço)
        enderecos_similares = Endereco.objects.filter(
            codigo__icontains=endereco_raw
        ).select_related('armazem')[:5]
        
        if enderecos_similares.exists():
            sugestoes = [f"{e.codigo} ({e.armazem.nome})" for e in enderecos_similares]
            return JsonResponse({
                'valido': False,
                'erro': f'Endereço não encontrado. Você quis dizer: {", ".join(sugestoes)}?',
                'sugestoes': sugestoes
            })
        
        # Não encontrou nenhum
        return JsonResponse({
            'valido': False,
            'erro': f'❌ Endereço "{endereco_raw}" não cadastrado. Por favor, cadastre-o nas configurações primeiro.'
        })
        
    except Exception as e:
        return JsonResponse({
            'valido': False,
            'erro': f'Erro na validação: {str(e)}'
        })
    

def validar_edicao_empenho_solicitacao(solicitacao):
    """
    Impede criação, alteração ou remoção de empenho
    depois que a movimentação da solicitação começou.
    """

    status_bloqueados = {
        'MOVIMENTACAO_PARCIAL',
        'CONCLUIDO',
        'CANCELADO',
    }

    if solicitacao.status in status_bloqueados:
        raise ValueError(
            'Não é possível alterar o empenho porque '
            'a movimentação desta solicitação já foi iniciada.'
        )

    quantidade_movimentada = Decimal(
        str(
            solicitacao.quantidade_movimentada
            or 0
        )
    )

    if quantidade_movimentada > 0:
        raise ValueError(
            'Não é possível alterar o empenho porque '
            'esta solicitação já possui itens movimentados.'
        )


@login_required
@permission_required('sapp.pode_ver_estoque', raise_exception=True)
def buscar_origens(request):
    """
    Busca origens/destinos para autocomplete (mantida igual)
    """
    termo = request.GET.get('term', '').strip()
    if len(termo) < 2:
        return JsonResponse([], safe=False)
    
    origens = OrigemDestino.objects.filter(nome__icontains=termo)[:10]
    resultados = [{'id': o.id, 'nome': o.nome} for o in origens]
    
    return JsonResponse(resultados, safe=False)


@login_required
@permission_required('sapp.pode_ver_estoque', raise_exception=True) 
def api_buscar_enderecos(request):
    """
    Busca ENDEREÇOS para autocomplete (NOVA)
    """
    termo = request.GET.get('termo', '').strip().upper()
    
    if not termo or len(termo) < 2:
        return JsonResponse([], safe=False)
    
    enderecos = Endereco.objects.filter(
        codigo__icontains=termo
    ).select_related('armazem')[:10]
    
    resultados = []
    for end in enderecos:
        resultados.append({
            'id': end.id,
            'codigo': end.codigo,
            'armazem': end.armazem.nome,
            'label': f"{end.codigo} ({end.armazem.nome})",
            'value': end.codigo
        })
    
    return JsonResponse(resultados, safe=False)


@login_required
@permission_required('sapp.pode_ver_estoque', raise_exception=True)  # Mude
def api_listar_enderecos(request):
    """
    Lista TODOS os endereços para o frontend (NOVA)
    """
    enderecos = Endereco.objects.select_related('armazem').all().order_by('codigo')
    
    dados = []
    for end in enderecos:
        dados.append({
            'id': end.id,
            'codigo': end.codigo,
            'armazem': end.armazem.nome,
            'armazem_id': end.armazem.id
        })
    
    return JsonResponse(dados, safe=False)




# sapp/views.py - Adicione ou substitua esta função

@login_required
def redirecionar_usuario(request):
    """
    Redireciona o usuário para a primeira página que ele tem permissão
    """
    user = request.user
    
    print(f"\n🔍 REDIRECIONANDO USUÁRIO: {user.username}")
    print(f"Superusuário: {user.is_superuser}")
    
    # Mostrar todas as permissões para debug
    all_perms = list(user.get_all_permissions())
    print(f"Permissões totais: {all_perms}")
    
    # Superusuário vai para dashboard
    if user.is_superuser:
        print("✅ Superusuário -> Dashboard")
        return redirect('sapp:dashboard')
    
    # 🔥 PRIORIDADE 1: ALMOXARIFADO (APENAS VISUALIZAÇÃO)
    if user.has_perm('almoxarifado.pode_ver_almoxarifado'):
        print("✅ Usuário tem permissão de almoxarifado -> Redirecionando para Almoxarifado")
        return redirect('almoxarifado:lista_itens')
    
    # 🔥 PRIORIDADE 2: ALMOXARIFADO (GERENCIAR)
    if user.has_perm('almoxarifado.pode_gerenciar_almoxarifado'):
        print("✅ Usuário tem permissão de gerenciar almoxarifado -> Redirecionando para Almoxarifado")
        return redirect('almoxarifado:lista_itens')
    
    # 🔥 PRIORIDADE 3: EMPENHO
    if user.has_perm('sapp.pode_ver_empenhos') or user.has_perm('sapp.pode_criar_empenhos'):
        print("✅ Usuário tem permissão de empenho -> Redirecionando para Empenho")
        return redirect('sapp:pagina_rascunho')
    
    # 🔥 PRIORIDADE 4: ESTOQUE (visualização)
    if user.has_perm('sapp.pode_ver_estoque'):
        print("✅ Usuário tem permissão de estoque -> Redirecionando para Estoque")
        return redirect('sapp:lista_estoque')
    
    # 🔥 PRIORIDADE 5: MOVIMENTAR ESTOQUE
    if user.has_perm('sapp.pode_movimentar_estoque'):
        print("✅ Usuário tem permissão de movimentar -> Redirecionando para Gestão")
        return redirect('sapp:gestao_estoque')
    
    # 🔥 PRIORIDADE 6: MAPA
    if user.has_perm('sapp.pode_ver_mapa'):
        print("✅ Usuário tem permissão de mapa -> Redirecionando para Mapa")
        return redirect('sapp:mapa_canvas', armazem_numero=1)
    
    # 🔥 PRIORIDADE 7: DASHBOARD (apenas se tiver permissão específica)
    if user.has_perm('sapp.pode_ver_dashboard'):
        print("✅ Usuário tem permissão de dashboard -> Redirecionando para Dashboard")
        return redirect('sapp:dashboard')
    
    # Se não tiver nenhuma permissão, fazer logout com mensagem
    print("❌ Usuário sem nenhuma permissão! Fazendo logout...")
    from django.contrib.auth import logout
    messages.error(request, "❌ Você não tem permissão para acessar nenhuma página do sistema!")
    logout(request)
    return redirect('sapp:login')



# sapp/views.py - Exportação SEM saldo 0

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .models import Estoque
from datetime import datetime
from django.db.models import Q

def exportar_estoque_excel(request):
    """Exporta o estoque para Excel com os filtros aplicados (apenas saldo > 0)"""
    
    try:
        # Cria o workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Estoque"
        
        # Estilos
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2F8F4E', end_color='2F8F4E', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        cell_font = Font(name='Arial', size=10)
        cell_alignment = Alignment(vertical='center')
        
        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        # Cabeçalhos
        headers = [
            'Status', 'AZ', 'Lote', 'Produto', 'Cultivar', 'Peneira', 
            'Categoria', 'Endereço', 'Saldo', 'Peso Unit.', 'Peso Total',
            'Espécie', 'Tratamento', 'Embalagem', 'Cliente', 'Empresa', 
            'Conferente', 'Observação'
        ]
        
        # Aplica estilos nos cabeçalhos
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Query base - FILTRA APENAS SALDO > 0
        queryset = Estoque.objects.select_related(
            'cultivar', 'peneira', 'categoria', 'tratamento', 
            'especie', 'status_sistemico', 'conferente'
        ).filter(saldo__gt=0)  # 🔥 AQUI: Apenas itens com saldo maior que 0
        
        # Aplica busca
        busca = request.GET.get('busca', '')
        if busca:
            queryset = queryset.filter(
                Q(lote__icontains=busca) |
                Q(produto__icontains=busca) |
                Q(endereco__icontains=busca) |
                Q(cliente__icontains=busca) |
                Q(empresa__icontains=busca) |
                Q(az__icontains=busca) |
                Q(observacao__icontains=busca) |
                Q(conferente__username__icontains=busca) |
                Q(conferente__first_name__icontains=busca) |
                Q(conferente__last_name__icontains=busca)
            )
        
        # Aplica filtros de coluna (exceto page, page_size, busca, export)
        for key, values in request.GET.lists():
            if key in ['page', 'page_size', 'busca', 'export']:
                continue
            
            # Filtros numéricos
            if key.startswith('min_'):
                campo = key.replace('min_', '')
                for val in values:
                    if val:
                        try:
                            queryset = queryset.filter(**{f"{campo}__gte": float(val)})
                        except (ValueError, TypeError):
                            pass
                continue
            
            if key.startswith('max_'):
                campo = key.replace('max_', '')
                for val in values:
                    if val:
                        try:
                            queryset = queryset.filter(**{f"{campo}__lte": float(val)})
                        except (ValueError, TypeError):
                            pass
                continue
            
            # Filtros de seleção múltipla
            if values and values != ['']:
                filtro_q = Q()
                for valor in values:
                    if valor == '__null__':
                        # Filtra por valores vazios/nulos
                        if key in ['cultivar', 'peneira', 'categoria', 'tratamento', 'especie', 'status_sistemico']:
                            filtro_q |= Q(**{f"{key}__isnull": True})
                        else:
                            filtro_q |= Q(**{f"{key}__exact": ''}) | Q(**{f"{key}__isnull": True})
                    else:
                        # Mapeia campos relacionados
                        if key == 'cultivar':
                            filtro_q |= Q(cultivar__nome=valor)
                        elif key == 'peneira':
                            filtro_q |= Q(peneira__nome=valor)
                        elif key == 'categoria':
                            filtro_q |= Q(categoria__nome=valor)
                        elif key == 'tratamento':
                            filtro_q |= Q(tratamento__nome=valor)
                        elif key == 'especie':
                            filtro_q |= Q(especie__nome=valor)
                        elif key == 'status_sistemico':
                            filtro_q |= Q(status_sistemico__nome=valor)
                        elif key == 'conferente':
                            filtro_q |= Q(conferente__username=valor)
                        else:
                            filtro_q |= Q(**{key: valor})
                
                queryset = queryset.filter(filtro_q)
        
        # Ordena
        queryset = queryset.order_by('lote')
        
        total_registros = queryset.count()
        print(f"📊 Exportando {total_registros} registros (apenas saldo > 0)")
        
        # Preenche os dados
        for row_idx, item in enumerate(queryset, 2):
            # Nome do conferente
            nome_conferente = ''
            if item.conferente:
                nome_conferente = item.conferente.get_full_name().strip()
                if not nome_conferente:
                    nome_conferente = item.conferente.first_name.strip()
                if not nome_conferente:
                    nome_conferente = item.conferente.username
            
            # Nome do status
            nome_status = item.status_sistemico.nome if item.status_sistemico else 'Indefinido'
            
            data_row = [
                nome_status,
                item.az or '',
                item.lote or '',
                item.produto or '',
                item.cultivar.nome if item.cultivar else '',
                item.peneira.nome if item.peneira else '',
                item.categoria.nome if item.categoria else '',
                item.endereco or '',
                item.saldo if item.saldo is not None else 0,
                float(item.peso_unitario) if item.peso_unitario else 0.0,
                float(item.peso_total) if item.peso_total else 0.0,
                item.especie.nome if item.especie else '',
                item.tratamento.nome if item.tratamento else '',
                item.get_embalagem_display() if item.embalagem else '',
                item.cliente or '',
                item.empresa or '',
                nome_conferente,
                item.observacao or '',
            ]
            
            for col, value in enumerate(data_row, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = thin_border
        
        # Ajusta largura das colunas
        column_widths = {
            1: 15,   # Status
            2: 8,    # AZ
            3: 20,   # Lote
            4: 30,   # Produto
            5: 15,   # Cultivar
            6: 10,   # Peneira
            7: 12,   # Categoria
            8: 18,   # Endereço
            9: 10,   # Saldo
            10: 12,  # Peso Unit.
            11: 12,  # Peso Total
            12: 12,  # Espécie
            13: 15,  # Tratamento
            14: 12,  # Embalagem
            15: 25,  # Cliente
            16: 25,  # Empresa
            17: 25,  # Conferente
            18: 30,  # Observação
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Congela o cabeçalho
        ws.freeze_panes = 'A2'
        
        # Filtro automático
        if total_registros > 0:
            ultima_linha = total_registros + 1
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ultima_linha}"
        
        # Altura das linhas
        ws.row_dimensions[1].height = 30
        
        # Prepara resposta
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'estoque_{timestamp}.xlsx'
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        wb.save(response)
        
        print(f"✅ Arquivo Excel gerado: {filename} com {total_registros} registros")
        
        return response
        
    except Exception as e:
        import traceback
        print(f"❌ Erro na exportação Excel: {e}")
        print(traceback.format_exc())
        
        return HttpResponse(
            f"Erro ao gerar arquivo Excel: {str(e)}",
            content_type='text/plain',
            status=500
        )




# ============================================================================
# FASE 4 - ATUALIZAÇÃO AO VIVO, FEED E SOM
# ============================================================================

@login_required
def api_versao_cards(request):
    """
    Retorna um hash/versão dos cards para verificar se houve alteração.
    Usado pelo polling de 30 segundos.
    """
    from django.core.cache import cache
    from hashlib import md5
    
    cache_key = 'cards_version_hash'
    version = cache.get(cache_key)
    
    if not version:
        # Gerar hash baseado nos cards ativos
        cards_data = Solicitacao.objects.exclude(
            status='CONCLUIDO'
        ).values_list('id', 'data_atualizacao', 'status', 'quantidade_empenhada')
        
        hash_input = str(list(cards_data)).encode('utf-8')
        version = md5(hash_input).hexdigest()
        cache.set(cache_key, version, 5)  # Cache por 30 segundos
    
    return JsonResponse({
        'success': True,
        'version': version,
        'timestamp': timezone.now().isoformat(),
    })




@login_required
def api_configuracao_atualizacao(request):
    """
    GET: Retorna configuração do usuário
    POST: Salva configuração do usuário
    """
    config, created = ConfiguracaoAtualizacao.objects.get_or_create(
        usuario=request.user,
        defaults={
            'som_ativo': True,
            'volume': 50,
            'intervalo_atualizacao': 30,
        }
    )
    
    if request.method == 'GET':
        return JsonResponse({
            'success': True,
            'config': {
                'som_ativo': config.som_ativo,
                'volume': config.volume,
                'intervalo_atualizacao': config.intervalo_atualizacao,
            }
        })
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            if 'som_ativo' in data:
                config.som_ativo = bool(data['som_ativo'])
            if 'volume' in data:
                config.volume = max(0, min(100, int(data['volume'])))
            if 'intervalo_atualizacao' in data:
                config.intervalo_atualizacao = max(10, min(300, int(data['intervalo_atualizacao'])))
            
            config.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Configuração salva com sucesso!',
                'config': {
                    'som_ativo': config.som_ativo,
                    'volume': config.volume,
                    'intervalo_atualizacao': config.intervalo_atualizacao,
                }
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})


@login_required
def api_html_cards_atualizados(request):
    """
    Retorna o HTML atualizado dos cards para substituição no frontend.
    Usado pelo polling para atualizar sem recarregar a página.
    """
    solicitacoes = Solicitacao.objects.select_related(
        'criador', 'armazem', 'especie', 'coluna_kanban'
    ).order_by('-data_criacao')
    
    data = []
    for sol in solicitacoes:
        data.append({
            'id': sol.id,
            'titulo': sol.titulo,
            'criador_nome': sol.criador.get_full_name() or sol.criador.username,
            'data_criacao': sol.data_criacao.strftime('%d/%m/%Y %H:%M'),
            'status': sol.status,
            'unidade_controle': sol.unidade_controle,
            'quantidade_solicitada': float(sol.quantidade_solicitada),
            'quantidade_empenhada': float(sol.quantidade_empenhada),
            'percentual_empenhado': float(sol.percentual_empenhado),
            'coluna_kanban': sol.coluna_kanban.nome if sol.coluna_kanban else 'Início',
            'prioridade': sol.prioridade,
            'criterios': {
                'armazem': sol.armazem.nome if sol.armazem else '',
                'produto': sol.produto or '',
                'especie': sol.especie.nome if sol.especie else '',
                'cliente': sol.cliente or '',
            }
        })
    
    return JsonResponse({
        'success': True,
        'cards': data,
        'timestamp': timezone.now().isoformat(),
    })


# ============================================================================
# FASE 5 - KANBAN COMPLETO, REGRAS AUTOMÁTICAS E WORKFLOW
# ============================================================================







@login_required
@permission_required('sapp.pode_configuracoes', raise_exception=True)
def pagina_config_workflow(request):
    """Página de configuração do workflow"""
    return render(request, 'sapp/pagina_config_workflow.html')

from sapp.models import ColunaKanban, RegraWorkflow
def criar_regras_workflow_padrao():
    """
    Cria as regras de workflow padrão se não existirem.
    Chamar via manage.py ou na primeira migração.
    """
    # Garantir que colunas existem
    ColunaKanban.criar_colunas_padrao()
    
    coluna_inicio = ColunaKanban.objects.get(nome='Início')
    coluna_meio = ColunaKanban.objects.get(nome='Meio')
    coluna_fim = ColunaKanban.objects.get(nome='Fim')
    
    regras_padrao = [
        # Início
        {'coluna': coluna_inicio, 'evento': 'CRIACAO', 'status': 'AGUARDANDO_EMPENHO', 'auto': True},
        
        # Meio
        {'coluna': coluna_meio, 'evento': 'PRIMEIRO_EMPENHO', 'status': 'EMPENHO_PARCIAL', 'auto': True},
        {'coluna': coluna_meio, 'evento': 'EMPENHO_PARCIAL', 'status': 'EMPENHO_PARCIAL', 'auto': True},
        {'coluna': coluna_meio, 'evento': 'EMPENHO_COMPLETO', 'status': 'EMPENHO_COMPLETO', 'auto': True},
        {'coluna': coluna_meio, 'evento': 'PRIMEIRA_MOVIMENTACAO', 'status': 'MOVIMENTACAO_PARCIAL', 'auto': True},
        {'coluna': coluna_meio, 'evento': 'MOVIMENTACAO_PARCIAL', 'status': 'MOVIMENTACAO_PARCIAL', 'auto': True},
        
        # Fim
        {'coluna': coluna_fim, 'evento': 'TRANSFERENCIA_COMPLETA', 'status': 'CONCLUIDO', 'auto': True},
        {'coluna': coluna_fim, 'evento': 'EXPEDICAO_COMPLETA', 'status': 'CONCLUIDO', 'auto': True},
        {'coluna': coluna_fim, 'evento': 'CONCLUSAO', 'status': 'CONCLUIDO', 'auto': True},
        {'coluna': coluna_fim, 'evento': 'CANCELAMENTO', 'status': 'CANCELADO', 'auto': True},
    ]
    
    created_count = 0
    for regra_data in regras_padrao:
        regra, created = RegraWorkflow.objects.get_or_create(
            coluna=regra_data['coluna'],
            evento=regra_data['evento'],
            defaults={
                'status_resultante': regra_data['status'],
                'movimentacao_automatica': regra_data['auto'],
            }
        )
        if created:
            created_count += 1
    
    return created_count



@login_required
def api_remover_itens_solicitacao(
    request,
    solicitacao_id
):
    """
    Remove todos os itens pendentes do empenho vinculado
    exclusivamente à solicitação informada.
    """
    if request.method != 'POST':
        return JsonResponse(
            {
                'success': False,
                'error': 'Método não permitido.'
            },
            status=405
        )

    try:
        with transaction.atomic():
            solicitacao = (
                Solicitacao.objects
                .select_for_update()
                .get(id=solicitacao_id)
            )

            if solicitacao.status in [
                'CONCLUIDO',
                'CANCELADO',
            ]:
                raise ValueError(
                    'Este card já está concluído ou cancelado.'
                )

            empenho = (
                Empenho.objects
                .select_for_update()
                .filter(
                    solicitacao_id=solicitacao.id,
                    status__nome='Rascunho'
                )
                .first()
            )

            if not empenho:
                return JsonResponse({
                    'success': True,
                    'message': 'Nenhum item para remover.'
                })

            itens = list(
                empenho.itens.select_related(
                    'estoque'
                )
            )

            quantidade_removida = sum(
                item.quantidade
                for item in itens
            )

            quantidade_itens = len(itens)

            # Exclusão individual para executar
            # ItemEmpenho.delete() e liberar Estoque.empenhado.
            for item in itens:
                item.delete()

            solicitacao.quantidade_empenhada = Decimal('0')
            solicitacao.status = 'AGUARDANDO_EMPENHO'

            solicitacao.save(
                update_fields=[
                    'quantidade_empenhada',
                    'status',
                    'data_atualizacao',
                ]
            )

            avaliar_workflow(
                solicitacao,
                'CRIACAO',
                request.user
            )

            HistoricoCard.objects.create(
                solicitacao=solicitacao,
                usuario=request.user,
                acao='REMOCAO_ITEM',
                quantidade=quantidade_removida,
                unidade=(
                    'KG'
                    if solicitacao.unidade_controle
                    == 'QUILOGRAMA'
                    else 'BAG'
                ),
                observacao=(
                    f'{quantidade_itens} item(ns) removido(s) '
                    f'do Empenho #{empenho.id}'
                )
            )

            from django.core.cache import cache
            cache.delete('cards_version_hash')

            return JsonResponse({
                'success': True,
                'message': (
                    f'{quantidade_itens} item(ns) '
                    f'removido(s) com sucesso.'
                ),
                'quantidade_empenhada': 0,
                'status': solicitacao.status,
            })

    except Solicitacao.DoesNotExist:
        return JsonResponse(
            {
                'success': False,
                'error': 'Solicitação não encontrada.'
            },
            status=404
        )

    except ValueError as erro:
        return JsonResponse(
            {
                'success': False,
                'error': str(erro)
            },
            status=400
        )

    except Exception as erro:
        logger.exception(
            'Erro ao remover itens da solicitação %s',
            solicitacao_id
        )

        return JsonResponse(
            {
                'success': False,
                'error': str(erro)
            },
            status=500
        )









# ============================================================================
# IMPORTS NECESSÁRIOS (verifique se estão no topo do arquivo)
# ============================================================================
from collections import defaultdict
from decimal import Decimal, InvalidOperation
import json
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.db import transaction, models
from django.http import JsonResponse
from django.db.models import Q, Sum, Prefetch
from django.utils import timezone
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from .models import (
    Empenho, EmpenhoStatus, ItemEmpenho, Estoque, 
    HistoricoMovimentacao, HistoricoItemEmpenho, StatusSistemico,
    Cultivar, Peneira, Categoria, Tratamento, Especie,
    Armazem, Produto, Solicitacao, ColunaKanban, RegraWorkflow,
    HistoricoCard, ConfiguracaoAtualizacao,
)






def get_coluna_por_posicao(posicao):
    """
    Retorna a coluna do Kanban pela posição (0=primeira, 1=segunda, etc.)
    Se não existir coluna suficiente, retorna a primeira disponível.
    """
    colunas = list(ColunaKanban.objects.filter(ativa=True).order_by('ordem'))
    if not colunas:
        # Criar colunas padrão se não existirem
        ColunaKanban.criar_colunas_padrao()
        colunas = list(ColunaKanban.objects.filter(ativa=True).order_by('ordem'))
    
    if posicao < len(colunas):
        return colunas[posicao]
    return colunas[0] if colunas else None




# ============================================================================
# IMPORTS (VERIFIQUE SE ESTÃO TODOS NO TOPO DO ARQUIVO)
# ============================================================================
from collections import defaultdict
from decimal import Decimal, InvalidOperation
import json
import logging
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.db import transaction, models
from django.http import JsonResponse
from django.db.models import Q, Sum, Prefetch
from django.utils import timezone
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from .models import (
    Empenho, EmpenhoStatus, ItemEmpenho, Estoque, 
    HistoricoMovimentacao, HistoricoItemEmpenho, StatusSistemico,
    Cultivar, Peneira, Categoria, Tratamento, Especie,
    Armazem, Produto, Solicitacao, ColunaKanban, RegraWorkflow,
    HistoricoCard, ConfiguracaoAtualizacao,
)

logger = logging.getLogger(__name__)

def obter_empenho_da_solicitacao(
    solicitacao,
    status_nome='Rascunho'
):
    """
    Retorna exclusivamente o empenho vinculado à solicitação
    recebida.

    Não utiliza título ou observação para localizar o empenho.
    """

    if not solicitacao or not solicitacao.id:
        return None

    filtros = {
        'solicitacao_id': solicitacao.id,
    }

    if status_nome:
        filtros['status__nome'] = status_nome

    return (
        Empenho.objects
        .filter(**filtros)
        .order_by('-data_criacao', '-id')
        .first()
    )
# ============================================================================
# FUNÇÃO CENTRAL DO WORKFLOW
# ============================================================================
def avaliar_workflow(solicitacao, evento, usuario=None):
    """
    Avalia as regras de workflow configuradas e move o card se necessário.
    Usa EXCLUSIVAMENTE as regras da tabela RegraWorkflow.
    """
    try:
        regra = RegraWorkflow.objects.filter(
            evento=evento,
            coluna__ativa=True
        ).select_related('coluna').first()
        
        if not regra:
            return False
        
        if not regra.movimentacao_automatica:
            return False
        
        coluna_anterior = solicitacao.coluna_kanban.nome if solicitacao.coluna_kanban else 'Nenhuma'
        
        if solicitacao.coluna_kanban and solicitacao.coluna_kanban.id == regra.coluna.id:
            if solicitacao.status != regra.status_resultante:
                solicitacao.status = regra.status_resultante
                solicitacao.save(update_fields=['status', 'data_atualizacao'])
            return False
        
        solicitacao.coluna_kanban = regra.coluna
        solicitacao.status = regra.status_resultante
        solicitacao.save(update_fields=['coluna_kanban', 'status', 'data_atualizacao'])
        
        HistoricoCard.objects.create(
            solicitacao=solicitacao,
            usuario=usuario or solicitacao.criador,
            acao='MOVIMENTACAO_KANBAN',
            coluna_anterior=coluna_anterior,
            coluna_nova=regra.coluna.nome,
            observacao=f'Automático: {regra.get_evento_display()} → {regra.coluna.nome}'
        )
        
        from django.core.cache import cache
        cache.delete('cards_version_hash')
        
        return True
        
    except Exception as e:
        logger.error(f"Erro workflow solicitação {solicitacao.id}, evento {evento}: {e}")
        return False



# ============================================================
# SUBSTITUA APENAS ESTAS DUAS VIEWS NO SEU views.py
# ============================================================





@login_required
@permission_required(
    'sapp.pode_criar_solicitacao',
    raise_exception=True
)
def criar_solicitacao(request):
    """
    Cria uma nova solicitação.
    Salva corretamente Destino, Armazém e Observação.
    """

    if request.method == 'POST':

        titulo = request.POST.get(
            'titulo',
            ''
        ).strip().upper()

        armazem_id = request.POST.get(
            'armazem'
        )

        produto = request.POST.get(
            'produto',
            ''
        ).strip()

        especie_id = request.POST.get(
            'especie'
        )

        cliente = request.POST.get(
            'cliente',
            ''
        ).strip()

        destino = request.POST.get(
            'destino',
            ''
        ).strip()

        unidade_controle = request.POST.get(
            'unidade_controle',
            'EMBALAGEM'
        )

        quantidade_texto = request.POST.get(
            'quantidade_solicitada',
            '0'
        )

        observacao = request.POST.get(
            'observacao',
            ''
        ).strip()

        prioridade = request.POST.get(
            'prioridade',
            'MEDIA'
        )

        if not titulo:
            messages.error(
                request,
                'Título é obrigatório.'
            )

            return redirect(
                'sapp:pagina_solicitacoes'
            )

        try:
            quantidade = Decimal(
                str(
                    quantidade_texto
                ).replace(',', '.')
            )

            if quantidade <= 0:
                raise ValueError

        except (
            ValueError,
            InvalidOperation
        ):
            messages.error(
                request,
                'Quantidade inválida.'
            )

            return redirect(
                'sapp:pagina_solicitacoes'
            )

        try:
            with transaction.atomic():

                armazem = (
                    Armazem.objects
                    .filter(
                        id=armazem_id
                    )
                    .first()
                    if armazem_id
                    else None
                )

                especie = (
                    Especie.objects
                    .filter(
                        id=especie_id
                    )
                    .first()
                    if especie_id
                    else None
                )

                solicitacao = (
                    Solicitacao.objects.create(
                        titulo=titulo,

                        criador=request.user,

                        armazem=armazem,

                        produto=(
                            produto
                            or None
                        ),

                        especie=especie,

                        cliente=(
                            cliente
                            or None
                        ),

                        destino=destino,

                        unidade_controle=(
                            unidade_controle
                        ),

                        quantidade_solicitada=(
                            quantidade
                        ),

                        observacao=(
                            observacao
                            or None
                        ),

                        prioridade=prioridade,

                        status=(
                            'AGUARDANDO_EMPENHO'
                        ),
                    )
                )

                avaliar_workflow(
                    solicitacao,
                    'CRIACAO',
                    request.user
                )

                HistoricoCard.objects.create(
                    solicitacao=solicitacao,

                    usuario=request.user,

                    acao='CRIACAO',

                    quantidade=quantidade,

                    unidade=(
                        'KG'
                        if unidade_controle
                        == 'QUILOGRAMA'
                        else 'BAG'
                    ),

                    observacao=(
                        f'Solicitação criada: '
                        f'{titulo}'
                        + (
                            f' | Destino: '
                            f'{destino}'
                            if destino
                            else ''
                        )
                    )
                )

                from django.core.cache import cache

                cache.delete(
                    'cards_version_hash'
                )

            messages.success(
                request,
                (
                    f'Solicitação '
                    f'"{titulo}" '
                    f'criada com sucesso!'
                )
            )

            return redirect(
                'sapp:pagina_solicitacoes'
            )

        except Exception as erro:

            logger.exception(
                'Erro ao criar solicitação'
            )

            messages.error(
                request,
                (
                    'Erro ao criar solicitação: '
                    f'{erro}'
                )
            )

            return redirect(
                'sapp:pagina_solicitacoes'
            )

    return render(
        request,
        'sapp/criar_solicitacao.html',
        {
            'armazens': (
                Armazem.objects
                .all()
                .order_by('nome')
            ),

            'especies': (
                Especie.objects
                .all()
                .order_by('nome')
            ),
        }
    )


def _data_local_formatada(valor):
    """
    Converte a data para o fuso horário
    configurado no Django.
    """

    if not valor:
        return ''

    return timezone.localtime(
        valor
    ).strftime(
        '%d/%m/%Y %H:%M'
    )


@login_required
def api_dados_impressao_solicitacao(
    request,
    solicitacao_id
):
    """
    API da impressão.

    A coluna Armazém de cada item usa o AZ real do lote:
    - pendente: item.estoque.az
    - processado: historico.estoque_origem.az
      com fallback para estoque_destino.az
    """

    if request.method != 'GET':
        return JsonResponse(
            {
                'success': False,
                'error': 'Método não permitido.'
            },
            status=405
        )

    solicitacao = get_object_or_404(
        Solicitacao.objects
        .select_related(
            'criador',
            'responsavel',
            'armazem',
            'especie'
        ),
        id=solicitacao_id
    )

    empenhos = (
        Empenho.objects
        .filter(
            solicitacao_id=solicitacao.id
        )
        .prefetch_related(
            'itens__estoque',
            'itens__estoque__cultivar',
            'itens__estoque__peneira',
            'itens__estoque__categoria',
            'itens__estoque__especie',
            'itens__estoque__tratamento',
            Prefetch(
                'historico_itens',
                queryset=(
                    HistoricoItemEmpenho.objects
                    .select_related(
                        'estoque_origem',
                        'estoque_destino'
                    )
                    .order_by(
                        'processado_em',
                        'id'
                    )
                )
            )
        )
        .order_by(
            'data_criacao',
            'id'
        )
    )

    itens_pendentes = []
    itens_processados = []

    for empenho in empenhos:

        for item in empenho.itens.all():

            estoque = item.estoque

            itens_pendentes.append({
                'item_id': item.id,
                'situacao': 'PENDENTE',
                'lote': item.lote or '',
                'quantidade': float(
                    item.quantidade or 0
                ),

                # AZ REAL DO LOTE
                'armazem': (
                    estoque.az
                    if (
                        estoque
                        and estoque.az
                    )
                    else ''
                ),

                'endereco': (
                    item.endereco_origem
                    or (
                        estoque.endereco
                        if estoque
                        else ''
                    )
                ),

                'produto': (
                    estoque.produto
                    if estoque
                    else ''
                ),

                'cultivar': (
                    item.cultivar
                    or (
                        estoque.cultivar.nome
                        if (
                            estoque
                            and estoque.cultivar
                        )
                        else ''
                    )
                ),

                'peneira': (
                    item.peneira
                    or (
                        estoque.peneira.nome
                        if (
                            estoque
                            and estoque.peneira
                        )
                        else ''
                    )
                ),

                'categoria': (
                    item.categoria
                    or (
                        estoque.categoria.nome
                        if (
                            estoque
                            and estoque.categoria
                        )
                        else ''
                    )
                ),

                'especie': (
                    estoque.especie.nome
                    if (
                        estoque
                        and estoque.especie
                    )
                    else ''
                ),

                'tratamento': (
                    estoque.tratamento.nome
                    if (
                        estoque
                        and estoque.tratamento
                    )
                    else ''
                ),

                'embalagem': (
                    estoque.embalagem
                    if estoque
                    else ''
                ),

                'empresa': (
                    estoque.empresa
                    if estoque
                    else ''
                ),

                'cliente': (
                    estoque.cliente
                    if estoque
                    else ''
                ),
            })

        for historico in empenho.historico_itens.all():

            tipo = str(
                historico.tipo or ''
            ).lower()

            if tipo == 'transferencia':
                situacao = 'TRANSFERIDO'

            elif tipo == 'expedicao':
                situacao = 'EXPEDIDO'

            else:
                situacao = (
                    historico.get_tipo_display()
                    if hasattr(
                        historico,
                        'get_tipo_display'
                    )
                    else str(
                        historico.tipo or ''
                    )
                ).upper()

            estoque_origem = (
                historico.estoque_origem
                if hasattr(
                    historico,
                    'estoque_origem'
                )
                else None
            )

            estoque_destino = (
                historico.estoque_destino
                if hasattr(
                    historico,
                    'estoque_destino'
                )
                else None
            )

            armazem_lote = (
                estoque_origem.az
                if (
                    estoque_origem
                    and estoque_origem.az
                )
                else (
                    estoque_destino.az
                    if (
                        estoque_destino
                        and estoque_destino.az
                    )
                    else ''
                )
            )

            itens_processados.append({
                'item_id': historico.id,
                'situacao': situacao,
                'lote': historico.lote or '',
                'quantidade': float(
                    historico.quantidade or 0
                ),

                # AZ REAL DO LOTE ORIGINAL
                'armazem': (
                    armazem_lote
                    or ''
                ),

                'endereco': (
                    historico.endereco_origem
                    or ''
                ),

                'endereco_destino': (
                    historico.endereco_destino
                    or ''
                ),

                'produto': (
                    historico.produto
                    or ''
                ),

                'cultivar': (
                    historico.cultivar
                    or ''
                ),

                'peneira': (
                    historico.peneira
                    or ''
                ),

                'categoria': (
                    historico.categoria
                    or ''
                ),

                'especie': (
                    historico.especie
                    or ''
                ),

                'tratamento': (
                    historico.tratamento
                    or ''
                ),

                'embalagem': (
                    historico.embalagem
                    or ''
                ),

                'empresa': (
                    historico.empresa
                    or ''
                ),

                'cliente': (
                    historico.cliente
                    or ''
                ),

                'processado_em': (
                    _data_local_formatada(
                        historico.processado_em
                    )
                ),
            })

    status_display = (
        solicitacao.get_status_display()
        if hasattr(
            solicitacao,
            'get_status_display'
        )
        else solicitacao.status
    )

    emitido_em = (
        _data_local_formatada(
            timezone.now()
        )
    )

    if (
        solicitacao.unidade_controle
        == 'QUILOGRAMA'
    ):
        quantidade_empenhada_display = (
            solicitacao.quantidade_empenhada_kg
            or Decimal('0')
        )
    else:
        quantidade_empenhada_display = (
            solicitacao.quantidade_empenhada
            or Decimal('0')
        )

    return JsonResponse({
        'success': True,
        'emitido_em': emitido_em,

        'solicitacao': {
            'id': solicitacao.id,
            'titulo': solicitacao.titulo,

            'criador': (
                solicitacao.criador.get_full_name()
                or solicitacao.criador.username
            ),

            'responsavel': (
                (
                    solicitacao.responsavel.get_full_name()
                    or solicitacao.responsavel.username
                )
                if solicitacao.responsavel
                else ''
            ),

            'data_criacao': (
                _data_local_formatada(
                    solicitacao.data_criacao
                )
            ),

            'data_atualizacao': (
                _data_local_formatada(
                    solicitacao.data_atualizacao
                )
            ),

            'data_finalizacao': (
                _data_local_formatada(
                    solicitacao.data_finalizacao
                )
                if (
                    solicitacao.status
                    == 'CONCLUIDO'
                    and solicitacao.data_finalizacao
                )
                else ''
            ),

            'destino': (
                solicitacao.destino
                or ''
            ),

            'observacao': (
                solicitacao.observacao
                or ''
            ),

            'prioridade': (
                solicitacao.get_prioridade_display()
                if hasattr(
                    solicitacao,
                    'get_prioridade_display'
                )
                else solicitacao.prioridade
            ),

            'quantidade_solicitada': float(
                solicitacao.quantidade_solicitada
                or 0
            ),

            'quantidade_empenhada': float(
                solicitacao.quantidade_empenhada
                or 0
            ),

            'quantidade_empenhada_display': float(
                quantidade_empenhada_display
            ),

            'quantidade_movimentada': float(
                solicitacao.quantidade_movimentada
                or 0
            ),

            'unidade_controle': (
                solicitacao.unidade_controle
            ),

            'status': (
                solicitacao.status
            ),

            'status_display': (
                status_display
            ),

            'criterios': {
                'armazem': (
                    solicitacao.armazem.nome
                    if solicitacao.armazem
                    else ''
                ),

                'produto': (
                    solicitacao.produto
                    or ''
                ),

                'especie': (
                    solicitacao.especie.nome
                    if solicitacao.especie
                    else ''
                ),

                'cliente': (
                    solicitacao.cliente
                    or ''
                ),

                'destino': (
                    solicitacao.destino
                    or ''
                ),
            },
        },

        'itens_pendentes': (
            itens_pendentes
        ),

        'itens_processados': (
            itens_processados
        ),
    })

# ============================================================================
# PÁGINA DE SOLICITAÇÕES
# ============================================================================
@login_required
@permission_required('sapp.pode_ver_empenhos', raise_exception=True)
def pagina_solicitacoes(request):
    """Página principal de solicitações"""
    return render(request, 'sapp/pagina_solicitacoes.html')


# ============================================================================
# API LISTAR SOLICITAÇÕES
# ============================================================================
@login_required
def api_listar_solicitacoes(request):

    solicitacoes = (
        Solicitacao.objects
        .select_related(
            'criador',
            'armazem',
            'especie',
            'coluna_kanban',
        )
        .prefetch_related(
            'empenhos__itens',
            'empenhos__historico_itens',
        )
        .order_by('-data_criacao')
    )

    data = []

    for sol in solicitacoes:

        # =====================================================
        # QUANTIDADE EMPENHADA
        # =====================================================

        if sol.unidade_controle == 'QUILOGRAMA':
            qtd_emp = (
                sol.quantidade_empenhada_kg
                or Decimal('0')
            )
        else:
            qtd_emp = Decimal(
                str(sol.quantidade_empenhada)
            )

        if sol.quantidade_solicitada > 0:
            percentual = (
                qtd_emp
                / sol.quantidade_solicitada
            ) * 100
        else:
            percentual = 0


        # =====================================================
        # LOTES RELACIONADOS AO CARD
        # =====================================================

        lotes = set()
        embalagens = set()
        # Itens que ainda estão empenhados
        for empenho in sol.empenhos.all():

            for item in empenho.itens.all():

                lote = str(
                    item.lote or ''
                ).strip()

                if lote:
                    lotes.add(lote)

                embalagem = str(
                    getattr(
                        item,
                        'embalagem_snapshot',
                        ''
                    )
                    or ''
                ).strip().upper()

                if embalagem:
                    embalagens.add(embalagem)


            # Itens que já foram transferidos ou expedidos
            for historico in empenho.historico_itens.all():

                lote = str(
                    historico.lote or ''
                ).strip()

                if lote:
                    lotes.add(lote)

                embalagem = str(
                    historico.embalagem
                    or ''
                ).strip().upper()

                if embalagem:
                    embalagens.add(embalagem)


        # Histórico do próprio card.
        # Também ajuda com cards antigos.
        lotes_historico_card = (
            HistoricoCard.objects
            .filter(
                solicitacao_id=sol.id
            )
            .exclude(
                lote__isnull=True
            )
            .exclude(
                lote=''
            )
            .values_list(
                'lote',
                flat=True
            )
        )

        for lote in lotes_historico_card:

            lote = str(
                lote or ''
            ).strip()

            if lote:
                lotes.add(lote)


        # =====================================================
        # JSON
        # =====================================================

        data.append({

            'id': sol.id,

            'titulo': sol.titulo,

            'criador_nome': (
                sol.criador.get_full_name()
                or sol.criador.username
            ),

            'data_criacao': (
                sol.data_criacao.strftime(
                    '%d/%m/%Y %H:%M'
                )
            ),

            'status': sol.status,

            'unidade_controle': (
                sol.unidade_controle
            ),

            'quantidade_solicitada': float(
                sol.quantidade_solicitada
            ),

            # Valor bruto em unidades
            'quantidade_empenhada': float(
                sol.quantidade_empenhada
            ),

            # Valor convertido conforme unidade
            'quantidade_empenhada_display': float(
                qtd_emp
            ),

            'quantidade_movimentada': float(
                sol.quantidade_movimentada
            ),

            'percentual_empenhado': float(
                percentual
            ),

            'percentual_movimentado': float(
                sol.percentual_movimentado
            ),

            'coluna_kanban': (
                sol.coluna_kanban.nome
                if sol.coluna_kanban
                else 'Sem coluna'
            ),

            'coluna_kanban_id': (
                sol.coluna_kanban.id
                if sol.coluna_kanban
                else None
            ),

            'prioridade': sol.prioridade,


            # =============================================
            # NOVO
            # =============================================

            'lotes': sorted(lotes),
            'embalagens': sorted(embalagens),

            'criterios': {

                'armazem': (
                    sol.armazem.nome
                    if sol.armazem
                    else ''
                ),

                'produto': (
                    sol.produto
                    or ''
                ),

                'especie': (
                    sol.especie.nome
                    if sol.especie
                    else ''
                ),

                'cliente': (
                    sol.cliente
                    or ''
                ),
            }

        })

    return JsonResponse({
        'success': True,
        'solicitacoes': data,
    })

    


@login_required
@permission_required(
    'sapp.pode_empenhar_solicitacao',
    raise_exception=True
)
def api_lotes_disponiveis_para_solicitacao(
    request,
    solicitacao_id
):
    """
    Lista lotes compatíveis com a solicitação.

    Os itens salvos no empenho do card atual aparecem primeiro.

    Um item empenhado por outro card:
    - não aparece como empenho deste card;
    - não fica selecionado;
    - não sobe para o topo;
    - continua descontado do saldo disponível.
    """

    solicitacao = get_object_or_404(
        Solicitacao.objects.select_related(
            'armazem',
            'especie',
        ),
        id=solicitacao_id
    )

    # Busca exclusivamente o empenho salvo deste card.
    empenho = obter_empenho_da_solicitacao(
        solicitacao=solicitacao,
        status_nome='Rascunho'
    )

    ids_empenhados_no_card = []

    if empenho:
        ids_empenhados_no_card = list(
            empenho.itens.values_list(
                'estoque_id',
                flat=True
            )
        )

    # Mostra:
    # 1. lotes que ainda têm disponibilidade;
    # 2. lotes do empenho do card atual, mesmo se a
    #    disponibilidade geral estiver zerada.
    qs = (
        Estoque.objects
        .filter(
            Q(saldo__gt=F('empenhado'))
            |
            Q(id__in=ids_empenhados_no_card)
        )
        .select_related(
            'cultivar',
            'peneira',
            'categoria',
            'especie',
            'tratamento',
            'status_sistemico',
            'conferente',
        )
    )

    # ------------------------------------------------------------------
    # FILTROS DA SOLICITAÇÃO
    # ------------------------------------------------------------------
    if solicitacao.armazem:
        qs = qs.filter(
            az=solicitacao.armazem.nome
        )

    if solicitacao.produto:
        qs = qs.filter(
            produto__iexact=solicitacao.produto
        )

    if solicitacao.especie:
        qs = qs.filter(
            especie=solicitacao.especie
        )

    if solicitacao.cliente:
        qs = qs.filter(
            cliente__iexact=solicitacao.cliente
        )

    # ------------------------------------------------------------------
    # BUSCA
    # ------------------------------------------------------------------
    busca = request.GET.get(
        'busca',
        ''
    ).strip()

    if busca:
        qs = qs.filter(
            Q(lote__icontains=busca)
            |
            Q(produto__icontains=busca)
            |
            Q(endereco__icontains=busca)
            |
            Q(cliente__icontains=busca)
            |
            Q(cultivar__nome__icontains=busca)
            |
            Q(az__icontains=busca)
        )

    # ------------------------------------------------------------------
    # FILTROS POR COLUNA
    # ------------------------------------------------------------------
    filter_map = {
        'az': 'az__in',
        'lote': 'lote__in',
        'produto': 'produto__in',
        'cultivar': 'cultivar__nome__in',
        'peneira': 'peneira__nome__in',
        'categoria': 'categoria__nome__in',
        'endereco': 'endereco__in',
        'especie': 'especie__nome__in',
        'tratamento': 'tratamento__nome__in',
        'embalagem': 'embalagem__in',
        'cliente': 'cliente__in',
        'empresa': 'empresa__in',
        'conferente': 'conferente__username__in',
    }

    for param, lookup in filter_map.items():
        valores = [
            valor.strip()
            for valor in request.GET.getlist(param)
            if valor and valor.strip()
        ]

        if valores:
            qs = qs.filter(**{
                lookup: valores
            })

    # ------------------------------------------------------------------
    # MARCAR SOMENTE OS ITENS DO EMPENHO DO CARD ATUAL
    # ------------------------------------------------------------------
    if ids_empenhados_no_card:
        qs = qs.annotate(
            empenho_do_card_atual=Case(
                When(
                    id__in=ids_empenhados_no_card,
                    then=Value(1)
                ),
                default=Value(0),
                output_field=IntegerField(),
            )
        )
    else:
        qs = qs.annotate(
            empenho_do_card_atual=Value(
                0,
                output_field=IntegerField()
            )
        )

    # Os itens salvos no empenho deste card ficam no topo.
    qs = qs.order_by(
        '-empenho_do_card_atual',
        'lote',
        'endereco',
        'id',
    )

    # ------------------------------------------------------------------
    # PAGINAÇÃO
    # ------------------------------------------------------------------
    try:
        page = max(
            1,
            int(
                request.GET.get(
                    'page',
                    1
                )
            )
        )

        page_size = int(
            request.GET.get(
                'page_size',
                100
            )
        )

        page_size = max(
            1,
            min(
                page_size,
                1000
            )
        )

    except (TypeError, ValueError):
        page = 1
        page_size = 100

    total = qs.count()

    start = (
        page - 1
    ) * page_size

    end = start + page_size

    lotes_qs = qs[start:end]

    # Mapa dos itens salvos no empenho atual.
    itens_empenhados_por_estoque = {}

    if empenho:
        itens_empenhados_por_estoque = {
            item.estoque_id: item
            for item in empenho.itens.all()
        }

    lotes = []

    for lote in lotes_qs:
        item_empenhado = (
            itens_empenhados_por_estoque.get(
                lote.id
            )
        )

        pertence_ao_empenho_atual = (
            item_empenhado is not None
        )

        # Quantidade que pertence ao empenho deste card.
        quantidade_empenhada_card = (
            item_empenhado.quantidade
            if item_empenhado
            else 0
        )

        # O disponível normal já desconta empenhos de todos os cards.
        disponivel_geral = Decimal(
            str(
                lote.disponivel
                or 0
            )
        )

        # Para um item que já pertence ao card atual,
        # devolvemos sua própria reserva ao limite editável.
        disponivel_para_card = (
            disponivel_geral
            + Decimal(
                str(
                    quantidade_empenhada_card
                    or 0
                )
            )
        )

        lotes.append({
            'id': lote.id,

            'lote': lote.lote,
            'produto': lote.produto or '',

            'cultivar': (
                lote.cultivar.nome
                if lote.cultivar
                else ''
            ),

            'peneira': (
                lote.peneira.nome
                if lote.peneira
                else ''
            ),

            'categoria': (
                lote.categoria.nome
                if lote.categoria
                else ''
            ),

            'especie': (
                lote.especie.nome
                if lote.especie
                else ''
            ),

            'tratamento': (
                lote.tratamento.nome
                if lote.tratamento
                else ''
            ),

            'endereco': lote.endereco or '',

            'saldo': float(
                lote.saldo or 0
            ),

            # Quantidade reservada no estoque por todos os empenhos.
            'empenhado': float(
                lote.empenhado or 0
            ),

            'disponivel': float(
                disponivel_geral
            ),

            # Limite que o card atual pode utilizar.
            'disponivel_para_card': float(
                disponivel_para_card
            ),

            'peso_unitario': float(
                lote.peso_unitario or 0
            ),

            'peso_total': float(
                lote.peso_total or 0
            ),

            'embalagem': lote.embalagem or '',
            'cliente': lote.cliente or '',
            'empresa': lote.empresa or '',
            'az': lote.az or '',

            'conferente': (
                (
                    lote.conferente.get_full_name()
                    or lote.conferente.username
                )
                if lote.conferente
                else ''
            ),

            'observacao': (
                lote.observacao or ''
            ),

            'status_sistemico': (
                {
                    'nome': (
                        lote.status_sistemico.nome
                    ),
                    'cor': (
                        lote.status_sistemico.cor
                    ),
                    'icone': (
                        lote.status_sistemico.icone
                    ),
                }
                if lote.status_sistemico
                else None
            ),

            # Estes campos dizem respeito somente
            # ao empenho salvo do card aberto.
            'empenho_do_card_atual': (
                pertence_ao_empenho_atual
            ),

            'item_empenho_id': (
                item_empenhado.id
                if item_empenhado
                else None
            ),

            'quantidade_empenhada_card': float(
                quantidade_empenhada_card or 0
            ),
        })

    # ------------------------------------------------------------------
    # ITENS SALVOS EXCLUSIVAMENTE NO EMPENHO DO CARD ATUAL
    # ------------------------------------------------------------------
    itens_empenhados = []

    if empenho:
        itens_qs = (
            empenho.itens
            .select_related(
                'estoque'
            )
            .order_by(
                'lote',
                'endereco_origem',
                'id',
            )
        )

        for item in itens_qs:
            estoque = item.estoque

            itens_empenhados.append({
                'id': item.id,
                'empenho_id': empenho.id,
                'solicitacao_id': solicitacao.id,

                'estoque_id': item.estoque_id,
                'lote': item.lote,

                'quantidade': float(
                    item.quantidade or 0
                ),

                'endereco': (
                    item.endereco_origem
                    or (
                        estoque.endereco
                        if estoque
                        else ''
                    )
                ),

                'peso_unitario': float(
                    estoque.peso_unitario or 0
                ) if estoque else 0,

                'empenho_salvo': True,
                'empenho_do_card_atual': True,
            })

    quantidade_movimentada = Decimal(
        str(
            solicitacao.quantidade_movimentada
            or 0
        )
    )

    empenho_bloqueado = (
        solicitacao.status in {
            'MOVIMENTACAO_PARCIAL',
            'CONCLUIDO',
            'CANCELADO',
        }
        or quantidade_movimentada > 0
    )

    return JsonResponse({
        'success': True,

        'lotes': lotes,

        'total': total,
        'page': page,
        'page_size': page_size,
        'has_more': end < total,

        'solicitacao': {
            'id': solicitacao.id,
            'titulo': solicitacao.titulo,

            'quantidade_solicitada': float(
                solicitacao.quantidade_solicitada
                or 0
            ),

            'quantidade_empenhada': float(
                solicitacao.quantidade_empenhada
                or 0
            ),

            'quantidade_movimentada': float(
                solicitacao.quantidade_movimentada
                or 0
            ),

            'unidade_controle': (
                solicitacao.unidade_controle
            ),

            'status': solicitacao.status,
            'destino': solicitacao.destino or '',
            'observacao': solicitacao.observacao or '',

            'empenho_bloqueado': empenho_bloqueado,
        },

        'empenho_id': (
            empenho.id
            if empenho
            else None
        ),

        # Somente itens do empenho salvo deste card.
        'itens_empenhados': itens_empenhados,
    })
# ============================================================================
# EMPENHAR NA SOLICITAÇÃO
# ============================================================================

@login_required
@permission_required(
    'sapp.pode_empenhar_solicitacao',
    raise_exception=True
)
def empenhar_na_solicitacao(
    request,
    solicitacao_id
):
    """
    Cria ou atualiza o empenho vinculado exclusivamente
    à solicitação informada.

    Depois que existir movimentação, o empenho fica congelado.
    """

    if request.method != 'POST':
        return JsonResponse(
            {
                'success': False,
                'error': 'Método não permitido.'
            },
            status=405
        )

    try:
        data = json.loads(
            request.body or '{}'
        )

    except json.JSONDecodeError:
        return JsonResponse(
            {
                'success': False,
                'error': 'JSON inválido.'
            },
            status=400
        )

    itens_selecionados = data.get(
        'itens',
        []
    )

    if not isinstance(
        itens_selecionados,
        list
    ):
        return JsonResponse(
            {
                'success': False,
                'error': 'Lista de itens inválida.'
            },
            status=400
        )

    if not itens_selecionados:
        return JsonResponse(
            {
                'success': False,
                'error': 'Nenhum item selecionado.'
            },
            status=400
        )

    try:
        with transaction.atomic():

            # Bloqueia somente a solicitação.
            solicitacao = (
                Solicitacao.objects
                .select_for_update(
                    of=('self',)
                )
                .get(
                    id=solicitacao_id
                )
            )

            # Depois que começou a movimentar,
            # não pode criar nem atualizar empenho.
            validar_edicao_empenho_solicitacao(
                solicitacao
            )

            status_rascunho, _ = (
                EmpenhoStatus.objects
                .get_or_create(
                    nome='Rascunho',
                    defaults={
                        'descricao': (
                            'Card em elaboração'
                        )
                    }
                )
            )

            # Busca somente o empenho deste card.
            empenho = (
                Empenho.objects
                .select_for_update(
                    of=('self',)
                )
                .filter(
                    solicitacao_id=solicitacao.id,
                    status_id=status_rascunho.id,
                )
                .order_by(
                    '-data_criacao',
                    '-id',
                )
                .first()
            )

            if not empenho:
                empenho = Empenho.objects.create(
                    solicitacao=solicitacao,
                    usuario=request.user,
                    status=status_rascunho,
                    observacao=solicitacao.titulo,
                )

            elif (
                empenho.observacao
                != solicitacao.titulo
            ):
                empenho.observacao = (
                    solicitacao.titulo
                )

                empenho.save(
                    update_fields=[
                        'observacao',
                        'data_atualizacao',
                    ]
                )

            # ----------------------------------------------------------
            # TOTAL ATUAL DO EMPENHO
            # ----------------------------------------------------------
            total_kg_empenhado = Decimal('0')

            if (
                solicitacao.unidade_controle
                == 'QUILOGRAMA'
            ):
                itens_atuais = (
                    empenho.itens
                    .select_related(
                        'estoque'
                    )
                )

                for item_existente in itens_atuais:
                    peso = Decimal(
                        str(
                            item_existente
                            .estoque
                            .peso_unitario
                            or 0
                        )
                    )

                    quantidade_atual = Decimal(
                        str(
                            item_existente.quantidade
                            or 0
                        )
                    )

                    total_kg_empenhado += (
                        quantidade_atual
                        * peso
                    )

            itens_validos = 0

            # ----------------------------------------------------------
            # PROCESSAR ITENS RECEBIDOS
            # ----------------------------------------------------------
            for item_data in itens_selecionados:

                try:
                    lote_id = int(
                        item_data.get(
                            'lote_id'
                        )
                    )

                    quantidade_adicionar = Decimal(
                        str(
                            item_data.get(
                                'quantidade',
                                0
                            )
                        )
                    )

                except (
                    TypeError,
                    ValueError,
                    ArithmeticError,
                ):
                    raise ValueError(
                        'Lote ou quantidade inválidos.'
                    )

                if quantidade_adicionar <= 0:
                    continue

                # Bloqueia somente o estoque.
                lote = (
                    Estoque.objects
                    .select_for_update(
                        of=('self',)
                    )
                    .get(
                        id=lote_id
                    )
                )

                # Busca o item exclusivamente no empenho
                # vinculado ao card atual.
                item_existente = (
                    ItemEmpenho.objects
                    .select_for_update(
                        of=('self',)
                    )
                    .filter(
                        empenho_id=empenho.id,
                        estoque_id=lote.id,
                    )
                    .first()
                )

                quantidade_anterior = Decimal(
                    str(
                        item_existente.quantidade
                        if item_existente
                        else 0
                    )
                )

                saldo = Decimal(
                    str(
                        lote.saldo
                        or 0
                    )
                )

                empenhado_geral = Decimal(
                    str(
                        lote.empenhado
                        or 0
                    )
                )

                # O empenhado geral contém reservas de todos os cards.
                # Recolocamos somente a reserva deste card para permitir
                # editar seu próprio empenho.
                disponivel_para_card = (
                    saldo
                    - empenhado_geral
                    + quantidade_anterior
                )

                quantidade_final = (
                    quantidade_anterior
                    + quantidade_adicionar
                )

                if (
                    quantidade_final
                    > disponivel_para_card
                ):
                    raise ValueError(
                        f'Quantidade indisponível para o lote '
                        f'{lote.lote}. Disponível para este '
                        f'card: {disponivel_para_card}.'
                    )

                # ------------------------------------------------------
                # VALIDAÇÃO EM KG
                # ------------------------------------------------------
                if (
                    solicitacao.unidade_controle
                    == 'QUILOGRAMA'
                ):
                    peso = Decimal(
                        str(
                            lote.peso_unitario
                            or 0
                        )
                    )

                    if peso <= 0:
                        raise ValueError(
                            f'O lote {lote.lote} não possui '
                            f'peso unitário válido.'
                        )

                    kg_adicionados = (
                        quantidade_adicionar
                        * peso
                    )

                    total_kg_empenhado += (
                        kg_adicionados
                    )

                    quantidade_solicitada = Decimal(
                        str(
                            solicitacao
                            .quantidade_solicitada
                            or 0
                        )
                    )

                    if (
                        total_kg_empenhado
                        > quantidade_solicitada
                    ):
                        raise ValueError(
                            f'O total de '
                            f'{total_kg_empenhado:.2f} KG '
                            f'excede os '
                            f'{quantidade_solicitada:.2f} KG '
                            f'solicitados.'
                        )

                # ------------------------------------------------------
                # SALVAR ITEM NO EMPENHO DO CARD ATUAL
                # ------------------------------------------------------
                if item_existente:
                    item_existente.quantidade = (
                        quantidade_final
                    )

                    item_existente.save(
                        update_fields=[
                            'quantidade',
                        ]
                    )

                else:
                    ItemEmpenho.objects.create(
                        empenho=empenho,
                        estoque=lote,
                        quantidade=quantidade_adicionar,

                        lote=lote.lote,

                        cultivar=(
                            lote.cultivar.nome
                            if lote.cultivar
                            else ''
                        ),

                        peneira=(
                            lote.peneira.nome
                            if lote.peneira
                            else ''
                        ),

                        categoria=(
                            lote.categoria.nome
                            if lote.categoria
                            else ''
                        ),

                        endereco_origem=(
                            lote.endereco
                        ),

                        saldo_anterior=(
                            lote.saldo
                        ),
                    )

                itens_validos += 1

                # ------------------------------------------------------
                # HISTÓRICO
                # ------------------------------------------------------
                if (
                    solicitacao.unidade_controle
                    == 'QUILOGRAMA'
                ):
                    quantidade_historico = (
                        quantidade_adicionar
                        * Decimal(
                            str(
                                lote.peso_unitario
                                or 0
                            )
                        )
                    )

                    unidade_historico = 'KG'

                else:
                    quantidade_historico = (
                        quantidade_adicionar
                    )

                    unidade_historico = 'BAG'

                HistoricoCard.objects.create(
                    solicitacao=solicitacao,
                    usuario=request.user,
                    acao='EMPENHO',
                    lote=lote.lote,

                    quantidade=(
                        quantidade_historico
                    ),

                    unidade=unidade_historico,

                    observacao=(
                        f'Item salvo no Empenho '
                        f'#{empenho.id} deste card.'
                    )
                )

            if itens_validos == 0:
                raise ValueError(
                    'Nenhum item possui quantidade válida.'
                )

            # ----------------------------------------------------------
            # RECALCULAR TOTAL DO EMPENHO
            # ----------------------------------------------------------
            total_unidades = (
                empenho.itens.aggregate(
                    total=Sum(
                        'quantidade'
                    )
                )['total']
                or Decimal('0')
            )

            if (
                solicitacao.unidade_controle
                == 'QUILOGRAMA'
            ):
                total_kg_final = Decimal('0')

                itens_finais = (
                    empenho.itens
                    .select_related(
                        'estoque'
                    )
                )

                for item_final in itens_finais:
                    quantidade_item = Decimal(
                        str(
                            item_final.quantidade
                            or 0
                        )
                    )

                    peso_item = Decimal(
                        str(
                            item_final
                            .estoque
                            .peso_unitario
                            or 0
                        )
                    )

                    total_kg_final += (
                        quantidade_item
                        * peso_item
                    )

                solicitacao.quantidade_empenhada = (
                    total_kg_final
                )

                quantidade_para_status = (
                    total_kg_final
                )

            else:
                solicitacao.quantidade_empenhada = (
                    Decimal(
                        str(
                            total_unidades
                        )
                    )
                )

                quantidade_para_status = (
                    solicitacao.quantidade_empenhada
                )

            quantidade_solicitada = Decimal(
                str(
                    solicitacao.quantidade_solicitada
                    or 0
                )
            )

            # ----------------------------------------------------------
            # ATUALIZAR STATUS
            # ----------------------------------------------------------
            if quantidade_para_status <= 0:
                solicitacao.status = (
                    'AGUARDANDO_EMPENHO'
                )

                evento = 'CRIACAO'

            elif (
                quantidade_para_status
                >= quantidade_solicitada
            ):
                solicitacao.status = (
                    'EMPENHO_COMPLETO'
                )

                evento = 'EMPENHO_COMPLETO'

            else:
                solicitacao.status = (
                    'EMPENHO_PARCIAL'
                )

                evento = 'EMPENHO_PARCIAL'

            solicitacao.save(
                update_fields=[
                    'quantidade_empenhada',
                    'status',
                    'data_atualizacao',
                ]
            )

            avaliar_workflow(
                solicitacao,
                evento,
                request.user
            )

            cache.delete(
                'cards_version_hash'
            )

            coluna_kanban_nome = ''

            if solicitacao.coluna_kanban_id:
                coluna_kanban_nome = (
                    ColunaKanban.objects
                    .filter(
                        id=solicitacao.coluna_kanban_id
                    )
                    .values_list(
                        'nome',
                        flat=True
                    )
                    .first()
                    or ''
                )

            return JsonResponse({
                'success': True,

                'message': (
                    f'{itens_validos} item(ns) '
                    f'salvo(s) no empenho deste card.'
                ),

                'empenho_id': empenho.id,
                'solicitacao_id': solicitacao.id,

                'total_empenhado': float(
                    solicitacao.quantidade_empenhada
                    or 0
                ),

                'percentual': float(
                    solicitacao.percentual_empenhado
                    or 0
                ),

                'status': solicitacao.status,
                'coluna_kanban': coluna_kanban_nome,

                'empenho_bloqueado': False,
            })

    except Solicitacao.DoesNotExist:
        return JsonResponse(
            {
                'success': False,
                'error': 'Solicitação não encontrada.'
            },
            status=404
        )

    except Estoque.DoesNotExist:
        return JsonResponse(
            {
                'success': False,
                'error': 'Um dos lotes não foi encontrado.'
            },
            status=404
        )

    except ValueError as erro:
        return JsonResponse(
            {
                'success': False,
                'error': str(erro)
            },
            status=400
        )

    except Exception as erro:
        logger.exception(
            'Erro ao empenhar solicitação %s',
            solicitacao_id
        )

        return JsonResponse(
            {
                'success': False,
                'error': f'Erro ao empenhar: {erro}'
            },
            status=500
        )
# ============================================================================
# REMOVER ITEM DO EMPENHO
# ============================================================================
@login_required
@permission_required(
    'sapp.pode_empenhar_solicitacao',
    raise_exception=True
)
def api_remover_item_empenho(
    request,
    solicitacao_id,
    item_id
):
    """
    Remove um item exclusivamente do empenho do card atual.

    Não permite remoção depois que a movimentação começou.
    """

    if request.method != 'POST':
        return JsonResponse(
            {
                'success': False,
                'error': 'Método não permitido.'
            },
            status=405
        )

    try:
        with transaction.atomic():

            solicitacao = (
                Solicitacao.objects
                .select_for_update(
                    of=('self',)
                )
                .get(
                    id=solicitacao_id
                )
            )

            validar_edicao_empenho_solicitacao(
                solicitacao
            )

            empenho = (
                Empenho.objects
                .select_for_update(
                    of=('self',)
                )
                .filter(
                    solicitacao_id=solicitacao.id,
                    status__nome='Rascunho',
                )
                .order_by(
                    '-data_criacao',
                    '-id',
                )
                .first()
            )

            if not empenho:
                raise ValueError(
                    'Nenhum empenho salvo foi encontrado '
                    'para este card.'
                )

            item = (
                ItemEmpenho.objects
                .select_for_update(
                    of=('self',)
                )
                .get(
                    id=item_id,
                    empenho_id=empenho.id,
                )
            )

            quantidade_removida = Decimal(
                str(
                    item.quantidade
                    or 0
                )
            )

            lote_nome = item.lote

            estoque_id = item.estoque_id

            peso_unitario = Decimal('0')

            if estoque_id:
                estoque = (
                    Estoque.objects
                    .select_for_update(
                        of=('self',)
                    )
                    .get(
                        id=estoque_id
                    )
                )

                peso_unitario = Decimal(
                    str(
                        estoque.peso_unitario
                        or 0
                    )
                )

            # Executa o delete personalizado do item.
            item.delete()

            # ----------------------------------------------------------
            # RECALCULAR EMPENHO DO CARD
            # ----------------------------------------------------------
            if (
                solicitacao.unidade_controle
                == 'QUILOGRAMA'
            ):
                total_restante = Decimal('0')

                itens_restantes = (
                    empenho.itens
                    .select_related(
                        'estoque'
                    )
                )

                for item_restante in itens_restantes:
                    quantidade = Decimal(
                        str(
                            item_restante.quantidade
                            or 0
                        )
                    )

                    peso = Decimal(
                        str(
                            item_restante
                            .estoque
                            .peso_unitario
                            or 0
                        )
                    )

                    total_restante += (
                        quantidade
                        * peso
                    )

                solicitacao.quantidade_empenhada = (
                    total_restante
                )

                quantidade_historico = (
                    quantidade_removida
                    * peso_unitario
                )

                unidade_historico = 'KG'

            else:
                total_restante = (
                    empenho.itens.aggregate(
                        total=Sum(
                            'quantidade'
                        )
                    )['total']
                    or Decimal('0')
                )

                solicitacao.quantidade_empenhada = (
                    Decimal(
                        str(
                            total_restante
                        )
                    )
                )

                quantidade_historico = (
                    quantidade_removida
                )

                unidade_historico = 'BAG'

            quantidade_solicitada = Decimal(
                str(
                    solicitacao.quantidade_solicitada
                    or 0
                )
            )

            if (
                solicitacao.quantidade_empenhada
                <= 0
            ):
                solicitacao.status = (
                    'AGUARDANDO_EMPENHO'
                )

                evento = 'CRIACAO'

            elif (
                solicitacao.quantidade_empenhada
                >= quantidade_solicitada
            ):
                solicitacao.status = (
                    'EMPENHO_COMPLETO'
                )

                evento = 'EMPENHO_COMPLETO'

            else:
                solicitacao.status = (
                    'EMPENHO_PARCIAL'
                )

                evento = 'EMPENHO_PARCIAL'

            solicitacao.save(
                update_fields=[
                    'quantidade_empenhada',
                    'status',
                    'data_atualizacao',
                ]
            )

            HistoricoCard.objects.create(
                solicitacao=solicitacao,
                usuario=request.user,
                acao='REMOCAO_EMPENHO',
                lote=lote_nome,
                quantidade=quantidade_historico,
                unidade=unidade_historico,
                observacao=(
                    f'Item removido do Empenho '
                    f'#{empenho.id} deste card.'
                )
            )

            avaliar_workflow(
                solicitacao,
                evento,
                request.user
            )

            cache.delete(
                'cards_version_hash'
            )

            return JsonResponse({
                'success': True,

                'message': (
                    f'Item do lote {lote_nome} '
                    f'removido do empenho.'
                ),

                'item_id': item_id,
                'estoque_id': estoque_id,

                'total_empenhado': float(
                    solicitacao.quantidade_empenhada
                    or 0
                ),

                'status': solicitacao.status,
            })

    except Solicitacao.DoesNotExist:
        return JsonResponse(
            {
                'success': False,
                'error': 'Solicitação não encontrada.'
            },
            status=404
        )

    except ItemEmpenho.DoesNotExist:
        return JsonResponse(
            {
                'success': False,
                'error': (
                    'Item não encontrado no empenho '
                    'deste card.'
                )
            },
            status=404
        )

    except Estoque.DoesNotExist:
        return JsonResponse(
            {
                'success': False,
                'error': 'Estoque do item não encontrado.'
            },
            status=404
        )

    except ValueError as erro:
        return JsonResponse(
            {
                'success': False,
                'error': str(erro)
            },
            status=400
        )

    except Exception as erro:
        logger.exception(
            'Erro ao remover o item %s do card %s',
            item_id,
            solicitacao_id
        )

        return JsonResponse(
            {
                'success': False,
                'error': str(erro)
            },
            status=500
        )
# ============================================================================
# MOVIMENTAR SOLICITAÇÃO (TRANSFERIR / EXPEDIR)
# ============================================================================
# ============================================================================
# MOVIMENTAR (TRANSFERIR / EXPEDIR)
# ============================================================================
@login_required
def api_movimentar_solicitacao(request, solicitacao_id):
    """
    Transfere ou expede itens pertencentes exclusivamente
    ao empenho vinculado à solicitação informada.
    """

    if request.method != 'POST':
        return JsonResponse(
            {
                'success': False,
                'error': 'Método não permitido.'
            },
            status=405
        )

    # ------------------------------------------------------------------------
    # LER JSON
    # ------------------------------------------------------------------------
    try:
        data = json.loads(request.body or '{}')
    except json.JSONDecodeError:
        return JsonResponse(
            {
                'success': False,
                'error': 'JSON inválido.'
            },
            status=400
        )

    acao = str(data.get('acao') or '').strip().lower()
    itens_ids = data.get('itens_ids', [])
    observacao = str(
        data.get('observacao') or ''
    ).strip()

    # ------------------------------------------------------------------------
    # VALIDAR AÇÃO
    # ------------------------------------------------------------------------
    if acao not in [
        'transferir',
        'expedir',
    ]:
        return JsonResponse(
            {
                'success': False,
                'error': 'Ação inválida.'
            },
            status=400
        )

    # ------------------------------------------------------------------------
    # VALIDAR ITENS
    # ------------------------------------------------------------------------
    if not isinstance(itens_ids, list) or not itens_ids:
        return JsonResponse(
            {
                'success': False,
                'error': 'Nenhum item selecionado.'
            },
            status=400
        )

    try:
        itens_ids = [
            int(item_id)
            for item_id in itens_ids
        ]
    except (TypeError, ValueError):
        return JsonResponse(
            {
                'success': False,
                'error': 'Lista de itens inválida.'
            },
            status=400
        )

    # Remove IDs repetidos.
    itens_ids = list(dict.fromkeys(itens_ids))

    try:
        with transaction.atomic():

            # ================================================================
            # BLOQUEAR SOMENTE A SOLICITAÇÃO
            #
            # Não utilizar select_related aqui.
            # coluna_kanban pode aceitar NULL e gerar LEFT OUTER JOIN.
            # ================================================================
            solicitacao = (
                Solicitacao.objects
                .select_for_update(of=('self',))
                .get(id=solicitacao_id)
            )

            # Impede movimentação em cards finalizados.
            if solicitacao.status in [
                'CONCLUIDO',
                'CANCELADO',
            ]:
                raise ValueError(
                    'Este card já está concluído ou cancelado.'
                )

            # ================================================================
            # BLOQUEAR SOMENTE O EMPENHO
            #
            # of=('self',) impede que o PostgreSQL tente bloquear
            # outras tabelas usadas na consulta.
            # ================================================================
            empenho = (
                Empenho.objects
                .select_for_update(of=('self',))
                .filter(
                    solicitacao_id=solicitacao.id,
                    status__nome='Rascunho',
                )
                .order_by('id')
                .first()
            )

            if not empenho:
                raise ValueError(
                    'Nenhum empenho em rascunho foi encontrado '
                    'para esta solicitação.'
                )

            # ================================================================
            # BLOQUEAR SOMENTE OS ITENS
            #
            # Não usar select_related junto com select_for_update.
            # ================================================================
            itens = list(
                ItemEmpenho.objects
                .select_for_update(of=('self',))
                .filter(
                    id__in=itens_ids,
                    empenho_id=empenho.id,
                )
                .order_by('id')
            )

            if not itens:
                raise ValueError(
                    'Nenhum item válido foi encontrado.'
                )

            # Confirma que todos os itens enviados pertencem
            # ao empenho desta solicitação.
            ids_encontrados = {
                item.id
                for item in itens
            }

            ids_solicitados = set(itens_ids)

            if ids_encontrados != ids_solicitados:
                raise ValueError(
                    'Um ou mais itens não pertencem '
                    'a esta solicitação.'
                )

            # ================================================================
            # DADOS DA AÇÃO
            # ================================================================
            novo_endereco = ''
            novo_az = ''
            numero_carga = ''
            cliente_expedicao = ''
            placa = ''

            if acao == 'transferir':
                novo_endereco = str(
                    data.get('novo_endereco') or ''
                ).strip().upper()

                novo_az = str(
                    data.get('az') or ''
                ).strip().upper()

                if not novo_endereco:
                    raise ValueError(
                        'Novo endereço não informado '
                        'para transferência.'
                    )

                # ============================================================
                # VALIDAR ENDEREÇO NO BACKEND
                #
                # Não confiar apenas no JavaScript.
                # Transferência só pode ir para um endereço cadastrado.
                # O AZ também é obtido do cadastro e não do valor enviado
                # pelo navegador.
                # ============================================================
                endereco_destino = (
                    Endereco.objects
                    .select_related('armazem')
                    .filter(
                        codigo__iexact=novo_endereco
                    )
                    .first()
                )

                if not endereco_destino:
                    raise ValueError(
                        f'O endereço "{novo_endereco}" não está '
                        'cadastrado no sistema. '
                        'Cadastre o endereço antes de transferir.'
                    )

                novo_endereco = (
                    endereco_destino.codigo
                    .strip()
                    .upper()
                )

                novo_az = (
                    endereco_destino.armazem.nome
                    if endereco_destino.armazem
                    else ''
                )

            else:
                numero_carga = str(
                    data.get('numero_carga') or ''
                ).strip()

                cliente_expedicao = str(
                    data.get('cliente') or ''
                ).strip()

                placa = str(
                    data.get('placa') or ''
                ).strip().upper()

            total_movimentado_unidades = Decimal('0')
            total_movimentado_kg = Decimal('0')

            # ================================================================
            # PROCESSAR ITENS
            #
            # IMPORTANTE:
            # - origem (Estoque) é usada para movimentar o saldo físico atual.
            # - item (ItemEmpenho) é usado para auditoria/impressão, pois guarda
            #   o retrato do lote no momento em que foi empenhado.
            # ================================================================
            for item in itens:

                # ------------------------------------------------------------
                # BLOQUEAR SOMENTE O ESTOQUE DE ORIGEM
                #
                # Não utilizar select_related aqui.
                # cultivar, peneira, tratamento etc. podem aceitar NULL.
                # ------------------------------------------------------------
                origem = (
                    Estoque.objects
                    .select_for_update(of=('self',))
                    .get(id=item.estoque_id)
                )

                quantidade = Decimal(
                    str(item.quantidade or 0)
                )

                saldo_atual = Decimal(
                    str(origem.saldo or 0)
                )

                quantidade_reservada = Decimal(
                    str(origem.empenhado or 0)
                )

                peso_unitario = Decimal(
                    str(origem.peso_unitario or 0)
                )

                if quantidade <= 0:
                    raise ValueError(
                        f'Quantidade inválida para '
                        f'o lote {item.lote}.'
                    )

                if quantidade > saldo_atual:
                    raise ValueError(
                        f'Saldo insuficiente para o lote '
                        f'{item.lote}. Saldo atual: '
                        f'{origem.saldo}.'
                    )

                if quantidade > quantidade_reservada:
                    raise ValueError(
                        f'A quantidade reservada do lote '
                        f'{item.lote} é insuficiente.'
                    )

                # ============================================================
                # TRANSFERÊNCIA
                # ============================================================
                if acao == 'transferir':

                    az_destino = (
                        novo_az
                        or origem.az
                        or ''
                    )

                    endereco_origem = (
                        origem.endereco or ''
                    )

                    az_origem = (
                        origem.az or ''
                    )

                    # Impede transferência para o mesmo local.
                    if (
                        endereco_origem.strip().upper()
                        == novo_endereco
                        and
                        az_origem.strip().upper()
                        == az_destino.strip().upper()
                    ):
                        raise ValueError(
                            f'O lote {origem.lote} já está no '
                            f'endereço {novo_endereco}, '
                            f'AZ {az_destino or "-"}.'
                        )

                    # --------------------------------------------------------
                    # BUSCAR ESTOQUE DE DESTINO
                    #
                    # Não há select_related nesta consulta.
                    # O bloqueio fica somente em Estoque.
                    # --------------------------------------------------------
                    destino = (
                        Estoque.objects
                        .select_for_update(of=('self',))
                        .filter(
                            lote=origem.lote,
                            produto=origem.produto,
                            cultivar=origem.cultivar,
                            peneira=origem.peneira,
                            categoria=origem.categoria,
                            tratamento=origem.tratamento,
                            especie=origem.especie,
                            endereco=novo_endereco,
                            az=az_destino,
                            empresa=origem.empresa,
                            embalagem=origem.embalagem,
                            cliente=origem.cliente,
                            peso_unitario=origem.peso_unitario,
                        )
                        .order_by('id')
                        .first()
                    )

                    # Caso não exista estoque no destino,
                    # cria um novo registro.
                    if not destino:
                        destino = Estoque.objects.create(
                            lote=origem.lote,
                            produto=origem.produto,
                            cultivar=origem.cultivar,
                            peneira=origem.peneira,
                            categoria=origem.categoria,
                            tratamento=origem.tratamento,
                            especie=origem.especie,
                            endereco=novo_endereco,
                            az=az_destino,
                            entrada=Decimal('0'),
                            saida=Decimal('0'),
                            empenhado=Decimal('0'),
                            peso_unitario=origem.peso_unitario,
                            embalagem=origem.embalagem,
                            conferente=request.user,
                            empresa=origem.empresa,
                            cliente=origem.cliente,
                            observacao=origem.observacao,
                            status_sistemico=(
                                origem.status_sistemico
                            ),
                        )

                    # Entrada no destino.
                    destino.entrada = (
                        Decimal(
                            str(destino.entrada or 0)
                        )
                        + quantidade
                    )

                    # Salva normalmente para o model.save()
                    # recalcular campos derivados.
                    destino.save()

                    # Saída da origem.
                    origem.saida = (
                        Decimal(
                            str(origem.saida or 0)
                        )
                        + quantidade
                    )

                    origem.save()

                    # --------------------------------------------------------
                    # HISTÓRICO DA SAÍDA
                    # --------------------------------------------------------
                    HistoricoMovimentacao.objects.create(
                        estoque=origem,
                        usuario=request.user,
                        quantidade=quantidade,
                        tipo='Transferência (Saída)',
                        descricao=(
                            f'Transferido {quantidade} un '
                            f'de {endereco_origem or "-"} '
                            f'(AZ {az_origem or "-"}) para '
                            f'{novo_endereco} '
                            f'(AZ {az_destino or "-"}).'
                        )
                    )

                    # --------------------------------------------------------
                    # HISTÓRICO DA ENTRADA
                    # --------------------------------------------------------
                    HistoricoMovimentacao.objects.create(
                        estoque=destino,
                        usuario=request.user,
                        quantidade=quantidade,
                        tipo='Transferência (Entrada)',
                        descricao=(
                            f'Recebido {quantidade} un '
                            f'em {novo_endereco} '
                            f'(AZ {az_destino or "-"}), '
                            f'vindo de {endereco_origem or "-"} '
                            f'(AZ {az_origem or "-"}).'
                        )
                    )

                    # --------------------------------------------------------
                    # HISTÓRICO DO ITEM DO EMPENHO
                    # --------------------------------------------------------
                    HistoricoItemEmpenho.objects.create(
                        empenho=empenho,
                        item_empenho_id_original=item.id,
                        estoque_origem=origem,
                        estoque_destino=destino,

                        # SNAPSHOT DO MOMENTO DO EMPENHO.
                        lote=(
                            item.lote
                            or origem.lote
                            or ''
                        ),

                        produto=(
                            item.produto_snapshot
                            or origem.produto
                            or ''
                        ),

                        cultivar=(
                            item.cultivar
                            or (
                                origem.cultivar.nome
                                if origem.cultivar
                                else ''
                            )
                        ),

                        peneira=(
                            item.peneira
                            or (
                                origem.peneira.nome
                                if origem.peneira
                                else ''
                            )
                        ),

                        categoria=(
                            item.categoria
                            or (
                                origem.categoria.nome
                                if origem.categoria
                                else ''
                            )
                        ),

                        tratamento=(
                            item.tratamento_snapshot
                            or (
                                origem.tratamento.nome
                                if origem.tratamento
                                else ''
                            )
                        ),

                        especie=(
                            item.especie_snapshot
                            or (
                                origem.especie.nome
                                if origem.especie
                                else ''
                            )
                        ),

                        embalagem=(
                            item.embalagem_snapshot
                            or origem.embalagem
                            or ''
                        ),

                        empresa=(
                            item.empresa_snapshot
                            or origem.empresa
                            or ''
                        ),

                        cliente=(
                            item.cliente_snapshot
                            or origem.cliente
                            or ''
                        ),

                        endereco_origem=(
                            item.endereco_origem
                            or origem.endereco
                            or ''
                        ),

                        az_origem=(
                            item.az_origem
                            or origem.az
                            or ''
                        ),

                        saldo_anterior=(
                            item.saldo_anterior
                            or 0
                        ),

                        peso_unitario=(
                            item.peso_unitario_snapshot
                            or origem.peso_unitario
                            or 0
                        ),

                        observacao_origem=(
                            item.observacao_snapshot
                            or ''
                        ),

                        conferente=(
                            item.conferente_snapshot
                            or ''
                        ),

                        # Dado da movimentação posterior.
                        endereco_destino=novo_endereco,

                        quantidade=quantidade,
                        tipo='transferencia',

                        # Observação digitada na transferência.
                        observacao=observacao,

                        processado_por=request.user,
                    )

                    descricao_feed = (
                        f'Transferido {quantidade} un '
                        f'de {endereco_origem or "-"} '
                        f'(AZ {az_origem or "-"}) para '
                        f'{novo_endereco} '
                        f'(AZ {az_destino or "-"}).'
                    )

                    if observacao:
                        descricao_feed += (
                            f' Observação: {observacao}'
                        )

                # ============================================================
                # EXPEDIÇÃO
                # ============================================================
                else:
                    endereco_origem = (
                        origem.endereco or ''
                    )

                    az_origem = (
                        origem.az or ''
                    )

                    origem.saida = (
                        Decimal(
                            str(origem.saida or 0)
                        )
                        + quantidade
                    )

                    origem.save()

                    descricao_movimentacao = (
                        f'Expedido {quantidade} un '
                        f'do lote {origem.lote}, '
                        f'endereço {endereco_origem or "-"}, '
                        f'AZ {az_origem or "-"}.'
                    )

                    if numero_carga:
                        descricao_movimentacao += (
                            f' Carga: {numero_carga}.'
                        )

                    if cliente_expedicao:
                        descricao_movimentacao += (
                            f' Cliente: {cliente_expedicao}.'
                        )

                    if placa:
                        descricao_movimentacao += (
                            f' Placa: {placa}.'
                        )

                    if observacao:
                        descricao_movimentacao += (
                            f' Observação: {observacao}'
                        )

                    HistoricoMovimentacao.objects.create(
                        estoque=origem,
                        usuario=request.user,
                        quantidade=quantidade,
                        tipo='Expedição',
                        descricao=descricao_movimentacao,
                        numero_carga=(
                            numero_carga or None
                        ),
                        cliente=(
                            cliente_expedicao
                            or origem.cliente
                        ),
                        placa=placa or None,
                    )

                    HistoricoItemEmpenho.objects.create(
                        empenho=empenho,
                        item_empenho_id_original=item.id,
                        estoque_origem=origem,

                        # SNAPSHOT DO MOMENTO DO EMPENHO.
                        lote=(
                            item.lote
                            or origem.lote
                            or ''
                        ),

                        produto=(
                            item.produto_snapshot
                            or origem.produto
                            or ''
                        ),

                        cultivar=(
                            item.cultivar
                            or (
                                origem.cultivar.nome
                                if origem.cultivar
                                else ''
                            )
                        ),

                        peneira=(
                            item.peneira
                            or (
                                origem.peneira.nome
                                if origem.peneira
                                else ''
                            )
                        ),

                        categoria=(
                            item.categoria
                            or (
                                origem.categoria.nome
                                if origem.categoria
                                else ''
                            )
                        ),

                        tratamento=(
                            item.tratamento_snapshot
                            or (
                                origem.tratamento.nome
                                if origem.tratamento
                                else ''
                            )
                        ),

                        especie=(
                            item.especie_snapshot
                            or (
                                origem.especie.nome
                                if origem.especie
                                else ''
                            )
                        ),

                        embalagem=(
                            item.embalagem_snapshot
                            or origem.embalagem
                            or ''
                        ),

                        empresa=(
                            item.empresa_snapshot
                            or origem.empresa
                            or ''
                        ),

                        # Cliente abaixo é o cliente original do empenho.
                        # O cliente informado na expedição continua no
                        # HistoricoMovimentacao/feed.
                        cliente=(
                            item.cliente_snapshot
                            or origem.cliente
                            or ''
                        ),

                        endereco_origem=(
                            item.endereco_origem
                            or origem.endereco
                            or ''
                        ),

                        az_origem=(
                            item.az_origem
                            or origem.az
                            or ''
                        ),

                        saldo_anterior=(
                            item.saldo_anterior
                            or 0
                        ),

                        peso_unitario=(
                            item.peso_unitario_snapshot
                            or origem.peso_unitario
                            or 0
                        ),

                        observacao_origem=(
                            item.observacao_snapshot
                            or ''
                        ),

                        conferente=(
                            item.conferente_snapshot
                            or ''
                        ),

                        quantidade=quantidade,
                        tipo='expedicao',

                        # Observação da operação.
                        observacao=observacao,

                        numero_carga=numero_carga or '',
                        placa=placa or '',
                        processado_por=request.user,
                    )

                    descricao_feed = (
                        f'Expedido {quantidade} un '
                        f'de {endereco_origem or "-"} '
                        f'(AZ {az_origem or "-"}).'
                    )

                    if numero_carga:
                        descricao_feed += (
                            f' Carga: {numero_carga}.'
                        )

                    if cliente_expedicao:
                        descricao_feed += (
                            f' Cliente: {cliente_expedicao}.'
                        )

                    if placa:
                        descricao_feed += (
                            f' Placa: {placa}.'
                        )

                    if observacao:
                        descricao_feed += (
                            f' Observação: {observacao}'
                        )

                # ============================================================
                # SOMAR MOVIMENTAÇÃO
                # ============================================================
                total_movimentado_unidades += quantidade

                total_movimentado_kg += (
                    quantidade
                    * peso_unitario
                )

                # ------------------------------------------------------------
                # HISTÓRICO DO CARD
                #
                # Para solicitação em KG, registra a quantidade em KG.
                # ------------------------------------------------------------
                if (
                    solicitacao.unidade_controle
                    == 'QUILOGRAMA'
                ):
                    quantidade_historico = (
                        quantidade
                        * peso_unitario
                    )
                    unidade_historico = 'KG'
                else:
                    quantidade_historico = quantidade
                    unidade_historico = 'BAG'

                HistoricoCard.objects.create(
                    solicitacao=solicitacao,
                    usuario=request.user,
                    acao=(
                        'TRANSFERENCIA'
                        if acao == 'transferir'
                        else 'EXPEDICAO'
                    ),
                    lote=origem.lote,
                    quantidade=quantidade_historico,
                    unidade=unidade_historico,
                    observacao=descricao_feed,
                )

                # ------------------------------------------------------------
                # EXCLUIR ITEM DO EMPENHO
                #
                # Executa o delete personalizado e libera a reserva
                # existente em Estoque.empenhado.
                # ------------------------------------------------------------
                item.delete()

            # ================================================================
            # QUANTIDADE MOVIMENTADA NESTA OPERAÇÃO
            # ================================================================
            if (
                solicitacao.unidade_controle
                == 'QUILOGRAMA'
            ):
                quantidade_movimentada_agora = (
                    total_movimentado_kg
                )
            else:
                quantidade_movimentada_agora = (
                    total_movimentado_unidades
                )

            solicitacao.quantidade_movimentada = (
                Decimal(
                    str(
                        solicitacao.quantidade_movimentada
                        or 0
                    )
                )
                + quantidade_movimentada_agora
            )

            # ================================================================
            # RECALCULAR QUANTIDADE AINDA EMPENHADA
            # ================================================================
            if (
                solicitacao.unidade_controle
                == 'QUILOGRAMA'
            ):
                total_restante_kg = Decimal('0')

                itens_restantes = (
                    ItemEmpenho.objects
                    .filter(
                        empenho_id=empenho.id
                    )
                    .values(
                        'quantidade',
                        'estoque__peso_unitario',
                    )
                )

                for item_restante in itens_restantes:
                    quantidade_restante = Decimal(
                        str(
                            item_restante.get(
                                'quantidade'
                            )
                            or 0
                        )
                    )

                    peso_restante = Decimal(
                        str(
                            item_restante.get(
                                'estoque__peso_unitario'
                            )
                            or 0
                        )
                    )

                    total_restante_kg += (
                        quantidade_restante
                        * peso_restante
                    )

                solicitacao.quantidade_empenhada = (
                    total_restante_kg
                )

            else:
                total_restante_unidades = (
                    empenho.itens.aggregate(
                        total=Sum('quantidade')
                    )['total']
                    or 0
                )

                solicitacao.quantidade_empenhada = (
                    Decimal(
                        str(total_restante_unidades)
                    )
                )

            # ================================================================
            # ATUALIZAR STATUS DA SOLICITAÇÃO
            # ================================================================
            quantidade_solicitada = Decimal(
                str(
                    solicitacao.quantidade_solicitada
                    or 0
                )
            )

            quantidade_movimentada = Decimal(
                str(
                    solicitacao.quantidade_movimentada
                    or 0
                )
            )

            if (
                quantidade_movimentada
                >= quantidade_solicitada
            ):
                solicitacao.quantidade_movimentada = (
                    quantidade_solicitada
                )

                solicitacao.status = 'CONCLUIDO'

                evento = (
                    'TRANSFERENCIA_COMPLETA'
                    if acao == 'transferir'
                    else 'EXPEDICAO_COMPLETA'
                )

            elif quantidade_movimentada > 0:
                solicitacao.status = (
                    'MOVIMENTACAO_PARCIAL'
                )

                evento = 'MOVIMENTACAO_PARCIAL'

            else:
                evento = None

            solicitacao.save(
                update_fields=[
                    'quantidade_movimentada',
                    'quantidade_empenhada',
                    'status',
                    'data_atualizacao',
                ]
            )

            # ================================================================
            # EXECUTAR WORKFLOW
            # ================================================================
            if evento:
                avaliar_workflow(
                    solicitacao,
                    evento,
                    request.user
                )

            # ================================================================
            # LIMPAR CACHE
            # ================================================================
            from django.core.cache import cache

            cache.delete(
                'cards_version_hash'
            )

            # ================================================================
            # COLUNA DO KANBAN
            #
            # Ela é buscada somente agora, depois dos bloqueios.
            # ================================================================
            coluna_kanban_nome = ''

            if solicitacao.coluna_kanban_id:
                coluna_kanban_nome = (
                    ColunaKanban.objects
                    .filter(
                        id=solicitacao.coluna_kanban_id
                    )
                    .values_list(
                        'nome',
                        flat=True
                    )
                    .first()
                    or ''
                )

            return JsonResponse(
                {
                    'success': True,

                    'message': (
                        f'{len(itens)} item(ns) '
                        f'{"transferido(s)" if acao == "transferir" else "expedido(s)"} '
                        f'com sucesso.'
                    ),

                    'movimentado_agora': float(
                        quantidade_movimentada_agora
                    ),

                    'total_movimentado': float(
                        solicitacao.quantidade_movimentada
                    ),

                    'quantidade_empenhada': float(
                        solicitacao.quantidade_empenhada
                    ),

                    'quantidade_solicitada': float(
                        solicitacao.quantidade_solicitada
                    ),

                    'percentual': float(
                        solicitacao.percentual_movimentado
                    ),

                    'status': solicitacao.status,

                    'concluido': (
                        solicitacao.status
                        == 'CONCLUIDO'
                    ),

                    'coluna_kanban': (
                        coluna_kanban_nome
                    ),
                }
            )

    except Solicitacao.DoesNotExist:
        return JsonResponse(
            {
                'success': False,
                'error': 'Solicitação não encontrada.'
            },
            status=404
        )

    except Estoque.DoesNotExist:
        return JsonResponse(
            {
                'success': False,
                'error': (
                    'O estoque de um dos itens '
                    'não foi encontrado.'
                )
            },
            status=404
        )

    except ValueError as erro:
        return JsonResponse(
            {
                'success': False,
                'error': str(erro)
            },
            status=400
        )

    except Exception as erro:
        logger.exception(
            'Erro ao movimentar a solicitação %s',
            solicitacao_id
        )

        return JsonResponse(
            {
                'success': False,
                'error': str(erro)
            },
            status=500
        )


@login_required
def api_dados_impressao_solicitacao(
    request,
    solicitacao_id
):
    """
    Dados para:
    - impressão da Solicitação;
    - modal de Transferência/Expedição.

    REGRA DE AUDITORIA:
    os dados exibidos representam o estado do lote no momento
    do EMPENHO.

    Nunca usamos o endereço de destino da transferência como
    endereço original do item.
    """

    solicitacao = get_object_or_404(
        Solicitacao.objects.select_related(
            'criador',
            'armazem',
            'especie',
        ),
        id=solicitacao_id
    )

    empenho = (
        Empenho.objects
        .filter(
            solicitacao_id=solicitacao.id
        )
        .order_by('-data_criacao')
        .first()
    )

    itens_pendentes = []
    itens_processados = []

    def nome_usuario(usuario):
        if not usuario:
            return ''

        return (
            usuario.get_full_name()
            or usuario.username
            or ''
        )

    if empenho:
        # ----------------------------------------------------
        # PENDENTES
        #
        # A FK estoque continua disponível para operação
        # física, mas a exibição usa prioritariamente snapshot.
        # ----------------------------------------------------
        itens = (
            empenho.itens
            .select_related(
                'estoque',
                'estoque__cultivar',
                'estoque__peneira',
                'estoque__categoria',
                'estoque__especie',
                'estoque__tratamento',
                'estoque__conferente',
            )
            .order_by('id')
        )

        for item in itens:
            estoque = item.estoque

            peso_unitario = Decimal(
                str(
                    item.peso_unitario_snapshot
                    or 0
                )
            )

            if (
                peso_unitario <= 0
                and estoque
            ):
                peso_unitario = Decimal(
                    str(
                        estoque.peso_unitario
                        or 0
                    )
                )

            saldo_no_empenho = Decimal(
                str(
                    item.saldo_anterior
                    or 0
                )
            )

            endereco_empenho = (
                item.endereco_origem
                or (
                    estoque.endereco
                    if estoque
                    else ''
                )
            )

            armazem_empenho = (
                item.az_origem
                or (
                    estoque.az
                    if estoque
                    else ''
                )
            )

            produto_empenho = (
                item.produto_snapshot
                or (
                    estoque.produto
                    if estoque
                    else ''
                )
            )

            especie_empenho = (
                item.especie_snapshot
                or (
                    estoque.especie.nome
                    if (
                        estoque
                        and estoque.especie
                    )
                    else ''
                )
            )

            tratamento_empenho = (
                item.tratamento_snapshot
                or (
                    estoque.tratamento.nome
                    if (
                        estoque
                        and estoque.tratamento
                    )
                    else ''
                )
            )

            embalagem_empenho = (
                item.embalagem_snapshot
                or (
                    estoque.embalagem
                    if estoque
                    else ''
                )
            )

            empresa_empenho = (
                item.empresa_snapshot
                or (
                    estoque.empresa
                    if estoque
                    else ''
                )
            )

            cliente_empenho = (
                item.cliente_snapshot
                or (
                    estoque.cliente
                    if estoque
                    else ''
                )
            )

            observacao_empenho = (
                item.observacao_snapshot
                or item.observacao
                or ''
            )

            conferente_empenho = (
                item.conferente_snapshot
                or (
                    nome_usuario(
                        estoque.conferente
                    )
                    if estoque
                    else ''
                )
            )

            quantidade = Decimal(
                str(
                    item.quantidade
                    or 0
                )
            )

            itens_pendentes.append({
                'item_id': item.id,

                'lote': (
                    item.lote
                    or (
                        estoque.lote
                        if estoque
                        else ''
                    )
                ),

                'quantidade': float(
                    quantidade
                ),

                # "saldo_atual" é mantido por compatibilidade
                # com o JS antigo, mas agora significa
                # SALDO NO MOMENTO DO EMPENHO.
                'saldo_atual': float(
                    saldo_no_empenho
                ),

                'saldo_empenho': float(
                    saldo_no_empenho
                ),

                'endereco': endereco_empenho,

                # Dois nomes para compatibilidade.
                'az': armazem_empenho,
                'armazem': armazem_empenho,

                'produto': produto_empenho,

                'cultivar': (
                    item.cultivar
                    or (
                        estoque.cultivar.nome
                        if (
                            estoque
                            and estoque.cultivar
                        )
                        else ''
                    )
                ),

                'peneira': (
                    item.peneira
                    or (
                        estoque.peneira.nome
                        if (
                            estoque
                            and estoque.peneira
                        )
                        else ''
                    )
                ),

                'categoria': (
                    item.categoria
                    or (
                        estoque.categoria.nome
                        if (
                            estoque
                            and estoque.categoria
                        )
                        else ''
                    )
                ),

                'especie': especie_empenho,
                'tratamento': tratamento_empenho,
                'embalagem': embalagem_empenho,
                'empresa': empresa_empenho,
                'cliente': cliente_empenho,

                'peso_unitario': str(
                    peso_unitario
                ),

                'peso_total': str(
                    quantidade
                    * peso_unitario
                ),

                'observacao': (
                    observacao_empenho
                ),

                'conferente': (
                    conferente_empenho
                ),

                'situacao': 'pendente',

                'data_empenho': (
                    item.data_criacao.strftime(
                        '%d/%m/%Y %H:%M'
                    )
                    if item.data_criacao
                    else ''
                ),
            })

        # ----------------------------------------------------
        # PROCESSADOS
        #
        # Estes registros já são históricos. Não usamos
        # estoque_destino para substituir o endereço original.
        # ----------------------------------------------------
        historicos = (
            empenho.historico_itens
            .select_related(
                'estoque_origem',
                'estoque_origem__conferente',
            )
            .order_by(
                'processado_em',
                'id',
            )
        )

        for historico in historicos:
            origem_legada = (
                historico.estoque_origem
            )

            peso_unitario = Decimal(
                str(
                    historico.peso_unitario
                    or 0
                )
            )

            if (
                peso_unitario <= 0
                and origem_legada
            ):
                peso_unitario = Decimal(
                    str(
                        origem_legada.peso_unitario
                        or 0
                    )
                )

            saldo_no_empenho = Decimal(
                str(
                    historico.saldo_anterior
                    or 0
                )
            )

            armazem_empenho = (
                historico.az_origem
                or (
                    origem_legada.az
                    if origem_legada
                    else ''
                )
            )

            observacao_origem = (
                historico.observacao_origem
                or ''
            )

            quantidade = Decimal(
                str(
                    historico.quantidade
                    or 0
                )
            )

            itens_processados.append({
                'item_id': historico.id,

                'lote': (
                    historico.lote
                    or ''
                ),

                'quantidade': float(
                    quantidade
                ),

                'saldo_atual': float(
                    saldo_no_empenho
                ),

                'saldo_empenho': float(
                    saldo_no_empenho
                ),

                # ORIGEM DO EMPENHO.
                'endereco': (
                    historico.endereco_origem
                    or ''
                ),

                'az': armazem_empenho,
                'armazem': armazem_empenho,

                'produto': (
                    historico.produto
                    or ''
                ),

                'cultivar': (
                    historico.cultivar
                    or ''
                ),

                'peneira': (
                    historico.peneira
                    or ''
                ),

                'categoria': (
                    historico.categoria
                    or ''
                ),

                'especie': (
                    historico.especie
                    or ''
                ),

                'tratamento': (
                    historico.tratamento
                    or ''
                ),

                'embalagem': (
                    historico.embalagem
                    or ''
                ),

                'empresa': (
                    historico.empresa
                    or ''
                ),

                'cliente': (
                    historico.cliente
                    or ''
                ),

                'peso_unitario': str(
                    peso_unitario
                ),

                'peso_total': str(
                    quantidade
                    * peso_unitario
                ),

                # Observação original do lote no empenho.
                'observacao': (
                    observacao_origem
                ),

                # Observação digitada na operação.
                'observacao_movimentacao': (
                    historico.observacao
                    or ''
                ),

                'conferente': (
                    historico.conferente
                    or ''
                ),

                'tipo': (
                    historico.get_tipo_display()
                ),

                'processado_em': (
                    historico.processado_em.strftime(
                        '%d/%m/%Y %H:%M'
                    )
                    if historico.processado_em
                    else ''
                ),

                'situacao': (
                    'transferido'
                    if (
                        historico.tipo
                        == 'transferencia'
                    )
                    else 'expedido'
                ),

                # Destino continua disponível separadamente
                # para auditoria, mas NÃO substitui endereço.
                'endereco_destino': (
                    historico.endereco_destino
                    if (
                        historico.tipo
                        == 'transferencia'
                    )
                    else ''
                ),
            })

    return JsonResponse({
        'success': True,

        'empenho_id': (
            empenho.id
            if empenho
            else None
        ),

        'solicitacao': {
            'id': solicitacao.id,
            'titulo': solicitacao.titulo,

            'criador': (
                solicitacao.criador.get_full_name()
                or solicitacao.criador.username
            ),

            'data_criacao': (
                solicitacao.data_criacao.strftime(
                    '%d/%m/%Y %H:%M'
                )
            ),

            'data_atualizacao': (
                solicitacao.data_atualizacao.strftime(
                    '%d/%m/%Y %H:%M'
                )
                if solicitacao.data_atualizacao
                else ''
            ),

            'data_finalizacao': (
                solicitacao.data_finalizacao.strftime(
                    '%d/%m/%Y %H:%M'
                )
                if getattr(
                    solicitacao,
                    'data_finalizacao',
                    None
                )
                else ''
            ),

            'quantidade_solicitada': float(
                solicitacao.quantidade_solicitada
            ),

            'quantidade_empenhada': float(
                solicitacao.quantidade_empenhada
            ),

            'quantidade_movimentada': float(
                solicitacao.quantidade_movimentada
            ),

            'unidade_controle': (
                solicitacao.unidade_controle
            ),

            'status': solicitacao.status,

            'destino': (
                getattr(
                    solicitacao,
                    'destino',
                    ''
                )
                or ''
            ),

            'observacao': (
                solicitacao.observacao
                or ''
            ),

            'criterios': {
                'armazem': (
                    solicitacao.armazem.nome
                    if solicitacao.armazem
                    else ''
                ),

                'produto': (
                    solicitacao.produto
                    or ''
                ),

                'especie': (
                    solicitacao.especie.nome
                    if solicitacao.especie
                    else ''
                ),

                'cliente': (
                    solicitacao.cliente
                    or ''
                ),
            },
        },

        'itens_pendentes': (
            itens_pendentes
        ),

        'itens_processados': (
            itens_processados
        ),
    })

# ============================================================================
# ============================================================================


# BLOCO COMPLETO PARA sapp/views.py
# Remova/substitua as versões antigas das funções com o mesmo nome.

import json
from decimal import Decimal

from django.contrib.auth.decorators import login_required, permission_required
from django.db import transaction
from django.db.models import Q, Prefetch
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, render
from django.utils import timezone
from django.views.decorators.http import require_GET, require_POST

from .models import (
    ColunaKanban,
    Empenho,
    HistoricoCard,
    HistoricoItemEmpenho,
    ItemEmpenho,
    Solicitacao,
    TagKanban,
)


def _formatar_data(valor):
    if not valor:
        return ''
    return timezone.localtime(valor).strftime('%d/%m/%Y %H:%M')


def _nome_usuario(usuario):
    if not usuario:
        return ''
    return usuario.get_full_name() or usuario.username


def _lotes_da_solicitacao(solicitacao):
    """
    Retorna lotes pendentes e já processados usando o vínculo
    Empenho.solicitacao. Não depende do título do card.
    """
    lotes = set()

    for empenho in solicitacao.empenhos.all():
        for item in empenho.itens.all():
            lote = str(item.lote or '').strip()
            if lote:
                lotes.add(lote)

        for item in empenho.historico_itens.all():
            lote = str(item.lote or '').strip()
            if lote:
                lotes.add(lote)

    return sorted(lotes)


def _serializar_tag(tag):
    return {
        'id': tag.id,
        'nome': tag.nome,
        'icone': tag.icone,
        'cor': tag.cor,
        'cor_texto': tag.cor_texto,
    }


def _serializar_card(solicitacao):
    if solicitacao.unidade_controle == 'QUILOGRAMA':
        qtd_empenhada_display = (
            solicitacao.quantidade_empenhada_kg
            or Decimal('0')
        )
    else:
        qtd_empenhada_display = Decimal(
            str(solicitacao.quantidade_empenhada or 0)
        )

    quantidade_solicitada = Decimal(
        str(solicitacao.quantidade_solicitada or 0)
    )

    percentual = (
        (qtd_empenhada_display / quantidade_solicitada) * 100
        if quantidade_solicitada > 0
        else Decimal('0')
    )

    embalagens = set()

    for empenho in solicitacao.empenhos.all():

        for item in empenho.itens.all():
            embalagem = str(
                getattr(
                    item,
                    'embalagem_snapshot',
                    ''
                )
                or ''
            ).strip().upper()

            if embalagem:
                embalagens.add(embalagem)

        for historico in empenho.historico_itens.all():
            embalagem = str(
                historico.embalagem
                or ''
            ).strip().upper()

            if embalagem:
                embalagens.add(embalagem)
                
    return {
        'id': solicitacao.id,
        'titulo': solicitacao.titulo,
        'observacao': solicitacao.observacao or '',
        'destino': solicitacao.destino or '',
        'criador_nome': _nome_usuario(solicitacao.criador),
        'responsavel_nome': _nome_usuario(solicitacao.responsavel),
        'data_criacao': _formatar_data(solicitacao.data_criacao),
        'data_finalizacao': (
            _formatar_data(solicitacao.data_finalizacao)
            if solicitacao.status == 'CONCLUIDO'
            else ''
        ),
        'status': solicitacao.status,
        'status_display': (
            solicitacao.get_status_display()
            if hasattr(solicitacao, 'get_status_display')
            else solicitacao.status
        ),
        'unidade_controle': solicitacao.unidade_controle,
        'quantidade_solicitada': float(quantidade_solicitada),
        'quantidade_empenhada': float(
            solicitacao.quantidade_empenhada or 0
        ),
        'quantidade_empenhada_display': float(
            qtd_empenhada_display
        ),
        'quantidade_movimentada': float(
            solicitacao.quantidade_movimentada or 0
        ),
        'percentual_empenhado': float(percentual),
        'percentual_movimentado': float(
            solicitacao.percentual_movimentado
        ),
        'prioridade': solicitacao.prioridade,
        'coluna_id': solicitacao.coluna_kanban_id,
        'coluna_nome': (
            solicitacao.coluna_kanban.nome
            if solicitacao.coluna_kanban
            else ''
        ),
        'criterios': {
            'armazem': (
                solicitacao.armazem.nome
                if solicitacao.armazem else ''
            ),
            'produto': solicitacao.produto or '',
            'especie': (
                solicitacao.especie.nome
                if solicitacao.especie else ''
            ),
            'cliente': solicitacao.cliente or '',
            'destino': solicitacao.destino or '',
        },
        'lotes': _lotes_da_solicitacao(solicitacao),


        'embalagens': sorted(
            embalagens
        ),

        'tags': [
            _serializar_tag(tag)
            for tag in solicitacao.tags_kanban.all()
        ],
    }


def _queryset_kanban():
    return (
        Solicitacao.objects
        .select_related(
            'criador',
            'responsavel',
            'armazem',
            'especie',
            'coluna_kanban',
        )
        .prefetch_related(
            'tags_kanban',
            Prefetch(
                'empenhos',
                queryset=(
                    Empenho.objects
                    .prefetch_related('itens', 'historico_itens')
                    .order_by('id')
                ),
            ),
        )
    )


@login_required
def pagina_kanban(request):
    return render(
        request,
        'sapp/kanban.html',
        {
            'tags_kanban': TagKanban.objects.filter(
                ativa=True
            ).order_by('ordem', 'nome'),
        },
    )


@login_required
@require_GET
def api_kanban_dados(request):
    colunas = list(
        ColunaKanban.objects
        .filter(ativa=True)
        .order_by('ordem', 'nome')
    )

    solicitacoes = list(
        _queryset_kanban()
        .filter(coluna_kanban__in=colunas)
        .order_by('-prioridade', '-data_criacao')
    )

    por_coluna = {coluna.id: [] for coluna in colunas}

    for solicitacao in solicitacoes:
        if solicitacao.coluna_kanban_id in por_coluna:
            por_coluna[solicitacao.coluna_kanban_id].append(
                _serializar_card(solicitacao)
            )

    return JsonResponse({
        'success': True,
        'colunas': [
            {
                'id': coluna.id,
                'nome': coluna.nome,
                'cor': coluna.cor,
                'total': len(por_coluna[coluna.id]),
                'cards': por_coluna[coluna.id],
            }
            for coluna in colunas
        ],
        'colunas_disponiveis': [
            {
                'id': coluna.id,
                'nome': coluna.nome,
                'cor': coluna.cor,
            }
            for coluna in colunas
        ],
        'tags_disponiveis': [
            _serializar_tag(tag)
            for tag in TagKanban.objects.filter(
                ativa=True
            ).order_by('ordem', 'nome')
        ],
        'timestamp': timezone.now().isoformat(),
    })


@login_required
@require_GET
def api_pesquisar_kanban(request):
    termo = request.GET.get('q', '').strip()

    if len(termo) < 2:
        return JsonResponse({
            'success': True,
            'resultados': [],
        })

    queryset = (
        _queryset_kanban()
        .filter(
            Q(titulo__icontains=termo)
            | Q(observacao__icontains=termo)
            | Q(destino__icontains=termo)
            | Q(cliente__icontains=termo)
            | Q(produto__icontains=termo)
            | Q(empenhos__itens__lote__icontains=termo)
            | Q(
                empenhos__historico_itens__lote__icontains=termo
            )
        )
        .distinct()
        .order_by('-data_criacao')[:50]
    )

    return JsonResponse({
        'success': True,
        'resultados': [
            {
                'id': solicitacao.id,
                'titulo': solicitacao.titulo,
                'status': solicitacao.status,
                'coluna_id': solicitacao.coluna_kanban_id,
                'coluna_nome': (
                    solicitacao.coluna_kanban.nome
                    if solicitacao.coluna_kanban else ''
                ),
                'lotes': _lotes_da_solicitacao(solicitacao),
                'destino': solicitacao.destino or '',
                'data_criacao': _formatar_data(
                    solicitacao.data_criacao
                ),
            }
            for solicitacao in queryset
        ],
    })


# views.py (ou api_views.py)
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_protect
import json
from .models import Solicitacao, ColunaKanban, HistoricoCard

@login_required
@require_POST
@csrf_protect
def mover_card_kanban(request, solicitacao_id):
    try:
        data = json.loads(request.body)
        coluna_id = data.get('coluna_id')
        observacao = data.get('observacao', '')

        if not coluna_id:
            return JsonResponse({'success': False, 'error': 'ID da coluna não informado.'}, status=400)

        solicitacao = get_object_or_404(Solicitacao, pk=solicitacao_id)
        coluna_destino = get_object_or_404(ColunaKanban, pk=coluna_id, ativa=True)

        coluna_anterior = solicitacao.coluna_kanban.nome if solicitacao.coluna_kanban else 'Nenhuma'

        # Atualiza a coluna
        solicitacao.coluna_kanban = coluna_destino
        solicitacao.save(update_fields=['coluna_kanban', 'data_atualizacao'])

        # Registra no histórico
        HistoricoCard.objects.create(
            solicitacao=solicitacao,
            usuario=request.user,
            acao='MOVIMENTACAO_KANBAN',
            coluna_anterior=coluna_anterior,
            coluna_nova=coluna_destino.nome,
            observacao=observacao,
        )

        return JsonResponse({
            'success': True,
            'coluna_nova': coluna_destino.nome,
        })

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'JSON inválido.'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@require_POST
def api_atualizar_tags_card(request, solicitacao_id):
    try:
        dados = json.loads(request.body or '{}')
    except json.JSONDecodeError:
        return JsonResponse(
            {'success': False, 'error': 'JSON inválido.'},
            status=400,
        )

    ids = dados.get('tags_ids', [])

    if not isinstance(ids, list):
        return JsonResponse(
            {'success': False, 'error': 'Lista de tags inválida.'},
            status=400,
        )

    ids_validos = [
        int(valor)
        for valor in ids
        if str(valor).isdigit()
    ]

    solicitacao = get_object_or_404(
        Solicitacao,
        id=solicitacao_id,
    )

    tags = TagKanban.objects.filter(
        id__in=ids_validos,
        ativa=True,
    )

    solicitacao.tags_kanban.set(tags)

    return JsonResponse({
        'success': True,
        'tags': [
            _serializar_tag(tag)
            for tag in solicitacao.tags_kanban.all()
        ],
    })


@login_required
@require_POST
def api_criar_tag_kanban(request):
    try:
        dados = json.loads(request.body or '{}')
    except json.JSONDecodeError:
        return JsonResponse(
            {'success': False, 'error': 'JSON inválido.'},
            status=400,
        )

    nome = str(dados.get('nome', '')).strip()
    icone = str(dados.get('icone', '')).strip()
    cor = str(dados.get('cor', '#2f8f4e')).strip()

    if not nome:
        return JsonResponse(
            {'success': False, 'error': 'Informe o texto da tag.'},
            status=400,
        )

    tag, criada = TagKanban.objects.get_or_create(
        nome=nome,
        defaults={
            'icone': icone,
            'cor': cor,
            'cor_texto': '#ffffff',
            'ativa': True,
        },
    )

    if not criada:
        tag.icone = icone
        tag.cor = cor
        tag.ativa = True
        tag.save(
            update_fields=['icone', 'cor', 'ativa']
        )

    return JsonResponse({
        'success': True,
        'tag': _serializar_tag(tag),
    })


# ============================================================================
# EXCLUIR SOLICITAÇÃO
# ============================================================================
@login_required
def api_excluir_solicitacao(
    request,
    solicitacao_id
):
    """
    Exclui somente cards sem movimentação e sem histórico
    processado.

    Cards concluídos permanecem preservados para auditoria.
    """
    if request.method != 'POST':
        return JsonResponse(
            {
                'success': False,
                'error': 'Método não permitido.'
            },
            status=405
        )

    try:
        with transaction.atomic():
            solicitacao = (
                Solicitacao.objects
                .select_for_update()
                .get(id=solicitacao_id)
            )

            empenhos = list(
                Empenho.objects
                .select_for_update()
                .filter(
                    solicitacao_id=solicitacao.id
                )
            )

            possui_historico_processado = any(
                empenho.historico_itens.exists()
                for empenho in empenhos
            )

            if solicitacao.status in [
                'CONCLUIDO',
                'CANCELADO',
            ]:
                raise ValueError(
                    'Cards concluídos ou cancelados não '
                    'podem ser excluídos. Eles devem '
                    'permanecer no histórico.'
                )

            if solicitacao.quantidade_movimentada > 0:
                raise ValueError(
                    'Este card possui movimentações e não '
                    'pode ser excluído.'
                )

            if possui_historico_processado:
                raise ValueError(
                    'Este card possui itens transferidos ou '
                    'expedidos e não pode ser excluído.'
                )

            titulo = solicitacao.titulo

            # Importante:
            # removemos um item por vez para executar
            # ItemEmpenho.delete() e liberar Estoque.empenhado.
            for empenho in empenhos:
                for item in list(empenho.itens.all()):
                    item.delete()

                empenho.delete()

            HistoricoCard.objects.filter(
                solicitacao=solicitacao
            ).delete()

            solicitacao.delete()

            from django.core.cache import cache
            cache.delete('cards_version_hash')

            return JsonResponse({
                'success': True,
                'message': (
                    f'Card "{titulo}" excluído com sucesso!'
                )
            })

    except Solicitacao.DoesNotExist:
        return JsonResponse(
            {
                'success': False,
                'error': 'Card não encontrado.'
            },
            status=404
        )

    except ValueError as erro:
        return JsonResponse(
            {
                'success': False,
                'error': str(erro)
            },
            status=400
        )

    except Exception as erro:
        logger.exception(
            'Erro ao excluir solicitação %s',
            solicitacao_id
        )

        return JsonResponse(
            {
                'success': False,
                'error': str(erro)
            },
            status=500
        )

# ============================================================================
# CONFIGURAÇÃO DO WORKFLOW
# ============================================================================
@login_required
def api_config_workflow(request):
    """GET: Retorna configurações. POST: Salva configurações."""
    if request.method == 'GET':
        colunas = ColunaKanban.objects.filter(ativa=True).order_by('ordem')
        regras = RegraWorkflow.objects.select_related('coluna').all()
        
        return JsonResponse({
            'success': True,
            'workflow': {
                'colunas': [{'id': c.id, 'nome': c.nome, 'cor': c.cor, 'ordem': c.ordem, 'ativa': c.ativa} for c in colunas],
                'regras': [{'id': r.id, 'coluna_id': r.coluna.id, 'coluna_nome': r.coluna.nome, 'evento': r.evento, 'evento_display': r.get_evento_display(), 'status_resultante': r.status_resultante, 'movimentacao_automatica': r.movimentacao_automatica} for r in regras],
                'eventos_disponiveis': [{'value': e[0], 'label': e[1]} for e in RegraWorkflow.EVENTO_CHOICES],
                'status_disponiveis': ['AGUARDANDO_EMPENHO', 'EMPENHO_PARCIAL', 'EMPENHO_COMPLETO', 'MOVIMENTACAO_PARCIAL', 'CONCLUIDO', 'CANCELADO'],
            }
        })
    
    elif request.method == 'POST':
        try:
            payload = json.loads(request.body)
            
            # Processar colunas
            if 'colunas' in payload:
                for col_data in payload['colunas']:
                    col_id = col_data.get('id')
                    
                    if col_id and int(col_id) > 0 and ColunaKanban.objects.filter(id=col_id).exists():
                        coluna = ColunaKanban.objects.get(id=col_id)
                        if 'nome' in col_data: coluna.nome = col_data['nome']
                        if 'cor' in col_data: coluna.cor = col_data['cor']
                        if 'ordem' in col_data: coluna.ordem = col_data['ordem']
                        if 'ativa' in col_data: coluna.ativa = col_data['ativa']
                        coluna.save()
                    else:
                        nome = col_data.get('nome', 'Nova Coluna')
                        if not ColunaKanban.objects.filter(nome=nome).exists():
                            ColunaKanban.objects.create(
                                nome=nome, cor=col_data.get('cor', '#6c757d'),
                                ordem=col_data.get('ordem', ColunaKanban.objects.count() + 1),
                                ativa=col_data.get('ativa', True)
                            )
            
            # Excluir colunas
            if 'colunas_excluir' in payload:
                for col_id in payload['colunas_excluir']:
                    if int(col_id) > 0:
                        ColunaKanban.objects.filter(id=col_id).delete()
            
            # Processar regras
            if 'regras' in payload:
                regra_ids_enviados = [int(r.get('id')) for r in payload['regras'] if r.get('id') and int(r.get('id')) > 0]
                if regra_ids_enviados:
                    RegraWorkflow.objects.exclude(id__in=regra_ids_enviados).delete()
                else:
                    RegraWorkflow.objects.all().delete()
                
                for regra_data in payload['regras']:
                    coluna_id = regra_data.get('coluna_id')
                    if not coluna_id: continue
                    
                    try:
                        coluna = ColunaKanban.objects.get(id=int(coluna_id))
                    except (ValueError, ColunaKanban.DoesNotExist):
                        continue
                    
                    evento = regra_data.get('evento', 'CRIACAO')
                    status_resultante = regra_data.get('status_resultante', 'AGUARDANDO_EMPENHO')
                    auto = regra_data.get('movimentacao_automatica', True)
                    if isinstance(auto, str): auto = auto.lower() in ['true', '1', 'yes', 'on']
                    
                    regra_id = regra_data.get('id')
                    if regra_id and int(regra_id) > 0:
                        try:
                            regra = RegraWorkflow.objects.get(id=int(regra_id))
                            regra.coluna = coluna
                            regra.evento = evento
                            regra.status_resultante = status_resultante
                            regra.movimentacao_automatica = auto
                            regra.save()
                        except RegraWorkflow.DoesNotExist:
                            RegraWorkflow.objects.get_or_create(
                                coluna=coluna, evento=evento,
                                defaults={'status_resultante': status_resultante, 'movimentacao_automatica': auto}
                            )
                    else:
                        RegraWorkflow.objects.get_or_create(
                            coluna=coluna, evento=evento,
                            defaults={'status_resultante': status_resultante, 'movimentacao_automatica': auto}
                        )
            
            return JsonResponse({'success': True, 'message': 'Workflow salvo com sucesso!'})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})


# ============================================================================
# FEED DE ATUALIZAÇÕES
# ============================================================================
@login_required
def api_atualizacoes_recentes(request):
    """Retorna o feed de atualizações recentes"""
    limite = int(request.GET.get('limite', 20))
    desde = request.GET.get('desde')
    
    query = HistoricoCard.objects.select_related('usuario', 'solicitacao')
    
    if desde:
        query = query.filter(id__gt=int(desde))
    
    atualizacoes = query.order_by('-data')[:limite]
    
    data = []
    for a in atualizacoes:
        data.append({
            'id': a.id,
            'usuario': a.usuario.get_full_name() or a.usuario.username,
            'acao': a.get_acao_display(),
            'descricao': a.descricao_completa(),
            'lote': a.lote,
            'quantidade': float(a.quantidade) if a.quantidade else None,
            'unidade': a.unidade,
            'card_id': a.solicitacao.id,
            'card_titulo': a.solicitacao.titulo,
            'data': a.data.strftime('%d/%m/%Y %H:%M:%S'),
            'data_iso': a.data.isoformat(),
        })
    
    return JsonResponse({'success': True, 'atualizacoes': data, 'total': len(data)})