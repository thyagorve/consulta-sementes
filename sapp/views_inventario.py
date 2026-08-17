from decimal import Decimal, InvalidOperation
from io import BytesIO
import re
import unicodedata

from django.contrib.auth.decorators import login_required, permission_required
from django.http import HttpResponse
from django.shortcuts import render
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import Estoque


CAMPOS_MODELO = [
    'Código', 'Lote', 'Endereço', 'AZ', 'Quantidade Física',
    'Peneira', 'Categoria', 'Cultivar', 'Espécie', 'Tratamento',
    'Embalagem', 'Cliente', 'Empresa', 'Peso Unitário (kg)',
]

ALIASES = {
    'codigo': ['codigo', 'código', 'cod', 'produto', 'sku'],
    'lote': ['lote'],
    'endereco': ['endereco', 'endereço', 'localizacao', 'localização', 'local'],
    'az': ['az', 'armazem', 'armazém'],
    'quantidade': ['quantidade fisica', 'quantidade física', 'quantidade', 'qtd', 'saldo', 'estoque'],
    'peneira': ['peneira'],
    'categoria': ['categoria'],
    'cultivar': ['cultivar', 'variedade'],
    'especie': ['especie', 'espécie'],
    'tratamento': ['tratamento'],
    'embalagem': ['embalagem', 'unidade', 'un'],
    'cliente': ['cliente'],
    'empresa': ['empresa'],
    'peso_unitario': ['peso unitario kg', 'peso unitário kg', 'peso unitario', 'peso unitário', 'peso'],
}


def _norm(value):
    text = str(value or '').strip()
    text = ''.join(
        c for c in unicodedata.normalize('NFKD', text)
        if not unicodedata.combining(c)
    )
    return re.sub(r'[^a-z0-9]+', '', text.lower())


def _txt(value):
    return str(value or '').strip()


def _decimal(value):
    if value in (None, ''):
        return Decimal('0')
    text = str(value).strip().replace(' ', '')
    if ',' in text and '.' in text:
        text = text.replace('.', '').replace(',', '.')
    else:
        text = text.replace(',', '.')
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError, TypeError):
        return Decimal('0')


def _map_header(row):
    normalized = [_norm(v) for v in row]
    result = {}
    for field, aliases in ALIASES.items():
        candidates = {_norm(a) for a in aliases}
        for idx, col in enumerate(normalized):
            if col in candidates:
                result[field] = idx
                break
    return result


def _cell(row, mapping, field):
    idx = mapping.get(field)
    if idx is None or idx >= len(row):
        return ''
    return row[idx]


def _stock_dict(item):
    return {
        'codigo': _txt(item.produto),
        'lote': _txt(item.lote),
        'endereco': _txt(item.endereco),
        'az': _txt(item.az),
        'quantidade': Decimal(str(item.saldo or 0)),
        'empenhado': Decimal(str(item.empenhado or 0)),
        'disponivel': Decimal(str(item.disponivel or 0)),
        'peneira': _txt(item.peneira.nome if item.peneira else ''),
        'categoria': _txt(item.categoria.nome if item.categoria else ''),
        'cultivar': _txt(item.cultivar.nome if item.cultivar else ''),
        'especie': _txt(item.especie.nome if item.especie else ''),
        'tratamento': _txt(item.tratamento.nome if item.tratamento else ''),
        'embalagem': _txt(item.embalagem),
        'cliente': _txt(item.cliente),
        'empresa': _txt(item.empresa),
        'peso_unitario': Decimal(str(item.peso_unitario or 0)),
        'id': item.id,
    }


def _compare_text(label, excel_value, system_value, diffs):
    excel_text = _txt(excel_value)
    if not excel_text:
        return
    if _norm(excel_text) != _norm(system_value):
        diffs.append({
            'campo': label,
            'excel': excel_text or '—',
            'sistema': _txt(system_value) or '—',
        })


@login_required
@permission_required('sapp.pode_ver_estoque', raise_exception=True)
def inventario_estoque(request):
    context = {
        'resultado': None,
        'resumo': None,
        'nome_arquivo': '',
    }

    if request.method == 'POST':
        arquivo = request.FILES.get('arquivo')
        if not arquivo:
            context['erro'] = 'Selecione um arquivo Excel (.xlsx).'
            return render(request, 'sapp/inventario_estoque.html', context)

        if not arquivo.name.lower().endswith('.xlsx'):
            context['erro'] = 'Formato inválido. Use o modelo .xlsx disponibilizado pelo sistema.'
            return render(request, 'sapp/inventario_estoque.html', context)

        try:
            wb = load_workbook(arquivo, data_only=True, read_only=True)
            ws = wb['Inventario'] if 'Inventario' in wb.sheetnames else wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                raise ValueError('A planilha está vazia.')

            mapping = _map_header(rows[0])
            required = ['codigo', 'lote', 'endereco', 'quantidade']
            missing = [f for f in required if f not in mapping]
            if missing:
                raise ValueError(
                    'Colunas obrigatórias ausentes: ' + ', '.join(missing) + '. Baixe o modelo do sistema.'
                )

            estoque_qs = (
                Estoque.objects
                .select_related('peneira', 'categoria', 'cultivar', 'especie', 'tratamento')
                .all()
            )
            stock = [_stock_dict(item) for item in estoque_qs]

            exact_index = {}
            lot_index = {}
            for item in stock:
                exact_index.setdefault(
                    (_norm(item['codigo']), _norm(item['lote']), _norm(item['endereco'])), []
                ).append(item)
                lot_index.setdefault(
                    (_norm(item['codigo']), _norm(item['lote'])), []
                ).append(item)

            results = []
            counted_ids = set()

            for excel_row_number, row in enumerate(rows[1:], start=2):
                codigo = _txt(_cell(row, mapping, 'codigo'))
                lote = _txt(_cell(row, mapping, 'lote'))
                endereco = _txt(_cell(row, mapping, 'endereco'))
                qtd_fisica = _decimal(_cell(row, mapping, 'quantidade'))

                if not any([codigo, lote, endereco]) and qtd_fisica == 0:
                    continue

                key_exact = (_norm(codigo), _norm(lote), _norm(endereco))
                candidates = exact_index.get(key_exact, [])
                address_mismatch = False

                if candidates:
                    system_item = candidates[0]
                else:
                    by_lot = lot_index.get((_norm(codigo), _norm(lote)), [])
                    system_item = by_lot[0] if by_lot else None
                    address_mismatch = bool(system_item)

                diffs = []
                if system_item:
                    counted_ids.add(system_item['id'])
                    if qtd_fisica != system_item['quantidade']:
                        diffs.append({
                            'campo': 'Quantidade / Saldo',
                            'excel': str(qtd_fisica),
                            'sistema': str(system_item['quantidade']),
                        })

                    if address_mismatch or _norm(endereco) != _norm(system_item['endereco']):
                        diffs.append({
                            'campo': 'Endereço',
                            'excel': endereco or '—',
                            'sistema': system_item['endereco'] or '—',
                        })

                    field_pairs = [
                        ('AZ', 'az'), ('Peneira', 'peneira'), ('Categoria', 'categoria'),
                        ('Cultivar', 'cultivar'), ('Espécie', 'especie'), ('Tratamento', 'tratamento'),
                        ('Embalagem', 'embalagem'), ('Cliente', 'cliente'), ('Empresa', 'empresa'),
                    ]
                    for label, field in field_pairs:
                        _compare_text(label, _cell(row, mapping, field), system_item[field], diffs)

                    excel_weight = _cell(row, mapping, 'peso_unitario')
                    if _txt(excel_weight):
                        excel_weight = _decimal(excel_weight)
                        if excel_weight != system_item['peso_unitario']:
                            diffs.append({
                                'campo': 'Peso Unitário',
                                'excel': str(excel_weight),
                                'sistema': str(system_item['peso_unitario']),
                            })

                    status = 'OK' if not diffs else 'DIVERGENTE'
                else:
                    status = 'NAO_ENCONTRADO'
                    diffs.append({
                        'campo': 'Cadastro',
                        'excel': f'{codigo} / {lote} / {endereco}',
                        'sistema': 'Não encontrado',
                    })

                results.append({
                    'linha': excel_row_number,
                    'status': status,
                    'codigo': codigo,
                    'lote': lote,
                    'endereco_excel': endereco,
                    'quantidade_excel': qtd_fisica,
                    'sistema': system_item,
                    'diferencas': diffs,
                })

            # Itens do sistema com saldo físico > 0 que não apareceram na contagem.
            for item in stock:
                if item['quantidade'] > 0 and item['id'] not in counted_ids:
                    results.append({
                        'linha': None,
                        'status': 'NAO_CONTADO',
                        'codigo': item['codigo'],
                        'lote': item['lote'],
                        'endereco_excel': '—',
                        'quantidade_excel': None,
                        'sistema': item,
                        'diferencas': [{
                            'campo': 'Contagem física',
                            'excel': 'Não informado',
                            'sistema': f"Saldo {item['quantidade']} em {item['endereco']}",
                        }],
                    })

            context['resultado'] = results
            context['nome_arquivo'] = arquivo.name
            context['resumo'] = {
                'total': len(results),
                'ok': sum(1 for r in results if r['status'] == 'OK'),
                'divergentes': sum(1 for r in results if r['status'] == 'DIVERGENTE'),
                'nao_encontrados': sum(1 for r in results if r['status'] == 'NAO_ENCONTRADO'),
                'nao_contados': sum(1 for r in results if r['status'] == 'NAO_CONTADO'),
            }

        except Exception as exc:
            context['erro'] = f'Não foi possível comparar o arquivo: {exc}'

    return render(request, 'sapp/inventario_estoque.html', context)


@login_required
@permission_required('sapp.pode_ver_estoque', raise_exception=True)
def baixar_modelo_inventario(request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Inventario'

    green = '1E5F34'
    light_green = 'EAF7EE'
    border_color = 'D6DEE3'
    thin = Side(style='thin', color=border_color)

    for col, header in enumerate(CAMPOS_MODELO, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = PatternFill('solid', fgColor=green)
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Exemplos claros, removíveis pelo usuário.
    examples = [
        ['4702000002', '12345', 'R-C LN01 P01', 'AZ 01', 50, '6.5', 'S1', 'EXEMPLO', 'SOJA', 'PADRÃO', 'BAG', '', '', 1000],
        ['4702000002', '12345', 'R-C LN01 P02', 'AZ 01', 25, '6.5', 'S1', 'EXEMPLO', 'SOJA', 'PADRÃO', 'BAG', '', '', 1000],
    ]
    for r_idx, data in enumerate(examples, start=2):
        for c_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.fill = PatternFill('solid', fgColor='F7FBF8')
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    widths = [16, 16, 22, 12, 18, 13, 14, 22, 16, 18, 13, 20, 20, 20]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(CAMPOS_MODELO))}3'

    info = wb.create_sheet('Instrucoes')
    info.merge_cells('A1:D1')
    info['A1'] = 'INVENTÁRIO DE ESTOQUE - COMO PREENCHER'
    info['A1'].fill = PatternFill('solid', fgColor=green)
    info['A1'].font = Font(color='FFFFFF', bold=True, size=14)
    info['A1'].alignment = Alignment(horizontal='center')

    instructions = [
        ['Campo', 'Obrigatório', 'Comparação', 'Observação'],
        ['Código', 'Sim', 'Código/produto', 'Use exatamente o código cadastrado no sistema.'],
        ['Lote', 'Sim', 'Lote', 'É a principal referência operacional do empenho.'],
        ['Endereço', 'Sim', 'Endereço atual', 'Divergências de localização são destacadas.'],
        ['Quantidade Física', 'Sim', 'Saldo físico do sistema', 'Informe o que foi contado fisicamente.'],
        ['Peneira, Categoria, Cultivar', 'Não', 'Cadastro', 'Se preenchidos, serão conferidos campo a campo.'],
        ['Espécie, Tratamento, Embalagem', 'Não', 'Cadastro', 'Pode deixar vazio para não comparar o campo.'],
        ['Cliente, Empresa, AZ, Peso', 'Não', 'Cadastro', 'Também entram na auditoria quando preenchidos.'],
    ]
    for r_idx, row in enumerate(instructions, start=3):
        for c_idx, value in enumerate(row, start=1):
            cell = info.cell(row=r_idx, column=c_idx, value=value)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if r_idx == 3:
                cell.fill = PatternFill('solid', fgColor=light_green)
                cell.font = Font(bold=True, color=green)

    for col, width in zip('ABCD', [28, 16, 30, 55]):
        info.column_dimensions[col].width = width

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    response = HttpResponse(
        stream.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="modelo_inventario_estoque.xlsx"'
    return response
